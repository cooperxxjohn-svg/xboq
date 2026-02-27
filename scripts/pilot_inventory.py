#!/usr/bin/env python3
"""
Pilot Inventory — lightweight metadata-only scan of tender folders.

NO OCR, NO rendering, NO pipeline execution.
Only reads file metadata, PDF page counts, and Excel sheet headers.

For each tender folder produces:
  - manifest.json  (machine-readable file inventory)
  - summary.md     (human-readable findings)

Plus a master rollup:
  - inventory_rollup.md   (one-row-per-tender table)
  - inventory_rollup.csv  (same, machine-readable)

Usage:
    python3 scripts/pilot_inventory.py \\
        --input /path/to/parent_dir \\
        --output /path/to/output

    python3 scripts/pilot_inventory.py \\
        --input /path/to/parent_dir \\
        --output /path/to/output \\
        --include "Tender Documents" "Tender Documents 3"
"""

import argparse
import csv
import json
import logging
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

# Repo root
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from src.pilot.file_router import (
    classify_directory,
    classify_file,
    RoutingSummary,
    ClassifiedFile,
    CATEGORY_DRAWINGS_PDF,
    CATEGORY_BOQ_PDF,
    CATEGORY_CONDITIONS_PDF,
    CATEGORY_ADDENDA_PDF,
    CATEGORY_BOQ_XLSX,
    CATEGORY_CONDITIONS_DOC,
    CATEGORY_UNKNOWN,
    ALL_CATEGORIES,
    _SKIP_PATTERNS,
)

logger = logging.getLogger("pilot_inventory")


# ═════════════════════════════════════════════════════════════════════════
# LIGHTWEIGHT METADATA EXTRACTORS (no OCR, no rendering)
# ═════════════════════════════════════════════════════════════════════════

def _sanitize_str(s: str) -> str:
    """Remove Unicode surrogates and control chars that break JSON serialization."""
    if not s:
        return ""
    # Encode to utf-8, replacing surrogates, then decode back
    return s.encode("utf-8", errors="replace").decode("utf-8", errors="replace").strip()


def get_pdf_metadata(path: Path) -> Dict[str, Any]:
    """
    Extract lightweight PDF metadata: page count, title, author, producer.
    Uses PyMuPDF (fitz) if available, falls back to pypdf.
    Does NOT render pages or run OCR.
    """
    meta: Dict[str, Any] = {"page_count": 0, "title": "", "author": "", "producer": ""}
    try:
        import fitz
        doc = fitz.open(str(path))
        meta["page_count"] = doc.page_count
        pdf_meta = doc.metadata or {}
        meta["title"] = _sanitize_str(pdf_meta.get("title") or "")
        meta["author"] = _sanitize_str(pdf_meta.get("author") or "")
        meta["producer"] = _sanitize_str(pdf_meta.get("producer") or "")
        doc.close()
        return meta
    except Exception:
        pass

    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        meta["page_count"] = len(reader.pages)
        info = reader.metadata
        if info:
            meta["title"] = _sanitize_str(str(info.get("/Title", "") or ""))
            meta["author"] = _sanitize_str(str(info.get("/Author", "") or ""))
            meta["producer"] = _sanitize_str(str(info.get("/Producer", "") or ""))
        return meta
    except Exception:
        pass

    return meta


def get_excel_preview(path: Path) -> Dict[str, Any]:
    """
    Inspect Excel workbook structure: sheet names + first 3 header rows.
    Does NOT parse all data rows.
    """
    preview: Dict[str, Any] = {
        "sheet_names": [],
        "sheets": {},
        "likely_boq": False,
        "boq_indicators": [],
    }
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        preview["sheet_names"] = wb.sheetnames

        boq_kw = {"BOQ", "SOQ", "PRICE", "BID", "RATE", "ESTIMATE", "BILL",
                   "QTY", "QUANTITY", "AMOUNT", "UNIT", "TOTAL"}

        for sheet_name in wb.sheetnames:
            sheet_info: Dict[str, Any] = {"header_rows": [], "row_count_estimate": 0}
            try:
                ws = wb[sheet_name]
                # Read first 3 rows as header preview
                row_num = 0
                for row in ws.iter_rows(min_row=1, max_row=3, max_col=20, values_only=True):
                    cells = [_sanitize_str(str(c)) if c is not None else "" for c in row]
                    sheet_info["header_rows"].append(cells)
                    row_num += 1

                # Estimate total rows (read_only mode doesn't give max_row reliably)
                # Just count up to 500 to get a rough estimate
                count = 0
                for _ in ws.iter_rows(min_row=1, max_row=500, max_col=1, values_only=True):
                    count += 1
                sheet_info["row_count_estimate"] = count if count < 500 else "500+"

                # Check for BOQ indicators
                all_header_text = " ".join(
                    " ".join(r) for r in sheet_info["header_rows"]
                ).upper()

                matched = [kw for kw in boq_kw if kw in all_header_text]
                if matched:
                    preview["likely_boq"] = True
                    preview["boq_indicators"].extend(
                        f"sheet '{sheet_name}': {kw}" for kw in matched
                    )

                # Also check sheet name itself
                sn_upper = sheet_name.upper()
                sn_matched = [kw for kw in boq_kw if kw in sn_upper]
                if sn_matched:
                    preview["likely_boq"] = True
                    for kw in sn_matched:
                        indicator = f"sheet name '{sheet_name}': {kw}"
                        if indicator not in preview["boq_indicators"]:
                            preview["boq_indicators"].append(indicator)

            except Exception:
                pass

            preview["sheets"][sheet_name] = sheet_info

        wb.close()
    except Exception as e:
        preview["error"] = str(e)

    return preview


# ═════════════════════════════════════════════════════════════════════════
# INVENTORY BUILDER
# ═════════════════════════════════════════════════════════════════════════

def _format_size(size_bytes: int) -> str:
    """Human-readable file size."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


def build_tender_inventory(folder: Path) -> Dict[str, Any]:
    """
    Build a full inventory for one tender folder.

    Returns a dict with:
      - folder_name, folder_path
      - file_type_counts
      - total_size_bytes
      - classification (from file_router)
      - pdf_metadata (page counts, titles)
      - excel_previews (sheet names, headers)
      - analysis (likely_boq_source, conditions_files, etc.)
    """
    t0 = time.perf_counter()
    inventory: Dict[str, Any] = {
        "folder_name": folder.name,
        "folder_path": str(folder),
        "scan_timestamp": datetime.now().isoformat(),
        "files": [],
        "file_type_counts": {},
        "total_files": 0,
        "total_size_bytes": 0,
        "pdf_total_pages": 0,
        "classification_summary": {},
        "pdf_metadata": {},
        "excel_previews": {},
        "analysis": {},
    }

    # Classify via file_router
    routing = classify_directory(folder, zip_name=folder.name)

    # Also scan ALL files (including unsupported extensions for full inventory)
    all_files: List[Dict[str, Any]] = []
    type_counts: Dict[str, int] = {}
    total_size = 0

    for path in sorted(folder.rglob("*")):
        if not path.is_file():
            continue
        if any(skip in path.parts for skip in _SKIP_PATTERNS):
            continue
        if path.name.startswith("."):
            continue

        ext = path.suffix.lower() or "(none)"
        size = path.stat().st_size
        total_size += size
        type_counts[ext] = type_counts.get(ext, 0) + 1

        try:
            rel_path = str(path.relative_to(folder))
        except ValueError:
            rel_path = path.name

        file_entry: Dict[str, Any] = {
            "name": path.name,
            "relative_path": rel_path,
            "extension": ext,
            "size_bytes": size,
            "size_human": _format_size(size),
        }

        # Find classification from routing
        cf_match = None
        for cf in routing.classified:
            if cf.path == path:
                cf_match = cf
                break

        if cf_match:
            file_entry["category"] = cf_match.category
            file_entry["classification_reason"] = cf_match.reason
        else:
            file_entry["category"] = "unscanned"
            file_entry["classification_reason"] = "extension not in scan set"

        # PDF metadata (lightweight — no OCR/rendering)
        if ext == ".pdf":
            pdf_meta = get_pdf_metadata(path)
            file_entry["page_count"] = pdf_meta["page_count"]
            if pdf_meta["title"]:
                file_entry["pdf_title"] = pdf_meta["title"]
            if pdf_meta["author"]:
                file_entry["pdf_author"] = pdf_meta["author"]
            inventory["pdf_total_pages"] += pdf_meta["page_count"]
            inventory["pdf_metadata"][rel_path] = pdf_meta

        # Excel preview (lightweight — headers only)
        if ext in (".xlsx", ".xls"):
            excel_prev = get_excel_preview(path)
            file_entry["sheet_names"] = excel_prev["sheet_names"]
            file_entry["likely_boq_excel"] = excel_prev["likely_boq"]
            if excel_prev["boq_indicators"]:
                file_entry["boq_indicators"] = excel_prev["boq_indicators"]
            inventory["excel_previews"][rel_path] = excel_prev

        all_files.append(file_entry)

    inventory["files"] = all_files
    inventory["total_files"] = len(all_files)
    inventory["total_size_bytes"] = total_size
    inventory["total_size_human"] = _format_size(total_size)
    inventory["file_type_counts"] = type_counts

    # Classification summary
    by_cat = routing.by_category
    cls_summary: Dict[str, Any] = {}
    for cat in ALL_CATEGORIES:
        files_in_cat = by_cat.get(cat, [])
        if files_in_cat:
            cls_summary[cat] = {
                "count": len(files_in_cat),
                "files": [cf.path.name for cf in files_in_cat],
                "total_size": sum(cf.size_bytes for cf in files_in_cat),
            }
    inventory["classification_summary"] = cls_summary

    # Analysis — key findings
    analysis = _analyze_tender(inventory, routing)
    inventory["analysis"] = analysis
    inventory["scan_duration_sec"] = round(time.perf_counter() - t0, 2)

    return inventory


def _analyze_tender(inventory: Dict[str, Any], routing: RoutingSummary) -> Dict[str, Any]:
    """Produce analysis findings for a tender inventory."""
    by_cat = routing.by_category
    analysis: Dict[str, Any] = {}

    # BOQ source
    boq_excels = by_cat.get(CATEGORY_BOQ_XLSX, [])
    boq_pdfs = by_cat.get(CATEGORY_BOQ_PDF, [])
    if boq_excels:
        analysis["likely_boq_source"] = "Excel"
        analysis["boq_candidates"] = [cf.path.name for cf in boq_excels]
        # Enrich with sheet info
        for cf in boq_excels:
            try:
                rel = str(cf.path.relative_to(Path(inventory["folder_path"])))
            except ValueError:
                rel = cf.path.name
            prev = inventory["excel_previews"].get(rel, {})
            if prev.get("likely_boq"):
                analysis["boq_excel_confirmed"] = True
                analysis["boq_excel_indicators"] = prev.get("boq_indicators", [])
    elif boq_pdfs:
        analysis["likely_boq_source"] = "PDF"
        analysis["boq_candidates"] = [cf.path.name for cf in boq_pdfs]
    else:
        analysis["likely_boq_source"] = "NOT FOUND"
        analysis["boq_candidates"] = []

    # Commercial terms / conditions
    cond_pdfs = by_cat.get(CATEGORY_CONDITIONS_PDF, [])
    cond_docs = by_cat.get(CATEGORY_CONDITIONS_DOC, [])
    cond_all = [cf.path.name for cf in cond_pdfs + cond_docs]
    # Identify GCC/SCC specifically
    gcc_scc = [n for n in cond_all if any(kw in n.upper() for kw in ["GCC", "SCC", "IRGCC"])]
    analysis["conditions_files"] = cond_all
    analysis["gcc_scc_candidates"] = gcc_scc or ["(none found — check conditions PDFs)"]

    # Addenda / amendments
    addenda = by_cat.get(CATEGORY_ADDENDA_PDF, [])
    analysis["addenda_files"] = [cf.path.name for cf in addenda]
    analysis["addenda_count"] = len(addenda)

    # Drawings
    drawings = by_cat.get(CATEGORY_DRAWINGS_PDF, [])
    analysis["drawings_count"] = len(drawings)
    analysis["drawings_total_pages"] = sum(
        inventory["pdf_metadata"].get(
            _safe_relpath(cf.path, inventory["folder_path"]), {}
        ).get("page_count", 0)
        for cf in drawings
    )

    # Missing / unclear items
    missing = []
    if not boq_excels and not boq_pdfs:
        missing.append("No BOQ file identified (Excel or PDF)")
    if not cond_pdfs and not cond_docs:
        missing.append("No conditions/specs files found")
    if not drawings:
        missing.append("No drawing files identified")
    analysis["missing_items"] = missing

    # Duplicates / versions
    name_counts: Dict[str, int] = {}
    for f in inventory["files"]:
        base = f["name"].rsplit(".", 1)[0].lower().strip()
        # Normalize " (1)" suffixes
        import re
        base = re.sub(r'\s*\(\d+\)$', '', base)
        name_counts[base] = name_counts.get(base, 0) + 1
    duplicates = {name: cnt for name, cnt in name_counts.items() if cnt > 1}
    analysis["potential_duplicates"] = duplicates if duplicates else {}

    # Suggested processing plan
    plan = []
    if boq_excels:
        plan.append(f"1. Parse Excel BOQ ({', '.join(cf.path.name for cf in boq_excels)}) for structured line items")
    elif boq_pdfs:
        plan.append(f"1. OCR+extract BOQ PDFs ({', '.join(cf.path.name for cf in boq_pdfs)})")

    if cond_pdfs:
        plan.append(f"2. Extract conditions from {len(cond_pdfs)} PDF(s) — look for GCC/SCC clauses, LD, retention, defects liability")
    if addenda:
        plan.append(f"3. Process {len(addenda)} addenda/amendment file(s) — check for BOQ changes, date extensions")
    if drawings:
        plan.append(f"4. Index {len(drawings)} drawing PDF(s) ({analysis['drawings_total_pages']} pages) — classify by discipline")

    if not plan:
        plan.append("1. Manual review needed — no clear targets identified")

    analysis["suggested_processing_plan"] = plan

    return analysis


def _safe_relpath(path: Path, root_str: str) -> str:
    """Safely compute relative path, returning name on failure."""
    try:
        return str(path.relative_to(root_str))
    except ValueError:
        return path.name


# ═════════════════════════════════════════════════════════════════════════
# OUTPUT WRITERS
# ═════════════════════════════════════════════════════════════════════════

def write_manifest_json(inventory: Dict[str, Any], output_path: Path):
    """Write per-tender manifest.json."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(inventory, f, indent=2, default=str, ensure_ascii=False)


def write_summary_md(inventory: Dict[str, Any], output_path: Path):
    """Write per-tender summary.md."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    a = inventory["analysis"]
    lines = [
        f"# Tender Inventory: {inventory['folder_name']}",
        "",
        f"*Scanned: {inventory['scan_timestamp']}*",
        "",
        "---",
        "",
        "## File Counts and Sizes",
        "",
        f"- **Total files**: {inventory['total_files']}",
        f"- **Total size**: {inventory['total_size_human']}",
        f"- **PDF pages**: {inventory['pdf_total_pages']}",
        "",
        "| Extension | Count |",
        "|-----------|-------|",
    ]
    for ext, count in sorted(inventory["file_type_counts"].items()):
        lines.append(f"| {ext} | {count} |")

    # Classification breakdown
    lines += [
        "",
        "## What We Have",
        "",
        "| Category | Count | Files |",
        "|----------|-------|-------|",
    ]
    for cat, info in inventory["classification_summary"].items():
        file_list = ", ".join(info["files"][:4])
        if len(info["files"]) > 4:
            file_list += f" +{len(info['files']) - 4} more"
        lines.append(f"| {cat} | {info['count']} | {file_list} |")

    # Likely BOQ source
    lines += [
        "",
        "## Likely BOQ Source",
        "",
        f"**Source type**: {a.get('likely_boq_source', 'unknown')}",
        "",
    ]
    if a.get("boq_candidates"):
        lines.append("Candidate files:")
        for name in a["boq_candidates"]:
            lines.append(f"- `{name}`")
    if a.get("boq_excel_confirmed"):
        lines.append("")
        lines.append("**Confirmed BOQ Excel** — headers contain pricing keywords:")
        for ind in a.get("boq_excel_indicators", []):
            lines.append(f"- {ind}")

    # Commercial terms
    lines += [
        "",
        "## Likely Commercial Terms Location",
        "",
    ]
    gcc = a.get("gcc_scc_candidates", [])
    if gcc and gcc != ["(none found — check conditions PDFs)"]:
        lines.append("GCC/SCC candidates:")
        for name in gcc:
            lines.append(f"- `{name}`")
    else:
        lines.append("No explicit GCC/SCC files found. Check conditions PDFs for embedded clauses.")

    if a.get("conditions_files"):
        lines.append("")
        lines.append("All conditions/specs files:")
        for name in a["conditions_files"]:
            lines.append(f"- `{name}`")

    # Addenda
    lines += [
        "",
        "## Addenda / Corrigendum / Amendments",
        "",
    ]
    if a.get("addenda_files"):
        lines.append(f"Found **{a['addenda_count']}** addenda file(s):")
        for name in a["addenda_files"]:
            lines.append(f"- `{name}`")
    else:
        lines.append("No addenda files found.")

    # Drawings
    lines += [
        "",
        "## Drawings",
        "",
        f"- **{a.get('drawings_count', 0)}** drawing PDF(s)",
        f"- **{a.get('drawings_total_pages', 0)}** total pages",
    ]

    # Missing items
    lines += [
        "",
        "## Missing / Unclear Items",
        "",
    ]
    if a.get("missing_items"):
        for item in a["missing_items"]:
            lines.append(f"- {item}")
    else:
        lines.append("All key components identified.")

    # Duplicates
    if a.get("potential_duplicates"):
        lines += [
            "",
            "## Potential Duplicates / Versions",
            "",
        ]
        for name, cnt in a["potential_duplicates"].items():
            lines.append(f"- `{name}` appears **{cnt}** times")

    # Processing plan
    lines += [
        "",
        "## Suggested Processing Plan",
        "",
    ]
    for step in a.get("suggested_processing_plan", []):
        lines.append(f"- {step}")

    # File listing
    lines += [
        "",
        "---",
        "",
        "## Complete File Listing",
        "",
        "| File | Category | Size | Pages |",
        "|------|----------|------|-------|",
    ]
    for f in inventory["files"]:
        pages = f.get("page_count", "")
        pages_str = str(pages) if pages else "-"
        lines.append(
            f"| {f['relative_path']} | {f.get('category', '?')} | {f['size_human']} | {pages_str} |"
        )

    lines.append("")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_rollup_md(inventories: List[Dict[str, Any]], output_path: Path):
    """Write master rollup markdown."""
    lines = [
        "# Pilot Tender Inventory — Master Rollup",
        "",
        f"*Generated: {datetime.now().isoformat()}*",
        "",
        f"**{len(inventories)} tenders scanned**",
        "",
        "## Summary Table",
        "",
        "| Tender | Files | PDF Pages | Excel | BOQ Source | Conditions | Addenda | Drawings |",
        "|--------|-------|-----------|-------|-----------|------------|---------|----------|",
    ]
    total_files = 0
    total_pages = 0

    for inv in inventories:
        a = inv["analysis"]
        cls = inv["classification_summary"]
        total_files += inv["total_files"]
        total_pages += inv["pdf_total_pages"]

        excel_count = inv["file_type_counts"].get(".xls", 0) + inv["file_type_counts"].get(".xlsx", 0)
        cond_count = cls.get(CATEGORY_CONDITIONS_PDF, {}).get("count", 0) + cls.get(CATEGORY_CONDITIONS_DOC, {}).get("count", 0)
        add_count = a.get("addenda_count", 0)
        draw_count = a.get("drawings_count", 0)
        boq_src = a.get("likely_boq_source", "?")

        lines.append(
            f"| {inv['folder_name']} | {inv['total_files']} | {inv['pdf_total_pages']} "
            f"| {excel_count} | {boq_src} | {cond_count} | {add_count} | {draw_count} |"
        )

    lines += [
        "",
        f"**Totals**: {total_files} files, {total_pages} PDF pages across {len(inventories)} tenders",
        "",
    ]

    # Top findings
    lines += [
        "## Key Findings",
        "",
    ]
    for inv in inventories:
        a = inv["analysis"]
        lines.append(f"### {inv['folder_name']}")
        lines.append("")
        if a.get("missing_items"):
            for item in a["missing_items"]:
                lines.append(f"- **Missing**: {item}")
        if a.get("potential_duplicates"):
            for name, cnt in a["potential_duplicates"].items():
                lines.append(f"- **Duplicate**: `{name}` ({cnt} copies)")
        if not a.get("missing_items") and not a.get("potential_duplicates"):
            lines.append("- All key components present")
        lines.append("")

    # What to process first
    lines += [
        "## Recommended Processing Order",
        "",
    ]
    # Sort by: tenders with Excel BOQ first (fastest win), then by page count
    sorted_inv = sorted(inventories, key=lambda i: (
        0 if i["analysis"].get("likely_boq_source") == "Excel" else 1,
        i["pdf_total_pages"],
    ))
    for i, inv in enumerate(sorted_inv, 1):
        a = inv["analysis"]
        lines.append(
            f"{i}. **{inv['folder_name']}** — "
            f"{inv['pdf_total_pages']} pages, "
            f"BOQ: {a.get('likely_boq_source', '?')}, "
            f"{a.get('drawings_count', 0)} drawings"
        )

    lines.append("")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


ROLLUP_CSV_FIELDS = [
    "tender_name", "total_files", "total_size_mb", "pdf_count", "pdf_total_pages",
    "excel_count", "doc_count", "rar_count",
    "drawings_count", "boq_pdf_count", "conditions_count", "addenda_count",
    "unknown_count", "likely_boq_source", "boq_candidates",
    "gcc_scc_candidates", "missing_items", "scan_sec",
]


def write_rollup_csv(inventories: List[Dict[str, Any]], output_path: Path):
    """Write master rollup CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ROLLUP_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for inv in inventories:
            a = inv["analysis"]
            cls = inv["classification_summary"]
            tc = inv["file_type_counts"]
            row = {
                "tender_name": inv["folder_name"],
                "total_files": inv["total_files"],
                "total_size_mb": f"{inv['total_size_bytes'] / (1024*1024):.1f}",
                "pdf_count": tc.get(".pdf", 0),
                "pdf_total_pages": inv["pdf_total_pages"],
                "excel_count": tc.get(".xls", 0) + tc.get(".xlsx", 0),
                "doc_count": tc.get(".doc", 0) + tc.get(".docx", 0),
                "rar_count": tc.get(".rar", 0),
                "drawings_count": a.get("drawings_count", 0),
                "boq_pdf_count": cls.get(CATEGORY_BOQ_PDF, {}).get("count", 0),
                "conditions_count": (
                    cls.get(CATEGORY_CONDITIONS_PDF, {}).get("count", 0)
                    + cls.get(CATEGORY_CONDITIONS_DOC, {}).get("count", 0)
                ),
                "addenda_count": a.get("addenda_count", 0),
                "unknown_count": cls.get(CATEGORY_UNKNOWN, {}).get("count", 0),
                "likely_boq_source": a.get("likely_boq_source", "?"),
                "boq_candidates": "; ".join(a.get("boq_candidates", [])),
                "gcc_scc_candidates": "; ".join(a.get("gcc_scc_candidates", [])),
                "missing_items": "; ".join(a.get("missing_items", [])),
                "scan_sec": inv.get("scan_duration_sec", ""),
            }
            writer.writerow(row)


# ═════════════════════════════════════════════════════════════════════════
# FOLDER DISCOVERY (reused from pilot_batch_ingest)
# ═════════════════════════════════════════════════════════════════════════

def find_tender_folders(
    input_dir: Path,
    include: Optional[List[str]] = None,
) -> List[Path]:
    """Find tender subdirectories in input_dir."""
    _skip = {".git", "__MACOSX", "__pycache__", ".DS_Store", "node_modules"}
    folders = sorted(
        p for p in input_dir.iterdir()
        if p.is_dir() and p.name not in _skip and not p.name.startswith(".")
    )
    if include:
        include_lower = {n.lower() for n in include}
        folders = [f for f in folders if f.name.lower() in include_lower]
    return folders


# ═════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Pilot Inventory — lightweight metadata scan of tender folders (no OCR/pipeline)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--input", "-i", type=Path, required=True, dest="input_dir",
        help="Parent directory containing tender subfolders",
    )
    parser.add_argument(
        "--output", "-o", type=Path, required=True, dest="output_dir",
        help="Output directory for inventory reports",
    )
    parser.add_argument(
        "--include", nargs="+", type=str, default=None,
        help="Only inventory these tender folders (by name)",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Verbose output",
    )
    args = parser.parse_args()

    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    if not args.input_dir.exists():
        logger.error(f"Input directory not found: {args.input_dir}")
        return 1

    folders = find_tender_folders(args.input_dir, args.include)
    if not folders:
        logger.error(f"No tender folders found in {args.input_dir}")
        return 1

    logger.info(f"Found {len(folders)} tender folder(s) to scan")
    for f in folders:
        logger.info(f"  -> {f.name}")

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    inv_dir = output_dir / "inventory"

    all_inventories: List[Dict[str, Any]] = []

    for i, folder in enumerate(folders, 1):
        logger.info(f"\n[{i}/{len(folders)}] Scanning: {folder.name}")
        logger.info("=" * 60)

        inventory = build_tender_inventory(folder)
        all_inventories.append(inventory)

        # Write per-tender outputs
        tender_out = inv_dir / folder.name
        write_manifest_json(inventory, tender_out / "manifest.json")
        write_summary_md(inventory, tender_out / "summary.md")

        logger.info(
            f"  {inventory['total_files']} files, "
            f"{inventory['pdf_total_pages']} PDF pages, "
            f"{inventory['total_size_human']} total"
        )
        logger.info(f"  BOQ source: {inventory['analysis'].get('likely_boq_source', '?')}")
        logger.info(f"  Scanned in {inventory.get('scan_duration_sec', '?')}s")

    # Write master rollup
    write_rollup_md(all_inventories, output_dir / "inventory_rollup.md")
    write_rollup_csv(all_inventories, output_dir / "inventory_rollup.csv")

    # Print summary
    logger.info(f"\n{'=' * 60}")
    logger.info(f"INVENTORY COMPLETE: {len(all_inventories)} tenders scanned")
    logger.info(f"Per-tender reports: {inv_dir}")
    logger.info(f"Rollup:             {output_dir / 'inventory_rollup.md'}")
    logger.info(f"Rollup CSV:         {output_dir / 'inventory_rollup.csv'}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
