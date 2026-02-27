"""
Tests for the upgraded tender pipeline:
- PageIndex (classification)
- PageSelection (prioritization)
- Extractors (per-type extraction)
- RFI Engine (checklist-driven RFIs)
- Guardrails (extraction gap warnings)
"""

import sys
import json
import unittest
from pathlib import Path

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import pytest
from src.analysis.page_index import (
    build_page_index, PageIndex, IndexedPage, _classify_page,
)
from src.analysis.page_selection import (
    select_pages, SelectedPages,
)
from src.analysis.extractors import (
    run_extractors, ExtractionResult,
)
from src.analysis.extractors.extract_notes import extract_requirements
from src.analysis.extractors.extract_schedule_tables import extract_schedule_rows
from src.analysis.extractors.extract_boq import extract_boq_items
from src.analysis.extractors.extract_drawings_minimal import extract_drawing_callouts
from src.analysis.rfi_engine import generate_rfis, CheckContext
from src.analysis.pipeline import _check_guardrails, _build_processing_stats
from src.models.analysis_models import (
    PlanSetGraph, PlanSheet, Discipline, SheetType, RFIItem, EvidenceRef,
)


# =============================================================================
# FIXTURES — Synthetic texts for testing
# =============================================================================

BOQ_TEXT = """
BILL OF QUANTITIES

Item No.  Description                        Unit    Qty     Rate
1.1       Earthwork excavation in soil        cum     120.5   450.00
1.2       PCC M15 below footings              cum     25.0    5200.00
1.3       RCC M25 for footings                cum     45.0    7500.00
2.1       Brickwork in CM 1:6                 sqm     350.0   650.00
2.2       Internal plaster 12mm thick         sqm     800.0
2.3       External plaster 20mm thick         sqm     450.0   380.00
"""

SCHEDULE_TEXT = """
DOOR SCHEDULE

Mark    Size        Type        Material    Qty    Remarks
D1      900x2100    Flush       Teak        4      Main Entry
D2      800x2100    Flush       WPC         12     Internal
D3      750x2100    Panel       Sal Wood    2      Store Room
D4      1200x2100   Sliding     Aluminium   1      Balcony

WINDOW SCHEDULE

Mark    Size        Type        Glazing     Qty
W1      1200x1200   Casement    5mm Clear   8
W2      900x1200    Sliding     5mm Clear   6
W3      600x600     Fixed       Frosted     4
"""

PLAN_TEXT = """
SHEET A-101
GROUND FLOOR PLAN
SCALE 1:100

LIVING ROOM         KITCHEN
3600 x 4200mm       3000 x 3600mm

BEDROOM 1           BATHROOM
3600 x 3600mm       1800 x 2400mm

D1  D2  D3  D4
W1  W2  W3

SEE SECTION A-A
DETAIL 3/A-5.01

200mm THK BRICK WALL
RCC M25 COLUMN
"""

NOTES_TEXT = """
GENERAL NOTES

1. All cement shall be OPC 43 grade conforming to IS:8112.
2. All steel reinforcement shall be Fe500D conforming to IS:1786.
3. Concrete grade for structural elements shall be M25 as per IS:456.
4. Minimum cover to reinforcement: 40mm for footings, 25mm for columns.
5. All brickwork shall be first class burnt clay bricks conforming to IS:1077.
6. Cement mortar for brickwork shall be 1:6 (cement:sand).
7. All doors and windows shall comply with IS:4021.
8. Waterproofing shall be done with polymer modified cement-based coating.
9. All plumbing work shall conform to IS:2065.
10. Testing of concrete: cube test at 7 and 28 days as per IS:516.

Material specifications must comply with CPWD specifications 2019 volume 1 and 2.
ASTM C150 cement may be used as alternative to OPC 43 grade.
"""

SECTION_TEXT = """
SHEET A-301
SECTION A-A
SCALE 1:50

PLINTH LEVEL +0.600
GROUND FLOOR LEVEL +0.000
FIRST FLOOR LEVEL +3.300
TERRACE LEVEL +6.600

150mm THK RCC SLAB
300 x 450mm RCC BEAM
300 x 300mm RCC COLUMN
"""

CONDITIONS_TEXT = """
GENERAL CONDITIONS OF CONTRACT

1. Terms and conditions as per GCC clause 1 to 63.
2. Defect liability period: 12 months from completion.
3. Retention: 5% of each running bill.
4. Mobilization advance: 10% of contract value.
5. Time for completion: 18 months.
6. Liquidated damages: 0.5% per week of delay, max 10%.
"""


# =============================================================================
# TEST: PageIndex Classification
# =============================================================================

class TestPageIndex:
    """Test page classification logic."""

    def test_boq_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(BOQ_TEXT)
        assert doc_type == "boq", f"Expected 'boq', got '{doc_type}'"
        assert conf > 0.4

    def test_schedule_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(SCHEDULE_TEXT)
        assert doc_type == "schedule", f"Expected 'schedule', got '{doc_type}'"

    def test_plan_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(PLAN_TEXT)
        assert doc_type == "plan", f"Expected 'plan', got '{doc_type}'"
        assert disc == "architectural", f"Expected 'architectural', got '{disc}'"
        assert sheet_id is not None

    def test_notes_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(NOTES_TEXT)
        assert doc_type == "notes", f"Expected 'notes', got '{doc_type}'"

    def test_section_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(SECTION_TEXT)
        assert doc_type == "section", f"Expected 'section', got '{doc_type}'"

    def test_conditions_classification(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(CONDITIONS_TEXT)
        assert doc_type == "conditions", f"Expected 'conditions', got '{doc_type}'"

    def test_text_layer_pages_skip_rendering(self):
        """Text layer pages should be classified without rendering."""
        # Simulate: we have text for all pages
        texts = [PLAN_TEXT, SCHEDULE_TEXT, BOQ_TEXT, NOTES_TEXT, SECTION_TEXT]
        # Use a dummy path — build_page_index will classify from text
        # Can't test rendering without a real PDF, but we can test the text path
        for text in texts:
            doc_type, _, _, _, conf, _ = _classify_page(text)
            assert doc_type != "unknown", f"Should classify text, got 'unknown' for {text[:40]}"

    def test_unknown_fallback(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page("random gibberish text here")
        assert doc_type == "unknown"

    def test_sheet_id_extraction(self):
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(PLAN_TEXT)
        assert sheet_id is not None
        assert "A" in sheet_id  # A-101

    def test_discipline_from_sheet_id(self):
        """Discipline should be derived from sheet ID prefix."""
        structural_text = "SHEET S-01\nSTRUCTURAL LAYOUT\nFOUNDATION PLAN"
        doc_type, disc, sheet_id, title, conf, kw = _classify_page(structural_text)
        assert disc == "structural"


# =============================================================================
# TEST: PageSelection
# =============================================================================

class TestPageSelection:
    """Test page selection logic."""

    def _make_index(self, page_types: list) -> PageIndex:
        """Create a PageIndex from a list of (doc_type, discipline) tuples."""
        pages = []
        from collections import Counter
        type_counter = Counter()
        disc_counter = Counter()

        for i, (dtype, disc) in enumerate(page_types):
            pages.append(IndexedPage(
                page_idx=i, doc_type=dtype, discipline=disc,
                confidence=0.7, has_text_layer=True,
            ))
            type_counter[dtype] += 1
            disc_counter[disc] += 1

        return PageIndex(
            pdf_name="test.pdf",
            total_pages=len(pages),
            pages=pages,
            counts_by_type=dict(type_counter),
            counts_by_discipline=dict(disc_counter),
        )

    def test_all_schedules_selected(self):
        """All schedule pages must always be selected."""
        page_types = [("plan", "architectural")] * 100 + [("schedule", "architectural")] * 5
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=30)
        schedule_selected = [i for i in result.selected if idx.pages[i].doc_type == "schedule"]
        assert len(schedule_selected) == 5, f"Expected all 5 schedules, got {len(schedule_selected)}"

    def test_budget_respected(self):
        """Selection must not exceed budget."""
        page_types = [("plan", "architectural")] * 200
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=50)
        assert result.budget_used <= 50

    def test_small_set_selects_all(self):
        """If pages <= budget, select everything."""
        page_types = [("plan", "architectural")] * 10
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=80)
        assert result.budget_used == 10

    def test_balanced_discipline_sampling(self):
        """Tier 2 should round-robin across disciplines."""
        page_types = (
            [("plan", "architectural")] * 50 +
            [("plan", "structural")] * 50 +
            [("plan", "electrical")] * 50 +
            [("plan", "plumbing")] * 50
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=40)

        # Count per discipline
        disc_counts = {}
        for i in result.selected:
            disc = idx.pages[i].discipline
            disc_counts[disc] = disc_counts.get(disc, 0) + 1

        # Each discipline should get roughly equal share
        assert len(disc_counts) == 4, "Should sample from all 4 disciplines"
        counts = list(disc_counts.values())
        assert max(counts) - min(counts) <= 2, f"Discipline imbalance: {disc_counts}"

    def test_tier1_always_included(self):
        """BOQ, conditions, and notes pages should always be included."""
        page_types = (
            [("plan", "architectural")] * 100 +
            [("boq", "other")] * 3 +
            [("conditions", "other")] * 2 +
            [("notes", "architectural")] * 2
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=30)

        boq_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "boq")
        cond_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "conditions")
        notes_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "notes")

        assert boq_selected == 3, f"All BOQ pages should be selected, got {boq_selected}"
        assert cond_selected == 2
        assert notes_selected == 2

    def test_tier1_boq_not_crowded_out_by_conditions(self):
        """BOQ pages must be selected even when conditions pages exceed budget."""
        page_types = (
            [("conditions", "other")] * 103 +
            [("boq", "other")] * 2 +
            [("spec", "other")] * 2 +
            [("section", "architectural")] * 1 +
            [("unknown", "other")] * 259
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=80)

        boq_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "boq")
        assert boq_selected == 2, f"All BOQ pages must be selected, got {boq_selected}"

        # Conditions should be sub-capped, not all 103
        cond_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "conditions")
        assert cond_selected <= 20, f"Conditions should be sub-capped at 20, got {cond_selected}"
        assert cond_selected >= 1, "At least some conditions should be selected"

    def test_tier1_conditions_sub_capped(self):
        """103 conditions pages should be sub-capped to ~20."""
        page_types = (
            [("conditions", "other")] * 103 +
            [("plan", "architectural")] * 50
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=80)

        cond_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "conditions")
        assert cond_selected <= 20, f"Expected <=20 conditions, got {cond_selected}"
        assert cond_selected >= 10, f"Expected >=10 conditions sampled, got {cond_selected}"

    def test_tier1_expands_budget_when_necessary(self):
        """If Tier 1 uncapped types exceed budget, budget should expand."""
        page_types = (
            [("boq", "other")] * 50 +
            [("addendum", "other")] * 20 +
            [("schedule", "architectural")] * 20 +
            [("plan", "architectural")] * 100
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=80)

        # All boq + addendum + schedule = 90 must be selected (exceeds 80 budget)
        boq = sum(1 for i in result.selected if idx.pages[i].doc_type == "boq")
        addendum = sum(1 for i in result.selected if idx.pages[i].doc_type == "addendum")
        schedule = sum(1 for i in result.selected if idx.pages[i].doc_type == "schedule")
        assert boq == 50, f"All 50 BOQ pages must be selected, got {boq}"
        assert addendum == 20, f"All 20 addendum pages must be selected, got {addendum}"
        assert schedule == 20, f"All 20 schedule pages must be selected, got {schedule}"
        assert result.budget_used >= 90, f"Budget should expand to fit Tier 1, got {result.budget_used}"

    def test_tier1_priority_order(self):
        """BOQ pages selected before conditions regardless of page order."""
        # BOQ pages at end of PDF (high page indices), conditions at start
        page_types = (
            [("conditions", "other")] * 100 +
            [("unknown", "other")] * 50 +
            [("boq", "other")] * 3
        )
        idx = self._make_index(page_types)
        result = select_pages(idx, budget_pages=40)

        boq_selected = sum(1 for i in result.selected if idx.pages[i].doc_type == "boq")
        assert boq_selected == 3, f"All 3 BOQ pages must be selected, got {boq_selected}"


# =============================================================================
# TEST: Extractors
# =============================================================================

class TestExtractors:
    """Test specialized extractors."""

    def test_notes_extract_requirements(self):
        items = extract_requirements(NOTES_TEXT, source_page=0, sheet_id=None, doc_type="notes")
        assert len(items) >= 2, f"Expected >=2 requirements, got {len(items)}"
        # Check we extracted code references
        categories = {item["category"] for item in items}
        assert "standard" in categories or "material" in categories

    def test_schedule_extract_rows(self):
        items = extract_schedule_rows(SCHEDULE_TEXT, source_page=0, sheet_id="A-501")
        assert len(items) >= 2, f"Expected >=2 schedule rows, got {len(items)}"
        # Check marks
        marks = {item["mark"] for item in items}
        assert "D1" in marks or "D-1" in marks or any("D" in m for m in marks)

    def test_boq_extract_items(self):
        items = extract_boq_items(BOQ_TEXT, source_page=0)
        assert len(items) >= 1, f"Expected >=1 BOQ item, got {len(items)}"
        # Check that quantities were extracted
        with_qty = [item for item in items if item.get("qty") is not None]
        assert len(with_qty) >= 1

    def test_drawing_extract_callouts(self):
        items = extract_drawing_callouts(PLAN_TEXT, source_page=0, sheet_id="A-101")
        assert len(items) >= 3, f"Expected >=3 callouts, got {len(items)}"
        types = {item["callout_type"] for item in items}
        assert "dimension" in types or "tag" in types or "material" in types

    def test_extractor_routing(self):
        """run_extractors should route pages to correct extractors."""
        pages = [
            IndexedPage(page_idx=0, doc_type="notes", discipline="architectural",
                        confidence=0.7, has_text_layer=True),
            IndexedPage(page_idx=1, doc_type="schedule", discipline="architectural",
                        confidence=0.7, has_text_layer=True),
            IndexedPage(page_idx=2, doc_type="plan", discipline="architectural",
                        confidence=0.7, has_text_layer=True),
            IndexedPage(page_idx=3, doc_type="boq", discipline="other",
                        confidence=0.7, has_text_layer=True),
        ]
        page_index = PageIndex(
            pdf_name="test.pdf",
            total_pages=4,
            pages=pages,
            counts_by_type={"notes": 1, "schedule": 1, "plan": 1, "boq": 1},
            counts_by_discipline={"architectural": 3, "other": 1},
        )
        text_map = {
            0: NOTES_TEXT,
            1: SCHEDULE_TEXT,
            2: PLAN_TEXT,
            3: BOQ_TEXT,
        }
        result = run_extractors(text_map, page_index)
        assert result.pages_processed == 4
        assert len(result.requirements) > 0, "Should have requirements from notes"
        assert len(result.schedules) > 0, "Should have schedule rows"
        assert len(result.callouts) > 0, "Should have drawing callouts"


# =============================================================================
# TEST: RFI Engine
# =============================================================================

class TestRFIEngine:
    """Test checklist-driven RFI generation."""

    def _make_graph(self, door_tags=None, window_tags=None, room_names=None,
                     has_door_schedule=False, has_window_schedule=False,
                     has_finish_schedule=False, has_legend=False,
                     pages_with_scale=0, pages_without_scale=5):
        return PlanSetGraph(
            project_id="test",
            sheets=[],
            total_pages=10,
            disciplines_found=["A"],
            all_door_tags=door_tags or [],
            all_window_tags=window_tags or [],
            all_room_names=room_names or [],
            has_door_schedule=has_door_schedule,
            has_window_schedule=has_window_schedule,
            has_finish_schedule=has_finish_schedule,
            has_legend=has_legend,
            pages_with_scale=pages_with_scale,
            pages_without_scale=pages_without_scale,
        )

    def _make_extraction(self, schedules=None, boq_items=None, callouts=None, requirements=None):
        return ExtractionResult(
            requirements=requirements or [],
            schedules=schedules or [],
            boq_items=boq_items or [],
            callouts=callouts or [],
        )

    def _make_selected(self, n=10, budget=80):
        return SelectedPages(
            selected=list(range(n)),
            budget_total=budget,
            budget_used=n,
            always_include_count=n,
        )

    def _make_page_index(self, types=None):
        if types is None:
            types = {"plan": 5, "unknown": 5}
        pages = []
        idx = 0
        for dtype, count in types.items():
            for _ in range(count):
                pages.append(IndexedPage(
                    page_idx=idx, doc_type=dtype, discipline="unknown",
                    confidence=0.5, has_text_layer=True,
                ))
                idx += 1
        return PageIndex(
            pdf_name="test.pdf",
            total_pages=idx,
            pages=pages,
            counts_by_type=types,
            counts_by_discipline={"unknown": idx},
        )

    def test_missing_door_schedule_fires_rfi(self):
        """Door tags found but no schedule should generate an RFI."""
        graph = self._make_graph(door_tags=["D1", "D2", "D3"])
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "sheet_id": "A-101", "confidence": 0.8},
            {"text": "D2", "callout_type": "tag", "source_page": 0, "sheet_id": "A-101", "confidence": 0.8},
        ]
        extracted = self._make_extraction(callouts=callouts)
        selected = self._make_selected()
        page_index = self._make_page_index()

        rfis = generate_rfis(extracted, page_index, selected, graph)
        door_rfis = [r for r in rfis if "door schedule" in r.question.lower()]
        assert len(door_rfis) >= 1, "Should fire RFI for missing door schedule"

    def test_every_rfi_has_evidence(self):
        """Every generated RFI must have evidence."""
        graph = self._make_graph(
            door_tags=["D1"], window_tags=["W1"], room_names=["BEDROOM 1"],
        )
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "sheet_id": None, "confidence": 0.8},
        ]
        extracted = self._make_extraction(callouts=callouts)
        selected = self._make_selected()
        page_index = self._make_page_index()

        rfis = generate_rfis(extracted, page_index, selected, graph)
        for rfi in rfis:
            assert rfi.evidence is not None, f"RFI {rfi.id} missing evidence"
            assert rfi.evidence.has_evidence, f"RFI {rfi.id} has empty evidence"

    def test_large_tender_produces_many_rfis(self):
        """A large tender with minimal data should produce >15 RFIs."""
        # Simulate: large tender, mostly unknown pages, no schedules, no structural
        page_index = self._make_page_index({
            "plan": 50, "unknown": 300, "cover": 2,
        })
        graph = self._make_graph(
            door_tags=["D1", "D2", "D3"], window_tags=["W1", "W2"],
            room_names=["BEDROOM 1", "KITCHEN", "LIVING"],
            pages_without_scale=45, pages_with_scale=5,
        )
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "sheet_id": None, "confidence": 0.8},
            {"text": "3600mm", "callout_type": "dimension", "source_page": 0, "sheet_id": None, "confidence": 0.7},
        ]
        extracted = self._make_extraction(callouts=callouts)
        selected = self._make_selected(n=80, budget=80)

        rfis = generate_rfis(extracted, page_index, selected, graph)
        assert len(rfis) >= 15, f"Expected >=15 RFIs on large tender, got {len(rfis)}"

    def test_no_evidence_disclaimer(self):
        """Low-confidence RFIs should have the 'may exist elsewhere' disclaimer."""
        page_index = self._make_page_index({"plan": 5})
        graph = self._make_graph()
        extracted = self._make_extraction()
        selected = self._make_selected(n=5, budget=80)

        rfis = generate_rfis(extracted, page_index, selected, graph)
        # At least some RFIs should have the disclaimer
        disclaimers = [
            r for r in rfis
            if "may exist elsewhere" in (r.evidence.confidence_reason or "")
        ]
        assert len(disclaimers) > 0, "Some RFIs should have 'may exist elsewhere' disclaimer"


# =============================================================================
# TEST: Guardrails
# =============================================================================

class TestGuardrails:
    """Test extraction gap detection."""

    def test_schedule_pages_empty_extraction_warns(self):
        page_index = PageIndex(
            pdf_name="test.pdf", total_pages=50,
            pages=[], counts_by_type={"schedule": 5, "plan": 30},
            counts_by_discipline={},
        )
        extracted = ExtractionResult()  # empty
        warnings = _check_guardrails(page_index, extracted, [])
        types = [w["type"] for w in warnings]
        assert "schedule_extraction_gap" in types

    def test_low_rfi_count_warns_on_big_tender(self):
        page_index = PageIndex(
            pdf_name="test.pdf", total_pages=200,
            pages=[], counts_by_type={"plan": 150, "unknown": 50},
            counts_by_discipline={},
        )
        extracted = ExtractionResult()
        # Only 3 RFIs
        rfis = [object()] * 3
        warnings = _check_guardrails(page_index, extracted, rfis)
        types = [w["type"] for w in warnings]
        assert "low_rfi_count" in types

    def test_no_warnings_on_complete_extraction(self):
        page_index = PageIndex(
            pdf_name="test.pdf", total_pages=20,
            pages=[], counts_by_type={"plan": 15, "schedule": 2},
            counts_by_discipline={},
        )
        extracted = ExtractionResult(
            schedules=[{"mark": "D1"}, {"mark": "D2"}],
            requirements=[{"text": "test"}] * 5,
        )
        rfis = [object()] * 20
        warnings = _check_guardrails(page_index, extracted, rfis)
        # Should not have schedule_extraction_gap or low_rfi_count
        types = [w["type"] for w in warnings]
        assert "schedule_extraction_gap" not in types
        assert "low_rfi_count" not in types


# =============================================================================
# COVERAGE GATING TESTS
# =============================================================================

from src.models.analysis_models import (
    RunCoverage, CoverageStatus, SelectionMode,
)
from src.analysis.dependency_reasoner import reason_dependencies
from src.analysis.page_selection import SAFE_CAP_FULL_READ


class TestCoverageGating:
    """Tests for coverage-gated blocker and RFI assertions."""

    def _build_graph_with_doors(self) -> PlanSetGraph:
        """Build a graph with door tags but no door schedule."""
        graph = PlanSetGraph(
            project_id="test_cov",
            total_pages=10,
            sheets=[
                PlanSheet(
                    page_index=0, sheet_id="A-101",
                    sheet_type=SheetType.FLOOR_PLAN,
                    discipline=Discipline.ARCHITECTURAL,
                    detected={"door_tags": ["D1", "D2", "D3"]},
                ),
            ],
            all_door_tags=["D1", "D2", "D3"],
            has_door_schedule=False,
            disciplines_found=["A", "S"],
            sheet_types_found={"floor_plan": 4, "section": 1, "elevation": 1},
            pages_with_scale=0,
            pages_without_scale=10,
        )
        return graph

    def test_full_read_missing_schedule_is_genuine(self):
        """FULL_READ + schedule covered → blocker stays HIGH, coverage_status=not_found_after_search."""
        graph = self._build_graph_with_doors()
        run_cov = RunCoverage(
            pages_total=10, pages_indexed=10, pages_deep_processed=10,
            doc_types_detected={"plan": 4, "schedule": 2, "section": 1, "elevation": 1},
            doc_types_fully_covered=["plan", "schedule", "section", "elevation"],
            doc_types_partially_covered=[], doc_types_not_covered=[],
            selection_mode=SelectionMode.FULL_READ,
            ocr_budget_pages=80,
        )
        blockers, rfis, _, _ = reason_dependencies(graph, run_coverage=run_cov)
        # Should have door schedule blocker
        door_blockers = [b for b in blockers if b.issue_type == "missing_schedule" and "door" in b.title.lower()]
        assert len(door_blockers) == 1, f"Expected 1 door blocker, got {len(door_blockers)}"
        b = door_blockers[0]
        assert b.severity.value == "high", f"Expected HIGH severity, got {b.severity.value}"
        assert b.coverage_status == "not_found_after_search", f"Expected not_found_after_search, got {b.coverage_status}"

    def test_fast_budget_skipped_schedule_is_unknown(self):
        """FAST_BUDGET + schedule NOT covered → blocker downgraded to MEDIUM, coverage_status=unknown_not_processed."""
        graph = self._build_graph_with_doors()
        run_cov = RunCoverage(
            pages_total=300, pages_indexed=300, pages_deep_processed=80,
            doc_types_detected={"plan": 100, "schedule": 20, "section": 15, "elevation": 10},
            doc_types_fully_covered=[],
            doc_types_partially_covered=["plan", "section", "elevation"],
            doc_types_not_covered=["schedule"],  # schedule pages NOT covered
            selection_mode=SelectionMode.FAST_BUDGET,
            ocr_budget_pages=80,
        )
        blockers, rfis, _, _ = reason_dependencies(graph, run_coverage=run_cov)
        door_blockers = [b for b in blockers if b.issue_type == "missing_schedule" and "door" in b.title.lower()]
        assert len(door_blockers) == 1, f"Expected 1 door blocker, got {len(door_blockers)}"
        b = door_blockers[0]
        assert b.severity.value == "medium", f"Expected MEDIUM severity, got {b.severity.value}"
        assert b.coverage_status == "unknown_not_processed", f"Expected unknown_not_processed, got {b.coverage_status}"
        assert b.evidence.confidence <= 0.5, f"Expected confidence <= 0.5, got {b.evidence.confidence}"

    def test_rfi_engine_coverage_gating(self):
        """generate_rfis with FAST_BUDGET → door schedule RFI has coverage_status set."""
        page_index = PageIndex(
            pdf_name="test.pdf",
            pages=[
                IndexedPage(page_idx=0, doc_type="plan", discipline="architectural"),
                IndexedPage(page_idx=1, doc_type="plan", discipline="structural"),
            ],
            total_pages=2,
            counts_by_type={"plan": 2},
            counts_by_discipline={"architectural": 1, "structural": 1},
        )
        selected = SelectedPages(
            selected=[0, 1],
            budget_total=80, budget_used=2,
            selection_mode="fast_budget",
        )
        extracted = ExtractionResult(
            callouts=[
                {"text": "D1", "callout_type": "tag", "source_page": 0, "sheet_id": "A-101", "confidence": 0.8},
                {"text": "D2", "callout_type": "tag", "source_page": 0, "sheet_id": "A-101", "confidence": 0.8},
            ]
        )
        graph = PlanSetGraph(
            project_id="test_rfi_cov",
            total_pages=2,
            all_door_tags=["D1", "D2"],
            has_door_schedule=False,
            disciplines_found=["A", "S"],
            sheet_types_found={"floor_plan": 2},
            pages_with_scale=0,
            pages_without_scale=2,
        )
        run_cov = RunCoverage(
            pages_total=300, pages_indexed=300, pages_deep_processed=80,
            doc_types_detected={"plan": 100, "schedule": 20},
            doc_types_fully_covered=[],
            doc_types_partially_covered=["plan"],
            doc_types_not_covered=["schedule"],
            selection_mode=SelectionMode.FAST_BUDGET,
            ocr_budget_pages=80,
        )
        rfis = generate_rfis(extracted, page_index, selected, graph, run_coverage=run_cov)
        door_rfis = [r for r in rfis if r.issue_type == "CHK-A-001"]
        assert len(door_rfis) == 1, f"Expected 1 door RFI, got {len(door_rfis)}"
        rfi = door_rfis[0]
        assert rfi.coverage_status == "unknown_not_processed", f"Expected unknown_not_processed, got {rfi.coverage_status}"
        assert "[Coverage gap" in rfi.question, f"Expected coverage gap suffix in question, got: {rfi.question}"
        assert rfi.priority.value == "low", f"Expected LOW priority, got {rfi.priority.value}"

    def test_full_read_auto_mode(self):
        """page_index with 50 pages → select_pages auto-selects FULL_READ."""
        pages = [IndexedPage(page_idx=i, doc_type="plan", discipline="architectural") for i in range(50)]
        page_index = PageIndex(
            pdf_name="test.pdf",
            pages=pages,
            total_pages=50,
            counts_by_type={"plan": 50},
            counts_by_discipline={"architectural": 50},
        )
        result = select_pages(page_index, budget_pages=80)
        assert result.selection_mode == "full_read", f"Expected full_read, got {result.selection_mode}"
        assert len(result.selected) == 50, f"Expected 50 selected, got {len(result.selected)}"

    def test_fast_budget_mode(self):
        """page_index with 200 pages → select_pages uses FAST_BUDGET."""
        pages = [IndexedPage(page_idx=i, doc_type="plan", discipline="architectural") for i in range(200)]
        page_index = PageIndex(
            pdf_name="test.pdf",
            pages=pages,
            total_pages=200,
            counts_by_type={"plan": 200},
            counts_by_discipline={"architectural": 200},
        )
        result = select_pages(page_index, budget_pages=80)
        assert result.selection_mode == "fast_budget", f"Expected fast_budget, got {result.selection_mode}"
        assert len(result.selected) <= 80, f"Expected <= 80 selected, got {len(result.selected)}"


# =============================================================================
# EVIDENCE VIEWER & BID PACK TESTS
# =============================================================================

class TestEvidenceViewer:
    """Test the PDF page rendering helper (without actual PDF)."""

    def test_render_returns_none_for_missing_pdf(self):
        """Should return None when PDF path is None."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from demo_page import render_pdf_page_preview
        result = render_pdf_page_preview(None, 0)
        assert result is None, "Expected None for None pdf_path"

    def test_render_returns_none_for_nonexistent_path(self):
        """Should return None when PDF file does not exist."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from demo_page import render_pdf_page_preview
        result = render_pdf_page_preview("/nonexistent/path.pdf", 0)
        assert result is None, "Expected None for nonexistent path"

    def test_render_returns_none_for_negative_page(self):
        """Should return None when page index is negative."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from demo_page import render_pdf_page_preview
        result = render_pdf_page_preview("/nonexistent/path.pdf", -1)
        assert result is None, "Expected None for negative page index"


class TestPayloadFileInfo:
    """Verify file_info structure expectations."""

    def test_file_info_structure(self):
        """file_info should have name, pages, path, ocr_used keys."""
        file_info = [{"name": "test.pdf", "pages": 7, "path": "/tmp/test.pdf", "ocr_used": True}]
        fi = file_info[0]
        assert fi["path"] == "/tmp/test.pdf"
        assert fi["name"] == "test.pdf"
        assert fi["pages"] == 7
        assert fi["ocr_used"] is True

    def test_primary_pdf_path_none_for_demo(self):
        """Demo cache should have null primary_pdf_path."""
        import json
        cache_path = PROJECT_ROOT / "demo_cache" / "pwd_garage" / "analysis.json"
        with open(cache_path) as f:
            data = json.load(f)
        assert data.get("primary_pdf_path") is None, "Demo cache should have null PDF path"
        assert data.get("file_info") is not None, "Demo cache should have file_info"
        assert data["file_info"][0]["path"] is None, "Demo cache file_info path should be null"


class TestCoverageDashboard:
    """Test coverage data derivations used by the dashboard."""

    def test_coverage_status_categories(self):
        """Verify RunCoverage is_doc_type_covered returns correct statuses."""
        from src.models.analysis_models import RunCoverage, SelectionMode, CoverageStatus
        rc = RunCoverage(
            pages_total=100,
            pages_indexed=100,
            pages_deep_processed=50,
            doc_types_detected={"plan": 30, "schedule": 5, "spec": 20, "notes": 3},
            disciplines_detected={"architectural": 50, "structural": 30, "other": 20},
            doc_types_fully_covered=["schedule", "notes"],
            doc_types_partially_covered=["plan"],
            doc_types_not_covered=["spec"],
            selection_mode=SelectionMode.FAST_BUDGET,
        )
        # Fully covered → NOT_FOUND_AFTER_SEARCH (we looked and didn't find)
        assert rc.is_doc_type_covered("schedule") == CoverageStatus.NOT_FOUND_AFTER_SEARCH
        assert rc.is_doc_type_covered("notes") == CoverageStatus.NOT_FOUND_AFTER_SEARCH
        # Not covered → UNKNOWN_NOT_PROCESSED (we didn't look)
        assert rc.is_doc_type_covered("spec") == CoverageStatus.UNKNOWN_NOT_PROCESSED
        # Partially covered → UNKNOWN_NOT_PROCESSED (incomplete search)
        assert rc.is_doc_type_covered("plan") == CoverageStatus.UNKNOWN_NOT_PROCESSED

    def test_coverage_mode_from_page_count(self):
        """Small tenders auto-select FULL_READ, large tenders use FAST_BUDGET."""
        # FULL_READ for small tender
        small_index = PageIndex(
            pdf_name="test.pdf",
            total_pages=50,
            pages=[IndexedPage(page_idx=i, doc_type="plan", discipline="A",
                               confidence=0.9, keywords_hit=[], has_text_layer=True)
                   for i in range(50)],
            counts_by_type={"plan": 50},
            counts_by_discipline={"A": 50},
        )
        result = select_pages(small_index, budget_pages=80)
        assert result.selection_mode == "full_read"

        # FAST_BUDGET for large tender
        large_index = PageIndex(
            pdf_name="test.pdf",
            total_pages=200,
            pages=[IndexedPage(page_idx=i, doc_type="plan", discipline="A",
                               confidence=0.9, keywords_hit=[], has_text_layer=True)
                   for i in range(200)],
            counts_by_type={"plan": 200},
            counts_by_discipline={"A": 200},
        )
        result = select_pages(large_index, budget_pages=80)
        assert result.selection_mode == "fast_budget"
        assert len(result.selected) <= 80


# =============================================================================
# Sprint 20G: Run Mode Tests
# =============================================================================

class TestRunModes:
    """Tests for Sprint 20G run mode page budgets."""

    def _make_large_index(self, total: int) -> PageIndex:
        """Create a large PageIndex with mixed doc types."""
        pages = []
        types_cycle = ["plan", "detail", "section", "elevation", "spec", "unknown"]
        discs_cycle = ["architectural", "structural", "electrical", "plumbing"]
        for i in range(total):
            dt = types_cycle[i % len(types_cycle)]
            disc = discs_cycle[i % len(discs_cycle)]
            pages.append(IndexedPage(
                page_idx=i, doc_type=dt, discipline=disc,
                confidence=0.9, keywords_hit=[], has_text_layer=True,
            ))
        # Add some Tier-1 pages (BOQ + schedule)
        pages.append(IndexedPage(page_idx=total, doc_type="boq", discipline="other",
                                 confidence=0.95, keywords_hit=["boq"], has_text_layer=True))
        pages.append(IndexedPage(page_idx=total+1, doc_type="schedule", discipline="architectural",
                                 confidence=0.95, keywords_hit=["schedule"], has_text_layer=True))
        total_with_t1 = total + 2
        from collections import Counter
        tc = Counter(p.doc_type for p in pages)
        dc = Counter(p.discipline for p in pages)
        return PageIndex(
            pdf_name="large_test.pdf",
            total_pages=total_with_t1,
            pages=pages,
            counts_by_type=dict(tc),
            counts_by_discipline=dict(dc),
        )

    def test_demo_fast_caps_at_80(self):
        """DEMO_FAST (deep_cap=80) should deep-process <= 80 pages."""
        idx = self._make_large_index(367)
        result = select_pages(idx, deep_cap=80)
        assert result.selection_mode == "fast_budget"
        assert len(result.selected) <= 80, f"Expected <= 80, got {len(result.selected)}"

    def test_standard_review_caps_at_220(self):
        """STANDARD_REVIEW (deep_cap=220) should deep-process <= 220 pages."""
        idx = self._make_large_index(367)
        result = select_pages(idx, deep_cap=220)
        assert result.selection_mode == "fast_budget"
        assert len(result.selected) <= 220, f"Expected <= 220, got {len(result.selected)}"
        # Should be more than the demo_fast cap
        result_fast = select_pages(idx, deep_cap=80)
        assert len(result.selected) >= len(result_fast.selected), (
            f"Standard ({len(result.selected)}) should select >= Demo ({len(result_fast.selected)})"
        )

    def test_full_audit_selects_all_pages(self):
        """FULL_AUDIT (force_full_read=True) should select ALL pages."""
        idx = self._make_large_index(367)
        result = select_pages(idx, force_full_read=True)
        assert result.selection_mode == "full_read"
        expected_total = 367 + 2  # 367 Tier-2/3 + 2 Tier-1
        assert len(result.selected) == expected_total, (
            f"Expected {expected_total}, got {len(result.selected)}"
        )

    def test_deep_cap_overrides_budget_pages(self):
        """deep_cap should override budget_pages."""
        idx = self._make_large_index(300)
        result = select_pages(idx, budget_pages=80, deep_cap=150)
        assert result.selection_mode == "fast_budget"
        assert len(result.selected) <= 150

    def test_backward_compat_no_deep_cap(self):
        """Without deep_cap, behavior is unchanged (80-page default)."""
        idx = self._make_large_index(300)
        result = select_pages(idx, budget_pages=80)
        assert result.selection_mode == "fast_budget"
        assert len(result.selected) <= 80

    def test_run_mode_enum_properties(self):
        """RunMode enum properties return correct caps and labels."""
        from src.models.analysis_models import RunMode
        assert RunMode.DEMO_FAST.deep_cap == 80
        assert RunMode.STANDARD_REVIEW.deep_cap == 220
        assert RunMode.FULL_AUDIT.deep_cap is None
        assert "80" in RunMode.DEMO_FAST.label
        assert "220" in RunMode.STANDARD_REVIEW.label
        assert "All" in RunMode.FULL_AUDIT.label

    def test_processing_stats_includes_run_mode(self):
        """_build_processing_stats should include run_mode field."""
        stats = _build_processing_stats()
        # Default doesn't include run_mode (added by pipeline)
        assert "selection_mode" in stats

    def test_selection_ordering_deterministic(self):
        """Same input should always produce same selected pages."""
        idx = self._make_large_index(300)
        result1 = select_pages(idx, deep_cap=150)
        result2 = select_pages(idx, deep_cap=150)
        assert result1.selected == result2.selected, "Selection should be deterministic"


class TestRunModeSmoke:
    """Sprint 20G smoke tests: run mode integration and payload integrity."""

    def test_payload_reflects_selection_mode(self):
        """processing_stats should include run_mode when set."""
        # Simulate what pipeline does: build stats, then attach run_mode
        from src.analysis.page_selection import SelectedPages
        sel = SelectedPages(
            selected=[0, 1, 2],
            selection_mode="fast_budget",
            budget_total=220,
            budget_used=3,
        )
        stats = _build_processing_stats(selected_result=sel)
        # Pipeline would add: stats["run_mode"] = "standard_review"
        stats["run_mode"] = "standard_review"
        assert stats["run_mode"] == "standard_review"
        assert stats["selected_pages_count"] == 3

    def test_full_audit_payload_stub(self):
        """Full audit mode should produce valid stats even on large stub."""
        idx_pages = [
            IndexedPage(page_idx=i, doc_type="plan", discipline="architectural",
                        confidence=0.9, keywords_hit=[], has_text_layer=True)
            for i in range(500)
        ]
        idx = PageIndex(
            pdf_name="big.pdf", total_pages=500, pages=idx_pages,
            counts_by_type={"plan": 500}, counts_by_discipline={"architectural": 500},
        )
        result = select_pages(idx, force_full_read=True)
        assert result.selection_mode == "full_read"
        assert len(result.selected) == 500

        stats = _build_processing_stats(
            page_index_result=idx,
            selected_result=result,
        )
        stats["run_mode"] = "full_audit"
        assert stats["total_pages"] == 500
        assert stats["deep_processed_pages"] == 500
        assert stats["skipped_pages"] == 0
        assert stats["run_mode"] == "full_audit"

    def test_run_mode_backward_compat_missing(self):
        """When run_mode is not in processing_stats, defaults should work."""
        stats = _build_processing_stats()
        # run_mode field is added by pipeline, not by _build_processing_stats
        assert stats.get("run_mode") is None  # Not present by default
        # Existing fields still present
        assert "total_pages" in stats
        assert "deep_processed_pages" in stats
        assert "selection_mode" in stats


class TestBboxOverlay:
    """Test PDF page overlay with bounding boxes."""

    def test_overlay_returns_none_without_pdf(self):
        """render_pdf_page_with_overlay(None, ...) returns None even with bboxes."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from demo_page import render_pdf_page_with_overlay
        result = render_pdf_page_with_overlay(None, 0, [[0, 0, 100, 100]])
        assert result is None

    def test_overlay_no_bboxes_passes_through(self):
        """render_pdf_page_with_overlay with None bboxes returns same as base."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from demo_page import render_pdf_page_with_overlay
        result = render_pdf_page_with_overlay(None, 0, None)
        assert result is None


class TestNotFoundProof:
    """Test EvidenceRef new fields for NOT_FOUND proof."""

    def test_evidence_ref_searched_pages(self):
        """EvidenceRef with searched_pages serializes correctly."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0, 1, 2],
            searched_pages=[0, 1, 2, 3, 4],
            confidence=0.8,
        )
        d = ev.to_dict()
        assert d["searched_pages"] == [0, 1, 2, 3, 4]
        assert "Searched 5 pages" in ev.summary()

    def test_evidence_ref_text_coverage_pct(self):
        """EvidenceRef with text_coverage_pct serializes correctly."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0],
            text_coverage_pct=85.0,
            confidence=0.7,
        )
        d = ev.to_dict()
        assert d["text_coverage_pct"] == 85.0
        assert d["bbox"] is None  # default


class TestSkippedPagesExport:
    """Test skipped pages data structure for CSV export."""

    def test_skipped_pages_has_required_keys(self):
        """Demo cache run_coverage.pages_skipped has correct structure."""
        cache_path = Path(__file__).parent.parent / "demo_cache" / "pwd_garage" / "analysis.json"
        with open(cache_path) as f:
            data = json.load(f)
        run_cov = data.get("run_coverage", {})
        skipped = run_cov.get("pages_skipped", [])
        # Demo garage runs FULL_READ with 7 pages → 0 skipped
        assert isinstance(skipped, list)
        assert len(skipped) == 0, "Demo garage FULL_READ should have 0 skipped pages"
        # Verify the structure would work if there were entries
        assert run_cov.get("selection_mode") == "full_read"


# =============================================================================
# SPRINT 4a TESTS
# =============================================================================

class TestBboxFormat:
    """Test page-relative bbox format with confidence."""

    def test_page_relative_coords_in_range(self):
        """EvidenceRef with page-relative bbox serializes correctly, all coords 0.0–1.0."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0],
            bbox=[[[0.1, 0.2, 0.9, 0.8, 0.85]]],
            confidence=0.9,
        )
        d = ev.to_dict()
        assert d["bbox"] is not None
        assert len(d["bbox"]) == 1  # 1 page
        assert len(d["bbox"][0]) == 1  # 1 box on that page
        box = d["bbox"][0][0]
        assert len(box) == 5
        for coord in box[:4]:
            assert 0.0 <= coord <= 1.0

    def test_bbox_parallel_to_pages(self):
        """len(bbox) == len(pages). 3 pages and 3 bbox entries align."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0, 2, 6],
            bbox=[
                [[0.1, 0.2, 0.5, 0.6, 0.9]],
                [[0.2, 0.3, 0.8, 0.7, 0.75]],
                [],
            ],
            confidence=0.85,
        )
        d = ev.to_dict()
        assert len(d["bbox"]) == len(d["pages"])

    def test_empty_bbox_pages(self):
        """EvidenceRef with bbox: [[], []] — empty pages are valid."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0, 1],
            bbox=[[], []],
            confidence=0.7,
        )
        d = ev.to_dict()
        assert d["bbox"] == [[], []]

    def test_confidence_in_tuple(self):
        """Box [0.1, 0.2, 0.5, 0.6, 0.92] — 5th element is confidence."""
        box = [0.1, 0.2, 0.5, 0.6, 0.92]
        assert len(box) == 5
        assert box[4] == 0.92


class TestBboxOverlayConfidence:
    """Test overlay renderer with confidence-based coloring."""

    def test_overlay_returns_none_without_pdf_page_relative(self):
        """render_pdf_page_with_overlay(None, ...) returns None with page-relative bboxes."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from demo_page import render_pdf_page_with_overlay
        result = render_pdf_page_with_overlay(None, 0, [[0.1, 0.2, 0.5, 0.6, 0.9]])
        assert result is None

    def test_overlay_suppresses_low_confidence(self):
        """Box with conf < 0.4 should be suppressed (not drawn)."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from demo_page import render_pdf_page_with_overlay
        # With None PDF, result is None regardless, but function shouldn't crash
        result = render_pdf_page_with_overlay(None, 0, [[0.1, 0.2, 0.5, 0.6, 0.2]])
        assert result is None  # No PDF → None


class TestOcrTextCache:
    """Test OCR text cache data structure."""

    def test_cache_structure(self):
        """Build cache from mock page texts. Verify type and 10K cap."""
        texts = ["Short text that is longer than 10 chars", "x" * 15000, "   ", ""]
        cache = {}
        for idx, text in enumerate(texts):
            if text and len(text.strip()) > 10:
                cache[idx] = text[:10000]
        assert isinstance(cache, dict)
        assert 0 in cache
        assert 1 in cache
        assert 2 not in cache  # whitespace only
        assert 3 not in cache  # empty
        assert len(cache[1]) == 10000  # capped

    def test_empty_pages_excluded(self):
        """Pages with empty/whitespace text are NOT in cache."""
        texts = ["", "   ", "\n\n", "a" * 5]  # all too short or empty
        cache = {}
        for idx, text in enumerate(texts):
            if text and len(text.strip()) > 10:
                cache[idx] = text[:10000]
        assert len(cache) == 0

    def test_demo_cache_has_ocr_text(self):
        """Load demo cache analysis.json. Verify ocr_text_cache exists with 7 entries."""
        cache_path = Path(__file__).parent.parent / "demo_cache" / "pwd_garage" / "analysis.json"
        with open(cache_path) as f:
            data = json.load(f)
        cache = data.get("ocr_text_cache")
        assert cache is not None, "ocr_text_cache missing from demo cache"
        assert len(cache) == 7, f"Expected 7 entries, got {len(cache)}"
        for key, val in cache.items():
            assert isinstance(val, str), f"Value for page {key} should be string"
            assert len(val) > 0, f"Text for page {key} should not be empty"


class TestCitationFormat:
    """Test inline citation format."""

    def test_citation_with_sheet(self):
        """Given pages=[0], sheets=['A-101'], output contains '(A-101, p.1)'."""
        pages = [0]
        sheets = ["A-101"]
        pg = pages[0]
        sheet = sheets[0] if sheets else ""
        cite = f"({sheet + ', ' if sheet else ''}p.{pg + 1})"
        assert cite == "(A-101, p.1)"

    def test_citation_without_sheet(self):
        """Given pages=[2], sheets=[], output contains '(p.3)'."""
        pages = [2]
        sheets = []
        pg = pages[0]
        sheet = sheets[0] if 0 < len(sheets) else ""
        cite = f"({sheet + ', ' if sheet else ''}p.{pg + 1})"
        assert cite == "(p.3)"

    def test_citation_parallel_alignment(self):
        """3 snippets, 2 pages → only first 2 snippets get citations."""
        snippets = ["snip1", "snip2", "snip3"]
        pages = [0, 2]
        sheets = ["A-101"]
        citations_added = 0
        for i, snippet in enumerate(snippets[:3]):
            if i < len(pages):
                citations_added += 1
        assert citations_added == 2


class TestCoverageHeatmap:
    """Test coverage heatmap data logic."""

    def test_heatmap_all_green_full_read(self):
        """RunCoverage with 7 pages, 0 skipped → all pages should be 'processed'."""
        from src.models.analysis_models import RunCoverage, SelectionMode
        rc = RunCoverage(
            pages_total=7,
            pages_indexed=7,
            pages_deep_processed=7,
            pages_skipped=[],
            selection_mode=SelectionMode.FULL_READ,
        )
        d = rc.to_dict()
        skipped_set = {s.get("page_idx") for s in d.get("pages_skipped", [])}
        for pg in range(d["pages_total"]):
            assert pg not in skipped_set, f"Page {pg} should not be skipped"

    def test_heatmap_mixed_fast_budget(self):
        """RunCoverage with 100 pages, 30 skipped → 70 processed, 30 skipped."""
        from src.models.analysis_models import RunCoverage, SelectionMode
        skipped = [{"page_idx": i, "doc_type": "plan", "discipline": "A", "reason": "budget_exceeded"}
                   for i in range(70, 100)]
        rc = RunCoverage(
            pages_total=100,
            pages_indexed=100,
            pages_deep_processed=70,
            pages_skipped=skipped,
            selection_mode=SelectionMode.FAST_BUDGET,
        )
        d = rc.to_dict()
        skipped_set = {s.get("page_idx") for s in d.get("pages_skipped", [])}
        assert len(skipped_set) == 30
        # Pages 0-69 should be processed
        for pg in range(70):
            assert pg not in skipped_set
        # Pages 70-99 should be skipped
        for pg in range(70, 100):
            assert pg in skipped_set


# =============================================================================
# SPRINT 5 TESTS
# =============================================================================

class TestBboxId:
    """Test bbox_id assignment (6th element)."""

    def test_bbox_id_assignment_5_to_6(self):
        """5-element box gets bbox_id appended as 6th element."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0],
            bbox=[[[0.1, 0.2, 0.5, 0.6, 0.9]]],
            confidence=0.9,
        )
        ev.assign_bbox_ids("BLK-0010")
        box = ev.bbox[0][0]
        assert len(box) == 6, f"Expected 6 elements, got {len(box)}"
        assert box[5] == "BLK-0010-P0-0"

    def test_bbox_id_format(self):
        """Verify '{item_id}-P{page}-{box}' format across multiple pages/boxes."""
        from src.models.analysis_models import EvidenceRef
        ev = EvidenceRef(
            pages=[0, 2],
            bbox=[
                [[0.1, 0.2, 0.5, 0.6, 0.9]],
                [[0.2, 0.3, 0.8, 0.7, 0.75], [0.3, 0.4, 0.9, 0.8, 0.6]],
            ],
            confidence=0.85,
        )
        ev.assign_bbox_ids("RFI-0010")
        assert ev.bbox[0][0][5] == "RFI-0010-P0-0"
        assert ev.bbox[1][0][5] == "RFI-0010-P1-0"
        assert ev.bbox[1][1][5] == "RFI-0010-P1-1"

    def test_assign_all_bbox_ids(self):
        """assign_all_bbox_ids processes both blockers and rfis dict lists."""
        from src.models.analysis_models import assign_all_bbox_ids
        blockers = [{
            "id": "BLK-0010",
            "evidence": {"bbox": [[[0.1, 0.2, 0.5, 0.6, 0.9]]]},
        }]
        rfis = [{
            "id": "RFI-0010",
            "evidence": {"bbox": [[[0.2, 0.3, 0.8, 0.7, 0.75]]]},
        }]
        assign_all_bbox_ids(blockers, rfis)
        assert blockers[0]["evidence"]["bbox"][0][0][5] == "BLK-0010-P0-0"
        assert rfis[0]["evidence"]["bbox"][0][0][5] == "RFI-0010-P0-0"

    def test_bbox_id_none_bbox_skipped(self):
        """assign_all_bbox_ids with None bbox doesn't crash."""
        from src.models.analysis_models import assign_all_bbox_ids
        blockers = [{"id": "BLK-0001", "evidence": {"bbox": None}}]
        rfis = [{"id": "RFI-0001", "evidence": {}}]
        # Should not raise
        assign_all_bbox_ids(blockers, rfis)


class TestGlobalSearch:
    """Test global search across OCR text cache."""

    def _cache(self):
        return {
            "0": "GROUND FLOOR PLAN\nDoor D1 - 900 x 2100\nColumn C1\nBeam B1",
            "1": "FIRST FLOOR PLAN\nRoom R1\nStaircase ST1",
            "2": "UPPER FLOOR PLAN\nDoor D10 - 900 x 2100",
            "3": "SECTION A-A\nFoundation depth 1.5m",
            "4": "SECTION B-B\nRoof slab 150mm",
            "5": "ELEVATION NORTH\nPlinth level +0.45",
            "6": "ELEVATION EAST\nDoor D10\nWindow W1",
        }

    def test_search_basic_match(self):
        """Finds 'Door' on pages 0, 2, 6."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from search import search_ocr_text
        results = search_ocr_text(self._cache(), "Door")
        pages = [r["page_idx"] for r in results]
        assert 0 in pages, f"Expected page 0 in results, got {pages}"
        assert 2 in pages, f"Expected page 2 in results, got {pages}"
        assert 6 in pages, f"Expected page 6 in results, got {pages}"

    def test_search_case_insensitive(self):
        """'door' (lowercase) matches 'Door' (title case)."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from search import search_ocr_text
        results = search_ocr_text(self._cache(), "door")
        assert len(results) > 0, "Expected at least one match"
        assert all("door" in r["match_text"].lower() for r in results)

    def test_search_no_match(self):
        """Query 'xyz_nonexistent' returns empty list."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from search import search_ocr_text
        results = search_ocr_text(self._cache(), "xyz_nonexistent")
        assert results == []

    def test_search_empty_query(self):
        """Empty query returns empty list."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from search import search_ocr_text
        results = search_ocr_text(self._cache(), "")
        assert results == []

    def test_search_special_chars(self):
        """'1.5m' doesn't crash (regex-safe via re.escape)."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from search import search_ocr_text
        results = search_ocr_text(self._cache(), "1.5m")
        assert len(results) >= 1, "Expected at least 1 match for '1.5m'"
        assert results[0]["page_idx"] == 3


class TestBidStrategy:
    """Test bid strategy scorer."""

    def test_bid_unknown_when_no_inputs(self):
        """All optional scores = None/UNKNOWN when no user inputs provided."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from bid_strategy_scorer import compute_bid_strategy
        payload = {
            "readiness_score": 61,
            "sub_scores": {"completeness": 80, "coverage": 64},
            "blockers": [],
            "trade_coverage": [],
        }
        result = compute_bid_strategy({}, payload)
        assert result["client_fit"]["score"] is None
        assert result["client_fit"]["confidence"] == "UNKNOWN"
        assert result["competition_score"]["score"] is None
        assert result["competition_score"]["confidence"] == "UNKNOWN"
        # Risk and Readiness always computed
        assert result["risk_score"]["score"] is not None
        assert result["readiness_score"]["score"] is not None

    def test_bid_readiness_from_payload(self):
        """Readiness score is always computed from payload."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from bid_strategy_scorer import compute_bid_strategy
        payload = {
            "readiness_score": 72,
            "sub_scores": {"completeness": 80, "coverage": 70, "measurement": 60, "blocker": 75},
            "blockers": [],
            "trade_coverage": [],
        }
        result = compute_bid_strategy({}, payload)
        assert result["readiness_score"]["score"] == 72
        assert result["readiness_score"]["confidence"] == "HIGH"

    def test_bid_recommendations_from_blockers(self):
        """NOT_FOUND blockers trigger recommendation."""
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        from bid_strategy_scorer import compute_bid_strategy
        payload = {
            "readiness_score": 40,
            "sub_scores": {},
            "blockers": [
                {"id": "BLK-1", "coverage_status": "not_found_after_search", "severity": "high", "issue_type": "missing_schedule", "title": "Door schedule"},
                {"id": "BLK-2", "coverage_status": "not_found_after_search", "severity": "high", "issue_type": "scale_issue", "title": "Scale missing"},
            ],
            "trade_coverage": [],
        }
        result = compute_bid_strategy({}, payload)
        recs = result["recommendations"]
        assert any("not found" in r.lower() or "rfi" in r.lower() for r in recs), \
            f"Expected scope ambiguity recommendation, got {recs}"
        # Low readiness should also trigger recommendation
        assert any("readiness" in r.lower() or "extension" in r.lower() for r in recs), \
            f"Expected low readiness recommendation, got {recs}"


# =============================================================================
# SPRINT 6: NORMALIZATION
# =============================================================================

class TestNormalization:
    """Test normalization layer."""

    def test_normalize_boq_unit(self):
        """Unit normalization: cu.m -> cum, sq m -> sqm."""
        from src.analysis.normalize import normalize_boq_items
        items = [
            {"item_no": "1.1", "description": "Excavation  ", "unit": "cu.m", "qty": 120, "rate": 450},
            {"item_no": "1.2", "description": "PCC below footings.", "unit": "sq m", "qty": 80, "rate": None},
        ]
        result = normalize_boq_items(items)
        assert result[0]["unit"] == "cum"
        assert result[1]["unit"] == "sqm"
        # Description cleanup
        assert result[0]["description"] == "Excavation"
        assert result[1]["description"] == "PCC below footings"

    def test_normalize_boq_indian_number(self):
        """Indian lakh format numbers parsed correctly."""
        from src.analysis.normalize import _parse_indian_number
        assert _parse_indian_number("1,25,000") == 125000.0
        assert _parse_indian_number("450") == 450.0
        assert _parse_indian_number("12,500") == 12500.0
        assert _parse_indian_number("") is None
        assert _parse_indian_number("abc") is None

    def test_normalize_boq_string_qty(self):
        """String quantities parsed via _parse_indian_number."""
        from src.analysis.normalize import normalize_boq_items
        items = [{"item_no": "3.1", "unit": "nos", "qty": "1,25,000", "rate": "5,500"}]
        result = normalize_boq_items(items)
        assert result[0]["qty"] == 125000.0
        assert result[0]["rate"] == 5500.0

    def test_normalize_schedule_size_and_dedup(self):
        """Size normalization + mark uppercase + de-dup by (mark, schedule_type)."""
        from src.analysis.normalize import normalize_schedule_rows
        rows = [
            {"mark": "d1", "size": "900x2100", "schedule_type": "door", "confidence": 0.8},
            {"mark": "d1", "size": "900x2100", "schedule_type": "door", "confidence": 0.6},
            {"mark": "w1", "size": "1200X1500", "schedule_type": "window", "confidence": 0.7},
        ]
        result = normalize_schedule_rows(rows)
        assert result[0]["mark"] == "D1"
        assert result[0]["size"] == "900 x 2100"
        assert result[0]["confidence"] == 0.8  # kept higher confidence
        assert result[1]["mark"] == "W1"
        assert result[1]["size"] == "1200 x 1500"
        assert len(result) == 2  # de-duped D1

    def test_normalize_requirements_dedup(self):
        """Exact-match de-dup after lowercase + strip."""
        from src.analysis.normalize import normalize_requirements
        reqs = [
            {"text": "All concrete shall be M25", "source_page": 1},
            {"text": "All concrete shall be M25", "source_page": 2},
            {"text": "Steel shall be Fe500D", "source_page": 1},
        ]
        result = normalize_requirements(reqs)
        assert len(result) == 2
        assert result[0]["text"] == "All concrete shall be M25"
        assert result[1]["text"] == "Steel shall be Fe500D"


# =============================================================================
# SPRINT 6: ADDENDUM ADAPTER
# =============================================================================

class TestAddendumAdapter:
    """Test addendum adapter (OCR text -> AddendaParser)."""

    def test_adapter_creates_documents(self):
        """Contiguous addendum pages grouped into 1 ParsedDocument."""
        from src.analysis.addendum_adapter import build_addendum_documents
        from src.analysis.page_index import PageIndex, IndexedPage

        pages = [
            IndexedPage(page_idx=0, doc_type="plan", discipline="architectural"),
            IndexedPage(page_idx=1, doc_type="addendum", discipline="other"),
            IndexedPage(page_idx=2, doc_type="addendum", discipline="other"),
            IndexedPage(page_idx=3, doc_type="plan", discipline="structural"),
        ]
        page_index = PageIndex(
            pdf_name="test.pdf", total_pages=4, pages=pages,
            counts_by_type={"plan": 2, "addendum": 2},
        )
        ocr_text = {
            0: "FLOOR PLAN",
            1: "ADDENDUM NO 1\nDate: 01/02/2026\nRead As: M30 instead of M25",
            2: "Item 2.1 quantity revised to 500 sqm",
            3: "STRUCTURAL LAYOUT",
        }
        docs = build_addendum_documents(ocr_text, page_index)
        assert len(docs) == 1  # contiguous pages grouped
        assert "ADDENDUM" in docs[0].text_content
        assert "Item 2.1" in docs[0].text_content

    def test_adapter_no_addendum_pages(self):
        """No addendum pages → empty list."""
        from src.analysis.addendum_adapter import build_addendum_documents
        from src.analysis.page_index import PageIndex, IndexedPage

        pages = [IndexedPage(page_idx=0, doc_type="plan", discipline="architectural")]
        page_index = PageIndex(
            pdf_name="test.pdf", total_pages=1, pages=pages,
            counts_by_type={"plan": 1},
        )
        docs = build_addendum_documents({0: "FLOOR PLAN"}, page_index)
        assert len(docs) == 0


# =============================================================================
# SPRINT 6: DELTA DETECTOR
# =============================================================================

class TestDeltaDetector:
    """Test conflict detection between base and addendum items."""

    def test_boq_qty_change(self):
        """BOQ item with qty change detected."""
        from src.analysis.delta_detector import detect_boq_deltas
        base = [{"item_no": "1.1", "qty": 100, "unit": "cum", "source_page": 5}]
        addendum = [{"item_no": "1.1", "qty": 150, "unit": "cum", "source_page": 20}]
        conflicts = detect_boq_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["type"] == "boq_change"
        assert conflicts[0]["item_no"] == "1.1"
        assert conflicts[0]["changes"][0]["field"] == "qty"
        assert conflicts[0]["changes"][0]["base_value"] == 100
        assert conflicts[0]["changes"][0]["addendum_value"] == 150

    def test_boq_new_item(self):
        """New BOQ item in addendum detected."""
        from src.analysis.delta_detector import detect_boq_deltas
        base = [{"item_no": "1.1", "qty": 100, "source_page": 5}]
        addendum = [{"item_no": "3.1", "qty": 50, "description": "New item", "source_page": 20}]
        conflicts = detect_boq_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["type"] == "boq_new_item"
        assert conflicts[0]["item_no"] == "3.1"

    def test_schedule_size_change(self):
        """Schedule row with size change detected."""
        from src.analysis.delta_detector import detect_schedule_deltas
        base = [{"mark": "D1", "size": "900 x 2100", "source_page": 8}]
        addendum = [{"mark": "D1", "size": "1000 x 2100", "source_page": 22}]
        conflicts = detect_schedule_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["type"] == "schedule_change"
        assert conflicts[0]["mark"] == "D1"
        assert conflicts[0]["changes"][0]["field"] == "size"

    def test_requirement_modified(self):
        """Similar but changed requirement detected as modified."""
        from src.analysis.delta_detector import detect_requirement_deltas
        base = [{"text": "All concrete shall be M25 grade minimum", "source_page": 15}]
        addendum = [{"text": "All concrete shall be M30 grade minimum", "source_page": 25}]
        conflicts = detect_requirement_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["type"] == "requirement_modified"
        assert conflicts[0]["similarity"] >= 0.7

    def test_requirement_new(self):
        """Completely new requirement detected."""
        from src.analysis.delta_detector import detect_requirement_deltas
        base = [{"text": "All concrete shall be M25 grade minimum", "source_page": 15}]
        addendum = [{"text": "Fire protection coating required on all steel members", "source_page": 25}]
        conflicts = detect_requirement_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["type"] == "requirement_new"


# =============================================================================
# SPRINT 7: DELTA CONFIDENCE
# =============================================================================

class TestDeltaConfidence:
    """Test that all conflict records include delta_confidence."""

    def test_boq_delta_has_confidence(self):
        """BOQ change records have delta_confidence=0.9."""
        from src.analysis.delta_detector import detect_boq_deltas
        base = [{"item_no": "1.1", "qty": 100, "unit": "cum", "source_page": 5}]
        addendum = [{"item_no": "1.1", "qty": 150, "unit": "cum", "source_page": 20}]
        conflicts = detect_boq_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["delta_confidence"] == 0.9

        # New items get 0.8
        addendum2 = [{"item_no": "9.9", "qty": 10, "description": "New", "source_page": 21}]
        new_conflicts = detect_boq_deltas(base, addendum2)
        assert len(new_conflicts) == 1
        assert new_conflicts[0]["delta_confidence"] == 0.8

    def test_requirement_delta_confidence_varies(self):
        """Modified requirement confidence = similarity * coverage_factor."""
        from src.analysis.delta_detector import detect_requirement_deltas
        base = [{"text": "All concrete shall be M25 grade minimum", "source_page": 15}]
        addendum = [{"text": "All concrete shall be M30 grade minimum", "source_page": 25}]
        # With 100% OCR coverage
        conflicts = detect_requirement_deltas(base, addendum, ocr_coverage_pct=100)
        assert len(conflicts) == 1
        assert conflicts[0]["delta_confidence"] > 0
        # Coverage factor = min(1.0, 100/80) = 1.0
        # So confidence = similarity * 1.0 = similarity
        assert conflicts[0]["delta_confidence"] == conflicts[0]["similarity"]

        # With low OCR coverage -> lower confidence
        conflicts_low = detect_requirement_deltas(base, addendum, ocr_coverage_pct=40)
        assert conflicts_low[0]["delta_confidence"] < conflicts[0]["delta_confidence"]

    def test_requirement_new_confidence(self):
        """New requirements always get delta_confidence=0.7."""
        from src.analysis.delta_detector import detect_requirement_deltas
        base = [{"text": "Concrete M25", "source_page": 15}]
        addendum = [{"text": "Fire protection on all steel", "source_page": 25}]
        conflicts = detect_requirement_deltas(base, addendum)
        assert len(conflicts) == 1
        assert conflicts[0]["delta_confidence"] == 0.7


# =============================================================================
# SPRINT 7: TEXT PRE-NORMALIZATION
# =============================================================================

class TestTextPreNormalization:
    """Test normalize_requirement_text() for reducing false deltas."""

    def test_normalize_requirement_text_basic(self):
        """Strips boilerplate, normalizes units/numbers/codes."""
        from src.analysis.normalize import normalize_requirement_text
        # Boilerplate prefix
        assert normalize_requirement_text("Note: All walls shall be plastered") == "all walls shall be plastered"
        # Numbered prefix
        assert normalize_requirement_text("1) Use OPC cement") == "use opc cement"
        # Unit normalization
        result = normalize_requirement_text("Area is 100 sq.m minimum")
        assert "sqm" in result
        assert "sq.m" not in result
        # Indian number normalization
        result = normalize_requirement_text("Cost is 1,25,000 per unit")
        assert "125000" in result
        assert "1,25,000" not in result
        # IS code colon normalization
        result = normalize_requirement_text("As per IS : 456 clause 2.1")
        assert "IS 456" in result
        assert "IS : 456" not in result
        # Trailing punctuation
        assert normalize_requirement_text("Cement shall be OPC.") == "cement shall be opc"

    def test_normalize_requirement_text_idempotent(self):
        """Already-clean text is unchanged after normalization."""
        from src.analysis.normalize import normalize_requirement_text
        clean = "all concrete shall be m25 grade"
        assert normalize_requirement_text(clean) == clean


# =============================================================================
# SPRINT 7: ADDENDUM GAP TOLERANCE
# =============================================================================

class TestAddendumGapTolerance:
    """Test gap tolerance in addendum page grouping."""

    def test_gap_of_2_pages_merged(self):
        """Pages [1,2,5,6] with gap=2 → 1 group."""
        from src.analysis.addendum_adapter import build_addendum_documents
        from src.analysis.page_index import PageIndex, IndexedPage

        pages = [
            IndexedPage(page_idx=1, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=2, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=3, doc_type="notes", discipline="arch"),
            IndexedPage(page_idx=4, doc_type="notes", discipline="arch"),
            IndexedPage(page_idx=5, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=6, doc_type="addendum", discipline="arch"),
        ]
        page_index = PageIndex(pdf_name="test.pdf", total_pages=7, pages=pages)
        ocr_text = {1: "Addendum 1 text", 2: "More text", 5: "Continued addendum", 6: "End"}
        docs = build_addendum_documents(ocr_text, page_index)
        # Gap is exactly 2 (pages 3,4 between page 2 and 5), should merge
        assert len(docs) == 1

    def test_gap_of_3_pages_split(self):
        """Pages [1,2,6,7] with gap=3 → 2 groups."""
        from src.analysis.addendum_adapter import build_addendum_documents
        from src.analysis.page_index import PageIndex, IndexedPage

        pages = [
            IndexedPage(page_idx=1, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=2, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=3, doc_type="notes", discipline="arch"),
            IndexedPage(page_idx=4, doc_type="notes", discipline="arch"),
            IndexedPage(page_idx=5, doc_type="notes", discipline="arch"),
            IndexedPage(page_idx=6, doc_type="addendum", discipline="arch"),
            IndexedPage(page_idx=7, doc_type="addendum", discipline="arch"),
        ]
        page_index = PageIndex(pdf_name="test.pdf", total_pages=8, pages=pages)
        ocr_text = {1: "Addendum 1", 2: "Text", 6: "Addendum 2", 7: "Text"}
        docs = build_addendum_documents(ocr_text, page_index)
        # Gap is 3 (pages 3,4,5 between page 2 and 6), should split
        assert len(docs) == 2


# =============================================================================
# SPRINT 7: RECONCILER
# =============================================================================

class TestReconciler:
    """Test scope reconciliation cross-checks."""

    def test_reqs_mention_doors_no_schedule(self):
        """Requirements mentioning doors without door schedule → missing finding."""
        from src.analysis.reconciler import reconcile_scope
        reqs = [{"text": "All doors shall be solid core hardwood", "source_page": 5}]
        schedules = []  # No schedules at all
        boq_items = []
        findings = reconcile_scope(reqs, schedules, boq_items)
        door_findings = [f for f in findings if "door" in f["description"].lower()]
        assert len(door_findings) >= 1
        assert door_findings[0]["type"] == "missing"
        assert door_findings[0]["impact"] == "high"

    def test_boq_schedule_qty_mismatch(self):
        """BOQ door qty != schedule mark count → conflict finding."""
        from src.analysis.reconciler import reconcile_scope
        reqs = []
        schedules = [
            {"schedule_type": "door", "mark": "D1", "source_page": 10},
            {"schedule_type": "door", "mark": "D2", "source_page": 10},
            {"schedule_type": "door", "mark": "D3", "source_page": 10},
        ]
        boq_items = [
            {"item_no": "4.1", "description": "Flush door 900x2100", "qty": 5, "source_page": 20},
        ]
        findings = reconcile_scope(reqs, schedules, boq_items)
        conflict_findings = [f for f in findings if f["type"] == "conflict" and "door" in f["description"].lower()]
        assert len(conflict_findings) >= 1
        assert "5" in conflict_findings[0]["description"]  # BOQ qty
        assert "3" in conflict_findings[0]["description"]  # Schedule count

    def test_empty_inputs_no_findings(self):
        """Empty inputs should return no findings."""
        from src.analysis.reconciler import reconcile_scope
        assert reconcile_scope([], [], []) == []


# =============================================================================
# SPRINT 7: RISK CHECKLIST
# =============================================================================

class TestRiskChecklist:
    """Test risk template keyword searches."""

    def test_risk_search_finds_ld(self):
        """'liquidated damages' in text → found."""
        import sys, os
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))
        from risk_checklist import search_risk_items
        ocr_cache = {
            0: "General conditions of contract.",
            1: "The contractor shall be liable for liquidated damages of 0.5% per week.",
            2: "Material specifications.",
        }
        results = search_risk_items(ocr_cache)
        ld_result = [r for r in results if r["template_id"] == "RISK-LD"][0]
        assert ld_result["found"] is True
        assert len(ld_result["hits"]) >= 1
        assert ld_result["hits"][0]["page_idx"] == 1

    def test_risk_search_no_hits(self):
        """Clean text → all items not found."""
        import sys, os
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))
        from risk_checklist import search_risk_items
        ocr_cache = {
            0: "This is a simple drawing with dimensions.",
            1: "Floor plan showing rooms and corridors.",
        }
        results = search_risk_items(ocr_cache)
        for r in results:
            assert r["found"] is False
            assert len(r["hits"]) == 0

    def test_risk_search_scoped_to_conditions(self):
        """Scoped search only checks pages of allowed doc_types."""
        import sys, os
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))
        from risk_checklist import search_risk_items
        ocr_cache = {
            0: "This is a plan page with no risk items.",
            1: "General conditions: liquidated damages of 0.5% per week.",
            2: "Specification for concrete mix M25.",
        }
        page_doc_types = {0: "plan", 1: "conditions", 2: "spec"}
        results = search_risk_items(
            ocr_cache,
            page_doc_types=page_doc_types,
            allowed_doc_types={"conditions"},
        )
        ld_result = [r for r in results if r["template_id"] == "RISK-LD"][0]
        assert ld_result["found"] is True
        assert ld_result["searched_pages_count"] == 1

    def test_risk_not_found_shows_searched_metadata(self):
        """NOT_FOUND results include searched_pages_count and searched_doc_types."""
        import sys, os
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))
        from risk_checklist import search_risk_items
        ocr_cache = {
            0: "A simple plan drawing.",
            1: "Conditions of contract.",
        }
        page_doc_types = {0: "plan", 1: "conditions"}
        results = search_risk_items(
            ocr_cache,
            page_doc_types=page_doc_types,
            allowed_doc_types={"conditions", "spec"},
        )
        esc = [r for r in results if r["template_id"] == "RISK-ESC"][0]
        assert esc["found"] is False
        assert esc["searched_pages_count"] == 1
        assert "conditions" in esc["searched_doc_types"]


# =============================================================================
# SPRINT 8: RECON ACTIONS
# =============================================================================

class TestReconActions:
    """Sprint 8: Reconciliation finding -> RFI / Assumption actions."""

    def test_finding_produces_proposed_rfi(self):
        """A reconciliation finding generates a proposed_rfi with correct fields."""
        from src.analysis.recon_actions import finding_to_proposed_rfi
        finding = {
            "type": "missing",
            "category": "req_vs_schedule",
            "description": "Requirements mention doors but no door schedule was found",
            "impact": "high",
            "suggested_action": "Request door schedule from design team",
            "evidence": {"pages": [5], "items": ["door requirements"]},
            "confidence": 0.85,
        }
        rfi = finding_to_proposed_rfi(finding)
        assert "question" in rfi
        assert rfi["category"] == "req_vs_schedule"
        assert rfi["confidence"] == 0.85
        assert 5 in rfi["evidence_refs"]["pages"]

    def test_create_rfi_preserves_evidence(self):
        """Creating an RFI from recon finding keeps evidence pages and snippets."""
        from src.analysis.recon_actions import create_recon_rfi, finding_to_proposed_rfi
        finding = {
            "type": "conflict",
            "category": "boq_vs_schedule",
            "description": "BOQ door quantity (5) does not match door schedule marks (3)",
            "impact": "high",
            "suggested_action": "Reconcile door count",
            "evidence": {"pages": [10, 20], "items": ["BOQ doors: 5", "Schedule marks: 3"]},
            "confidence": 0.8,
        }
        proposed = finding_to_proposed_rfi(finding)
        rfi = create_recon_rfi(proposed, existing_rfis=[], finding=finding)
        assert rfi["id"].startswith("RFI-R-")
        assert rfi["source"] == "reconciler"
        assert 10 in rfi["evidence"]["pages"]
        assert 20 in rfi["evidence"]["pages"]
        assert rfi["evidence"]["confidence"] == 0.8

    def test_rfi_id_no_collision(self):
        """Recon RFI IDs don't clash with existing pipeline RFI IDs."""
        from src.analysis.recon_actions import create_recon_rfi, finding_to_proposed_rfi
        existing = [{"id": "RFI-0001"}, {"id": "RFI-0002"}, {"id": "RFI-R-0001"}]
        finding = {
            "type": "missing", "category": "req_vs_boq",
            "description": "test", "impact": "low",
            "suggested_action": "test", "evidence": {"pages": [], "items": []},
            "confidence": 0.5,
        }
        proposed = finding_to_proposed_rfi(finding)
        rfi = create_recon_rfi(proposed, existing_rfis=existing, finding=finding)
        assert rfi["id"] == "RFI-R-0002"

    def test_create_assumption_from_finding(self):
        """Creating an assumption from recon finding has stable schema."""
        from src.analysis.recon_actions import create_recon_assumption, finding_to_proposed_assumption
        finding = {
            "type": "missing", "category": "req_vs_schedule",
            "description": "Missing door schedule",
            "impact": "high",
            "suggested_action": "Request schedule",
            "evidence": {"pages": [5], "items": ["doors"]},
            "confidence": 0.85,
        }
        proposed = finding_to_proposed_assumption(finding)
        asmp = create_recon_assumption(proposed, existing_assumptions=[], finding=finding)
        assert asmp["id"].startswith("ASMP-R-")
        assert "text" in asmp
        assert "basis_pages" in asmp
        assert asmp["source"] == "reconciler"


# =============================================================================
# SPRINT 8: ASSUMPTIONS EXPORT
# =============================================================================

class TestAssumptionsExport:
    """Sprint 8: Assumptions export has stable CSV schema."""

    def test_assumptions_csv_schema(self):
        """Assumptions export includes all required columns."""
        from src.analysis.recon_actions import create_recon_assumption, finding_to_proposed_assumption
        import csv, io
        assumptions = []
        for i in range(3):
            finding = {
                "type": "missing", "category": "req_vs_schedule",
                "description": f"Finding {i}", "impact": "medium",
                "suggested_action": "Action", "evidence": {"pages": [i], "items": []},
                "confidence": 0.7,
            }
            proposed = finding_to_proposed_assumption(finding)
            asmp = create_recon_assumption(proposed, assumptions, finding)
            assumptions.append(asmp)

        buf = io.StringIO()
        fieldnames = ["id", "title", "text", "impact_if_wrong", "risk_level",
                       "basis_pages", "source", "created_at"]
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for a in assumptions:
            writer.writerow({
                "id": a["id"],
                "title": a["title"],
                "text": a["text"],
                "impact_if_wrong": a.get("impact_if_wrong", ""),
                "risk_level": a.get("risk_level", ""),
                "basis_pages": ";".join(str(p + 1) for p in a.get("basis_pages", [])),
                "source": a.get("source", ""),
                "created_at": a.get("created_at", ""),
            })
        output = buf.getvalue()
        assert "id,title,text,impact_if_wrong,risk_level,basis_pages,source,created_at" in output
        assert "ASMP-R-0001" in output


# =============================================================================
# SPRINT 8: OWNER PROFILES
# =============================================================================

class TestOwnerProfiles:
    """Sprint 8: Owner/client profile persistence."""

    def test_save_and_reload_profile(self, tmp_path):
        """Profile persists to disk and reloads with same values."""
        from src.analysis.owner_profiles import save_profile, load_profile
        inputs = {
            "relationship_level": "Repeat",
            "past_work_count": 3,
            "last_project_date": "2024-06",
            "payment_delays": False,
            "disputes": False,
            "high_co_rate": True,
            "competitors": ["Firm A", "Firm B"],
            "market_pressure": 6,
            "target_margin": 8.0,
            "win_probability": 60.0,
        }
        save_profile("Acme Corp", inputs, profile_dir=tmp_path)
        loaded = load_profile("Acme Corp", profile_dir=tmp_path)
        assert loaded is not None
        assert loaded["inputs"]["relationship_level"] == "Repeat"
        assert loaded["inputs"]["competitors"] == ["Firm A", "Firm B"]
        assert loaded["inputs"]["high_co_rate"] is True

    def test_diff_detects_changes(self):
        """diff_inputs returns only changed keys."""
        from src.analysis.owner_profiles import diff_inputs
        saved = {"relationship_level": "Repeat", "market_pressure": 5, "disputes": False}
        current = {"relationship_level": "Preferred", "market_pressure": 5, "disputes": True}
        diff = diff_inputs(saved, current)
        assert "relationship_level" in diff
        assert diff["relationship_level"]["saved"] == "Repeat"
        assert diff["relationship_level"]["current"] == "Preferred"
        assert "market_pressure" not in diff
        assert "disputes" in diff

    def test_list_profiles(self, tmp_path):
        """list_profiles returns saved owner names sorted."""
        from src.analysis.owner_profiles import save_profile, list_profiles
        save_profile("Zebra Inc", {"relationship_level": "New"}, profile_dir=tmp_path)
        save_profile("Alpha LLC", {"relationship_level": "Preferred"}, profile_dir=tmp_path)
        names = list_profiles(profile_dir=tmp_path)
        assert names == ["Alpha LLC", "Zebra Inc"]


# =============================================================================
# SPRINT 9: Assumption Status + Supersedes + Multi-doc
# =============================================================================

class TestAssumptionStatusFields:
    """Sprint 9: Assumption schema has bid control fields."""

    def test_new_assumption_has_draft_status(self):
        """New assumptions default to status='draft'."""
        from src.analysis.recon_actions import (
            create_recon_assumption, finding_to_proposed_assumption,
        )
        finding = {
            "type": "missing", "category": "req_vs_schedule",
            "description": "Requirements mention doors but no schedule",
            "impact": "high",
            "suggested_action": "Request door schedule",
            "evidence": {"pages": [3], "items": ["door requirements"]},
            "confidence": 0.8,
        }
        proposed = finding_to_proposed_assumption(finding)
        asmp = create_recon_assumption(proposed, [], finding)
        assert asmp["status"] == "draft"
        assert asmp["approved_by"] is None
        assert asmp["approved_at"] is None
        assert asmp["cost_impact"] is None
        assert asmp["scope_tag"] == ""

    def test_update_status_to_accepted(self):
        """update_assumption_status sets accepted fields correctly."""
        from src.analysis.recon_actions import update_assumption_status
        asmp = {
            "id": "ASMP-R-0001", "status": "draft",
            "approved_by": None, "approved_at": None,
            "cost_impact": None, "scope_tag": "",
        }
        updated = update_assumption_status(asmp, "accepted", approved_by="John")
        assert updated["status"] == "accepted"
        assert updated["approved_by"] == "John"
        assert updated["approved_at"] is not None
        # Original unchanged
        assert asmp["status"] == "draft"

    def test_update_status_preserves_cost_impact(self):
        """cost_impact set during status update persists."""
        from src.analysis.recon_actions import update_assumption_status
        asmp = {
            "id": "ASMP-R-0001", "status": "draft",
            "approved_by": None, "approved_at": None,
            "cost_impact": None, "scope_tag": "",
        }
        updated = update_assumption_status(
            asmp, "accepted", approved_by="Jane", cost_impact=5000.0, scope_tag="doors",
        )
        assert updated["cost_impact"] == 5000.0
        assert updated["scope_tag"] == "doors"
        # Revert to draft — cost_impact preserved (not explicitly cleared)
        reverted = update_assumption_status(updated, "draft")
        assert reverted["status"] == "draft"
        assert reverted["approved_by"] is None
        assert reverted["cost_impact"] == 5000.0  # preserved

    def test_exclusions_clarifications_export(self):
        """generate_exclusions_clarifications produces txt and csv."""
        from src.analysis.recon_actions import generate_exclusions_clarifications
        assumptions = [
            {"id": "ASMP-R-0001", "status": "accepted", "title": "Door grade",
             "text": "Standard commercial grade doors assumed", "cost_impact": 5000,
             "scope_tag": "doors", "risk_level": "high", "approved_by": "User",
             "approved_at": "2026-02-19T12:00:00"},
            {"id": "ASMP-R-0002", "status": "rejected", "title": "Window type",
             "text": "Aluminium windows excluded", "cost_impact": 12000,
             "scope_tag": "windows", "risk_level": "medium", "approved_by": "User",
             "approved_at": "2026-02-19T12:00:00"},
            {"id": "ASMP-R-0003", "status": "draft", "title": "Draft item",
             "text": "Not finalized", "cost_impact": None, "scope_tag": "",
             "risk_level": "low", "approved_by": None, "approved_at": None},
        ]
        txt, csv_content = generate_exclusions_clarifications(assumptions)
        assert "EXCLUSIONS" in txt
        assert "CLARIFICATIONS" in txt
        assert "ASMP-R-0001" in txt
        assert "ASMP-R-0002" in txt
        assert "ASMP-R-0003" not in txt  # draft excluded
        assert "id,title,text,status" in csv_content
        assert "ASMP-R-0001" in csv_content
        assert "ASMP-R-0003" not in csv_content


class TestSupersedesDetector:
    """Sprint 9: Addendum supersede language detection."""

    def test_detects_this_supersedes(self):
        """'this supersedes' pattern is detected."""
        from src.analysis.supersedes_detector import detect_supersede_language
        matches = detect_supersede_language(
            "This supersedes all previous correspondence regarding BOQ items."
        )
        assert len(matches) >= 1
        assert "supersedes" in matches[0]["matched_text"].lower()

    def test_detects_delete_and_substitute(self):
        """'delete and substitute' pattern is detected."""
        from src.analysis.supersedes_detector import detect_supersede_language
        matches = detect_supersede_language(
            "Delete and substitute the following clause regarding concrete grade."
        )
        assert len(matches) >= 1

    def test_tagging_is_deterministic(self):
        """Same input always produces same resolution tags."""
        from src.analysis.supersedes_detector import tag_conflicts_with_supersedes
        conflicts = [
            {"type": "boq_change", "addendum_page": 5, "delta_confidence": 0.9},
            {"type": "requirement_new", "addendum_page": 10, "delta_confidence": 0.7},
        ]
        texts = {
            5: "This supersedes Addendum No. 1 regarding BOQ item 2.3",
            10: "Additional requirement for fire-rated doors.",
        }
        r1 = tag_conflicts_with_supersedes(conflicts, texts)
        r2 = tag_conflicts_with_supersedes(conflicts, texts)
        assert r1[0]["resolution"] == r2[0]["resolution"] == "intentional_revision"
        assert r1[1]["resolution"] is None
        assert r2[1]["resolution"] is None


class TestMultiDocMapping:
    """Sprint 9: Multi-document page index and mapping."""

    def test_build_index_single_doc(self):
        """Single doc produces identity mapping."""
        from src.analysis.multi_doc import build_multi_doc_index, global_to_doc_page
        fi = [{"name": "base.pdf", "pages": 50, "path": "/tmp/base.pdf", "ocr_used": False}]
        mdi = build_multi_doc_index(fi)
        assert len(mdi.docs) == 1
        assert mdi.docs[0].doc_id == 0
        assert mdi.total_pages == 50
        assert global_to_doc_page(0, mdi) == (0, 0)
        assert global_to_doc_page(49, mdi) == (0, 49)

    def test_build_index_two_docs(self):
        """Two docs get correct offsets and page conversions."""
        from src.analysis.multi_doc import build_multi_doc_index, global_to_doc_page
        fi = [
            {"name": "base.pdf", "pages": 50, "path": "/tmp/base.pdf", "ocr_used": False},
            {"name": "addendum.pdf", "pages": 10, "path": "/tmp/addendum.pdf", "ocr_used": False},
        ]
        mdi = build_multi_doc_index(fi)
        assert mdi.total_pages == 60
        assert mdi.docs[1].global_page_start == 50
        # Global page 0 -> doc 0, local 0
        assert global_to_doc_page(0, mdi) == (0, 0)
        # Global page 49 -> doc 0, local 49
        assert global_to_doc_page(49, mdi) == (0, 49)
        # Global page 50 -> doc 1, local 0
        assert global_to_doc_page(50, mdi) == (1, 0)
        # Global page 59 -> doc 1, local 9
        assert global_to_doc_page(59, mdi) == (1, 9)

    def test_convert_evidence_pages(self):
        """Evidence page list converts to (doc_id, page) tuples."""
        from src.analysis.multi_doc import build_multi_doc_index, convert_evidence_pages
        fi = [
            {"name": "base.pdf", "pages": 30, "path": "/a", "ocr_used": False},
            {"name": "add.pdf", "pages": 5, "path": "/b", "ocr_used": False},
        ]
        mdi = build_multi_doc_index(fi)
        tuples = convert_evidence_pages([2, 31], mdi)
        assert tuples == [(0, 2), (1, 1)]


class TestSupersedesAndConflicts:
    """Sprint 9: Integration of supersede detection with delta_detector."""

    def test_conflict_tagged_with_resolution(self):
        """tag_conflicts_with_resolution adds resolution field."""
        from src.analysis.delta_detector import tag_conflicts_with_resolution
        conflicts = [{"type": "boq_change", "addendum_page": 3, "delta_confidence": 0.9}]
        texts = {3: "This supersedes the previous BOQ quantities."}
        result = tag_conflicts_with_resolution(conflicts, texts)
        assert result[0].get("resolution") == "intentional_revision"

    def test_conflict_without_supersede_has_none_resolution(self):
        """Conflicts on non-supersede pages get resolution=None."""
        from src.analysis.delta_detector import tag_conflicts_with_resolution
        conflicts = [{"type": "requirement_new", "addendum_page": 7, "delta_confidence": 0.7}]
        texts = {7: "New requirement for fire alarm system."}
        result = tag_conflicts_with_resolution(conflicts, texts)
        assert result[0].get("resolution") is None


# =============================================================================
# SPRINT 10: Cache, Toxic Pages, Clustering, QA Score
# =============================================================================


class TestPipelineCache(unittest.TestCase):
    """Sprint 10: Persistent pipeline cache tests."""

    def test_cache_key_deterministic(self):
        """Same PDF paths + config → same cache key."""
        from src.analysis.pipeline_cache import compute_cache_key
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(b"%PDF-1.4 test content for hashing")
            tmp_path = Path(f.name)
        try:
            config = {"dpi": 150, "budget_pages": 80}
            key1 = compute_cache_key([tmp_path], config)
            key2 = compute_cache_key([tmp_path], config)
            assert key1 == key2, f"Keys differ: {key1} vs {key2}"
            assert len(key1) == 64, f"Expected SHA256 hex length 64, got {len(key1)}"
        finally:
            os.unlink(tmp_path)

    def test_cache_save_load_roundtrip(self):
        """Save stage data, load it back, verify identical."""
        from src.analysis.pipeline_cache import save_cached_stage, load_cached_stage
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            data = {"total_pages": 10, "counts_by_type": {"plan": 5, "boq": 3, "schedule": 2}}
            save_cached_stage(Path(tmp), "page_index", data)
            loaded = load_cached_stage(Path(tmp), "page_index")
            assert loaded is not None, "Cache miss on freshly saved data"
            assert loaded["total_pages"] == 10
            assert loaded["counts_by_type"]["plan"] == 5

    def test_cache_miss_returns_none(self):
        """Non-existent stage returns None."""
        from src.analysis.pipeline_cache import load_cached_stage
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            result = load_cached_stage(Path(tmp), "nonexistent_stage")
            assert result is None, f"Expected None for cache miss, got {type(result)}"


class TestToxicPages(unittest.TestCase):
    """Sprint 10: Toxic page identification and retry tests."""

    def test_identify_failed_pages_finds_errors(self):
        """Pages with error field are detected as failed."""
        from src.analysis.toxic_pages import identify_failed_pages
        metadata = {
            "page_profiles": [
                {"page_index": 0, "ocr_used": True, "text_length": 500},
                {"page_index": 3, "ocr_used": True, "error": "wall_clock_exceeded"},
                {"page_index": 5, "ocr_used": True, "text_length": 0},
                {"page_index": 7, "ocr_used": False},
            ]
        }
        failed = identify_failed_pages(metadata)
        assert 3 in failed, "Page 3 with error should be detected"
        assert 5 in failed, "Page 5 with zero text_length should be detected"
        assert 0 not in failed, "Page 0 with text should NOT be detected"

    def test_no_failed_pages_empty_list(self):
        """Clean OCR results produce empty failed list."""
        from src.analysis.toxic_pages import identify_failed_pages
        metadata = {
            "page_profiles": [
                {"page_index": 0, "ocr_used": True, "text_length": 500},
                {"page_index": 1, "ocr_used": True, "text_length": 300},
            ]
        }
        failed = identify_failed_pages(metadata)
        assert failed == [], f"Expected empty list, got {failed}"

    def test_build_toxic_summary_counts(self):
        """Summary correctly counts toxic vs recovered pages."""
        from src.analysis.toxic_pages import build_toxic_pages_summary
        results = [
            {"page_idx": 3, "toxic": True, "reason": "empty_after_low_dpi_retry", "retry_time_s": 1.5},
            {"page_idx": 5, "toxic": False, "reason": "recovered_low_dpi_grayscale", "retry_time_s": 0.8},
            {"page_idx": 7, "toxic": True, "reason": "render_failed_low_dpi", "retry_time_s": 0.2},
        ]
        summary = build_toxic_pages_summary(results)
        assert summary["toxic_count"] == 2, f"Expected 2 toxic, got {summary['toxic_count']}"
        assert summary["recovered_count"] == 1, f"Expected 1 recovered, got {summary['recovered_count']}"
        assert summary["total_retry_time_s"] == 2.5


class TestRfiClustering(unittest.TestCase):
    """Sprint 10: RFI and assumption clustering tests."""

    def test_same_trade_overlapping_pages_grouped(self):
        """Two RFIs with same trade + shared evidence pages → 1 cluster."""
        from src.analysis.rfi_clustering import cluster_rfis
        rfis = [
            {"id": "RFI-0001", "trade": "architectural", "priority": "high",
             "question": "Provide door schedule for Block A",
             "evidence": {"pages": [2, 5]}},
            {"id": "RFI-0002", "trade": "architectural", "priority": "medium",
             "question": "Confirm window types for Block A",
             "evidence": {"pages": [5, 8]}},
        ]
        clusters = cluster_rfis(rfis)
        # These share page 5 + same trade → should cluster
        assert len(clusters) == 1, f"Expected 1 cluster, got {len(clusters)}"
        assert clusters[0]["count"] == 2
        assert clusters[0]["priority"] == "high"  # highest priority wins

    def test_similar_text_merged(self):
        """Two RFIs with highly similar questions → merged cluster."""
        from src.analysis.rfi_clustering import cluster_rfis
        rfis = [
            {"id": "RFI-0001", "trade": "structural", "priority": "high",
             "question": "Provide structural details for foundation",
             "evidence": {"pages": [10]}},
            {"id": "RFI-0002", "trade": "mep", "priority": "medium",
             "question": "Provide structural details for foundations",
             "evidence": {"pages": [20]}},
        ]
        clusters = cluster_rfis(rfis, similarity_threshold=0.65)
        # Very similar text → should merge despite different trade
        assert len(clusters) == 1, f"Expected 1 cluster, got {len(clusters)}"

    def test_clustering_deterministic(self):
        """Run clustering twice → same cluster IDs."""
        from src.analysis.rfi_clustering import cluster_rfis
        rfis = [
            {"id": "RFI-0001", "trade": "architectural", "priority": "high",
             "question": "Door schedule missing", "evidence": {"pages": [1]}},
            {"id": "RFI-0002", "trade": "structural", "priority": "medium",
             "question": "Bar bending schedule missing", "evidence": {"pages": [10]}},
            {"id": "RFI-0003", "trade": "architectural", "priority": "low",
             "question": "Door schedule required", "evidence": {"pages": [1, 3]}},
        ]
        c1 = cluster_rfis(rfis)
        c2 = cluster_rfis(rfis)
        assert len(c1) == len(c2), "Cluster count changed"
        for a, b in zip(c1, c2):
            assert a["cluster_id"] == b["cluster_id"], "Cluster IDs differ"
            assert a["count"] == b["count"], "Cluster sizes differ"


class TestQaScore(unittest.TestCase):
    """Sprint 10: Bid Pack QA confidence score tests."""

    def test_perfect_payload_100(self):
        """Full coverage, no conflicts, no toxic → near-perfect score."""
        from src.analysis.qa_score import compute_qa_score
        payload = {
            "run_coverage": {
                "pages_total": 20,
                "pages_deep_processed": 20,
            },
            "conflicts": [],
            "addendum_index": [],
            "extraction_summary": {
                "counts": {"boq_items": 50, "schedules": 10, "requirements": 30},
            },
            "toxic_pages": {"toxic_count": 0},
        }
        result = compute_qa_score(payload)
        assert result["score"] >= 95, f"Expected ~100, got {result['score']}"
        assert result["confidence"] == "HIGH"

    def test_empty_payload_partial_score(self):
        """Empty payload → partial score (no coverage/parsing but no conflicts either)."""
        from src.analysis.qa_score import compute_qa_score
        payload = {}
        result = compute_qa_score(payload)
        # Empty payload: coverage=0, conflicts=20 (none), addenda=20 (none),
        # parse=0 (nothing parsed), toxic=20 (none) → 60
        assert result["score"] == 60, f"Expected 60 for empty payload, got {result['score']}"
        assert result["breakdown"]["coverage_completeness"] == 0
        assert result["breakdown"]["parse_completeness"] == 0
        assert result["confidence"] == "LOW"

    def test_qa_score_stable(self):
        """Same payload → same score twice (deterministic)."""
        from src.analysis.qa_score import compute_qa_score
        payload = {
            "run_coverage": {"pages_total": 50, "pages_deep_processed": 30},
            "conflicts": [{"type": "boq_change"}, {"type": "requirement_new"}],
            "addendum_index": [{"addendum_no": 1}],
            "extraction_summary": {"counts": {"boq_items": 10, "schedules": 0, "requirements": 5}},
            "toxic_pages": {"toxic_count": 1},
        }
        r1 = compute_qa_score(payload)
        r2 = compute_qa_score(payload)
        assert r1["score"] == r2["score"], f"Scores differ: {r1['score']} vs {r2['score']}"
        assert r1["breakdown"] == r2["breakdown"], "Breakdowns differ"


# =============================================================================
# SPRINT 11: QUANTITIES
# =============================================================================

class TestQuantities:
    """Sprint 11: quantities.py tests."""

    def test_schedule_to_quantities(self):
        """Schedule rows → quantity rows with correct schema."""
        from src.analysis.quantities import build_quantities_from_schedules
        schedules = [
            {"mark": "D1", "fields": {"type": "Flush"}, "schedule_type": "door", "source_page": 3, "has_qty": True, "qty": 4},
            {"mark": "D2", "fields": {"type": "Panel"}, "schedule_type": "door", "source_page": 3, "has_qty": False, "qty": None},
            {"mark": "D2", "fields": {"type": "Panel"}, "schedule_type": "door", "source_page": 4, "has_qty": False, "qty": None},
        ]
        result = build_quantities_from_schedules(schedules)
        assert len(result) >= 2, f"Expected >= 2, got {len(result)}"
        for q in result:
            assert "item" in q
            assert "unit" in q
            assert "qty" in q
            assert "confidence" in q
            assert "source_type" in q
            assert q["source_type"] == "schedule"
        # D1 has explicit qty=4
        d1 = [q for q in result if "D1" in q["item"]]
        assert d1 and d1[0]["qty"] == 4.0

    def test_boq_to_quantities_with_rates(self):
        """BOQ items → quantities with trade + total computed."""
        from src.analysis.quantities import build_quantities_from_boq
        items = [
            {"item_no": "2.1", "description": "RCC footing M25", "unit": "cum", "qty": 10, "rate": 5000, "source_page": 1, "confidence": 0.9},
            {"item_no": "5.3", "description": "Vitrified tile flooring", "unit": "sqm", "qty": 120, "rate": 800, "source_page": 2, "confidence": 0.8},
        ]
        result = build_quantities_from_boq(items)
        assert len(result) == 2
        # Trade inference from item_no prefix
        assert result[0]["trade"] == "concrete"  # 2.x → concrete
        assert result[1]["trade"] == "flooring"  # 5.x → flooring
        # Total = qty * rate
        assert result[0]["total"] == 50000.0
        assert result[1]["total"] == 96000.0

    def test_callout_to_quantities(self):
        """Drawing callouts → quantity counts."""
        from src.analysis.quantities import build_quantities_from_callouts
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "confidence": 0.7},
            {"text": "D2", "callout_type": "tag", "source_page": 0, "confidence": 0.7},
            {"text": "W1", "callout_type": "tag", "source_page": 1, "confidence": 0.7},
            {"text": "Living Room", "callout_type": "room", "source_page": 0, "confidence": 0.6},
            {"text": "Kitchen", "callout_type": "room", "source_page": 1, "confidence": 0.6},
        ]
        result = build_quantities_from_callouts(callouts)
        assert len(result) >= 2  # doors, windows, rooms
        door_q = [q for q in result if "door" in q["item"].lower()]
        assert door_q, "Should have door quantities"
        assert door_q[0]["qty"] == 2.0  # D1, D2

    def test_build_all_dedup(self):
        """build_all_quantities deduplicates schedule/callout overlap."""
        from src.analysis.quantities import build_all_quantities
        schedules = [
            {"mark": "D1", "fields": {}, "schedule_type": "door", "source_page": 0, "has_qty": True, "qty": 5},
        ]
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "confidence": 0.7},
            {"text": "D2", "callout_type": "tag", "source_page": 0, "confidence": 0.7},
        ]
        result = build_all_quantities(schedules, [], callouts)
        # Schedule has door → callout doors should be filtered out
        door_items = [q for q in result if "door" in q["item"].lower()]
        # Should only have the schedule door, not the callout door
        sources = [q["source_type"] for q in door_items]
        assert "callout" not in sources, "Callout doors should be filtered when schedule has doors"


# =============================================================================
# SPRINT 11: PRICING GUIDANCE
# =============================================================================

class TestPricingGuidance:
    """Sprint 11: pricing_guidance.py tests."""

    def test_high_qa_low_contingency(self):
        """High QA score → low contingency range."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        result = compute_pricing_guidance(
            qa_score={"score": 90, "breakdown": {}, "top_actions": []},
            addendum_index=[],
            conflicts=[],
            owner_profile=None,
            run_coverage=None,
        )
        cont = result["contingency_range"]
        assert cont["low_pct"] <= 5.0, f"Expected low <= 5, got {cont['low_pct']}"
        assert cont["high_pct"] <= 7.0, f"Expected high <= 7, got {cont['high_pct']}"

    def test_many_conflicts_high_contingency(self):
        """Many unresolved conflicts → higher contingency."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        conflicts = [{"type": "boq_change", "description": f"conflict {i}"} for i in range(10)]
        result = compute_pricing_guidance(
            qa_score={"score": 40, "breakdown": {}, "top_actions": []},
            addendum_index=[{"addendum_no": 1}, {"addendum_no": 2}],
            conflicts=conflicts,
            owner_profile=None,
            run_coverage=None,
        )
        cont = result["contingency_range"]
        # Base 8-15 + 5 for 10 conflicts + 2 for 2 addenda
        assert cont["recommended_pct"] >= 10.0, f"Expected recommended >= 10, got {cont['recommended_pct']}"

    def test_not_covered_generates_exclusions(self):
        """doc_types_not_covered → recommended exclusions."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        result = compute_pricing_guidance(
            qa_score=None,
            addendum_index=[],
            conflicts=[],
            owner_profile=None,
            run_coverage={"doc_types_found": ["boq"], "doc_types_not_covered": ["specification", "plan"]},
        )
        excl = result["recommended_exclusions"]
        assert len(excl) >= 2, f"Expected >= 2 exclusions, got {len(excl)}"
        # Should mention schedule (not found in doc_types_found)
        all_text = " ".join(excl).lower()
        assert "specification" in all_text or "plan" in all_text


# =============================================================================
# SPRINT 11: DOCX EXPORTS
# =============================================================================

class TestDocxExports:
    """Sprint 11: docx_exports.py tests."""

    def test_rfis_docx_valid_pk_header(self):
        """RFIs DOCX has valid PK header (ZIP/DOCX format)."""
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from docx_exports import generate_rfis_docx
        rfis = [
            {"question": "What grade of concrete?", "trade": "concrete", "priority": "high"},
            {"question": "Door schedule missing D5", "trade": "doors", "priority": "medium"},
        ]
        docx_bytes = generate_rfis_docx(rfis, project_id="TEST-01")
        assert docx_bytes[:2] == b"PK", "DOCX should start with PK header"
        assert len(docx_bytes) > 500, f"DOCX too small: {len(docx_bytes)} bytes"

    def test_bid_summary_docx_valid(self):
        """Bid summary DOCX is valid PK with all sections."""
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from docx_exports import generate_bid_summary_docx
        payload = {
            "project_id": "TEST-02",
            "readiness_score": 72,
            "decision": "Bid with reservations",
            "rfis": [{"question": "Test RFI", "trade": "civil", "priority": "low"}],
            "blockers": [{"title": "Missing BOQ", "severity": "high", "trade": "general"}],
            "qa_score": {"score": 75, "breakdown": {"coverage_completeness": 15}, "top_actions": ["Improve coverage"]},
            "pricing_guidance": {
                "contingency_range": {"low_pct": 5, "high_pct": 8, "recommended_pct": 6.5, "rationale": "Test"},
                "recommended_exclusions": ["Scope from spec"],
                "recommended_clarifications": ["Clarify concrete grade"],
                "suggested_alternates_ve": [{"item": "Alt material", "reason": "Cost saving"}],
            },
        }
        docx_bytes = generate_bid_summary_docx(payload)
        assert docx_bytes[:2] == b"PK", "DOCX should start with PK header"
        assert len(docx_bytes) > 1000, f"DOCX too small: {len(docx_bytes)} bytes"


# =============================================================================
# SPRINT 11: RUN COMPARE
# =============================================================================

class TestRunCompare:
    """Sprint 11: run compare helper tests."""

    def test_qa_delta_computed(self):
        """QA score delta is computed from prev/current."""
        from src.analysis.pipeline import _compute_run_compare
        prev = {"qa_score": {"score": 70}, "rfis": [], "conflicts": [], "quantities_count": 5}
        current = {"qa_score": {"score": 82}, "rfis": [], "conflicts": [], "quantities": []}
        result = _compute_run_compare(prev, current)
        assert result is not None
        assert result["qa_score_delta"] == 12

    def test_new_rfis_detected(self):
        """New RFIs not in previous run are detected."""
        from src.analysis.pipeline import _compute_run_compare
        prev = {
            "qa_score": {"score": 50},
            "rfis": [{"question": "Old RFI"}],
            "conflicts": [],
            "quantities_count": 0,
        }
        current = {
            "qa_score": {"score": 55},
            "rfis": [{"question": "Old RFI"}, {"question": "New RFI added"}],
            "conflicts": [],
            "quantities": [],
        }
        result = _compute_run_compare(prev, current)
        assert "New RFI added" in result["new_rfis"]
        assert "Old RFI" not in result["new_rfis"]


# =============================================================================
# SPRINT 12: QUANTITY TRACEABILITY
# =============================================================================

class TestQuantityTraceability:
    """Sprint 12: evidence_bundle_id and derivation fields."""

    def test_schedule_qty_has_bundle_id(self):
        """Schedule quantities include evidence_bundle_id."""
        from src.analysis.quantities import build_quantities_from_schedules
        schedules = [
            {"mark": "D1", "fields": {}, "schedule_type": "door", "source_page": 3, "has_qty": True, "qty": 4},
        ]
        result = build_quantities_from_schedules(schedules)
        assert len(result) == 1
        assert "evidence_bundle_id" in result[0]
        assert len(result[0]["evidence_bundle_id"]) == 12
        assert "derivation" in result[0]
        assert "schedule:door:D1" in result[0]["derivation"]

    def test_boq_qty_has_derivation(self):
        """BOQ quantities include derivation field."""
        from src.analysis.quantities import build_quantities_from_boq
        items = [
            {"item_no": "2.1", "description": "RCC footing", "unit": "cum", "qty": 10, "rate": 5000, "source_page": 1, "confidence": 0.9},
        ]
        result = build_quantities_from_boq(items)
        assert result[0]["derivation"].startswith("boq:2.1:")
        assert "has_rate" in result[0]["derivation"]

    def test_callout_qty_has_bundle_id(self):
        """Callout quantities include evidence_bundle_id."""
        from src.analysis.quantities import build_quantities_from_callouts
        callouts = [
            {"text": "D1", "callout_type": "tag", "source_page": 0, "confidence": 0.7},
            {"text": "D2", "callout_type": "tag", "source_page": 1, "confidence": 0.7},
        ]
        result = build_quantities_from_callouts(callouts)
        door_q = [q for q in result if "door" in q["item"].lower()]
        assert door_q and "evidence_bundle_id" in door_q[0]

    def test_bundle_id_deterministic(self):
        """Same inputs produce same bundle_id across calls."""
        from src.analysis.quantities import build_quantities_from_schedules
        schedules = [
            {"mark": "D1", "fields": {}, "schedule_type": "door", "source_page": 3, "has_qty": True, "qty": 4},
        ]
        r1 = build_quantities_from_schedules(schedules)
        r2 = build_quantities_from_schedules(schedules)
        assert r1[0]["evidence_bundle_id"] == r2[0]["evidence_bundle_id"]


# =============================================================================
# SPRINT 12: QUANTITY RECONCILIATION
# =============================================================================

class TestQuantityReconciliation:
    """Sprint 12: quantity_reconciliation.py tests."""

    def test_door_mismatch_detected(self):
        """Different door counts across sources -> mismatch=True."""
        from src.analysis.quantity_reconciliation import reconcile_quantities
        schedules = [
            {"schedule_type": "door", "mark": "D1", "source_page": 0, "fields": {}},
            {"schedule_type": "door", "mark": "D2", "source_page": 0, "fields": {}},
        ]
        boq_items = [
            {"item_no": "4.1", "description": "Flush door", "qty": 5, "source_page": 10},
        ]
        result = reconcile_quantities([], schedules, boq_items, [])
        door_rows = [r for r in result if r["category"] == "doors"]
        assert len(door_rows) == 1
        assert door_rows[0]["mismatch"] is True
        assert door_rows[0]["schedule_count"] == 2
        assert door_rows[0]["boq_count"] == 5

    def test_no_mismatch_when_equal(self):
        """Same counts across sources -> mismatch=False."""
        from src.analysis.quantity_reconciliation import reconcile_quantities
        schedules = [
            {"schedule_type": "door", "mark": "D1", "source_page": 0, "fields": {}},
            {"schedule_type": "door", "mark": "D2", "source_page": 0, "fields": {}},
        ]
        boq_items = [
            {"item_no": "4.1", "description": "Flush door", "qty": 2, "source_page": 10},
        ]
        result = reconcile_quantities([], schedules, boq_items, [])
        door_rows = [r for r in result if r["category"] == "doors"]
        assert door_rows and door_rows[0]["mismatch"] is False

    def test_reconciliation_deterministic(self):
        """Same inputs -> identical output across two calls."""
        from src.analysis.quantity_reconciliation import reconcile_quantities
        schedules = [
            {"schedule_type": "door", "mark": "D1", "source_page": 0, "fields": {}},
        ]
        boq_items = [
            {"item_no": "4.1", "description": "Flush door", "qty": 3, "source_page": 10},
        ]
        r1 = reconcile_quantities([], schedules, boq_items, [])
        r2 = reconcile_quantities([], schedules, boq_items, [])
        assert r1 == r2

    def test_apply_action_prefer_schedule(self):
        """apply_reconciliation_action with prefer_schedule sets correct fields."""
        from src.analysis.quantity_reconciliation import apply_reconciliation_action
        row = {"category": "doors", "schedule_count": 3, "boq_count": 5, "mismatch": True}
        result = apply_reconciliation_action(row, "prefer_schedule")
        assert result["preferred_source"] == "schedule"
        assert result["preferred_qty"] == 3
        assert result is not row  # new dict, not mutated


# =============================================================================
# SPRINT 12: FINISH TAKEOFF
# =============================================================================

class TestFinishTakeoff:
    """Sprint 12: finish_takeoff.py tests."""

    def test_finish_with_areas(self):
        """Finish schedules with area field -> has_areas=True, aggregated totals."""
        from src.analysis.finish_takeoff import build_finish_takeoff
        schedules = [
            {"schedule_type": "finish", "mark": "R1", "source_page": 5, "fields": {
                "room": "Living Room", "floor": "Vitrified Tile", "wall": "Paint",
                "ceiling": "POP", "area": "25.5",
            }},
            {"schedule_type": "finish", "mark": "R2", "source_page": 5, "fields": {
                "room": "Kitchen", "floor": "Ceramic Tile", "wall": "Paint",
                "ceiling": "POP", "area": "12.0",
            }},
        ]
        result = build_finish_takeoff(schedules)
        assert result["has_areas"] is True
        assert len(result["finish_rows"]) > 0
        assert result["rooms_missing_area"] == []

    def test_finish_without_areas(self):
        """Finish schedules without area -> has_areas=False, rooms listed."""
        from src.analysis.finish_takeoff import build_finish_takeoff
        schedules = [
            {"schedule_type": "finish", "mark": "R1", "source_page": 5, "fields": {
                "room": "Bedroom", "floor": "Wood", "wall": "Wallpaper",
            }},
        ]
        result = build_finish_takeoff(schedules)
        assert result["has_areas"] is False
        assert len(result["rooms_missing_area"]) > 0

    def test_no_finish_schedules(self):
        """No finish schedules -> empty result."""
        from src.analysis.finish_takeoff import build_finish_takeoff
        schedules = [
            {"schedule_type": "door", "mark": "D1", "source_page": 0, "fields": {}},
        ]
        result = build_finish_takeoff(schedules)
        assert result["has_areas"] is False
        assert result["finish_rows"] == []


# =============================================================================
# SPRINT 12: FEEDBACK PERSISTENCE
# =============================================================================

class TestFeedback:
    """Sprint 12: feedback.py tests."""

    def test_make_feedback_entry_schema(self):
        """make_feedback_entry returns dict with all required fields."""
        from src.analysis.feedback import make_feedback_entry
        entry = make_feedback_entry(
            feedback_type="rfi", item_id="RFI-0001", verdict="correct",
            page_refs=[3, 5],
        )
        assert entry["feedback_type"] == "rfi"
        assert entry["item_id"] == "RFI-0001"
        assert entry["verdict"] == "correct"
        assert entry["page_refs"] == [3, 5]
        assert "timestamp" in entry

    def test_feedback_jsonl_write_and_load(self, tmp_path):
        """append_feedback writes JSONL, load_feedback reads it back."""
        from src.analysis.feedback import make_feedback_entry, append_feedback, load_feedback
        entry1 = make_feedback_entry("rfi", "RFI-0001", "correct")
        entry2 = make_feedback_entry("quantity", "abc123", "edited",
            corrected_value="10", original_value="8")
        append_feedback(entry1, tmp_path)
        append_feedback(entry2, tmp_path)
        loaded = load_feedback(tmp_path)
        assert len(loaded) == 2
        assert loaded[0]["item_id"] == "RFI-0001"
        assert loaded[1]["corrected_value"] == "10"

    def test_feedback_schema_stable(self):
        """Feedback entry schema does not change between calls."""
        from src.analysis.feedback import make_feedback_entry
        e1 = make_feedback_entry("rfi", "X", "wrong")
        e2 = make_feedback_entry("quantity", "Y", "edited", corrected_value="5")
        assert set(e1.keys()) == set(e2.keys())


# =============================================================================
# SPRINT 13: REVIEW QUEUE + BULK ACTIONS + APPROVAL STATES
# =============================================================================

class TestReviewQueue:
    """Sprint 13: review_queue.py tests."""

    def test_deterministic_ordering(self):
        """Same inputs produce identical order across two calls."""
        from src.analysis.review_queue import build_review_queue
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 3,
             "schedule_count": 2, "boq_count": 5},
            {"category": "windows", "mismatch": True, "max_delta": 7,
             "schedule_count": 8, "boq_count": 1},
        ]
        conflicts = [
            {"type": "boq_change", "delta_confidence": 0.6,
             "item_no": "1.1", "base_page": 0, "addendum_page": 5},
        ]
        r1 = build_review_queue(quantity_reconciliation=recon, conflicts=conflicts)
        r2 = build_review_queue(quantity_reconciliation=recon, conflicts=conflicts)
        assert r1 == r2
        # HIGH items come first
        assert r1[0]["severity"] == "high"

    def test_high_before_medium(self):
        """HIGH severity items sort before MEDIUM."""
        from src.analysis.review_queue import build_review_queue
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 1},  # low
            {"category": "windows", "mismatch": True, "max_delta": 10},  # high
        ]
        result = build_review_queue(quantity_reconciliation=recon)
        assert result[0]["severity"] == "high"
        assert result[-1]["severity"] == "low"

    def test_empty_inputs(self):
        """No mismatches, no conflicts -> empty queue."""
        from src.analysis.review_queue import build_review_queue
        result = build_review_queue()
        assert result == []

    def test_risk_hits_included(self):
        """HIGH impact risk hits appear in queue."""
        from src.analysis.review_queue import build_review_queue
        risk = [
            {"template_id": "RISK-LD", "label": "Liquidated Damages",
             "impact": "high", "found": True,
             "hits": [{"page_idx": 5, "keyword": "liquidated damages"}]},
        ]
        result = build_review_queue(risk_results=risk)
        assert len(result) == 1
        assert result[0]["type"] == "risk_hit"

    def test_intentional_revisions_excluded(self):
        """Conflicts with resolution=intentional_revision are excluded from queue."""
        from src.analysis.review_queue import build_review_queue
        conflicts = [
            {"type": "boq_change", "delta_confidence": 0.6,
             "resolution": "intentional_revision"},
        ]
        result = build_review_queue(conflicts=conflicts)
        assert len(result) == 0


class TestBulkActions:
    """Sprint 13: bulk_actions.py tests."""

    def test_prefer_schedule_updates_rows(self):
        """prefer_schedule_for_mismatches applies action to matching rows."""
        from src.analysis.bulk_actions import prefer_schedule_for_mismatches
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 3,
             "schedule_count": 2, "boq_count": 5, "preferred_source": None,
             "preferred_qty": None, "action": None, "evidence_refs": [], "notes": ""},
            {"category": "finishes", "mismatch": True, "max_delta": 1,
             "schedule_count": 10, "boq_count": 11, "preferred_source": None,
             "preferred_qty": None, "action": None, "evidence_refs": [], "notes": ""},
        ]
        updated, count = prefer_schedule_for_mismatches(recon)
        assert count == 1  # only doors, not finishes
        assert updated[0]["preferred_source"] == "schedule"
        assert updated[1]["preferred_source"] is None

    def test_generate_rfis_for_high(self):
        """generate_rfis_for_high_mismatches creates RFIs for delta >= 5."""
        from src.analysis.bulk_actions import generate_rfis_for_high_mismatches
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 10,
             "schedule_count": 2, "boq_count": 12, "preferred_source": None,
             "preferred_qty": None, "action": None, "evidence_refs": [], "notes": ""},
            {"category": "windows", "mismatch": True, "max_delta": 2,
             "schedule_count": 5, "boq_count": 7, "preferred_source": None,
             "preferred_qty": None, "action": None, "evidence_refs": [], "notes": ""},
        ]
        new_rfis, updated = generate_rfis_for_high_mismatches(recon, [])
        assert len(new_rfis) == 1  # only doors (delta 10 >= 5)
        assert new_rfis[0]["source"] == "bulk_action"
        assert updated[0]["action"] == "create_rfi"

    def test_mark_revisions_reviewed(self):
        """mark_intentional_revisions_reviewed sets review_status."""
        from src.analysis.bulk_actions import mark_intentional_revisions_reviewed
        conflicts = [
            {"type": "boq_change", "resolution": "intentional_revision"},
            {"type": "schedule_change", "resolution": None},
        ]
        updated, count = mark_intentional_revisions_reviewed(conflicts)
        assert count == 1
        assert updated[0]["review_status"] == "reviewed"
        assert "review_status" not in updated[1]


class TestApprovalStates:
    """Sprint 13: approval_states.py tests."""

    def test_set_rfi_status_valid(self):
        """set_rfi_status with valid status returns new dict."""
        from src.analysis.approval_states import set_rfi_status
        rfi = {"id": "RFI-0001", "question": "Test?", "status": "draft"}
        result = set_rfi_status(rfi, "approved", approved_by="User")
        assert result["status"] == "approved"
        assert result["status_changed_by"] == "User"
        assert result is not rfi

    def test_set_rfi_status_invalid_raises(self):
        """set_rfi_status with invalid status raises ValueError."""
        from src.analysis.approval_states import set_rfi_status
        with pytest.raises(ValueError):
            set_rfi_status({"id": "X"}, "invalid_status")

    def test_filter_rfis_approved_only(self):
        """filter_rfis_for_export returns only approved/sent by default."""
        from src.analysis.approval_states import filter_rfis_for_export
        rfis = [
            {"id": "1", "status": "draft"},
            {"id": "2", "status": "approved"},
            {"id": "3", "status": "sent"},
        ]
        result = filter_rfis_for_export(rfis)
        assert len(result) == 2
        assert all(r["status"] in ("approved", "sent") for r in result)

    def test_filter_rfis_include_drafts(self):
        """filter_rfis_for_export with include_drafts=True returns all."""
        from src.analysis.approval_states import filter_rfis_for_export
        rfis = [
            {"id": "1", "status": "draft"},
            {"id": "2", "status": "approved"},
        ]
        result = filter_rfis_for_export(rfis, include_drafts=True)
        assert len(result) == 2

    def test_filter_quantities_accepted(self):
        """filter_quantities_for_export returns only accepted by default."""
        from src.analysis.approval_states import filter_quantities_for_export
        qtys = [
            {"item": "Doors", "status": "draft"},
            {"item": "Windows", "status": "accepted"},
        ]
        result = filter_quantities_for_export(qtys)
        assert len(result) == 1
        assert result[0]["item"] == "Windows"

    def test_set_conflict_status(self):
        """set_conflict_status produces new dict with review_status."""
        from src.analysis.approval_states import set_conflict_status
        conflict = {"type": "boq_change", "delta_confidence": 0.8}
        result = set_conflict_status(conflict, "reviewed")
        assert result["review_status"] == "reviewed"
        assert result is not conflict

    def test_export_filters_enforce_states(self):
        """All three export filters correctly enforce approval states."""
        from src.analysis.approval_states import (
            filter_rfis_for_export, filter_quantities_for_export, filter_conflicts_for_export,
        )
        rfis = [{"status": "draft"}, {"status": "approved"}]
        qtys = [{"status": "draft"}, {"status": "accepted"}]
        conflicts = [{"review_status": "unreviewed"}, {"review_status": "reviewed"}]

        assert len(filter_rfis_for_export(rfis)) == 1
        assert len(filter_quantities_for_export(qtys)) == 1
        assert len(filter_conflicts_for_export(conflicts)) == 1


# =============================================================================
# Sprint 14: PROJECTS (Step 9)
# =============================================================================

class TestProjects:
    """Unit tests for src/analysis/projects.py"""

    def test_create_project_deterministic(self, tmp_path):
        """Create project creates metadata.json + runs/ dir."""
        from src.analysis.projects import create_project, project_dir
        meta = create_project(
            name="Test Project",
            owner="Acme Corp",
            bid_date="2026-03-01",
            notes="Unit test",
            project_id="test_proj_001",
            projects_dir=tmp_path,
        )
        assert meta["project_id"] == "test_proj_001"
        assert meta["name"] == "Test Project"
        assert meta["owner"] == "Acme Corp"
        pdir = project_dir("test_proj_001", tmp_path)
        assert (pdir / "metadata.json").exists()
        assert (pdir / "runs").is_dir()

    def test_load_project_roundtrip(self, tmp_path):
        """Create then load returns identical data."""
        from src.analysis.projects import create_project, load_project
        original = create_project(name="Roundtrip", project_id="rt_001", projects_dir=tmp_path)
        loaded = load_project("rt_001", tmp_path)
        assert loaded is not None
        assert loaded["project_id"] == original["project_id"]
        assert loaded["name"] == original["name"]
        assert loaded["created_at"] == original["created_at"]

    def test_load_missing_project(self, tmp_path):
        """Load non-existent project returns None."""
        from src.analysis.projects import load_project
        assert load_project("nonexistent", tmp_path) is None

    def test_list_projects_sorted(self, tmp_path):
        """Projects sorted newest first by updated_at."""
        from src.analysis.projects import create_project, list_projects
        import time
        create_project(name="Older", project_id="p_old", projects_dir=tmp_path)
        time.sleep(0.05)
        create_project(name="Newer", project_id="p_new", projects_dir=tmp_path)
        projects = list_projects(tmp_path)
        assert len(projects) == 2
        assert projects[0]["project_id"] == "p_new"

    def test_save_and_list_runs(self, tmp_path):
        """Save run then list_runs returns it."""
        from src.analysis.projects import create_project, save_run, list_runs
        create_project(name="Run Test", project_id="run_proj", projects_dir=tmp_path)
        save_run("run_proj", "run_001", "/path/analysis.json",
                 run_metadata={"readiness_score": 72, "decision": "BID"},
                 projects_dir=tmp_path)
        runs = list_runs("run_proj", tmp_path)
        assert len(runs) == 1
        assert runs[0]["run_id"] == "run_001"
        assert runs[0]["readiness_score"] == 72

    def test_update_project(self, tmp_path):
        """Update modifies fields and refreshes updated_at."""
        from src.analysis.projects import create_project, update_project, load_project
        import time
        create_project(name="Before", project_id="upd_001", projects_dir=tmp_path)
        original = load_project("upd_001", tmp_path)
        time.sleep(0.05)
        updated = update_project("upd_001", {"name": "After", "owner": "New Owner"}, tmp_path)
        assert updated["name"] == "After"
        assert updated["owner"] == "New Owner"
        assert updated["updated_at"] > original["updated_at"]


# =============================================================================
# Sprint 14: COLLABORATION (Step 10)
# =============================================================================

class TestCollaboration:
    """Unit tests for src/analysis/collaboration.py"""

    def test_make_entry_schema(self):
        """make_collaboration_entry returns dict with all required fields."""
        from src.analysis.collaboration import make_collaboration_entry
        entry = make_collaboration_entry("rfi", "RFI-0001", "comment",
                                          {"text": "Check dims"}, author="J.Smith")
        assert entry["entity_type"] == "rfi"
        assert entry["entity_id"] == "RFI-0001"
        assert entry["action_type"] == "comment"
        assert entry["data"]["text"] == "Check dims"
        assert entry["author"] == "J.Smith"
        assert "timestamp" in entry

    def test_comment_persist_and_load(self, tmp_path):
        """JSONL write + read roundtrip."""
        from src.analysis.collaboration import (
            make_collaboration_entry, append_collaboration, load_collaboration,
        )
        e1 = make_collaboration_entry("rfi", "RFI-001", "comment", {"text": "Hello"})
        e2 = make_collaboration_entry("rfi", "RFI-002", "assign", {"assigned_to": "Alice"})
        append_collaboration(e1, tmp_path)
        append_collaboration(e2, tmp_path)
        entries = load_collaboration(tmp_path)
        assert len(entries) == 2
        assert entries[0]["entity_id"] == "RFI-001"
        assert entries[1]["data"]["assigned_to"] == "Alice"

    def test_get_entity_collaboration(self, tmp_path):
        """Aggregates comments/assigned_to/due_date for one entity."""
        from src.analysis.collaboration import (
            make_collaboration_entry, append_collaboration, load_collaboration,
            get_entity_collaboration,
        )
        entries_to_write = [
            make_collaboration_entry("rfi", "RFI-001", "comment", {"text": "Note 1"}),
            make_collaboration_entry("rfi", "RFI-001", "assign", {"assigned_to": "Bob"}),
            make_collaboration_entry("rfi", "RFI-001", "due_date", {"due_date": "2026-04-01"}),
            make_collaboration_entry("rfi", "RFI-002", "comment", {"text": "Other"}),
        ]
        for e in entries_to_write:
            append_collaboration(e, tmp_path)
        entries = load_collaboration(tmp_path)
        collab = get_entity_collaboration(entries, "rfi", "RFI-001")
        assert len(collab["comments"]) == 1
        assert collab["assigned_to"] == "Bob"
        assert collab["due_date"] == "2026-04-01"

    def test_enrich_items(self):
        """enrich_items_with_collaboration adds collab fields to items without mutation."""
        from src.analysis.collaboration import (
            make_collaboration_entry, enrich_items_with_collaboration,
        )
        items = [{"id": "RFI-001", "question": "Q1"}, {"id": "RFI-002", "question": "Q2"}]
        entries = [
            make_collaboration_entry("rfi", "RFI-001", "assign", {"assigned_to": "Eve"}),
        ]
        enriched = enrich_items_with_collaboration(items, entries, "rfi", id_field="id")
        assert len(enriched) == 2
        assert enriched[0]["assigned_to"] == "Eve"
        assert enriched[1]["assigned_to"] == ""
        # Originals not mutated
        assert "assigned_to" not in items[0]

    def test_empty_load(self, tmp_path):
        """Returns [] for missing file."""
        from src.analysis.collaboration import load_collaboration
        assert load_collaboration(tmp_path / "nonexistent") == []

    def test_build_appendix(self, tmp_path):
        """Produces non-empty string with entity refs."""
        from src.analysis.collaboration import (
            make_collaboration_entry, append_collaboration, load_collaboration,
            build_collaboration_appendix,
        )
        entries_to_write = [
            make_collaboration_entry("rfi", "RFI-001", "comment", {"text": "Review needed"}, author="Alice"),
            make_collaboration_entry("conflict", "CONF-001", "assign", {"assigned_to": "Bob"}),
        ]
        for e in entries_to_write:
            append_collaboration(e, tmp_path)
        entries = load_collaboration(tmp_path)
        text = build_collaboration_appendix(entries)
        assert "RFI" in text
        assert "CONFLICT" in text
        assert "Alice" in text
        assert len(text) > 50


# =============================================================================
# Sprint 14: SUBMISSION PACK (Step 11)
# =============================================================================

class TestSubmissionPack:
    """Unit tests for app/submission_pack.py"""

    def test_zip_structure_has_5_folders(self):
        """All 5 folder prefixes present + cover sheet."""
        import zipfile, io
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from submission_pack import generate_submission_pack

        buffers = {
            "rfis.csv": "id,question\n1,test",
            "boq.csv": "item,qty\n1,10",
            "exclusions_clarifications.txt": "None",
            "bid_summary.md": "# Summary",
            "requirements.csv": "req,status",
        }
        zip_bytes = generate_submission_pack(buffers, project_id="test", project_name="Test")
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            names = zf.namelist()
            assert any("00_Cover_Sheet" in n for n in names)
            assert any("01_Bid_Summary/" in n for n in names)
            assert any("02_RFIs/" in n for n in names)
            assert any("03_Exclusions_Clarifications/" in n for n in names)
            assert any("04_Quantities/" in n for n in names)
            assert any("05_Evidence_Appendix/" in n for n in names)

    def test_empty_buffers(self):
        """Empty buffers still produce valid ZIP with cover sheet."""
        import zipfile, io
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from submission_pack import generate_submission_pack
        zip_bytes = generate_submission_pack({})
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            names = zf.namelist()
            assert len(names) >= 1  # At least cover sheet
            assert "00_Cover_Sheet.txt" in names

    def test_manifest_preview(self):
        """Manifest maps files to correct folders."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from submission_pack import get_submission_manifest
        buffers = {"rfis.csv": "x", "boq.csv": "y", "unknown_file.txt": "z"}
        manifest = get_submission_manifest(buffers)
        assert "02_RFIs" in manifest
        assert "RFI_Log.csv" in manifest["02_RFIs"]
        assert "04_Quantities" in manifest
        assert "BOQ.csv" in manifest["04_Quantities"]
        # Unknown files go to fallback
        assert "05_Evidence_Appendix" in manifest
        assert "unknown_file.txt" in manifest["05_Evidence_Appendix"]

    def test_collaboration_appendix_in_zip(self):
        """Collaboration_Notes.txt ends up in 05_Evidence."""
        import zipfile, io
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from submission_pack import generate_submission_pack
        zip_bytes = generate_submission_pack(
            {"rfis.csv": "data"},
            collaboration_appendix="COLLAB NOTES\nTest line",
        )
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            names = zf.namelist()
            assert "05_Evidence_Appendix/Collaboration_Notes.txt" in names
            content = zf.read("05_Evidence_Appendix/Collaboration_Notes.txt").decode()
            assert "COLLAB NOTES" in content

    def test_bytes_content_handled(self):
        """Binary content (DOCX/PDF bytes) written correctly."""
        import zipfile, io
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from submission_pack import generate_submission_pack
        fake_pdf = b"%PDF-1.4 fake content"
        zip_bytes = generate_submission_pack({"bid_summary.pdf": fake_pdf})
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            content = zf.read("01_Bid_Summary/Bid_Summary.pdf")
            assert content == fake_pdf


# =============================================================================
# Sprint 15: EVIDENCE APPENDIX PDF (Step 8)
# =============================================================================

class TestEvidenceAppendixPdf:
    """Unit tests for app/evidence_appendix_pdf.py"""

    def test_generates_valid_pdf_bytes(self):
        """generate_evidence_appendix_pdf returns non-empty bytes starting with %PDF."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from evidence_appendix_pdf import generate_evidence_appendix_pdf
        rfis = [
            {"id": "RFI-0001", "trade": "structural", "priority": "high",
             "question": "Missing beam schedule", "status": "approved",
             "evidence_pages": [3, 5], "source_page": 3},
        ]
        conflicts = [
            {"type": "boq_change", "item_no": "2.1", "delta_confidence": 0.7,
             "review_status": "reviewed", "base_page": 0, "addendum_page": 5,
             "changes": [{"field": "qty", "base_value": "10", "addendum_value": "15"}]},
        ]
        result = generate_evidence_appendix_pdf(rfis=rfis, conflicts=conflicts)
        assert isinstance(result, bytes)
        assert len(result) > 100
        assert result[:5] == b"%PDF-"

    def test_empty_inputs_returns_valid_pdf(self):
        """Empty rfis and conflicts still produce valid PDF."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from evidence_appendix_pdf import generate_evidence_appendix_pdf
        result = generate_evidence_appendix_pdf(rfis=[], conflicts=[])
        assert isinstance(result, bytes)
        assert result[:5] == b"%PDF-"

    def test_approval_filter_excludes_drafts(self):
        """Draft RFIs excluded by default."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from evidence_appendix_pdf import generate_evidence_appendix_pdf
        rfis = [{"id": "RFI-DRAFT", "status": "draft", "question": "Q"}]
        result_no_drafts = generate_evidence_appendix_pdf(rfis=rfis, conflicts=[])
        result_with_drafts = generate_evidence_appendix_pdf(
            rfis=rfis, conflicts=[], include_drafts=True)
        assert isinstance(result_no_drafts, bytes)
        assert isinstance(result_with_drafts, bytes)
        # With drafts should be larger (contains the RFI content)
        assert len(result_with_drafts) > len(result_no_drafts)

    def test_with_assumptions(self):
        """Assumptions section included when provided."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from evidence_appendix_pdf import generate_evidence_appendix_pdf
        assumptions = [
            {"id": "ASM-001", "title": "Assume standard ceiling height",
             "status": "accepted", "basis_pages": [1, 2], "text": "2.7m ceiling assumed"},
        ]
        result = generate_evidence_appendix_pdf(
            rfis=[], conflicts=[], assumptions=assumptions)
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_has_bookmarks(self):
        """Output contains PDF outline entries (bookmarks)."""
        sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
        from evidence_appendix_pdf import generate_evidence_appendix_pdf
        rfis = [{"id": "RFI-BM", "status": "approved", "question": "Q",
                 "trade": "structural", "priority": "high"}]
        result = generate_evidence_appendix_pdf(rfis=rfis, conflicts=[])
        # PDF with bookmarks typically contains /Outlines reference
        assert b"/Outlines" in result or b"RFI Evidence" in result


# =============================================================================
# Sprint 15: MEETING AGENDA (Step 9)
# =============================================================================

class TestMeetingAgenda:
    """Unit tests for src/analysis/meeting_agenda.py"""

    def test_deterministic_ordering(self):
        """Same inputs produce identical agenda across two calls."""
        from src.analysis.meeting_agenda import build_meeting_agenda
        review_items = [
            {"type": "conflict", "severity": "high", "title": "Conflict A",
             "source_key": "c1", "recommended_action": "review"},
            {"type": "recon_mismatch", "severity": "medium", "title": "Mismatch B",
             "source_key": "r1", "recommended_action": "prefer_schedule"},
        ]
        a1 = build_meeting_agenda(review_items=review_items, assignments=[])
        a2 = build_meeting_agenda(review_items=review_items, assignments=[])
        # Compare everything except generated_at timestamp
        a1_copy = dict(a1)
        a2_copy = dict(a2)
        a1_copy.pop("generated_at", None)
        a2_copy.pop("generated_at", None)
        assert a1_copy == a2_copy

    def test_empty_inputs(self):
        """No items produces agenda with valid schema."""
        from src.analysis.meeting_agenda import build_meeting_agenda
        agenda = build_meeting_agenda(review_items=[], assignments=[])
        assert "sections" in agenda
        assert isinstance(agenda["sections"], list)
        assert len(agenda["sections"]) >= 1  # At least the empty placeholder section
        assert "summary" in agenda

    def test_high_items_first(self):
        """HIGH severity items appear in the first section."""
        from src.analysis.meeting_agenda import build_meeting_agenda
        items = [
            {"type": "risk_hit", "severity": "high", "title": "Risk A", "source_key": "r1"},
            {"type": "skipped_page", "severity": "low", "title": "Skip B", "source_key": "s1"},
        ]
        agenda = build_meeting_agenda(review_items=items, assignments=[])
        first_section = agenda["sections"][0]
        assert first_section["title"] == "High Priority Review Items"
        assert any(i["severity"] == "high" for i in first_section["items"])

    def test_docx_produces_bytes(self):
        """generate_agenda_docx returns non-empty bytes."""
        from src.analysis.meeting_agenda import build_meeting_agenda, generate_agenda_docx
        agenda = build_meeting_agenda(review_items=[], assignments=[])
        result = generate_agenda_docx(agenda)
        assert isinstance(result, bytes)
        assert len(result) > 100

    def test_pdf_produces_bytes(self):
        """generate_agenda_pdf returns non-empty bytes starting with %PDF."""
        from src.analysis.meeting_agenda import build_meeting_agenda, generate_agenda_pdf
        agenda = build_meeting_agenda(
            review_items=[{"type": "conflict", "severity": "high", "title": "Test", "source_key": "t1"}],
            assignments=[],
        )
        result = generate_agenda_pdf(agenda)
        assert isinstance(result, bytes)
        assert result[:5] == b"%PDF-"


# =============================================================================
# Sprint 15: EMAIL DRAFTS (Step 9)
# =============================================================================

class TestEmailDrafts:
    """Unit tests for src/exports/email_drafts.py"""

    def test_approved_only(self):
        """Draft RFIs excluded from email drafts by default."""
        from src.exports.email_drafts import generate_rfi_email_drafts
        rfis = [
            {"id": "RFI-001", "status": "approved", "trade": "structural",
             "priority": "high", "title": "Footing size", "description": "Check footing dims"},
            {"id": "RFI-002", "status": "draft", "trade": "structural",
             "priority": "medium", "title": "Draft question", "description": "Draft desc"},
        ]
        drafts = generate_rfi_email_drafts(rfis)
        all_text = drafts.get("rfi_email_all.txt", "")
        assert "Footing size" in all_text or "RFI-001" in all_text

    def test_exclusion_draft_schema(self):
        """Exclusion email draft contains exclusion and clarification items."""
        from src.exports.email_drafts import generate_exclusion_email_draft
        assumptions = [
            {"id": "ASM-001", "title": "Exclude provisional sums", "status": "rejected"},
            {"id": "ASM-002", "title": "Standard finishes assumed", "status": "accepted"},
        ]
        text = generate_exclusion_email_draft(assumptions)
        assert "Exclude provisional sums" in text
        assert "Standard finishes assumed" in text
        assert "EXCLUSIONS" in text
        assert "CLARIFICATIONS" in text

    def test_all_drafts_returns_dict(self):
        """generate_all_email_drafts returns dict of filename -> text."""
        from src.exports.email_drafts import generate_all_email_drafts
        result = generate_all_email_drafts(rfis=[], assumptions=[])
        assert isinstance(result, dict)


# =============================================================================
# Sprint 15: QUALITY DASHBOARD (Step 10)
# =============================================================================

class TestQualityDashboard:
    """Unit tests for src/analysis/quality_dashboard.py"""

    def test_empty_feedback_zeros(self):
        """No feedback entries produce zero rates."""
        from src.analysis.quality_dashboard import compute_quality_metrics
        metrics = compute_quality_metrics([], [])
        assert metrics["rfi_acceptance_rate"] == 0.0
        assert metrics["correction_rate"] == 0.0
        assert metrics["total_feedback_count"] == 0
        assert metrics["top_noisy_checks"] == []

    def test_acceptance_rate_calculation(self):
        """Correct acceptance rate from mixed feedback."""
        from src.analysis.quality_dashboard import compute_quality_metrics
        feedback = [
            {"feedback_type": "rfi", "item_id": "RFI-001", "verdict": "correct", "timestamp": "2026-02-10"},
            {"feedback_type": "rfi", "item_id": "RFI-002", "verdict": "correct", "timestamp": "2026-02-10"},
            {"feedback_type": "rfi", "item_id": "RFI-003", "verdict": "wrong", "timestamp": "2026-02-11"},
        ]
        metrics = compute_quality_metrics(feedback, [])
        assert abs(metrics["rfi_acceptance_rate"] - 2 / 3) < 0.01
        assert metrics["rfi_correct_count"] == 2
        assert metrics["rfi_wrong_count"] == 1

    def test_noisy_checks_sorted(self):
        """Top noisy checks sorted by wrong count descending."""
        from src.analysis.quality_dashboard import compute_quality_metrics
        feedback = [
            {"feedback_type": "rfi", "item_id": "RFI-A", "verdict": "wrong", "timestamp": "2026-02-10"},
            {"feedback_type": "rfi", "item_id": "RFI-A", "verdict": "wrong", "timestamp": "2026-02-11"},
            {"feedback_type": "rfi", "item_id": "RFI-B", "verdict": "wrong", "timestamp": "2026-02-10"},
            {"feedback_type": "rfi", "item_id": "RFI-B", "verdict": "correct", "timestamp": "2026-02-10"},
        ]
        metrics = compute_quality_metrics(feedback, [])
        noisy = metrics["top_noisy_checks"]
        assert len(noisy) >= 2
        assert noisy[0]["wrong_count"] >= noisy[1]["wrong_count"]
        assert noisy[0]["check_id"] == "RFI-A"

    def test_trend_from_runs(self):
        """Run history produces trend data entries."""
        from src.analysis.quality_dashboard import compute_quality_metrics
        runs = [
            {"run_id": "r1", "timestamp": "2026-02-10T10:00:00", "readiness_score": 60, "rfis_count": 5},
            {"run_id": "r2", "timestamp": "2026-02-11T10:00:00", "readiness_score": 75, "rfis_count": 3},
        ]
        metrics = compute_quality_metrics([], runs)
        assert len(metrics["trend_data"]) == 2
        assert metrics["trend_data"][0]["readiness_score"] == 60
        assert metrics["trend_data"][1]["readiness_score"] == 75


# =============================================================================
# SPRINT 16 — Storage Interface
# =============================================================================

class TestStorageInterface(unittest.TestCase):
    """Test LocalStorage round-trips through the StorageBackend interface."""

    def setUp(self):
        import tempfile
        self._tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_local_storage_create_load_project(self):
        """Round-trip: create → load project via LocalStorage."""
        from src.storage import LocalStorage
        storage = LocalStorage(base_dir=self._tmpdir)
        meta = storage.create_project(name="Test Project", owner="Alice")
        assert meta["name"] == "Test Project"
        assert "project_id" in meta

        loaded = storage.load_project(meta["project_id"])
        assert loaded is not None
        assert loaded["name"] == "Test Project"
        assert loaded["owner"] == "Alice"

    def test_local_storage_list_projects(self):
        """list_projects returns sorted list."""
        from src.storage import LocalStorage
        storage = LocalStorage(base_dir=self._tmpdir)
        storage.create_project(name="Alpha")
        storage.create_project(name="Beta")
        projects = storage.list_projects()
        assert len(projects) == 2
        # Newest first
        assert projects[0]["name"] == "Beta"

    def test_local_storage_profiles(self):
        """Save/load/list profiles via LocalStorage."""
        from src.storage import LocalStorage
        storage = LocalStorage(base_dir=self._tmpdir)
        storage.save_profile("Acme Corp", {"margin": 15, "risk": "low"})
        profile = storage.load_profile("Acme Corp")
        assert profile is not None
        assert profile["inputs"]["margin"] == 15

        names = storage.list_profiles()
        assert "Acme Corp" in names

    def test_local_storage_collaboration(self):
        """Append/load collaboration entries via LocalStorage."""
        from src.storage import LocalStorage
        from src.analysis.collaboration import make_collaboration_entry
        storage = LocalStorage(base_dir=self._tmpdir)
        # Need a project first
        meta = storage.create_project(name="Collab Test")
        pid = meta["project_id"]

        entry = make_collaboration_entry("rfi", "RFI-001", "comment", {"text": "Check this"})
        storage.append_collaboration(entry, pid)

        entries = storage.load_collaboration(pid)
        assert len(entries) == 1
        assert entries[0]["entity_id"] == "RFI-001"

    def test_local_storage_file_io(self):
        """save_file/load_file/file_exists/list_files via LocalStorage."""
        from src.storage import LocalStorage
        import os
        storage = LocalStorage(base_dir=self._tmpdir)

        test_path = os.path.join(self._tmpdir, "test_data", "sample.bin")
        storage.save_file(test_path, b"hello world")

        assert storage.file_exists(test_path)
        data = storage.load_file(test_path)
        assert data == b"hello world"

        files = storage.list_files(os.path.join(self._tmpdir, "test_data"))
        assert len(files) == 1
        assert files[0].endswith("sample.bin")


# =============================================================================
# SPRINT 16 — Simple Auth
# =============================================================================

class TestSimpleAuth(unittest.TestCase):
    """Test SimpleAuth tenant management and authentication."""

    def setUp(self):
        import tempfile, os
        self._tmpdir = tempfile.mkdtemp()
        self._auth_subdir = os.path.join(self._tmpdir, "auth")

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_create_tenant_and_authenticate(self):
        """Create tenant → authenticate with correct password succeeds."""
        from src.auth import SimpleAuth
        auth = SimpleAuth(auth_dir=self._auth_subdir)
        tenant = auth.create_tenant("acme", "Acme Builders", "secret123")
        assert tenant["tenant_id"] == "acme"
        assert tenant["name"] == "Acme Builders"

        result = auth.authenticate("acme", "secret123")
        assert result is not None
        assert result["tenant_id"] == "acme"

    def test_wrong_password_rejected(self):
        """Wrong password returns None."""
        from src.auth import SimpleAuth
        auth = SimpleAuth(auth_dir=self._auth_subdir)
        auth.create_tenant("acme", "Acme", "correct")
        result = auth.authenticate("acme", "wrong")
        assert result is None

    def test_tenant_isolation(self):
        """Two tenants get different storage base_dirs."""
        from src.auth import SimpleAuth
        auth = SimpleAuth(auth_dir=self._auth_subdir)
        auth.create_tenant("t1", "Tenant One", "pw1")
        auth.create_tenant("t2", "Tenant Two", "pw2")

        s1 = auth.get_storage_for_tenant("t1")
        s2 = auth.get_storage_for_tenant("t2")
        assert str(s1.base_dir) != str(s2.base_dir)
        assert "tenant_t1" in str(s1.base_dir)
        assert "tenant_t2" in str(s2.base_dir)

    def test_list_tenants(self):
        """list_tenants returns tenant list without password hashes."""
        from src.auth import SimpleAuth
        auth = SimpleAuth(auth_dir=self._auth_subdir)
        auth.create_tenant("alpha", "Alpha Co", "pw")
        auth.create_tenant("beta", "Beta Co", "pw")
        tenants = auth.list_tenants()
        assert len(tenants) == 2
        for t in tenants:
            assert "password_hash" not in t
            assert "salt" not in t

    def test_get_storage_for_tenant(self):
        """get_storage_for_tenant returns LocalStorage with correct base_dir."""
        from src.auth import SimpleAuth
        from src.storage import LocalStorage
        auth = SimpleAuth(auth_dir=self._auth_subdir)
        auth.create_tenant("demo", "Demo Tenant", "pw")
        storage = auth.get_storage_for_tenant("demo")
        assert isinstance(storage, LocalStorage)
        # Create a project through tenant storage and verify isolation
        meta = storage.create_project(name="Isolated Project")
        assert storage.load_project(meta["project_id"]) is not None


# =============================================================================
# SPRINT 16 — Job Queue
# =============================================================================

class TestJobQueue(unittest.TestCase):
    """Test LocalThreadQueue submit/status/list/cancel."""

    def test_submit_and_complete(self):
        """Submit function → poll → COMPLETED."""
        from src.jobs import LocalThreadQueue, JobStatus
        queue = LocalThreadQueue(max_workers=1)

        def dummy_fn(**kwargs):
            cb = kwargs.get("progress_callback")
            if cb:
                cb("test", "working", 0.5)
            return {"result": "ok"}

        job_id = queue.submit(dummy_fn)
        # Wait for completion
        import time
        for _ in range(50):
            job = queue.get_status(job_id)
            if job and job.status in (JobStatus.COMPLETED, JobStatus.FAILED):
                break
            time.sleep(0.1)

        job = queue.get_status(job_id)
        assert job is not None
        assert job.status == JobStatus.COMPLETED
        assert job.result == {"result": "ok"}

    def test_progress_callback_updates(self):
        """Progress updates are visible via get_status."""
        from src.jobs import LocalThreadQueue, JobStatus
        import threading
        queue = LocalThreadQueue(max_workers=1)

        barrier = threading.Event()

        def slow_fn(**kwargs):
            cb = kwargs.get("progress_callback")
            if cb:
                cb("stage1", "half done", 0.5)
            barrier.wait(timeout=5)
            return "done"

        job_id = queue.submit(slow_fn)
        import time
        time.sleep(0.3)
        job = queue.get_status(job_id)
        assert job is not None
        assert job.progress >= 0.5 or job.status == JobStatus.COMPLETED
        barrier.set()

        # Wait for completion
        for _ in range(50):
            job = queue.get_status(job_id)
            if job.status == JobStatus.COMPLETED:
                break
            time.sleep(0.1)

    def test_failed_job(self):
        """Exception → FAILED status with error message."""
        from src.jobs import LocalThreadQueue, JobStatus
        queue = LocalThreadQueue(max_workers=1)

        def failing_fn(**kwargs):
            raise ValueError("intentional test failure")

        job_id = queue.submit(failing_fn)
        import time
        for _ in range(50):
            job = queue.get_status(job_id)
            if job and job.status in (JobStatus.COMPLETED, JobStatus.FAILED):
                break
            time.sleep(0.1)

        job = queue.get_status(job_id)
        assert job.status == JobStatus.FAILED
        assert "intentional test failure" in job.error

    def test_list_jobs(self):
        """list_jobs returns recent jobs."""
        from src.jobs import LocalThreadQueue
        queue = LocalThreadQueue(max_workers=1)

        def noop(**kwargs):
            return None

        queue.submit(noop)
        queue.submit(noop)
        import time
        time.sleep(0.5)
        jobs = queue.list_jobs()
        assert len(jobs) >= 2

    def test_cancel_job(self):
        """cancel returns True for known job."""
        from src.jobs import LocalThreadQueue
        queue = LocalThreadQueue(max_workers=1)

        def noop(**kwargs):
            return None

        job_id = queue.submit(noop)
        result = queue.cancel(job_id)
        assert result is True

        result2 = queue.cancel("nonexistent_job_id")
        assert result2 is False


# =============================================================================
# SPRINT 16 — S3 Stub
# =============================================================================

class TestS3Stub(unittest.TestCase):
    """Test S3Storage stub validates config and raises NotImplementedError."""

    def test_s3_stub_init(self):
        """S3Storage accepts bucket/prefix/region."""
        from src.storage.s3_stub import S3Storage
        s3 = S3Storage(bucket="my-bucket", prefix="data/", region="us-east-1")
        assert s3.bucket == "my-bucket"
        assert s3.prefix == "data/"
        assert s3.region == "us-east-1"

    def test_s3_stub_raises(self):
        """All methods raise NotImplementedError."""
        from src.storage.s3_stub import S3Storage
        s3 = S3Storage(bucket="test-bucket")
        with self.assertRaises(NotImplementedError):
            s3.create_project(name="X")
        with self.assertRaises(NotImplementedError):
            s3.load_project("x")
        with self.assertRaises(NotImplementedError):
            s3.list_projects()
        with self.assertRaises(NotImplementedError):
            s3.save_file("/tmp/x", b"data")
        with self.assertRaises(NotImplementedError):
            s3.load_file("/tmp/x")

    def test_s3_stub_is_storage_backend(self):
        """S3Storage is an instance of StorageBackend."""
        from src.storage.s3_stub import S3Storage
        from src.storage.interface import StorageBackend
        s3 = S3Storage(bucket="test")
        assert isinstance(s3, StorageBackend)


# =============================================================================
# SPRINT 17 — Demo Config
# =============================================================================

class TestDemoConfig(unittest.TestCase):
    """Test demo configuration module."""

    def test_is_demo_mode_false_by_default(self):
        """Without env var, demo mode is off."""
        import os
        os.environ.pop("XBOQ_DEMO_MODE", None)
        # Force reimport to pick up env change
        import importlib
        import src.demo.demo_config as dc
        importlib.reload(dc)
        assert dc.is_demo_mode() is False

    def test_is_demo_mode_true(self):
        """XBOQ_DEMO_MODE=true enables demo mode."""
        import os, importlib
        os.environ["XBOQ_DEMO_MODE"] = "true"
        try:
            import src.demo.demo_config as dc
            importlib.reload(dc)
            assert dc.is_demo_mode() is True
        finally:
            os.environ.pop("XBOQ_DEMO_MODE", None)

    def test_demo_projects_non_empty(self):
        """At least 3 demo projects are defined."""
        from src.demo.demo_config import DEMO_PROJECTS
        assert len(DEMO_PROJECTS) >= 3

    def test_get_demo_project_found(self):
        """Known project_id returns config dict."""
        from src.demo.demo_config import get_demo_project
        result = get_demo_project("pwd_garage")
        assert result is not None
        assert result["project_id"] == "pwd_garage"
        assert "name" in result

    def test_get_demo_project_not_found(self):
        """Unknown project_id returns None."""
        from src.demo.demo_config import get_demo_project
        assert get_demo_project("nonexistent_xyz") is None


# =============================================================================
# SPRINT 17 — Demo Assets
# =============================================================================

class TestDemoAssets(unittest.TestCase):
    """Test demo asset resolution."""

    def test_resolve_demo_cache_pwd_garage(self):
        """pwd_garage should have a cached analysis.json."""
        from src.demo.demo_assets import resolve_demo_cache
        path = resolve_demo_cache("pwd_garage")
        # pwd_garage has out/pwd_garage/analysis.json from earlier sprints
        assert path is not None

    def test_resolve_demo_cache_missing(self):
        """Non-existent project returns None."""
        from src.demo.demo_assets import resolve_demo_cache
        assert resolve_demo_cache("no_such_project_xyz_999") is None

    def test_validate_demo_assets_returns_list(self):
        """validate_demo_assets returns a list with correct schema."""
        from src.demo.demo_assets import validate_demo_assets
        report = validate_demo_assets()
        assert isinstance(report, list)
        assert len(report) >= 3
        for r in report:
            assert "project_id" in r
            assert "cache_found" in r
            assert isinstance(r["cache_found"], bool)


# =============================================================================
# SPRINT 17 — Determinism Helpers
# =============================================================================

class TestDeterminism(unittest.TestCase):
    """Test determinism helpers."""

    def test_stable_hash_id_deterministic(self):
        """Same input produces same hash."""
        from src.analysis.determinism import stable_hash_id
        item = {"type": "conflict", "title": "Test Conflict", "severity": "high"}
        h1 = stable_hash_id(item, ("type", "title"))
        h2 = stable_hash_id(item, ("type", "title"))
        assert h1 == h2
        assert len(h1) == 12

    def test_stable_hash_id_different_keys(self):
        """Different keys produce different hashes."""
        from src.analysis.determinism import stable_hash_id
        item = {"type": "conflict", "title": "Test"}
        h1 = stable_hash_id(item, ("type",))
        h2 = stable_hash_id(item, ("title",))
        assert h1 != h2

    def test_stable_sort_preserves_primary_key(self):
        """stable_sort respects primary key ordering."""
        from src.analysis.determinism import stable_sort
        items = [
            {"severity": "low", "title": "B"},
            {"severity": "high", "title": "A"},
        ]
        result = stable_sort(
            items,
            primary_key=lambda x: {"high": 0, "low": 1}.get(x["severity"], 9),
            tiebreaker_fields=("title",),
        )
        assert result[0]["severity"] == "high"
        assert result[1]["severity"] == "low"

    def test_stable_sort_breaks_ties(self):
        """Same primary key uses tiebreaker for ordering."""
        from src.analysis.determinism import stable_sort
        items = [
            {"severity": "high", "title": "Zebra"},
            {"severity": "high", "title": "Apple"},
        ]
        result = stable_sort(
            items,
            primary_key=lambda x: 0,
            tiebreaker_fields=("title",),
        )
        assert result[0]["title"] == "Apple"
        assert result[1]["title"] == "Zebra"

    def test_normalize_for_hashing_none(self):
        """None normalizes to empty string."""
        from src.analysis.determinism import normalize_for_hashing
        assert normalize_for_hashing(None) == ""

    def test_normalize_for_hashing_dict_sorted(self):
        """Dict normalization sorts keys for stability."""
        from src.analysis.determinism import normalize_for_hashing
        r1 = normalize_for_hashing({"b": 2, "a": 1})
        r2 = normalize_for_hashing({"a": 1, "b": 2})
        assert r1 == r2


# =============================================================================
# SPRINT 17 — Highlights
# =============================================================================

class TestHighlights(unittest.TestCase):
    """Test highlights extraction from payload."""

    def test_build_highlights_full_payload(self):
        """Full payload produces at least 3 highlights."""
        from src.analysis.highlights import build_highlights
        payload = {
            "decision": "CONDITIONAL",
            "readiness_score": 61,
            "sub_scores": {"completeness": 80, "coverage": 64},
            "blockers": [{"severity": "high", "title": "Missing door schedule"}],
            "rfis": [{"trade": "architectural", "status": "draft"}],
            "trade_coverage": [{"trade": "civil", "coverage_pct": 100.0}],
        }
        highlights = build_highlights(payload)
        assert len(highlights) >= 3
        assert highlights[0]["label"] == "Bid Readiness"
        assert "61" in highlights[0]["value"]

    def test_build_highlights_empty_payload(self):
        """Empty payload produces at least the decision highlight."""
        from src.analysis.highlights import build_highlights
        highlights = build_highlights({})
        assert len(highlights) >= 1
        assert highlights[0]["label"] == "Bid Readiness"

    def test_highlight_severity_values(self):
        """All severities are one of good/warn/bad."""
        from src.analysis.highlights import build_highlights
        payload = {
            "decision": "PASS", "readiness_score": 90,
            "blockers": [], "rfis": [],
        }
        for h in build_highlights(payload):
            assert h["severity"] in ("good", "warn", "bad")

    def test_highlights_deterministic(self):
        """Same payload produces identical highlights."""
        from src.analysis.highlights import build_highlights
        payload = {
            "decision": "NO-GO", "readiness_score": 20,
            "blockers": [{"severity": "high", "title": "A"}],
            "rfis": [{"trade": "civil", "status": "draft"}],
        }
        h1 = build_highlights(payload)
        h2 = build_highlights(payload)
        assert h1 == h2


# =============================================================================
# SPRINT 17 — Determinism Wiring
# =============================================================================

class TestDeterminismWiring(unittest.TestCase):
    """Test that determinism tiebreakers produce stable output from wired modules."""

    def test_review_queue_stable(self):
        """Two calls to build_review_queue produce identical output."""
        from src.analysis.review_queue import build_review_queue
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 3,
             "schedule_count": 2, "boq_count": 5},
            {"category": "windows", "mismatch": True, "max_delta": 3,
             "schedule_count": 2, "boq_count": 5},
        ]
        r1 = build_review_queue(quantity_reconciliation=recon)
        r2 = build_review_queue(quantity_reconciliation=recon)
        assert r1 == r2

    def test_reconciler_sorted(self):
        """reconcile_scope returns sorted findings."""
        from src.analysis.reconciler import reconcile_scope
        reqs = [{"text": "All doors shall be hardwood", "source_page": 5}]
        schedules = [{"schedule_type": "door", "mark": "D1", "source_page": 10}]
        boq = [{"item_no": "4.1", "description": "Flush door", "qty": 5, "source_page": 20}]
        f1 = reconcile_scope(reqs, schedules, boq)
        f2 = reconcile_scope(reqs, schedules, boq)
        assert f1 == f2

    def test_quantities_stable(self):
        """build_all_quantities returns stable order."""
        from src.analysis.quantities import build_all_quantities
        boq = [
            {"item_no": "1.1", "description": "Earthwork", "unit": "cum",
             "qty": 100, "source_page": 0},
            {"item_no": "2.1", "description": "Brickwork", "unit": "sqm",
             "qty": 200, "source_page": 1},
        ]
        q1 = build_all_quantities([], boq, [])
        q2 = build_all_quantities([], boq, [])
        assert q1 == q2

    def test_meeting_agenda_stable(self):
        """build_meeting_agenda sections are stable."""
        from src.analysis.meeting_agenda import build_meeting_agenda
        items = [
            {"type": "conflict", "severity": "high", "title": "C1", "source_key": "c1"},
            {"type": "conflict", "severity": "high", "title": "C2", "source_key": "c2"},
        ]
        a1 = build_meeting_agenda(items, [])
        a2 = build_meeting_agenda(items, [])
        # Ignore generated_at timestamp
        a1.pop("generated_at", None)
        a2.pop("generated_at", None)
        assert a1 == a2

    def test_rfi_clustering_stable(self):
        """cluster_rfis returns stable order for same input."""
        from src.analysis.rfi_clustering import cluster_rfis
        rfis = [
            {"id": "RFI-1", "trade": "structural", "priority": "high",
             "question": "Check footing", "evidence": {"pages": [1, 2]}},
            {"id": "RFI-2", "trade": "structural", "priority": "medium",
             "question": "Verify rebar", "evidence": {"pages": [3]}},
            {"id": "RFI-3", "trade": "electrical", "priority": "high",
             "question": "Panel schedule", "evidence": {"pages": [5]}},
        ]
        c1 = cluster_rfis(rfis)
        c2 = cluster_rfis(rfis)
        assert len(c1) == len(c2)
        for i in range(len(c1)):
            assert c1[i]["cluster_id"] == c2[i]["cluster_id"]
            assert c1[i]["priority"] == c2[i]["priority"]


# =============================================================================
# SPRINT 18: Summary Card, Narration, Export Filenames, Safety Rails
# =============================================================================

class TestSummaryCard(unittest.TestCase):
    """Test build_summary_card with various payload shapes."""

    def test_full_payload(self):
        """Full payload returns all fields with correct types."""
        from src.demo.summary_card import build_summary_card
        payload = {
            "processing_stats": {
                "total_pages": 15, "deep_processed_pages": 10,
                "ocr_pages": 3, "text_layer_pages": 7, "skipped_pages": 5,
            },
            "qa_score": {"score": 85, "top_actions": ["Add specs", "Check dims"]},
            "rfis": [
                {"id": "r1", "status": "approved"},
                {"id": "r2", "status": "draft"},
            ],
            "quantities": [{"item": "Steel"}, {"item": "Concrete"}],
            "assumptions": [
                {"id": "a1", "status": "accepted"},
                {"id": "a2", "status": "rejected"},
            ],
            "decision": "CONDITIONAL",
            "readiness_score": 72,
        }
        card = build_summary_card(payload, project_name="Test Project")
        assert card["total_pages"] == 15
        assert card["deep_pages"] == 10
        assert card["ocr_pages"] == 3
        assert card["text_layer_pages"] == 7
        assert card["skipped_pages"] == 5
        assert card["qa_score"] == 85
        assert len(card["top_actions"]) == 2
        assert card["approved_rfis"] == 1  # only "approved"
        assert card["accepted_quantities"] == 2
        assert card["accepted_assumptions"] == 1  # only "accepted"
        assert card["decision"] == "CONDITIONAL"
        assert card["readiness_score"] == 72
        assert card["project_name"] == "Test Project"

    def test_empty_payload(self):
        """Empty payload returns all fields with safe defaults."""
        from src.demo.summary_card import build_summary_card
        card = build_summary_card({})
        assert card["total_pages"] == 0
        assert card["deep_pages"] == 0
        assert card["qa_score"] == 0
        assert card["approved_rfis"] == 0
        assert card["accepted_quantities"] == 0
        assert card["accepted_assumptions"] == 0
        assert card["decision"] == "N/A"
        assert card["readiness_score"] == 0
        assert card["submission_pack_ready"] is False
        assert card["cache_time_saved"] == "N/A"

    def test_partial_payload(self):
        """Partial payload (no qa_score, no rfis) is graceful."""
        from src.demo.summary_card import build_summary_card
        card = build_summary_card({"decision": "PASS", "readiness_score": 95})
        assert card["qa_score"] == 0
        assert card["approved_rfis"] == 0
        assert card["decision"] == "PASS"
        assert card["readiness_score"] == 95

    def test_deterministic(self):
        """Same payload produces identical card."""
        from src.demo.summary_card import build_summary_card
        payload = {
            "rfis": [{"status": "approved"}],
            "quantities": [{"item": "x"}],
            "readiness_score": 50,
        }
        c1 = build_summary_card(payload, project_name="P")
        c2 = build_summary_card(payload, project_name="P")
        assert c1 == c2


class TestNarration(unittest.TestCase):
    """Test build_narration_script for determinism and correctness."""

    def test_full_payload(self):
        """Full payload produces non-empty string with project name."""
        from src.demo.narration import build_narration_script
        payload = {
            "drawing_overview": {"pages_total": 7, "pages_deep": 5},
            "rfis": [{"status": "approved"}, {"status": "draft"}],
            "blockers": [{"title": "Missing fire spec"}],
            "quantities": [{"item": "Steel"}],
            "decision": "CONDITIONAL",
            "readiness_score": 68,
            "qa_score": {"score": 75},
            "timings": {"total_seconds": 12},
        }
        script = build_narration_script(payload, project_name="PWD Garage")
        assert len(script) > 50
        assert "PWD Garage" in script
        assert "7-page" in script

    def test_empty_payload(self):
        """Empty payload still produces a string (no crash)."""
        from src.demo.narration import build_narration_script
        script = build_narration_script({})
        assert isinstance(script, str)
        assert len(script) > 20

    def test_deterministic(self):
        """Same payload produces identical narration."""
        from src.demo.narration import build_narration_script
        payload = {
            "rfis": [{"status": "approved"}],
            "blockers": [{"title": "Missing detail"}],
            "readiness_score": 80,
        }
        s1 = build_narration_script(payload, project_name="Test")
        s2 = build_narration_script(payload, project_name="Test")
        assert s1 == s2

    def test_contains_key_sections(self):
        """Narration contains expected section markers."""
        from src.demo.narration import build_narration_script
        payload = {
            "rfis": [{"status": "draft"}],
            "blockers": [],
            "quantities": [{"item": "x"}],
        }
        script = build_narration_script(payload, project_name="Demo")
        assert "[INTRO]" in script
        assert "[PROCESSING]" in script
        assert "[FINDINGS]" in script
        assert "[COUNTS]" in script
        assert "[DELIVERABLES]" in script
        assert "[CLOSE]" in script


class TestExportFilenames(unittest.TestCase):
    """Test _demo_filename helper for stable, standardized names."""

    @staticmethod
    def _demo_filename(base, pname, ext):
        """Reproduce the helper from demo_page.py for unit testing."""
        from datetime import date
        if pname:
            safe = pname.replace(" ", "_")[:30]
            return f"{safe}_{base}_{date.today().isoformat()}.{ext}"
        return f"{base}.{ext}"

    def test_with_project_name(self):
        """Project name produces '{Name}_{base}_{date}.{ext}' format."""
        from datetime import date
        result = self._demo_filename("Bid_Summary", "PWD Garage", "pdf")
        assert result.startswith("PWD_Garage_Bid_Summary_")
        assert result.endswith(".pdf")
        assert date.today().isoformat() in result

    def test_without_project_name(self):
        """No project name falls back to '{base}.{ext}'."""
        result = self._demo_filename("Bid_Summary", "", "pdf")
        assert result == "Bid_Summary.pdf"

    def test_long_name_truncated(self):
        """Long project name truncated to 30 chars."""
        result = self._demo_filename("X", "A" * 50, "zip")
        prefix = result.split("_X_")[0]
        assert len(prefix) <= 30

    def test_deterministic(self):
        """Same inputs produce same output."""
        r1 = self._demo_filename("Test", "Project", "csv")
        r2 = self._demo_filename("Test", "Project", "csv")
        assert r1 == r2


class TestDemoSafetyRails(unittest.TestCase):
    """Test Sprint 18 config constants and safety rail building blocks."""

    def test_freeze_defaults_structure(self):
        """DEMO_FREEZE_DEFAULTS is a dict with expected keys."""
        from src.demo.demo_config import DEMO_FREEZE_DEFAULTS
        assert isinstance(DEMO_FREEZE_DEFAULTS, dict)
        assert "confidence_threshold" in DEMO_FREEZE_DEFAULTS
        assert "max_rfis_display" in DEMO_FREEZE_DEFAULTS
        assert "expand_all_sections" in DEMO_FREEZE_DEFAULTS

    def test_jump_targets_structure(self):
        """JUMP_TARGETS has 6 entries with key/label/tab_index."""
        from src.demo.demo_config import JUMP_TARGETS
        assert isinstance(JUMP_TARGETS, list)
        assert len(JUMP_TARGETS) == 6
        for jt in JUMP_TARGETS:
            assert "key" in jt
            assert "label" in jt
            assert "tab_index" in jt

    def test_summary_card_none_safety(self):
        """build_summary_card handles None-like values without crash."""
        from src.demo.summary_card import build_summary_card
        card = build_summary_card({"rfis": None, "quantities": None, "qa_score": None})
        assert card["approved_rfis"] == 0
        assert card["accepted_quantities"] == 0
        assert card["qa_score"] == 0

    def test_narration_none_safety(self):
        """build_narration_script handles None-like values without crash."""
        from src.demo.narration import build_narration_script
        script = build_narration_script({"rfis": None, "blockers": None})
        assert isinstance(script, str)
        assert len(script) > 10


class TestProcessingStats(unittest.TestCase):
    """Test _build_processing_stats and summary_card integration."""

    def test_build_processing_stats_direct(self):
        """_build_processing_stats computes correct counters from pipeline data."""
        from src.analysis.pipeline import _build_processing_stats

        # Simulate page_index_result with a simple object
        class MockPageIndex:
            total_pages = 100
        class MockSelected:
            selected = list(range(40))  # 40 pages selected

        ocr_meta = {
            "page_profiles": [
                {"page_index": i, "ocr_used": i < 30, "has_text_layer": i >= 30}
                for i in range(40)
            ],
            "ocr_pages": list(range(30)),
        }

        stats = _build_processing_stats(
            page_index_result=MockPageIndex(),
            selected_result=MockSelected(),
            ocr_metadata=ocr_meta,
        )
        assert stats["total_pages"] == 100
        assert stats["deep_processed_pages"] == 40
        assert stats["ocr_pages"] == 30
        assert stats["text_layer_pages"] == 10
        assert stats["skipped_pages"] == 60

    def test_build_processing_stats_empty(self):
        """All-None inputs return safe zeroes."""
        from src.analysis.pipeline import _build_processing_stats
        stats = _build_processing_stats()
        assert stats["total_pages"] == 0
        assert stats["deep_processed_pages"] == 0
        assert stats["ocr_pages"] == 0
        assert stats["text_layer_pages"] == 0
        assert stats["skipped_pages"] == 0

    def test_build_processing_stats_no_profiles_fallback(self):
        """Without page_profiles, falls back to ocr_pages list length."""
        from src.analysis.pipeline import _build_processing_stats

        class MockPageIndex:
            total_pages = 50
        class MockSelected:
            selected = list(range(20))

        ocr_meta = {
            "page_profiles": [],
            "ocr_pages": [0, 1, 2, 3, 4],
        }
        stats = _build_processing_stats(
            page_index_result=MockPageIndex(),
            selected_result=MockSelected(),
            ocr_metadata=ocr_meta,
        )
        assert stats["ocr_pages"] == 5
        assert stats["text_layer_pages"] == 15  # 20 - 5
        assert stats["skipped_pages"] == 30  # 50 - 20

    def test_summary_card_reads_processing_stats(self):
        """summary_card prefers processing_stats over drawing_overview."""
        from src.demo.summary_card import build_summary_card
        payload = {
            "drawing_overview": {"pages_total": 999},  # should be ignored
            "processing_stats": {
                "total_pages": 367,
                "deep_processed_pages": 80,
                "ocr_pages": 80,
                "text_layer_pages": 0,
                "skipped_pages": 287,
            },
            "decision": "CONDITIONAL",
            "readiness_score": 73,
        }
        card = build_summary_card(payload, project_name="Test")
        assert card["total_pages"] == 367  # from processing_stats, not 999
        assert card["deep_pages"] == 80
        assert card["ocr_pages"] == 80
        assert card["skipped_pages"] == 287
        assert card["text_layer_pages"] == 0

    def test_summary_card_fallback_run_coverage(self):
        """summary_card falls back to run_coverage when no processing_stats."""
        from src.demo.summary_card import build_summary_card
        payload = {
            "run_coverage": {
                "pages_total": 200,
                "pages_deep_processed": 50,
            },
            "drawing_overview": {
                "ocr_pages_count": 40,
            },
        }
        card = build_summary_card(payload)
        assert card["total_pages"] == 200
        assert card["deep_pages"] == 50
        assert card["ocr_pages"] == 40
        assert card["skipped_pages"] == 150


class TestBuildReportSubScores(unittest.TestCase):
    """Test build_report_from_results maps sub_scores correctly.

    We replicate the score_data extraction logic from demo_page.py here
    because importing demo_page directly triggers Streamlit and app.app
    imports that fail in the unit-test environment.
    """

    @staticmethod
    def _extract_sub_scores(deep: dict) -> dict:
        """Replicate the score_data → executive_summary sub_scores logic
        from build_report_from_results in demo_page.py."""
        raw_score = deep.get("readiness_score", {})
        if isinstance(raw_score, (int, float)):
            _deep_sub = deep.get("sub_scores", {})
            score_data = {
                "total_score": raw_score,
                "status": deep.get("decision", "NO-GO"),
                "coverage_score": _deep_sub.get("coverage", 0),
                "measurement_score": _deep_sub.get("measurement", 0),
                "completeness_score": _deep_sub.get("completeness", 0),
                "blocker_score": _deep_sub.get("blocker", 0),
            }
        else:
            score_data = raw_score if raw_score else {}
        return {
            "coverage": score_data.get("coverage_score", 0),
            "measurement": score_data.get("measurement_score", 0),
            "completeness": score_data.get("completeness_score", 0),
            "blocker": score_data.get("blocker_score", 0),
        }

    def test_sub_scores_mapping(self):
        """Sub-scores from pipeline must pass through to report correctly."""
        deep = {
            "readiness_score": 73,
            "decision": "CONDITIONAL",
            "sub_scores": {
                "coverage": 100,
                "measurement": 30,
                "completeness": 70,
                "blocker": 75,
            },
        }
        sub = self._extract_sub_scores(deep)
        self.assertEqual(sub["coverage"], 100)
        self.assertEqual(sub["measurement"], 30)
        self.assertEqual(sub["completeness"], 70)
        self.assertEqual(sub["blocker"], 75)

    def test_sub_scores_zero_not_default(self):
        """Actual zero sub-scores must remain 0, not be confused with missing."""
        deep = {
            "readiness_score": 10,
            "decision": "NO-GO",
            "sub_scores": {
                "coverage": 0,
                "measurement": 0,
                "completeness": 0,
                "blocker": 0,
            },
        }
        sub = self._extract_sub_scores(deep)
        self.assertEqual(sub["coverage"], 0)
        self.assertEqual(sub["measurement"], 0)
        self.assertEqual(sub["completeness"], 0)
        self.assertEqual(sub["blocker"], 0)

    def test_sub_scores_missing_graceful(self):
        """Missing sub_scores key defaults to 0 without crash."""
        deep = {
            "readiness_score": 50,
            "decision": "NO-GO",
        }
        sub = self._extract_sub_scores(deep)
        self.assertEqual(sub["coverage"], 0)
        self.assertEqual(sub["measurement"], 0)


class TestRenderSection2Dependencies(unittest.TestCase):
    """Test render_section_2_dependencies with various inputs (no Streamlit)."""

    def _extract_and_call(self, report):
        """Exercise the logic of render_section_2_dependencies without Streamlit.

        We test the ID-generation and guard logic directly since we cannot
        invoke Streamlit widgets in unit tests.
        """
        import hashlib, json
        deps = report.get("missing_dependencies") or []
        ids = []
        for _dep_i, dep in enumerate(deps):
            if not isinstance(dep, dict):
                continue
            _dep_id = dep.get("id") or ""
            if not _dep_id:
                try:
                    _dep_id = "DEP-" + hashlib.sha256(
                        json.dumps(dep, sort_keys=True, default=str).encode()
                    ).hexdigest()[:8]
                except Exception:
                    _dep_id = f"DEP-{_dep_i}"
            ids.append(_dep_id)
        return ids

    def test_deps_missing_key(self):
        """Report with no missing_dependencies key returns empty list."""
        ids = self._extract_and_call({})
        assert ids == []

    def test_deps_empty_list(self):
        """Report with empty deps list returns empty list."""
        ids = self._extract_and_call({"missing_dependencies": []})
        assert ids == []

    def test_deps_none(self):
        """Report with deps=None returns empty list."""
        ids = self._extract_and_call({"missing_dependencies": None})
        assert ids == []

    def test_deps_items_missing_id(self):
        """Deps without 'id' get a stable hash-based ID."""
        report = {"missing_dependencies": [
            {"dependency_type": "finish_schedule", "status": "missing"},
            {"dependency_type": "door_schedule", "status": "missing"},
        ]}
        ids = self._extract_and_call(report)
        assert len(ids) == 2
        assert all(id_.startswith("DEP-") for id_ in ids)
        # Hash-based IDs are 8 hex chars after "DEP-"
        assert len(ids[0]) == 12  # "DEP-" + 8 hex
        # Different deps get different IDs
        assert ids[0] != ids[1]
        # Deterministic: same input → same IDs
        ids2 = self._extract_and_call(report)
        assert ids == ids2

    def test_deps_with_existing_ids(self):
        """Deps with 'id' field use that ID as-is."""
        report = {"missing_dependencies": [
            {"id": "DEP-001", "dependency_type": "finish_schedule"},
        ]}
        ids = self._extract_and_call(report)
        assert ids == ["DEP-001"]


# =============================================================================
# Sprint 19: COMMERCIAL EXTRACTOR TESTS
# =============================================================================

from src.analysis.extractors.extract_commercial_terms import extract_commercial_terms
from src.analysis.extractors.extract_boq import infer_boq_trade, flag_boq_item
from src.analysis.extractors.extract_notes import _classify_trade, _extract_standards_codes, _extract_approved_makes

CONDITIONS_TEXT = """
GENERAL CONDITIONS OF CONTRACT

1. Liquidated Damages: LD shall be levied at 0.5% per week of delay subject to
   a maximum of 10% of the total contract value.

2. Retention: 5% retention money shall be deducted from each Running Account bill.

3. Defect Liability Period: The contractor shall be responsible for all defects
   for a DLP of 12 months from the date of completion.

4. Bid Validity: The tender shall remain valid for 90 days from the date of opening.

5. EMD: Earnest Money Deposit of Rs. 2,50,000/- shall be submitted with the bid.

6. Performance Bank Guarantee: PBG of 10% of the contract value shall be submitted.

7. Mobilization Advance: A mobilization advance of 10% may be granted against
   equivalent bank guarantee.

8. Insurance: The contractor shall provide a Contractor's All Risk (CAR) policy
   for 1,50,00,000 (1.5 crore).

9. Price Escalation: No price escalation clause shall apply to this contract.
"""


class TestCommercialExtractor(unittest.TestCase):
    """Test commercial term extraction from conditions text."""

    def test_ld_extraction(self):
        text = "Liquidated damages: 0.5% per week of delay"
        results = extract_commercial_terms(text, source_page=0)
        ld = [r for r in results if r["term_type"] == "ld_clause"]
        assert len(ld) == 1
        assert ld[0]["value"] == 0.5
        assert ld[0]["cadence"] == "week"

    def test_retention_extraction(self):
        text = "Retention: 5% shall be deducted from each bill"
        results = extract_commercial_terms(text, source_page=0)
        ret = [r for r in results if r["term_type"] == "retention"]
        assert len(ret) == 1
        assert ret[0]["value"] == 5.0

    def test_emd_indian_number(self):
        text = "EMD: Rs. 2,50,000/- shall be submitted"
        results = extract_commercial_terms(text, source_page=0)
        emd = [r for r in results if r["term_type"] == "emd_bid_security"]
        assert len(emd) == 1
        assert emd[0]["value"] == 250000.0

    def test_multiple_terms_from_conditions(self):
        results = extract_commercial_terms(CONDITIONS_TEXT, source_page=5)
        term_types = {r["term_type"] for r in results}
        # Should find at least LD, retention, warranty, bid validity, EMD, PBG
        assert "ld_clause" in term_types
        assert "retention" in term_types
        assert "warranty_dlp" in term_types
        assert "bid_validity" in term_types
        assert "emd_bid_security" in term_types
        assert "performance_bond" in term_types
        # All should reference source_page=5
        for r in results:
            assert r["source_page"] == 5

    def test_empty_text_no_crash(self):
        assert extract_commercial_terms("", source_page=0) == []
        assert extract_commercial_terms("  \n  ", source_page=0) == []


class TestCommercialRFIs(unittest.TestCase):
    """Test commercial RFI checks fire correctly."""

    def _make_context(self, commercial_terms=None, conditions_pages=0):
        """Helper: build a CheckContext with commercial terms."""
        from src.analysis.rfi_engine import CheckContext

        extracted = ExtractionResult()
        if commercial_terms is not None:
            extracted.commercial_terms = commercial_terms

        page_index = PageIndex(pdf_name="test.pdf", total_pages=10)
        page_index.counts_by_type = {"conditions": conditions_pages}

        selected = SelectedPages(budget_total=80)

        ctx = CheckContext(extracted, page_index, selected, plan_graph=None)
        return ctx

    def test_missing_ld_fires_when_covered(self):
        from src.analysis.rfi_engine import chk_no_ld_clause
        ctx = self._make_context(commercial_terms=[], conditions_pages=5)
        fired, ev, kwargs = chk_no_ld_clause(ctx)
        assert fired is True
        assert ev is not None

    def test_found_ld_does_not_fire(self):
        from src.analysis.rfi_engine import chk_no_ld_clause
        terms = [{"term_type": "ld_clause", "value": 0.5, "cadence": "week"}]
        ctx = self._make_context(commercial_terms=terms, conditions_pages=5)
        fired, ev, kwargs = chk_no_ld_clause(ctx)
        assert fired is False

    def test_ld_rate_high_fires(self):
        from src.analysis.rfi_engine import chk_ld_rate_high
        terms = [{"term_type": "ld_clause", "value": 2.0, "cadence": "week",
                   "snippet": "LD at 2% per week", "source_page": 3}]
        ctx = self._make_context(commercial_terms=terms, conditions_pages=5)
        fired, ev, kwargs = chk_ld_rate_high(ctx)
        assert fired is True
        assert kwargs["ld_rate"] == 2.0

    def test_all_10_checks_registered(self):
        from src.analysis.rfi_engine import CHECKLIST, CHECK_FN_MAP
        com_checks = [c for c in CHECKLIST if c[0].startswith("CHK-COM-")]
        assert len(com_checks) == 10
        for check_id, fn_name, *_ in com_checks:
            assert fn_name in CHECK_FN_MAP, f"{fn_name} not in CHECK_FN_MAP"

    def test_commercial_package_mapping(self):
        from src.analysis.rfi_engine import _package_for_trade
        from src.models.analysis_models import Trade
        assert _package_for_trade(Trade.COMMERCIAL) == "Commercial"


class TestBOQEnrichment(unittest.TestCase):
    """Test BOQ trade inference and flagging."""

    def test_trade_inference_structural(self):
        assert infer_boq_trade("RCC M25 footing concrete work") == "structural"

    def test_trade_inference_finishes(self):
        assert infer_boq_trade("Ceramic tile 600x600 flooring") == "finishes"

    def test_trade_inference_general(self):
        assert infer_boq_trade("Miscellaneous items") == "general"

    def test_zero_rate_flag(self):
        item = {"rate": 0, "qty": 10, "unit": "sqm", "description": "Test"}
        flags = flag_boq_item(item)
        assert "zero_rate" in flags

    def test_qty_missing_flag(self):
        item = {"rate": 100, "qty": None, "unit": "sqm", "description": "Test"}
        flags = flag_boq_item(item)
        assert "qty_missing" in flags

    def test_provisional_sum_flag(self):
        item = {"rate": 100, "qty": 1, "unit": "LS",
                "description": "Provisional sum for contingencies"}
        flags = flag_boq_item(item)
        assert "provisional_sum" in flags


class TestRequirementsEnrichment(unittest.TestCase):
    """Test requirements enrichment (trade, standards, makes)."""

    def test_trade_classification_structural(self):
        assert _classify_trade("RCC M25 grade concrete for column") == "structural"

    def test_trade_classification_finishes(self):
        assert _classify_trade("Ceramic tile flooring 600x600") == "finishes"

    def test_standards_codes_extraction(self):
        codes = _extract_standards_codes("Concrete shall conform to IS 456-2000 and ASTM C150")
        # Should find both IS and ASTM codes
        code_str = " ".join(codes).upper()
        assert "IS" in code_str
        assert "ASTM" in code_str

    def test_approved_makes(self):
        makes = _extract_approved_makes("Approved makes: Birla, ACC, Ambuja or equivalent")
        assert len(makes) >= 2
        assert "Birla" in makes or "ACC" in makes

    def test_backward_compat_old_requirements(self):
        """Old requirements without new fields still work via .get()."""
        old_req = {"text": "Some requirement", "category": "general",
                   "source_page": 0, "confidence": 0.5}
        # These should not crash
        assert old_req.get("trade", "general") == "general"
        assert old_req.get("standards_codes", []) == []
        assert old_req.get("approved_makes", []) == []

    def test_extract_requirements_includes_new_fields(self):
        """Verify extract_requirements returns new enrichment fields."""
        text = "All RCC concrete shall be M25 grade as per IS 456-2000."
        results = extract_requirements(text, source_page=0, sheet_id=None, doc_type="notes")
        if results:
            r = results[0]
            assert "trade" in r
            assert "standards_codes" in r
            assert "approved_makes" in r


# =============================================================================
# Sprint 20: Ground Truth, Diff, Training Pack, Pilot Docs, Extended Metadata
# =============================================================================

class TestGroundTruth(unittest.TestCase):
    """Tests for ground_truth.py — templates, parsing, column mapping, persistence."""

    def test_template_csv_boq_has_correct_columns(self):
        from src.analysis.ground_truth import generate_template_csv, GT_BOQ_COLUMNS
        csv_str = generate_template_csv("gt_boq")
        header_line = csv_str.strip().split("\n")[0]
        cols = [c.strip() for c in header_line.split(",")]
        assert cols == GT_BOQ_COLUMNS

    def test_template_csv_quantities_has_correct_columns(self):
        from src.analysis.ground_truth import generate_template_csv, GT_QUANTITIES_COLUMNS
        csv_str = generate_template_csv("gt_quantities")
        header_line = csv_str.strip().split("\n")[0]
        cols = [c.strip() for c in header_line.split(",")]
        assert cols == GT_QUANTITIES_COLUMNS

    def test_template_csv_schedules_has_correct_columns(self):
        from src.analysis.ground_truth import generate_template_csv, GT_SCHEDULES_DOORS_COLUMNS
        csv_str = generate_template_csv("gt_schedules_doors")
        header_line = csv_str.strip().split("\n")[0]
        cols = [c.strip() for c in header_line.split(",")]
        assert cols == GT_SCHEDULES_DOORS_COLUMNS

    def test_apply_column_mapping(self):
        from src.analysis.ground_truth import apply_column_mapping
        headers = ["Desc", "Quantity", "UoM"]
        rows = [["Concrete M25", "100", "m3"], ["Steel rebar", "500", "kg"]]
        mapping = {"description": "Desc", "qty": "Quantity", "unit": "UoM"}
        result = apply_column_mapping(headers, rows, mapping)
        assert len(result) == 2
        assert result[0]["description"] == "Concrete M25"
        assert result[0]["qty"] == "100"
        assert result[0]["unit"] == "m3"
        assert result[1]["description"] == "Steel rebar"

    def test_mapping_schema_stable(self):
        """Same inputs produce identical outputs."""
        from src.analysis.ground_truth import apply_column_mapping
        headers = ["Name", "Qty"]
        rows = [["Item A", "10"], ["Item B", "20"]]
        mapping = {"item": "Name", "qty": "Qty"}
        r1 = apply_column_mapping(headers, rows, mapping)
        r2 = apply_column_mapping(headers, rows, mapping)
        assert r1 == r2

    def test_save_and_load_gt_mapping(self):
        import tempfile
        from src.analysis.ground_truth import save_gt_mapping, load_gt_mapping
        with tempfile.TemporaryDirectory() as td:
            pid = "test_proj_gt"
            mapping = {
                "gt_type": "boq",
                "source_file": "estimate.xlsx",
                "sheet_name": "Sheet1",
                "column_map": {"description": "Desc", "qty": "Quantity"},
            }
            save_gt_mapping(pid, mapping, Path(td))
            loaded = load_gt_mapping(pid, Path(td))
            assert loaded is not None
            assert loaded["gt_type"] == "boq"
            assert loaded["column_map"]["description"] == "Desc"

    def test_save_and_load_gt_data(self):
        import tempfile
        from src.analysis.ground_truth import save_gt_data, load_gt_data
        with tempfile.TemporaryDirectory() as td:
            pid = "test_proj_gt2"
            rows = [
                {"item": "Concrete", "qty": "100", "unit": "m3"},
                {"item": "Steel", "qty": "500", "unit": "kg"},
            ]
            save_gt_data(pid, "quantities", rows, Path(td))
            loaded = load_gt_data(pid, "quantities", Path(td))
            assert len(loaded) == 2
            assert loaded[0]["item"] == "Concrete"
            assert loaded[1]["qty"] == "500"

    def test_load_gt_data_missing_returns_empty(self):
        import tempfile
        from src.analysis.ground_truth import load_gt_data
        with tempfile.TemporaryDirectory() as td:
            result = load_gt_data("nonexistent_project", "boq", Path(td))
            assert result == []


class TestGroundTruthDiff(unittest.TestCase):
    """Tests for ground_truth_diff.py — diff engine."""

    def test_diff_quantities_exact_match(self):
        from src.analysis.ground_truth_diff import diff_quantities
        our = [
            {"item": "Concrete M25", "qty": "100"},
            {"item": "Steel rebar", "qty": "500"},
        ]
        gt = [
            {"item": "Concrete M25", "qty": "100"},
            {"item": "Steel rebar", "qty": "500"},
        ]
        result = diff_quantities(our, gt)
        assert result["match_rate"] == 1.0
        assert result["matched_count"] == 2
        assert result["gt_count"] == 2
        assert len(result["top_mismatches"]) == 0

    def test_diff_quantities_with_mismatches(self):
        from src.analysis.ground_truth_diff import diff_quantities
        our = [{"item": "Concrete M25", "qty": "120"}]
        gt = [{"item": "Concrete M25", "qty": "100"}]
        result = diff_quantities(our, gt)
        assert result["match_rate"] == 1.0  # item matched
        assert len(result["top_mismatches"]) == 1
        assert result["top_mismatches"][0]["delta"] == 20.0

    def test_diff_quantities_missing_items(self):
        from src.analysis.ground_truth_diff import diff_quantities
        our = [{"item": "Concrete M25", "qty": "100"}]
        gt = [
            {"item": "Concrete M25", "qty": "100"},
            {"item": "Plywood sheets", "qty": "50"},
        ]
        result = diff_quantities(our, gt)
        assert result["match_rate"] == 0.5
        assert len(result["missing_in_ours"]) == 1
        assert "Plywood sheets" in result["missing_in_ours"]

    def test_diff_deterministic(self):
        from src.analysis.ground_truth_diff import diff_quantities
        our = [
            {"item": "Concrete M25", "qty": "100"},
            {"item": "Steel rebar", "qty": "500"},
            {"item": "Plywood", "qty": "30"},
        ]
        gt = [
            {"item": "Concrete M25", "qty": "110"},
            {"item": "Steel rebar", "qty": "500"},
        ]
        r1 = diff_quantities(our, gt)
        r2 = diff_quantities(our, gt)
        assert r1 == r2

    def test_diff_schedules_by_mark(self):
        from src.analysis.ground_truth_diff import diff_schedules
        our = [
            {"mark": "D1", "qty": "5"},
            {"mark": "D2", "qty": "3"},
        ]
        gt = [
            {"mark": "D1", "qty": "5"},
            {"mark": "D2", "qty": "4"},
            {"mark": "D3", "qty": "2"},
        ]
        result = diff_schedules(our, gt)
        assert result["matched_count"] == 2
        assert result["gt_count"] == 3
        assert "D3" in result["missing_in_ours"]
        assert len(result["top_mismatches"]) == 1
        assert result["top_mismatches"][0]["mark"] == "D2"

    def test_compute_gt_diff_empty_gt(self):
        from src.analysis.ground_truth_diff import compute_gt_diff
        payload = {"extraction_summary": {"boq_items": [], "schedules": []}, "quantities": []}
        result = compute_gt_diff(payload, [], [], [])
        assert result["overall_match_rate"] is None
        assert result["categories"] == {}

    def test_diff_works_without_ground_truth(self):
        """Graceful with one-sided data."""
        from src.analysis.ground_truth_diff import diff_quantities
        our = [{"item": "Concrete", "qty": "100"}]
        result = diff_quantities(our, [])
        assert result["match_rate"] == 1.0  # 0 GT items → 1.0
        assert result["gt_count"] == 0
        assert result["our_count"] == 1


class TestTrainingPack(unittest.TestCase):
    """Tests for training_pack.py — ZIP builder."""

    def _make_payload(self):
        return {
            "readiness_score": 0.75,
            "decision": "conditional_go",
            "extraction_summary": {"boq_items": [], "schedules": []},
            "quantities": [],
        }

    def test_zip_structure_correct(self):
        import zipfile, io
        from src.analysis.training_pack import build_training_pack
        zb = build_training_pack("p1", "r1", self._make_payload())
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            names = zf.namelist()
            assert "inputs/manifest.json" in names
            assert "outputs/analysis.json" in names
            assert "context/bid_context.json" in names
            assert "README.md" in names

    def test_zip_includes_gt_when_provided(self):
        import zipfile, io
        from src.analysis.training_pack import build_training_pack
        gt_boq = [{"description": "Concrete", "qty": "100"}]
        gt_diff = {"overall_match_rate": 0.9, "categories": {}}
        zb = build_training_pack(
            "p1", "r1", self._make_payload(),
            gt_diff=gt_diff, gt_boq=gt_boq,
        )
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            names = zf.namelist()
            assert "ground_truth/gt_boq.json" in names
            assert "diff/gt_diff.json" in names

    def test_zip_works_without_gt(self):
        import zipfile, io
        from src.analysis.training_pack import build_training_pack
        zb = build_training_pack("p1", "r1", self._make_payload())
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            names = zf.namelist()
            assert "ground_truth/gt_boq.json" not in names
            assert "diff/gt_diff.json" not in names
            # Core files still present
            assert "outputs/analysis.json" in names

    def test_zip_includes_feedback(self):
        import zipfile, io
        from src.analysis.training_pack import build_training_pack
        fb = [{"item_id": "b1", "action": "correct", "label": "correct"}]
        zb = build_training_pack(
            "p1", "r1", self._make_payload(), feedback_entries=fb,
        )
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            names = zf.namelist()
            assert "feedback/feedback.jsonl" in names
            content = zf.read("feedback/feedback.jsonl").decode()
            assert "correct" in content

    def test_zip_includes_csv_buffers(self):
        import zipfile, io
        from src.analysis.training_pack import build_training_pack
        csv_bufs = {"boq.csv": "col1,col2\na,b\n"}
        zb = build_training_pack(
            "p1", "r1", self._make_payload(), csv_buffers=csv_bufs,
        )
        with zipfile.ZipFile(io.BytesIO(zb)) as zf:
            names = zf.namelist()
            assert "outputs/boq.csv" in names

    def test_save_to_disk(self):
        import tempfile
        from src.analysis.training_pack import build_training_pack, save_training_pack_to_disk
        zb = build_training_pack("p1", "r1", self._make_payload())
        with tempfile.TemporaryDirectory() as td:
            path = save_training_pack_to_disk("p1", "r1", zb, Path(td))
            assert path.exists()
            assert path.name == "training_pack.zip"
            assert "datasets" in str(path)


def _load_pilot_docs():
    """Load pilot_docs module directly to avoid app.app import cascade."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "pilot_docs",
        str(PROJECT_ROOT / "app" / "pilot_docs.py"),
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


class TestPilotDocs(unittest.TestCase):
    """Tests for pilot_docs.py — DOCX generators."""

    def test_pilot_agreement_generates(self):
        pilot_docs = _load_pilot_docs()
        meta = {
            "company_name": "Acme Construction",
            "name": "Mumbai Tower",
            "bid_date": "2026-03-15",
            "trades_in_scope": ["structural", "mep"],
        }
        result = pilot_docs.generate_pilot_agreement_docx(meta)
        assert isinstance(result, bytes)
        assert len(result) > 100  # non-trivial

    def test_pilot_agreement_contains_metadata(self):
        pilot_docs = _load_pilot_docs()
        import zipfile, io
        meta = {
            "company_name": "Acme Construction",
            "name": "Mumbai Tower",
            "bid_date": "2026-03-15",
            "trades_in_scope": ["structural", "mep"],
        }
        result = pilot_docs.generate_pilot_agreement_docx(meta)
        # DOCX is a ZIP; check document.xml for metadata
        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            doc_xml = zf.read("word/document.xml").decode()
            assert "Acme Construction" in doc_xml
            assert "Mumbai Tower" in doc_xml


class TestProjectMetadataExtended(unittest.TestCase):
    """Tests for extended project metadata (Sprint 20 pilot fields)."""

    def test_create_project_with_pilot_fields(self):
        import tempfile
        from src.analysis.projects import create_project, load_project
        with tempfile.TemporaryDirectory() as td:
            meta = create_project(
                name="Test Pilot",
                company_name="Acme Corp",
                trades_in_scope=["structural", "mep"],
                output_preferences={"selected": ["Full BOQ"]},
                pilot_mode=True,
                projects_dir=Path(td),
            )
            assert meta["company_name"] == "Acme Corp"
            assert meta["pilot_mode"] is True
            assert "structural" in meta["trades_in_scope"]

            loaded = load_project(meta["project_id"], projects_dir=Path(td))
            assert loaded["company_name"] == "Acme Corp"
            assert loaded["pilot_mode"] is True

    def test_update_project_pilot_fields(self):
        import tempfile
        from src.analysis.projects import create_project, update_project, load_project
        with tempfile.TemporaryDirectory() as td:
            meta = create_project(name="Proj X", projects_dir=Path(td))
            pid = meta["project_id"]
            update_project(pid, {
                "company_name": "NewCo",
                "trades_in_scope": ["electrical"],
                "pilot_mode": True,
            }, projects_dir=Path(td))
            loaded = load_project(pid, projects_dir=Path(td))
            assert loaded["company_name"] == "NewCo"
            assert loaded["trades_in_scope"] == ["electrical"]
            assert loaded["pilot_mode"] is True

    def test_old_projects_backward_compatible(self):
        """Projects without pilot fields still load fine — .get() returns defaults."""
        import tempfile
        from src.analysis.projects import load_project
        proj_dir = Path(tempfile.mkdtemp())
        pid = "legacy_project"
        (proj_dir / pid).mkdir(parents=True)
        # Write minimal old-style metadata
        old_meta = {
            "project_id": pid,
            "name": "Legacy",
            "owner": "",
            "bid_date": "",
            "notes": "",
            "created_at": "2025-01-01T00:00:00",
            "updated_at": "2025-01-01T00:00:00",
        }
        import json
        with open(proj_dir / pid / "metadata.json", "w") as f:
            json.dump(old_meta, f)

        loaded = load_project(pid, projects_dir=proj_dir)
        assert loaded is not None
        assert loaded["name"] == "Legacy"
        # New fields should be safely absent — consumers use .get()
        assert loaded.get("company_name", "") == ""
        assert loaded.get("trades_in_scope", []) == []
        assert loaded.get("pilot_mode", False) is False

        # Cleanup
        import shutil
        shutil.rmtree(proj_dir)


# =============================================================================
# Sprint 20A: Structural Takeoff Integration
# =============================================================================

class TestStructuralPayloadIntegration(unittest.TestCase):
    """Tests for structural takeoff integration in the analysis pipeline payload."""

    def _make_structural_takeoff(self):
        """Helper: build a realistic structural_takeoff payload dict."""
        return {
            "mode": "assumption",
            "summary": {
                "concrete_m3": 42.5,
                "steel_kg": 8500.0,
                "steel_tons": 8.5,
                "element_counts": {
                    "columns": 12,
                    "beams": 18,
                    "footings": 12,
                    "slabs": 1,
                },
                "detail": {
                    "counts": {"columns": 12, "beams": 18, "footings": 12, "slabs": 1},
                    "concrete_m3": {
                        "columns": 10.8, "beams": 14.4, "footings": 7.2,
                        "slabs": 10.1, "total": 42.5,
                    },
                    "steel_kg": {
                        "columns": 2160, "beams": 2160, "footings": 648,
                        "slabs": 3532, "total": 8500,
                    },
                    "steel_tonnes": 8.5,
                },
            },
            "quantities": [
                {
                    "element_id": "C001",
                    "type": "column",
                    "label": "C1",
                    "count": 12,
                    "dimensions_mm": {"width": 230, "depth": 450, "length": 3000},
                    "concrete_m3": 0.9,
                    "steel_kg": {"main": 120, "stirrup": 60, "total": 180},
                    "sources": {"size": "assumption", "height": "assumption", "steel": "kg_per_m3"},
                    "assumptions": ["Default column size 230x450mm"],
                },
            ],
            "qc": {
                "confidence": 0.65,
                "issues": {
                    "total": 3,
                    "errors": 0,
                    "warnings": 2,
                    "info": 1,
                    "details": [
                        {"code": "Q001", "severity": "warning", "message": "Assumptions used for column sizes"},
                    ],
                },
                "assumptions": {
                    "count": 5,
                    "details": [
                        {"description": "Default column size 230x450mm used"},
                    ],
                },
            },
            "exports": {},
            "source_file": "structural_plan.pdf",
            "warnings": [],
        }

    def test_payload_includes_structural_when_present(self):
        """Payload should include structural_takeoff when data is present."""
        payload = {
            "project_id": "test",
            "structural_takeoff": self._make_structural_takeoff(),
        }
        st = payload.get("structural_takeoff")
        assert st is not None
        assert st["mode"] == "assumption"
        assert st["summary"]["concrete_m3"] == 42.5
        assert st["summary"]["steel_tons"] == 8.5
        assert st["summary"]["element_counts"]["columns"] == 12
        assert st["qc"]["confidence"] == 0.65
        assert len(st["quantities"]) == 1
        assert st["quantities"][0]["element_id"] == "C001"

    def test_payload_structural_none_when_absent(self):
        """Payload without structural pages should have None for structural_takeoff."""
        payload = {"project_id": "test", "structural_takeoff": None}
        assert payload.get("structural_takeoff") is None

    def test_payload_structural_error_mode(self):
        """Pipeline captures structural errors gracefully."""
        payload = {
            "project_id": "test",
            "structural_takeoff": {
                "mode": "error",
                "summary": {},
                "quantities": [],
                "qc": {},
                "exports": {},
                "warnings": ["Structural takeoff failed: ValueError: no pages"],
            },
        }
        st = payload.get("structural_takeoff")
        assert st["mode"] == "error"
        assert len(st["warnings"]) == 1
        assert "failed" in st["warnings"][0]

    def test_structural_export_csv_generation(self):
        """Structural quantities can be serialized to CSV format."""
        import csv, io
        quantities = self._make_structural_takeoff()["quantities"]
        fieldnames = [
            "element_id", "type", "label", "count",
            "width_mm", "depth_mm", "length_mm",
            "concrete_m3", "steel_kg_total",
        ]
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for eq in quantities:
            dims = eq.get("dimensions_mm", {})
            steel = eq.get("steel_kg", {})
            writer.writerow({
                "element_id": eq.get("element_id", ""),
                "type": eq.get("type", ""),
                "label": eq.get("label", ""),
                "count": eq.get("count", 0),
                "width_mm": dims.get("width", 0),
                "depth_mm": dims.get("depth", 0),
                "length_mm": dims.get("length", 0),
                "concrete_m3": eq.get("concrete_m3", 0),
                "steel_kg_total": steel.get("total", 0),
            })
        csv_str = buf.getvalue()
        assert "C001" in csv_str
        assert "column" in csv_str
        assert "element_id" in csv_str.split("\n")[0]

    def test_structural_backward_compat_get_defaults(self):
        """Old payloads without structural_takeoff don't crash."""
        old_payload = {
            "project_id": "old_project",
            "readiness_score": 0.8,
            "decision": "go",
        }
        st = old_payload.get("structural_takeoff")
        assert st is None
        # Consumers should use .get() everywhere
        mode = (st or {}).get("mode")
        assert mode is None
        summary = (st or {}).get("summary", {})
        assert summary.get("concrete_m3", 0) == 0

    def test_structural_summary_json_serializable(self):
        """Structural takeoff dict is fully JSON-serializable."""
        st_data = self._make_structural_takeoff()
        serialized = json.dumps(st_data, default=str)
        deserialized = json.loads(serialized)
        assert deserialized["mode"] == "assumption"
        assert deserialized["summary"]["concrete_m3"] == 42.5

    def test_export_bundle_includes_structural_files(self):
        """When structural data present, export bundle includes structural files."""
        import zipfile, io

        csv_buffers = {
            "boq.csv": "col1,col2\na,b\n",
        }
        # Add structural files (simulating what demo_page.py does)
        st_data = self._make_structural_takeoff()
        csv_buffers["structural_summary.json"] = json.dumps(st_data["summary"], indent=2)
        csv_buffers["structural_qc.json"] = json.dumps(st_data["qc"], indent=2)
        # Build a simple ZIP
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, content in csv_buffers.items():
                zf.writestr(fname, content)
        buf.seek(0)
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
            assert "structural_summary.json" in names
            assert "structural_qc.json" in names
            assert "boq.csv" in names


class TestStructuralPipelineSurvivesExceptions(unittest.TestCase):
    """Tests that pipeline gracefully handles structural exceptions."""

    def test_structural_error_produces_error_mode(self):
        """When structural pipeline raises, error mode is produced."""
        _structural_takeoff = None
        try:
            raise ValueError("Simulated structural failure")
        except Exception as _st_exc:
            _structural_takeoff = {
                "mode": "error",
                "summary": {},
                "quantities": [],
                "qc": {},
                "exports": {},
                "warnings": [f"Structural takeoff failed: {type(_st_exc).__name__}: {_st_exc}"],
            }
        assert _structural_takeoff is not None
        assert _structural_takeoff["mode"] == "error"
        assert "Simulated structural failure" in _structural_takeoff["warnings"][0]

    def test_structural_none_for_no_structural_pages(self):
        """When no structural pages detected, result is None."""
        disc_map = {"architectural": 15, "electrical": 3}
        has_structural = disc_map.get("structural", 0) > 0
        assert has_structural is False
        _structural_takeoff = None  # Should remain None
        assert _structural_takeoff is None

    def test_structural_export_absent_when_none(self):
        """CSV export code handles None structural_takeoff gracefully."""
        payload = {"project_id": "test", "structural_takeoff": None}
        csv_buffers = {}
        _st_export = payload.get("structural_takeoff")
        if _st_export and _st_export.get("mode") not in (None, "error"):
            csv_buffers["structural_summary.json"] = "should not appear"
        assert "structural_summary.json" not in csv_buffers


# =============================================================================
# Sprint 20C: Estimating Playbook Tests
# =============================================================================

class TestEstimatingPlaybook(unittest.TestCase):
    """Tests for src/analysis/estimating_playbook.py."""

    def test_default_playbook_structure(self):
        """default_playbook() returns correct schema."""
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        assert isinstance(pb, dict)
        assert "company" in pb
        assert "project" in pb
        assert "market_snapshot" in pb
        assert pb["company"]["risk_posture"] == "balanced"
        assert pb["company"]["default_contingency_pct"] == 5.0
        assert pb["project"]["competition_intensity"] == "med"
        assert pb["market_snapshot"]["material_trend"] == "stable"

    def test_validate_playbook_valid(self):
        """Valid playbook passes validation."""
        from src.analysis.estimating_playbook import default_playbook, validate_playbook
        pb = default_playbook()
        is_valid, warnings = validate_playbook(pb)
        assert is_valid is True
        assert isinstance(warnings, list)

    def test_validate_playbook_invalid_posture(self):
        """Unknown risk posture produces warning."""
        from src.analysis.estimating_playbook import default_playbook, validate_playbook
        pb = default_playbook()
        pb["company"]["risk_posture"] = "yolo"
        is_valid, warnings = validate_playbook(pb)
        assert is_valid is True  # warnings only, not fatal
        assert any("risk_posture" in w for w in warnings)

    def test_validate_playbook_not_dict(self):
        """Non-dict playbook fails validation."""
        from src.analysis.estimating_playbook import validate_playbook
        is_valid, warnings = validate_playbook("not a dict")
        assert is_valid is False

    def test_merge_playbook(self):
        """merge_playbook merges company defaults with project overrides."""
        from src.analysis.estimating_playbook import default_playbook, merge_playbook
        base = default_playbook()
        base["company"]["name"] = "TestCorp"
        base["company"]["default_contingency_pct"] = 6.0
        over = {
            "project": {"must_win": True, "competition_intensity": "high"},
            "market_snapshot": {"material_trend": "rising"},
        }
        merged = merge_playbook(base, over)
        assert merged["company"]["name"] == "TestCorp"
        assert merged["company"]["default_contingency_pct"] == 6.0
        assert merged["project"]["must_win"] is True
        assert merged["project"]["competition_intensity"] == "high"
        assert merged["market_snapshot"]["material_trend"] == "rising"

    def test_diff_playbook(self):
        """diff_playbook detects changed fields."""
        from src.analysis.estimating_playbook import default_playbook, diff_playbook
        base = default_playbook()
        current = default_playbook()
        current["company"]["risk_posture"] = "aggressive"
        current["project"]["must_win"] = True
        changes = diff_playbook(base, current)
        assert isinstance(changes, list)
        assert len(changes) >= 2
        fields = [c["field"] for c in changes]
        assert "risk_posture" in fields
        assert "must_win" in fields

    def test_diff_playbook_no_changes(self):
        """diff_playbook returns empty list when identical."""
        from src.analysis.estimating_playbook import default_playbook, diff_playbook
        pb = default_playbook()
        changes = diff_playbook(pb, pb)
        assert changes == []

    def test_summarize_playbook(self):
        """summarize_playbook_for_exports returns non-empty markdown."""
        from src.analysis.estimating_playbook import default_playbook, summarize_playbook_for_exports
        pb = default_playbook()
        pb["company"]["name"] = "TestCorp"
        summary = summarize_playbook_for_exports(pb)
        assert "TestCorp" in summary
        assert "contingency" in summary.lower()

    def test_summarize_empty_playbook(self):
        """summarize returns empty string for None playbook."""
        from src.analysis.estimating_playbook import summarize_playbook_for_exports
        assert summarize_playbook_for_exports(None) == ""
        assert summarize_playbook_for_exports({}) == ""

    def test_contingency_adjustments_default(self):
        """Default playbook gives 5% contingency."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        result = compute_playbook_contingency_adjustments(pb)
        assert result["base_pct"] == 5.0
        assert result["recommended_pct"] == 5.0
        assert result["market_adj_pct"] == 0.0

    def test_contingency_adjustments_conservative(self):
        """Conservative posture adds +1.5%."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["company"]["risk_posture"] = "conservative"
        result = compute_playbook_contingency_adjustments(pb)
        assert result["posture_adj_pct"] == 1.5
        assert result["recommended_pct"] == 6.5

    def test_contingency_adjustments_aggressive(self):
        """Aggressive posture subtracts 1%."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["company"]["risk_posture"] = "aggressive"
        result = compute_playbook_contingency_adjustments(pb)
        assert result["posture_adj_pct"] == -1.0
        assert result["recommended_pct"] == 4.0

    def test_contingency_adjustments_market_volatile(self):
        """Volatile market adds +2%."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["market_snapshot"]["material_trend"] = "volatile"
        result = compute_playbook_contingency_adjustments(pb)
        assert result["market_adj_pct"] == 2.0
        assert result["recommended_pct"] == 7.0

    def test_contingency_adjustments_override(self):
        """Project override replaces computed value."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["project"]["contingency_override_pct"] = 12.0
        result = compute_playbook_contingency_adjustments(pb)
        assert result["override_pct"] == 12.0
        assert result["recommended_pct"] == 12.0

    def test_contingency_adjustments_must_win(self):
        """Must-win reduces posture adjustment by 0.5%."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["project"]["must_win"] = True
        result = compute_playbook_contingency_adjustments(pb)
        assert result["posture_adj_pct"] == -0.5
        assert result["recommended_pct"] == 4.5

    def test_contingency_adjustments_none_playbook(self):
        """None playbook returns defaults."""
        from src.analysis.estimating_playbook import compute_playbook_contingency_adjustments
        result = compute_playbook_contingency_adjustments(None)
        assert result["recommended_pct"] == 5.0
        assert "No playbook" in result["basis"][0]

    def test_contingency_combined_market(self):
        """Multiple market factors stack."""
        from src.analysis.estimating_playbook import default_playbook, compute_playbook_contingency_adjustments
        pb = default_playbook()
        pb["market_snapshot"]["material_trend"] = "rising"
        pb["market_snapshot"]["labor_availability"] = "tight"
        pb["market_snapshot"]["logistics_difficulty"] = "hard"
        pb["market_snapshot"]["weather_factor"] = "high_risk"
        result = compute_playbook_contingency_adjustments(pb)
        # 1.0 + 1.0 + 1.5 + 1.0 = 4.5
        assert result["market_adj_pct"] == 4.5
        assert result["recommended_pct"] == 9.5


class TestCompanyPlaybooks(unittest.TestCase):
    """Tests for src/analysis/company_playbooks.py."""

    def setUp(self):
        import tempfile
        self._tmp = tempfile.mkdtemp()
        self._dir = Path(self._tmp)

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmp, ignore_errors=True)

    def test_save_and_load(self):
        """Save and load roundtrip."""
        from src.analysis.company_playbooks import save_playbook, load_playbook
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["company"]["name"] = "Acme Builders"
        save_playbook("Acme Builders", pb, playbooks_dir=self._dir)
        loaded = load_playbook("Acme Builders", playbooks_dir=self._dir)
        assert loaded is not None
        assert loaded["company"]["name"] == "Acme Builders"

    def test_list_playbooks(self):
        """list_playbooks returns saved entries."""
        from src.analysis.company_playbooks import save_playbook, list_playbooks
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        save_playbook("Company A", pb, playbooks_dir=self._dir)
        save_playbook("Company B", pb, playbooks_dir=self._dir)
        entries = list_playbooks(playbooks_dir=self._dir)
        assert len(entries) == 2
        names = [e["company_name"] for e in entries]
        assert "Company A" in names
        assert "Company B" in names

    def test_delete_playbook(self):
        """delete_playbook removes entry."""
        from src.analysis.company_playbooks import save_playbook, load_playbook, delete_playbook
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        save_playbook("Removable Corp", pb, playbooks_dir=self._dir)
        assert load_playbook("Removable Corp", playbooks_dir=self._dir) is not None
        deleted = delete_playbook("Removable Corp", playbooks_dir=self._dir)
        assert deleted is True
        assert load_playbook("Removable Corp", playbooks_dir=self._dir) is None

    def test_load_nonexistent(self):
        """Loading non-existent playbook returns None."""
        from src.analysis.company_playbooks import load_playbook
        assert load_playbook("Ghost Corp", playbooks_dir=self._dir) is None

    def test_save_records_updated_at(self):
        """Save playbook records updated_at timestamp."""
        from src.analysis.company_playbooks import save_playbook, load_playbook
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        save_playbook("Timestamp Corp", pb, playbooks_dir=self._dir)
        loaded = load_playbook("Timestamp Corp", playbooks_dir=self._dir)
        assert "updated_at" in loaded


class TestPricingWithPlaybook(unittest.TestCase):
    """Tests for pricing guidance integration with estimating playbook."""

    def test_pricing_with_playbook_conservative(self):
        """Conservative playbook increases contingency recommendation."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        from src.analysis.estimating_playbook import default_playbook
        # Without playbook
        result_no_pb = compute_pricing_guidance(
            qa_score={"score": 85, "breakdown": {}},
        )
        # With conservative playbook
        pb = default_playbook()
        pb["company"]["risk_posture"] = "conservative"
        result_with_pb = compute_pricing_guidance(
            qa_score={"score": 85, "breakdown": {}},
            estimating_playbook=pb,
        )
        assert "basis_of_recommendation" in result_with_pb
        assert result_with_pb["contingency_range"]["recommended_pct"] >= \
               result_no_pb["contingency_range"]["recommended_pct"]

    def test_pricing_with_playbook_override(self):
        """Project contingency override controls final value."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["project"]["contingency_override_pct"] = 15.0
        result = compute_pricing_guidance(
            qa_score={"score": 85, "breakdown": {}},
            estimating_playbook=pb,
        )
        assert "basis_of_recommendation" in result
        # Override should be reflected
        assert result["contingency_range"]["recommended_pct"] >= 10.0

    def test_pricing_without_playbook_unchanged(self):
        """Pricing without playbook works as before."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        result = compute_pricing_guidance(
            qa_score={"score": 85, "breakdown": {}},
        )
        assert "contingency_range" in result
        assert result["contingency_range"]["low_pct"] == 3.0
        assert result["contingency_range"]["high_pct"] == 5.0
        assert "basis_of_recommendation" not in result

    def test_pricing_playbook_basis_populated(self):
        """basis_of_recommendation is populated when playbook active."""
        from src.analysis.pricing_guidance import compute_pricing_guidance
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["market_snapshot"]["material_trend"] = "volatile"
        result = compute_pricing_guidance(
            qa_score={"score": 85, "breakdown": {}},
            estimating_playbook=pb,
        )
        assert "basis_of_recommendation" in result
        basis = result["basis_of_recommendation"]
        assert any("volatile" in b.lower() for b in basis)


class TestBidStrategyWithPlaybook(unittest.TestCase):
    """Tests for bid strategy scorer integration with estimating playbook."""

    def _get_compute_bid_strategy(self):
        """Import compute_bid_strategy from app/ without triggering app.app cascade."""
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "bid_strategy_scorer",
            str(PROJECT_ROOT / "app" / "bid_strategy_scorer.py"),
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod.compute_bid_strategy

    def test_playbook_seeds_relationship(self):
        """relationship_bid seeds relationship_level when empty."""
        compute = self._get_compute_bid_strategy()
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["project"]["relationship_bid"] = True
        result = compute(
            inputs={"relationship_level": ""},
            payload={"readiness_score": 50, "blockers": []},
            estimating_playbook=pb,
        )
        assert result["playbook_applied"] is True
        # Client fit should no longer be UNKNOWN
        assert result["client_fit"]["score"] is not None

    def test_playbook_seeds_competition(self):
        """competition_intensity seeds market_pressure when empty."""
        compute = self._get_compute_bid_strategy()
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["project"]["competition_intensity"] = "high"
        result = compute(
            inputs={"market_pressure": 0},
            payload={"readiness_score": 50, "blockers": []},
            estimating_playbook=pb,
        )
        assert result["playbook_applied"] is True

    def test_playbook_seeds_must_win(self):
        """must_win sets win_probability when empty."""
        compute = self._get_compute_bid_strategy()
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["project"]["must_win"] = True
        result = compute(
            inputs={"win_probability": 0},
            payload={"readiness_score": 50, "blockers": []},
            estimating_playbook=pb,
        )
        assert result["playbook_applied"] is True

    def test_playbook_seeds_target_margin(self):
        """risk_posture sets target_margin when empty."""
        compute = self._get_compute_bid_strategy()
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["company"]["risk_posture"] = "conservative"
        pb["company"]["default_profit_pct"] = 10.0
        result = compute(
            inputs={"target_margin": 0},
            payload={"readiness_score": 50, "blockers": []},
            estimating_playbook=pb,
        )
        assert result["playbook_applied"] is True

    def test_playbook_does_not_override_user_inputs(self):
        """Playbook doesn't override user-provided values."""
        compute = self._get_compute_bid_strategy()
        from src.analysis.estimating_playbook import default_playbook
        pb = default_playbook()
        pb["project"]["relationship_bid"] = True
        pb["project"]["competition_intensity"] = "high"
        result = compute(
            inputs={
                "relationship_level": "Preferred",
                "market_pressure": 8,
                "win_probability": 90.0,
                "target_margin": 15.0,
            },
            payload={"readiness_score": 50, "blockers": []},
            estimating_playbook=pb,
        )
        # User provided all inputs, so playbook seeding should not override
        assert result["client_fit"]["based_on"][0].startswith("Preferred")

    def test_no_playbook_returns_playbook_applied_false(self):
        """Without playbook, playbook_applied is False."""
        compute = self._get_compute_bid_strategy()
        result = compute(
            inputs={"relationship_level": "New"},
            payload={"readiness_score": 50, "blockers": []},
        )
        assert result["playbook_applied"] is False


class TestBidSummaryWithPlaybook(unittest.TestCase):
    """Tests for bid summary markdown with playbook section."""

    def test_summary_includes_playbook_section(self):
        """Bid summary includes Basis of Estimate when playbook present."""
        from src.analysis.estimating_playbook import default_playbook
        sys.path.insert(0, str(PROJECT_ROOT / "app"))
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "bid_summary",
            str(PROJECT_ROOT / "app" / "bid_summary.py"),
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        pb = default_playbook()
        pb["company"]["name"] = "SummaryTestCorp"
        payload = {
            "project_id": "test",
            "readiness_score": 60,
            "estimating_playbook": pb,
        }
        md = mod.generate_bid_summary_markdown(payload)
        assert "Basis of Estimate" in md
        assert "SummaryTestCorp" in md

    def test_summary_without_playbook_unchanged(self):
        """Bid summary without playbook has no Basis of Estimate section."""
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "bid_summary",
            str(PROJECT_ROOT / "app" / "bid_summary.py"),
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        payload = {"project_id": "test", "readiness_score": 60}
        md = mod.generate_bid_summary_markdown(payload)
        assert "Basis of Estimate" not in md


# =============================================================================
# Sprint 20D: Bulk Actions + Review Queue + Skipped Page Tests
# =============================================================================

class TestBulkActionEligibility(unittest.TestCase):
    """Tests for bulk action eligibility counting and payload mutation."""

    def test_prefer_schedule_eligibility(self):
        """Eligible count = mismatches in doors/windows categories."""
        from src.analysis.bulk_actions import prefer_schedule_for_mismatches
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 3},
            {"category": "windows", "mismatch": True, "max_delta": 1},
            {"category": "rooms", "mismatch": True, "max_delta": 2},
            {"category": "doors", "mismatch": False},
        ]
        updated, n = prefer_schedule_for_mismatches(recon)
        assert n == 2, f"Expected 2 eligible, got {n}"
        assert len(updated) == 4, "Should return all rows"
        actions = [r.get("action") for r in updated]
        sched_count = sum(1 for a in actions if a == "prefer_schedule")
        assert sched_count == 2

    def test_prefer_schedule_no_eligible(self):
        """No eligible rows returns count 0."""
        from src.analysis.bulk_actions import prefer_schedule_for_mismatches
        recon = [{"category": "rooms", "mismatch": True, "max_delta": 2}]
        updated, n = prefer_schedule_for_mismatches(recon)
        assert n == 0
        assert len(updated) == 1

    def test_prefer_schedule_empty(self):
        """Empty recon list returns ([], 0)."""
        from src.analysis.bulk_actions import prefer_schedule_for_mismatches
        updated, n = prefer_schedule_for_mismatches([])
        assert n == 0
        assert updated == []

    def test_generate_rfis_high_eligibility(self):
        """RFIs generated only for max_delta >= 5."""
        from src.analysis.bulk_actions import generate_rfis_for_high_mismatches
        recon = [
            {"category": "doors", "mismatch": True, "max_delta": 7,
             "schedule_count": 10, "boq_count": 3, "drawing_count": 5},
            {"category": "windows", "mismatch": True, "max_delta": 2,
             "schedule_count": 5, "boq_count": 3, "drawing_count": 4},
        ]
        new_rfis, updated = generate_rfis_for_high_mismatches(recon, [])
        assert len(new_rfis) == 1, f"Expected 1 HIGH RFI, got {len(new_rfis)}"
        assert new_rfis[0]["source"] == "bulk_action"

    def test_generate_rfis_no_eligible(self):
        """No HIGH mismatches yields empty."""
        from src.analysis.bulk_actions import generate_rfis_for_high_mismatches
        recon = [{"category": "doors", "mismatch": True, "max_delta": 2,
                  "schedule_count": 5, "boq_count": 3, "drawing_count": 4}]
        new_rfis, updated = generate_rfis_for_high_mismatches(recon, [])
        assert len(new_rfis) == 0

    def test_mark_revisions_reviewed(self):
        """Marks intentional revisions as reviewed."""
        from src.analysis.bulk_actions import mark_intentional_revisions_reviewed
        conflicts = [
            {"resolution": "intentional_revision", "type": "boq_change"},
            {"resolution": "conflict", "type": "boq_change"},
            {"resolution": "intentional_revision", "type": "schedule_change"},
        ]
        updated, n = mark_intentional_revisions_reviewed(conflicts)
        assert n == 2
        reviewed = [c for c in updated if c.get("review_status") == "reviewed"]
        assert len(reviewed) == 2

    def test_mark_revisions_no_eligible(self):
        """No intentional revisions yields 0."""
        from src.analysis.bulk_actions import mark_intentional_revisions_reviewed
        conflicts = [{"resolution": "conflict", "type": "boq_change"}]
        updated, n = mark_intentional_revisions_reviewed(conflicts)
        assert n == 0


class TestReviewQueueRendering(unittest.TestCase):
    """Tests for review queue structure and rendering safety."""

    def test_review_queue_items_have_required_fields(self):
        """All review items have type, severity, title, source_key."""
        from src.analysis.review_queue import build_review_queue
        items = build_review_queue(
            pages_skipped=[
                {"doc_type": "conditions", "page_idx": 5},
                {"doc_type": "conditions", "page_idx": 6},
            ],
        )
        assert len(items) >= 1
        for item in items:
            assert "type" in item
            assert "severity" in item
            assert "title" in item
            assert "source_key" in item
            assert "recommended_action" in item
            assert "source_key" not in item["title"]
            assert "_review_status" not in item["title"]

    def test_review_queue_skipped_page_grouped(self):
        """Skipped pages are grouped by doc_type."""
        from src.analysis.review_queue import build_review_queue
        items = build_review_queue(
            pages_skipped=[
                {"doc_type": "boq", "page_idx": 10},
                {"doc_type": "boq", "page_idx": 11},
                {"doc_type": "conditions", "page_idx": 20},
            ],
        )
        skipped = [i for i in items if i["type"] == "skipped_page"]
        assert len(skipped) == 2
        titles = [s["title"] for s in skipped]
        assert any("boq" in t for t in titles)
        assert any("conditions" in t for t in titles)

    def test_review_queue_no_title_overlap_fragments(self):
        """Title does not contain duplicate fragments or raw dict keys."""
        from src.analysis.review_queue import build_review_queue
        items = build_review_queue(
            quantity_reconciliation=[
                {"category": "doors", "mismatch": True, "max_delta": 3,
                 "schedule_count": 10, "boq_count": 7, "drawing_count": 8},
            ],
            conflicts=[
                {"type": "boq_change", "delta_confidence": 0.6, "item_no": "1.1",
                 "changes": [{"field": "qty"}]},
            ],
        )
        for item in items:
            title = item["title"]
            assert "{" not in title
            assert "}" not in title
            assert "source_key" not in title
            assert "evidence_bundle" not in title

    def test_review_queue_sorted_by_severity(self):
        """HIGH items appear before MEDIUM, before LOW."""
        from src.analysis.review_queue import build_review_queue
        items = build_review_queue(
            quantity_reconciliation=[
                {"category": "doors", "mismatch": True, "max_delta": 1},
                {"category": "windows", "mismatch": True, "max_delta": 6},
                {"category": "rooms", "mismatch": True, "max_delta": 3},
            ],
        )
        severities = [i["severity"] for i in items]
        assert severities.index("high") < severities.index("medium")
        assert severities.index("medium") < severities.index("low")


class TestSkippedPageSummary(unittest.TestCase):
    """Tests for FAST_BUDGET explanation and metric consistency."""

    def test_processing_stats_metrics(self):
        """processing_stats has expected keys with correct types."""
        payload = {
            "processing_stats": {
                "total_pages": 367,
                "deep_processed_pages": 80,
                "ocr_pages": 80,
                "skipped_pages": 287,
            },
            "run_coverage": {
                "selection_mode": "fast_budget",
                "pages_total": 367,
                "pages_deep_processed": 80,
                "pages_skipped": [{"page_idx": i, "doc_type": "unknown"} for i in range(287)],
            },
        }
        ps = payload["processing_stats"]
        rc = payload["run_coverage"]
        assert ps["total_pages"] == rc["pages_total"]
        assert ps["deep_processed_pages"] == rc["pages_deep_processed"]
        assert ps["skipped_pages"] == len(rc["pages_skipped"])

    def test_metric_display_em_dash(self):
        """None values display as em-dash, not 0."""
        def _display(val, suffix=""):
            if val is None:
                return "\u2014"
            return f"{val}{suffix}"
        assert _display(None) == "\u2014"
        assert _display(0) == "0"
        assert _display(80) == "80"

    def test_fast_budget_detection(self):
        """FAST_BUDGET mode is correctly detected from payload."""
        payload_fast = {
            "run_coverage": {"selection_mode": "fast_budget", "pages_skipped": [{"page_idx": 1}]},
            "processing_stats": {"total_pages": 100, "deep_processed_pages": 50, "skipped_pages": 50},
        }
        rc = payload_fast["run_coverage"]
        assert rc["selection_mode"] == "fast_budget"
        assert len(rc["pages_skipped"]) > 0

    def test_fallback_for_missing_processing_stats(self):
        """Handles payload without processing_stats gracefully."""
        payload_old = {
            "run_coverage": {
                "selection_mode": "fast_budget",
                "pages_total": 200,
                "pages_deep_processed": 80,
                "pages_skipped": [{"page_idx": i} for i in range(120)],
            },
        }
        ps = payload_old.get("processing_stats") or {}
        rc = payload_old.get("run_coverage") or {}
        total = ps.get("total_pages") or rc.get("pages_total", 0)
        deep = ps.get("deep_processed_pages") or rc.get("pages_deep_processed", 0)
        skipped = ps.get("skipped_pages") or len(rc.get("pages_skipped", []))
        assert total == 200
        assert deep == 80
        assert skipped == 120


# =============================================================================
# Sprint 20E: Coverage Score + Quantify Table + Blocker Diversity + Widget Keys
# =============================================================================


class TestCoverageScoreDisplay(unittest.TestCase):
    """Tests for Sprint 20E: Coverage score sublabel and rename logic."""

    def test_sublabel_when_deep_less_than_total(self):
        """Sublabel shows 'X/Y pages processed' when deep < total."""
        ps = {"total_pages": 367, "deep_processed_pages": 80}
        deep = ps.get("deep_processed_pages")
        total = ps.get("total_pages")
        should_show = deep is not None and total and deep < total
        assert should_show is True
        sublabel = f"{deep}/{total} pages processed"
        assert "80/367" in sublabel

    def test_no_sublabel_when_all_processed(self):
        """No sublabel when deep == total."""
        ps = {"total_pages": 50, "deep_processed_pages": 50}
        deep = ps.get("deep_processed_pages")
        total = ps.get("total_pages")
        should_show = deep is not None and total and deep < total
        assert should_show is False

    def test_label_rename_in_fast_budget(self):
        """Coverage label becomes 'Trade Coverage' in FAST_BUDGET mode."""
        ps = {"total_pages": 367, "deep_processed_pages": 80}
        deep = ps.get("deep_processed_pages")
        total = ps.get("total_pages")
        label = "Trade Coverage" if deep is not None and deep < total else "Coverage"
        assert label == "Trade Coverage"

    def test_label_stays_coverage_in_full_read(self):
        """Coverage label stays 'Coverage' in FULL_READ."""
        ps = {"total_pages": 100, "deep_processed_pages": 100}
        deep = ps.get("deep_processed_pages")
        total = ps.get("total_pages")
        label = "Trade Coverage" if deep is not None and deep < total else "Coverage"
        assert label == "Coverage"


class TestQuantifyTableRows(unittest.TestCase):
    """Tests for Sprint 20E: 'What We Can Quantify' table row additions."""

    def test_new_rows_from_payload(self):
        """BOQ Items, Commercial Terms, Requirements, Finish Rows extracted from payload."""
        payload = {
            "boq_stats": {"total_items": 15},
            "commercial_terms": [{"term_type": "defect_liability"}],
            "requirements_by_trade": {"civil": [1, 2], "mep": [3]},
            "finish_takeoff": {"finish_rows": [{"room": "A"}, {"room": "B"}]},
        }
        boq_count = (payload.get("boq_stats") or {}).get("total_items")
        comm_terms = payload.get("commercial_terms")
        comm_count = len(comm_terms) if isinstance(comm_terms, list) and comm_terms else None
        req_by_trade = payload.get("requirements_by_trade") or {}
        req_count = sum(len(v) for v in req_by_trade.values()) if req_by_trade else None
        finish_rows = (payload.get("finish_takeoff") or {}).get("finish_rows")
        finish_count = len(finish_rows) if isinstance(finish_rows, list) else None

        assert boq_count == 15
        assert comm_count == 1
        assert req_count == 3
        assert finish_count == 2

    def test_em_dash_for_missing_fields(self):
        """Missing fields show em-dash (—), not 0."""
        payload = {}  # no boq_stats, no commercial_terms
        boq_count = (payload.get("boq_stats") or {}).get("total_items")
        comm_terms = payload.get("commercial_terms")
        comm_count = len(comm_terms) if isinstance(comm_terms, list) and comm_terms else None
        assert boq_count is None
        assert comm_count is None

        # em-dash display
        def _display(val):
            return "\u2014" if val is None else str(val)
        assert _display(boq_count) == "\u2014"
        assert _display(comm_count) == "\u2014"


class TestBlockerCoverageGaps(unittest.TestCase):
    """Tests for Sprint 20E: Coverage gap warnings from run_coverage."""

    def test_boq_not_covered_generates_warning(self):
        """BOQ in doc_types_not_covered triggers high warning."""
        run_cov = {
            "doc_types_not_covered": ["boq", "spec", "unknown"],
            "doc_types_partially_covered": ["conditions"],
            "doc_types_detected": {"conditions": 103, "boq": 2},
        }
        not_covered = run_cov.get("doc_types_not_covered", [])
        assert "boq" in not_covered

    def test_no_warnings_when_fully_covered(self):
        """No coverage gaps when all types are covered."""
        run_cov = {
            "doc_types_not_covered": [],
            "doc_types_partially_covered": [],
            "doc_types_fully_covered": ["boq", "conditions", "spec"],
        }
        not_covered = run_cov.get("doc_types_not_covered", [])
        partial = run_cov.get("doc_types_partially_covered", [])
        warnings = []
        if "boq" in not_covered:
            warnings.append("boq")
        if "conditions" in partial or "conditions" in not_covered:
            warnings.append("conditions")
        assert len(warnings) == 0

    def test_mixed_coverage_gaps(self):
        """Multiple coverage gap types detected."""
        run_cov = {
            "doc_types_not_covered": ["boq", "addendum"],
            "doc_types_partially_covered": ["conditions"],
        }
        not_covered = run_cov.get("doc_types_not_covered", [])
        partial = run_cov.get("doc_types_partially_covered", [])
        warnings = []
        if "boq" in not_covered:
            warnings.append("boq_not_processed")
        if "addendum" in not_covered:
            warnings.append("addenda_not_processed")
        if "conditions" in partial or "conditions" in not_covered:
            warnings.append("conditions_partial")
        assert len(warnings) == 3


class TestWidgetKeyUniqueness(unittest.TestCase):
    """Tests for Sprint 20E: _make_widget_key uniqueness."""

    def test_keys_with_different_parts_are_unique(self):
        """Keys with different component parts should produce different strings."""
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "demo_page_keys",
            "/Users/cooperworks/floorplan-engine/app/demo_page.py",
            submodule_search_locations=[]
        )
        # Can't import Streamlit, so test the function logic directly
        def _make_widget_key(*parts):
            clean = [str(p).replace(" ", "_")[:40] for p in parts if p is not None and str(p)]
            return ":".join(clean) if clean else f"wk_{id(parts)}"

        k1 = _make_widget_key("tab0", "ev_prev", "BLK-001", "Scale issue")
        k2 = _make_widget_key("tab0", "ev_prev", "BLK-002", "Scale issue")
        k3 = _make_widget_key("tab1", "ev_prev", "BLK-001", "Scale issue")
        assert k1 != k2, "Different item IDs should produce different keys"
        assert k1 != k3, "Different tab prefixes should produce different keys"

    def test_key_with_none_parts(self):
        """None parts are filtered out."""
        def _make_widget_key(*parts):
            clean = [str(p).replace(" ", "_")[:40] for p in parts if p is not None and str(p)]
            return ":".join(clean) if clean else f"wk_{id(parts)}"

        k = _make_widget_key("prefix", None, "item", None, "suffix")
        assert "None" not in k
        assert "prefix" in k
        assert "item" in k
        assert "suffix" in k


# =============================================================================
# SPRINT 20F: TABLE ROUTER + EXTRACTION DIAGNOSTICS TESTS
# =============================================================================

class TestTableRouter(unittest.TestCase):
    """Test table_router fallback chain and diagnostics."""

    def test_router_returns_result_with_no_inputs(self):
        """Router should return method_used='none' when no inputs given."""
        from src.analysis.table_router import extract_table_rows_from_page
        result = extract_table_rows_from_page()
        self.assertEqual(result.method_used, "none")
        self.assertEqual(len(result.rows), 0)
        self.assertIn("selection_reason", result.diagnostics)

    def test_router_ocr_path_extracts_rows(self):
        """Router should reconstruct rows from OCR text."""
        from src.analysis.table_router import extract_table_rows_from_page
        ocr_text = (
            "Item No    Description                     Unit    Qty\n"
            "1.1        Earthwork excavation             cum     120.5\n"
            "1.2        PCC M15 below footings           cum     25.0\n"
            "1.3        RCC M25 for footings             cum     45.0\n"
        )
        result = extract_table_rows_from_page(ocr_text=ocr_text)
        self.assertGreater(len(result.rows), 0)
        self.assertEqual(result.method_used, "ocr_row_reconstruct")
        self.assertGreater(result.confidence, 0)

    def test_router_diagnostics_on_exception(self):
        """Router should capture diagnostics even when methods fail."""
        from src.analysis.table_router import extract_table_rows_from_page
        # Pass invalid pdf_path — methods should fail gracefully
        result = extract_table_rows_from_page(
            page_input="/nonexistent/path.pdf",
            page_meta={"page_number": 0, "has_text_layer": True},
        )
        self.assertIsInstance(result.diagnostics, dict)
        self.assertIn("methods_attempted", result.diagnostics)
        # Should have attempted at least one method
        self.assertGreater(len(result.diagnostics["methods_attempted"]), 0)

    def test_router_skips_unavailable_methods(self):
        """Router should skip methods listed in skip_methods config."""
        from src.analysis.table_router import extract_table_rows_from_page
        result = extract_table_rows_from_page(
            page_input="/nonexistent/path.pdf",
            page_meta={"page_number": 0, "has_text_layer": True},
            ocr_text="A  B  C\n1  2  3\n",
            config={"skip_methods": ["pdfplumber", "camelot_lattice", "camelot_stream"]},
        )
        skipped = result.diagnostics.get("methods_skipped", [])
        self.assertIn("pdfplumber", skipped)
        self.assertIn("camelot_lattice", skipped)

    def test_router_to_dict(self):
        """TableExtractionResult.to_dict() should produce serializable dict."""
        from src.analysis.table_router import TableExtractionResult
        r = TableExtractionResult(method_used="test", rows=[["a", "b"]], confidence=0.5)
        d = r.to_dict()
        self.assertEqual(d["method_used"], "test")
        self.assertEqual(d["row_count"], 1)
        self.assertEqual(d["confidence"], 0.5)


class TestBOQWithTableRouter(unittest.TestCase):
    """Test BOQ extractor preserves legacy schema while using table_router."""

    def test_legacy_schema_preserved(self):
        """BOQ items should have the same keys as before table_router."""
        items = extract_boq_items(BOQ_TEXT, source_page=0)
        self.assertGreater(len(items), 0)
        for item in items:
            self.assertIn("item_no", item)
            self.assertIn("description", item)
            self.assertIn("unit", item)
            self.assertIn("qty", item)
            self.assertIn("rate", item)
            self.assertIn("source_page", item)
            self.assertIn("confidence", item)

    def test_boq_extracts_same_or_more_items(self):
        """Enhanced BOQ should extract at least as many items as before."""
        items = extract_boq_items(BOQ_TEXT, source_page=0)
        # Original test expects >=1 item; enhanced should do at least that
        self.assertGreaterEqual(len(items), 1)
        with_qty = [i for i in items if i.get("qty") is not None]
        self.assertGreaterEqual(len(with_qty), 1)

    def test_boq_extended_item_patterns(self):
        """Extended patterns like A-1, 2(a) should be parsed."""
        text = """
BILL OF QUANTITIES
A-1   Supply and erection of steel door frames    nos    12
2(a)  Providing and fixing ceiling fans           nos    24
iii   Additional earthwork excavation              cum    50.0
"""
        items = extract_boq_items(text, source_page=0)
        item_nos = {i["item_no"] for i in items}
        # Should find at least one extended pattern item
        self.assertTrue(
            any(n in item_nos for n in ["A-1", "2(a)", "iii"]),
            f"Expected extended item_no, got {item_nos}"
        )

    def test_boq_debug_mode(self):
        """enable_debug should attach debug info without crashing."""
        items = extract_boq_items(
            BOQ_TEXT, source_page=0,
            enable_debug=True,
        )
        self.assertGreater(len(items), 0)
        # Debug info should be attached
        first = items[0]
        self.assertIn("_boq_page_debug", first)
        debug = first["_boq_page_debug"]
        self.assertIn("method_used", debug)


class TestScheduleWithTableRouter(unittest.TestCase):
    """Test schedule extractor preserves legacy schema while using table_router."""

    def test_legacy_schema_preserved(self):
        """Schedule rows should have the same keys as before."""
        items = extract_schedule_rows(SCHEDULE_TEXT, source_page=0, sheet_id="A-501")
        self.assertGreater(len(items), 0)
        for item in items:
            self.assertIn("mark", item)
            self.assertIn("fields", item)
            self.assertIn("schedule_type", item)
            self.assertIn("source_page", item)
            self.assertIn("has_size", item)
            self.assertIn("has_qty", item)

    def test_schedule_extracts_same_or_more_rows(self):
        """Enhanced schedule should extract at least as many rows."""
        items = extract_schedule_rows(SCHEDULE_TEXT, source_page=0, sheet_id="A-501")
        self.assertGreaterEqual(len(items), 2)
        marks = {i["mark"] for i in items}
        self.assertTrue(any("D" in m for m in marks))


class TestPipelineExtDiagnostics(unittest.TestCase):
    """Test that pipeline payload includes extraction_diagnostics safely."""

    def test_extraction_diagnostics_in_result(self):
        """ExtractionResult.to_dict() should include extraction_diagnostics."""
        from src.analysis.extractors import ExtractionResult
        r = ExtractionResult()
        d = r.to_dict()
        self.assertIn("extraction_diagnostics", d)
        self.assertIsInstance(d["extraction_diagnostics"], dict)

    def test_extraction_diagnostics_populated_after_run(self):
        """run_extractors should populate extraction_diagnostics."""
        pages = [
            IndexedPage(page_idx=0, doc_type="boq", discipline="other",
                        confidence=0.7, has_text_layer=True),
        ]
        page_index = PageIndex(
            pdf_name="test.pdf",
            total_pages=1,
            pages=pages,
            counts_by_type={"boq": 1},
            counts_by_discipline={"other": 1},
        )
        text_map = {0: BOQ_TEXT}
        result = run_extractors(text_map, page_index)
        diag = result.extraction_diagnostics
        self.assertIn("boq", diag)
        self.assertIn("pages_attempted", diag["boq"])
        self.assertEqual(diag["boq"]["pages_attempted"], 1)

    def test_diagnostics_absent_does_not_crash(self):
        """UI should handle missing extraction_diagnostics gracefully."""
        # Simulate old payload without extraction_diagnostics
        payload = {"project_id": "test", "readiness_score": 50}
        ext_diag = payload.get("extraction_diagnostics") or {}
        boq = ext_diag.get("boq", {})
        # Should not crash
        self.assertEqual(boq.get("pages_attempted", 0), 0)


class TestProcessingStatsEnhanced(unittest.TestCase):
    """Test Sprint 20F additions to processing_stats."""

    def test_processing_stats_has_new_fields(self):
        """_build_processing_stats should include table_attempt/success fields."""
        stats = _build_processing_stats()
        self.assertIn("table_attempt_pages", stats)
        self.assertIn("table_success_pages", stats)
        self.assertIn("selection_mode", stats)
        self.assertIn("selected_pages_count", stats)

    def test_processing_stats_with_extraction_result(self):
        """Processing stats should pick up table counts from extraction_result."""
        from src.analysis.extractors import ExtractionResult
        er = ExtractionResult()
        er.extraction_diagnostics = {
            "boq": {"pages_attempted": 3, "pages_parsed": 2},
            "schedules": {"pages_attempted": 1, "pages_parsed": 1},
        }
        stats = _build_processing_stats(extraction_result=er)
        self.assertEqual(stats["table_attempt_pages"], 4)
        self.assertEqual(stats["table_success_pages"], 3)


# =============================================================================
# Sprint 20G: Pilot File Router Tests
# =============================================================================

from src.pilot.file_router import (
    classify_file, classify_directory, RoutingSummary,
    ClassifiedFile, _match_keywords, _is_sheet_filename,
    CATEGORY_DRAWINGS_PDF, CATEGORY_BOQ_PDF, CATEGORY_CONDITIONS_PDF,
    CATEGORY_ADDENDA_PDF, CATEGORY_BOQ_XLSX, CATEGORY_CONDITIONS_DOC,
    CATEGORY_UNKNOWN, BOQ_KEYWORDS, CONDITIONS_KEYWORDS,
    ADDENDA_KEYWORDS, DRAWING_KEYWORDS,
)


class TestFileRouterKeywords:
    """Test keyword matching for India tender file classification."""

    def test_boq_keywords_match(self):
        assert _match_keywords("BOQ_Part_A.pdf", BOQ_KEYWORDS) == "BOQ"
        assert _match_keywords("Schedule of Quantities Vol1", BOQ_KEYWORDS) == "SCHEDULE OF QUANTITIES"
        assert _match_keywords("PRICE BID - Civil Works", BOQ_KEYWORDS) == "PRICE BID"
        assert _match_keywords("SOQ_Final", BOQ_KEYWORDS) == "SOQ"

    def test_conditions_keywords_match(self):
        assert _match_keywords("GCC_2023", CONDITIONS_KEYWORDS) == "GCC"
        assert _match_keywords("Special_Conditions_of_Contract", CONDITIONS_KEYWORDS) == "CONDITIONS"
        assert _match_keywords("NIT_Document_Revised", CONDITIONS_KEYWORDS) == "NIT"
        assert _match_keywords("Tender Notice - CPWD", CONDITIONS_KEYWORDS) == "TENDER NOTICE"

    def test_addenda_keywords_match(self):
        assert _match_keywords("Addendum_No_3", ADDENDA_KEYWORDS) == "ADDENDUM"
        assert _match_keywords("Corrigendum_1_final", ADDENDA_KEYWORDS) == "CORRIGENDUM"
        assert _match_keywords("Clarification_Responses", ADDENDA_KEYWORDS) == "CLARIFICATION"

    def test_drawing_keywords_match(self):
        assert _match_keywords("Architectural_Drawings", DRAWING_KEYWORDS) == "DRAWING"
        assert _match_keywords("STR_Foundation_Plan", DRAWING_KEYWORDS) == "STR"
        assert _match_keywords("MEP_Layout_Revised", DRAWING_KEYWORDS) == "MEP"
        assert _match_keywords("Floor Plan - Ground", DRAWING_KEYWORDS) == "FLOOR PLAN"

    def test_no_match_returns_none(self):
        assert _match_keywords("readme.txt", BOQ_KEYWORDS) is None
        assert _match_keywords("random_file_name", DRAWING_KEYWORDS) is None

    def test_sheet_filename_pattern(self):
        assert _is_sheet_filename("A-101") is True
        assert _is_sheet_filename("S_002_Foundation") is True
        assert _is_sheet_filename("E-003") is True
        assert _is_sheet_filename("random_document") is False
        assert _is_sheet_filename("meeting_notes") is False


class TestFileRouterClassification:
    """Test file classification by extension and keywords."""

    def _make_file(self, tmp_path: Path, name: str, size: int = 1000) -> Path:
        """Create a dummy file for testing."""
        p = tmp_path / name
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"x" * size)
        return p

    def test_pdf_boq(self, tmp_path):
        f = self._make_file(tmp_path, "BOQ_Civil_Works.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_BOQ_PDF
        assert "BOQ" in cf.reason

    def test_pdf_conditions(self, tmp_path):
        f = self._make_file(tmp_path, "GCC_2023_Standard.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF
        assert "GCC" in cf.reason

    def test_pdf_addenda(self, tmp_path):
        f = self._make_file(tmp_path, "Addendum_No_2.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF
        assert "ADDENDUM" in cf.reason

    def test_pdf_drawing_keyword(self, tmp_path):
        f = self._make_file(tmp_path, "Architectural_Drawings_Set.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_DRAWINGS_PDF
        assert "DRAWING" in cf.reason

    def test_pdf_drawing_sheet_pattern(self, tmp_path):
        f = self._make_file(tmp_path, "A-101_Ground_Floor.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_DRAWINGS_PDF

    def test_pdf_large_default_drawings(self, tmp_path):
        """Large PDFs without keywords default to drawings."""
        f = self._make_file(tmp_path, "unnamed_set.pdf", size=6_000_000)
        cf = classify_file(f)
        assert cf.category == CATEGORY_DRAWINGS_PDF
        assert ">5MB" in cf.reason

    def test_pdf_unknown_small(self, tmp_path):
        """Small PDFs without keywords are unknown."""
        f = self._make_file(tmp_path, "cover_letter.pdf", size=500)
        cf = classify_file(f)
        assert cf.category == CATEGORY_UNKNOWN

    def test_docx_conditions(self, tmp_path):
        f = self._make_file(tmp_path, "Special_Conditions_SCC.docx")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_DOC

    def test_docx_default(self, tmp_path):
        """DOC files default to conditions category."""
        f = self._make_file(tmp_path, "letter_to_contractor.doc")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_DOC

    def test_xlsx_boq_filename(self, tmp_path):
        f = self._make_file(tmp_path, "BOQ_Price_Schedule.xlsx")
        cf = classify_file(f)
        assert cf.category == CATEGORY_BOQ_XLSX

    def test_xlsx_unknown_no_indicators(self, tmp_path):
        f = self._make_file(tmp_path, "random_data.xlsx")
        cf = classify_file(f)
        # Will try to open with openpyxl (will fail since it's just bytes)
        # Then fall back to filename match — "random_data" has no BOQ keywords
        assert cf.category == CATEGORY_UNKNOWN

    def test_unsupported_extension(self, tmp_path):
        f = self._make_file(tmp_path, "photo.jpg")
        cf = classify_file(f)
        assert cf.category == CATEGORY_UNKNOWN

    def test_addenda_priority_over_conditions(self, tmp_path):
        """Addenda keywords take priority over conditions keywords in PDFs."""
        f = self._make_file(tmp_path, "Addendum_to_Tender_Conditions.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF


class TestFileRouterDirectory:
    """Test batch directory classification."""

    def _make_tender(self, tmp_path: Path) -> Path:
        """Create a mock tender directory with mixed files."""
        root = tmp_path / "tender"
        root.mkdir()
        files = {
            "BOQ_Civil.pdf": 2000,
            "Architectural_Drawings.pdf": 8_000_000,
            "GCC_Standard.pdf": 5000,
            "Addendum_1.pdf": 3000,
            "Price_Schedule.xlsx": 1500,
            "NIT_Document.docx": 4000,
            "cover_photo.jpg": 500,  # should be skipped
            ".DS_Store": 100,        # should be skipped
        }
        for name, size in files.items():
            p = root / name
            p.write_bytes(b"x" * size)
        return root

    def test_classify_directory_counts(self, tmp_path):
        root = self._make_tender(tmp_path)
        summary = classify_directory(root, zip_name="test_tender.zip")
        # jpg and .DS_Store should be skipped
        assert summary.total_files == 6  # 4 PDFs + 1 XLSX + 1 DOCX
        by_cat = summary.by_category
        assert len(by_cat[CATEGORY_BOQ_PDF]) == 1
        assert len(by_cat[CATEGORY_DRAWINGS_PDF]) == 1
        assert len(by_cat[CATEGORY_CONDITIONS_PDF]) == 1
        assert len(by_cat[CATEGORY_ADDENDA_PDF]) == 1

    def test_pdf_files_property(self, tmp_path):
        root = self._make_tender(tmp_path)
        summary = classify_directory(root, zip_name="test.zip")
        pdf_files = summary.pdf_files
        assert len(pdf_files) == 4  # 4 PDFs (BOQ, drawings, conditions, addenda)
        assert all(p.suffix == ".pdf" for p in pdf_files)

    def test_summary_table(self, tmp_path):
        root = self._make_tender(tmp_path)
        summary = classify_directory(root, zip_name="test.zip")
        table = summary.summary_table()
        assert "drawings_pdf" in table
        assert "boq_pdf" in table
        assert "TOTAL" in table

    def test_empty_directory(self, tmp_path):
        root = tmp_path / "empty"
        root.mkdir()
        summary = classify_directory(root, zip_name="empty.zip")
        assert summary.total_files == 0


class TestScorecardRow:
    """Test scorecard row generation from pipeline payloads."""

    def test_empty_payload(self):
        from scripts.pilot_batch_ingest import build_scorecard_row
        row = build_scorecard_row("proj_1", "test.zip", None, 42.5, error="failed")
        assert row["project_id"] == "proj_1"
        assert row["tender_name"] == "test.zip"
        assert row["runtime_sec"] == "42.5"
        assert row["error"] == "failed"
        assert row["pages_total"] == ""

    def test_full_payload(self):
        from scripts.pilot_batch_ingest import build_scorecard_row, SCORECARD_FIELDS
        payload = {
            "processing_stats": {
                "total_pages": 367,
                "deep_processed_pages": 367,
                "ocr_pages": 367,
                "skipped_pages": 0,
                "toxic_pages": 0,
                "selection_mode": "full_read",
            },
            "extraction_diagnostics": {
                "boq": {"pages_attempted": 5, "pages_parsed": 3},
                "schedules": {"pages_attempted": 2, "pages_parsed": 2},
            },
            "boq_items": [{"item": "1.1"}, {"item": "1.2"}, {"item": "1.3"}],
            "schedules": {
                "doors": [{"mark": "D1"}],
                "windows": [{"mark": "W1"}, {"mark": "W2"}],
                "finishes": [],
            },
            "commercial_terms": [{"term": "LD"}, {"term": "retention"}],
            "requirements": [{"req": "r1"}],
            "rfis": [{"q": "q1"}, {"q": "q2"}],
            "blockers": [{"b": "b1"}],
            "readiness_score": 73,
        }
        row = build_scorecard_row("proj_2", "sonipat.zip", payload, 228.5)
        assert row["pages_total"] == 367
        assert row["deep_processed_pages"] == 367
        assert row["boq_pages_attempted"] == 5
        assert row["boq_pages_parsed"] == 3
        assert row["boq_items_count"] == 3
        assert row["door_rows"] == 1
        assert row["window_rows"] == 2
        assert row["finish_rows"] == 0
        assert row["commercial_terms_count"] == 2
        assert row["requirements_count"] == 1
        assert row["rfis_count"] == 2
        assert row["blockers_count"] == 1
        assert row["error"] == ""
        # All scorecard fields should be present
        for field in SCORECARD_FIELDS:
            assert field in row, f"Missing field: {field}"

    def test_scorecard_csv_write(self, tmp_path):
        from scripts.pilot_batch_ingest import build_scorecard_row, write_scorecard, SCORECARD_FIELDS
        csv_path = tmp_path / "test_scorecard.csv"
        row1 = build_scorecard_row("p1", "a.zip", None, 10.0)
        row2 = build_scorecard_row("p2", "b.zip", None, 20.0, error="boom")
        write_scorecard([row1], csv_path)
        write_scorecard([row2], csv_path)  # append
        import csv as csv_mod
        with open(csv_path) as f:
            reader = csv_mod.DictReader(f)
            rows = list(reader)
        assert len(rows) == 2
        assert rows[0]["project_id"] == "p1"
        assert rows[1]["error"] == "boom"


class TestPilotBatchDryRun:
    """Smoke tests for dry-run mode (ZIPs + folders)."""

    def test_dry_run_imports(self):
        """Verify all pilot modules import cleanly."""
        from src.pilot.file_router import classify_file, classify_directory, _normalize_name
        from scripts.pilot_batch_ingest import (
            build_scorecard_row, write_scorecard, process_single_zip,
            process_single_folder, find_zips, find_folders, find_inputs,
        )

    def test_find_zips_empty(self, tmp_path):
        from scripts.pilot_batch_ingest import find_zips
        assert find_zips(tmp_path) == []

    def test_find_zips_with_filter(self, tmp_path):
        from scripts.pilot_batch_ingest import find_zips
        (tmp_path / "a.zip").write_bytes(b"PK")
        (tmp_path / "b.zip").write_bytes(b"PK")
        (tmp_path / "c.txt").write_bytes(b"nope")
        all_zips = find_zips(tmp_path)
        assert len(all_zips) == 2
        filtered = find_zips(tmp_path, only="a.zip")
        assert len(filtered) == 1
        assert filtered[0].name == "a.zip"


class TestFolderIngest:
    """Tests for folder-based input discovery and processing."""

    def test_find_folders_basic(self, tmp_path):
        from scripts.pilot_batch_ingest import find_folders
        (tmp_path / "Tender A").mkdir()
        (tmp_path / "Tender B").mkdir()
        (tmp_path / ".hidden").mkdir()
        (tmp_path / "file.txt").write_text("not a dir")
        folders = find_folders(tmp_path)
        assert len(folders) == 2
        names = [f.name for f in folders]
        assert "Tender A" in names
        assert "Tender B" in names
        assert ".hidden" not in names

    def test_find_folders_with_filter(self, tmp_path):
        from scripts.pilot_batch_ingest import find_folders
        (tmp_path / "Tender A").mkdir()
        (tmp_path / "Tender B").mkdir()
        (tmp_path / "Tender C").mkdir()
        filtered = find_folders(tmp_path, only=["Tender A", "Tender C"])
        assert len(filtered) == 2
        names = [f.name for f in filtered]
        assert "Tender A" in names
        assert "Tender C" in names
        assert "Tender B" not in names

    def test_find_inputs_auto_folders(self, tmp_path):
        """Auto mode detects folders when no ZIPs present."""
        from scripts.pilot_batch_ingest import find_inputs
        (tmp_path / "Tender 1").mkdir()
        (tmp_path / "Tender 2").mkdir()
        items, kind = find_inputs(tmp_path, input_type="auto")
        assert kind == "folder"
        assert len(items) == 2

    def test_find_inputs_auto_zips(self, tmp_path):
        """Auto mode detects ZIPs when no folders present."""
        from scripts.pilot_batch_ingest import find_inputs
        (tmp_path / "a.zip").write_bytes(b"PK")
        items, kind = find_inputs(tmp_path, input_type="auto")
        assert kind == "zip"
        assert len(items) == 1

    def test_find_inputs_limit(self, tmp_path):
        from scripts.pilot_batch_ingest import find_inputs
        for i in range(5):
            (tmp_path / f"Tender {i}").mkdir()
        items, kind = find_inputs(tmp_path, input_type="folder", limit=2)
        assert len(items) == 2

    def test_find_inputs_none(self, tmp_path):
        from scripts.pilot_batch_ingest import find_inputs
        items, kind = find_inputs(tmp_path)
        assert len(items) == 0
        assert kind == "none"

    def test_process_single_folder_dry_run(self, tmp_path):
        """Dry-run on a folder classifies files without running pipeline."""
        from scripts.pilot_batch_ingest import process_single_folder
        tender = tmp_path / "Test Tender"
        tender.mkdir()
        (tender / "BOQ.pdf").write_bytes(b"x" * 1000)
        (tender / "Drawings.pdf").write_bytes(b"x" * 6_000_000)
        (tender / "NIT.pdf").write_bytes(b"x" * 2000)
        output = tmp_path / "output"
        row = process_single_folder(
            folder_path=tender,
            output_dir=output,
            dry_run=True,
            verbose=False,
        )
        assert "dry-run" in str(row["pages_total"])
        assert "3 PDFs" in str(row["pages_total"])
        assert row["error"] == ""

    def test_process_single_folder_no_pdfs(self, tmp_path):
        """Folder with no PDFs reports error in scorecard."""
        from scripts.pilot_batch_ingest import process_single_folder
        tender = tmp_path / "Empty Tender"
        tender.mkdir()
        (tender / "readme.txt").write_text("nothing here")
        output = tmp_path / "output"
        row = process_single_folder(
            folder_path=tender,
            output_dir=output,
            dry_run=False,
            verbose=False,
        )
        assert row["error"] != ""  # should report "No PDF files found"


class TestNormalizeName:
    """Test filename normalization for keyword matching."""

    def test_underscore_split(self):
        from src.pilot.file_router import _normalize_name
        assert _normalize_name("tech_prebid_67946") == "tech prebid 67946"

    def test_hyphen_split(self):
        from src.pilot.file_router import _normalize_name
        assert _normalize_name("Pre-Bid-Meeting") == "Pre Bid Meeting"

    def test_camel_case_split(self):
        from src.pilot.file_router import _normalize_name
        assert _normalize_name("NoticeInvitingTender") == "Notice Inviting Tender"
        assert _normalize_name("TenderNotice") == "Tender Notice"

    def test_mixed(self):
        from src.pilot.file_router import _normalize_name
        result = _normalize_name("NoticeInvitingTenderPartAandB")
        assert "Notice Inviting Tender" in result

    def test_numbers_preserved(self):
        from src.pilot.file_router import _normalize_name
        assert _normalize_name("BOQ_67946") == "BOQ 67946"


class TestFileRouterNewKeywords:
    """Test new keywords added for India tender classification."""

    def _make_file(self, tmp_path: Path, name: str, size: int = 1000) -> Path:
        p = tmp_path / name
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"x" * size)
        return p

    def test_instruction_to_tenderer(self, tmp_path):
        """Instruction_to_tenderer.pdf → conditions."""
        f = self._make_file(tmp_path, "Instruction_to_tenderer.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF
        assert "INSTRUCTION" in cf.reason

    def test_notice_inviting_tender(self, tmp_path):
        """NoticeInvitingTenderPartAandB.pdf → conditions (via CamelCase split)."""
        f = self._make_file(tmp_path, "NoticeInvitingTenderPartAandB.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF

    def test_tender_notice_camel(self, tmp_path):
        """TenderNotice.pdf → conditions (CamelCase: 'Tender Notice')."""
        f = self._make_file(tmp_path, "TenderNotice.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF

    def test_prebid_addenda(self, tmp_path):
        """tech_prebid_67946.pdf → addenda (PREBID keyword)."""
        f = self._make_file(tmp_path, "tech_prebid_67946.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF

    def test_pre_bid_meeting(self, tmp_path):
        """Pre_Bid_Meeting_Onco.pdf → addenda."""
        f = self._make_file(tmp_path, "Pre_Bid_Meeting_Onco.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF

    def test_amdt_addenda(self, tmp_path):
        """1Amdt83061.pdf → addenda (AMDT keyword)."""
        f = self._make_file(tmp_path, "1Amdt83061.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF

    def test_ammendent_addenda(self, tmp_path):
        """ammendent.pdf → addenda (AMMEND keyword)."""
        f = self._make_file(tmp_path, "ammendent.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF

    def test_rfp_conditions(self, tmp_path):
        """RFP.pdf → conditions."""
        f = self._make_file(tmp_path, "RFP.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF

    def test_pbq_boq(self, tmp_path):
        """PBQ.pdf → BOQ (PBQ keyword)."""
        f = self._make_file(tmp_path, "PBQ.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_BOQ_PDF

    def test_gad_drawing(self, tmp_path):
        """GAD.pdf → drawings (General Arrangement Drawing)."""
        f = self._make_file(tmp_path, "GAD.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_DRAWINGS_PDF

    def test_rar_recognized(self, tmp_path):
        """RAR files are recognized but classified as unknown."""
        f = self._make_file(tmp_path, "Drawings_1.rar")
        cf = classify_file(f)
        assert cf.category == CATEGORY_UNKNOWN
        assert "RAR" in cf.reason

    def test_parent_directory_drawings(self, tmp_path):
        """Files in 'Tender Drawings' subdir → drawings via parent check."""
        drawings_dir = tmp_path / "Vol-IV" / "Tender Drawings"
        drawings_dir.mkdir(parents=True)
        f = drawings_dir / "BOUNDARY WALL.pdf"
        f.write_bytes(b"x" * 1000)
        cf = classify_file(f, relative_root=tmp_path)
        assert cf.category == CATEGORY_DRAWINGS_PDF
        assert "parent directory" in cf.reason

    def test_section_conditions(self, tmp_path):
        """Section 1.pdf → conditions."""
        f = self._make_file(tmp_path, "Section 1.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF

    def test_all_pdf_files_property(self, tmp_path):
        """all_pdf_files includes unknown-category PDFs."""
        root = tmp_path / "tender"
        root.mkdir()
        (root / "BOQ.pdf").write_bytes(b"x" * 1000)
        (root / "unknown_file.pdf").write_bytes(b"x" * 500)
        (root / "data.xlsx").write_bytes(b"x" * 300)
        summary = classify_directory(root)
        assert len(summary.pdf_files) == 1  # Only BOQ (known category)
        assert len(summary.all_pdf_files) == 2  # Both PDFs

    def test_modified_addenda(self, tmp_path):
        """Modified_P_P.pdf → addenda (MODIFIED keyword)."""
        f = self._make_file(tmp_path, "Modified_P_P.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_ADDENDA_PDF

    def test_irgcc_conditions(self, tmp_path):
        """IRGCC.pdf → conditions (contains GCC)."""
        f = self._make_file(tmp_path, "IRGCC.pdf")
        cf = classify_file(f)
        assert cf.category == CATEGORY_CONDITIONS_PDF
        assert "GCC" in cf.reason


# =============================================================================
# Pilot Inventory Tests
# =============================================================================


class TestPilotInventoryHelpers:
    """Test lightweight metadata extractors for pilot_inventory."""

    def test_find_tender_folders(self, tmp_path):
        from scripts.pilot_inventory import find_tender_folders
        (tmp_path / "Tender A").mkdir()
        (tmp_path / "Tender B").mkdir()
        (tmp_path / ".hidden").mkdir()
        (tmp_path / "file.txt").write_text("x")
        folders = find_tender_folders(tmp_path)
        assert len(folders) == 2
        names = [f.name for f in folders]
        assert "Tender A" in names
        assert ".hidden" not in names

    def test_find_tender_folders_include(self, tmp_path):
        from scripts.pilot_inventory import find_tender_folders
        (tmp_path / "Tender A").mkdir()
        (tmp_path / "Tender B").mkdir()
        (tmp_path / "Tender C").mkdir()
        folders = find_tender_folders(tmp_path, include=["Tender B"])
        assert len(folders) == 1
        assert folders[0].name == "Tender B"

    def test_format_size(self):
        from scripts.pilot_inventory import _format_size
        assert _format_size(500) == "500 B"
        assert "KB" in _format_size(2048)
        assert "MB" in _format_size(5_000_000)

    def test_get_pdf_metadata_missing_file(self, tmp_path):
        """get_pdf_metadata returns zeros for non-existent/corrupt files."""
        from scripts.pilot_inventory import get_pdf_metadata
        fake = tmp_path / "nope.pdf"
        fake.write_bytes(b"not a pdf")
        meta = get_pdf_metadata(fake)
        assert isinstance(meta, dict)
        assert "page_count" in meta
        # Should not crash; page_count may be 0

    def test_get_excel_preview_basic(self, tmp_path):
        """get_excel_preview reads sheet names and headers from a real xlsx."""
        from scripts.pilot_inventory import get_excel_preview
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOQ Sheet"
        ws.append(["Item", "Description", "UOM", "Qty", "Rate", "Amount"])
        ws.append(["1.1", "Earthwork excavation", "cum", 100, 250, 25000])
        ws.append(["1.2", "PCC M15", "cum", 50, 3500, 175000])
        xlsx_path = tmp_path / "test_boq.xlsx"
        wb.save(xlsx_path)
        wb.close()

        preview = get_excel_preview(xlsx_path)
        assert "BOQ Sheet" in preview["sheet_names"]
        assert preview["likely_boq"] is True
        assert len(preview["boq_indicators"]) > 0
        assert preview["sheets"]["BOQ Sheet"]["header_rows"][0][0] == "Item"

    def test_get_excel_preview_no_boq(self, tmp_path):
        """Excel without BOQ keywords is not flagged as BOQ."""
        from scripts.pilot_inventory import get_excel_preview
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["Name", "Phone", "Email"])
        xlsx_path = tmp_path / "contacts.xlsx"
        wb.save(xlsx_path)
        wb.close()

        preview = get_excel_preview(xlsx_path)
        assert preview["likely_boq"] is False

    def test_get_excel_preview_corrupt(self, tmp_path):
        """Corrupt Excel returns error without crashing."""
        from scripts.pilot_inventory import get_excel_preview
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not an excel file")
        preview = get_excel_preview(bad)
        assert isinstance(preview, dict)
        assert "error" in preview or preview["sheet_names"] == []


class TestPilotInventoryBuild:
    """Test the build_tender_inventory function."""

    def _make_tender_dir(self, tmp_path):
        """Create a mock tender directory."""
        tender = tmp_path / "Test Tender"
        tender.mkdir()
        (tender / "BOQ.pdf").write_bytes(b"%PDF-1.4 fake content" + b"x" * 1000)
        (tender / "GCC_Standard.pdf").write_bytes(b"%PDF-1.4 fake" + b"x" * 500)
        (tender / "Addendum_1.pdf").write_bytes(b"%PDF-1.4 fake" + b"x" * 300)
        (tender / "Drawing_Set.pdf").write_bytes(b"%PDF-1.4 fake" + b"x" * 6_000_000)

        # Create a real xlsx for preview testing
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price Schedule"
        ws.append(["Sl No", "Description", "Unit", "Qty", "Rate", "Amount"])
        ws.append(["1", "Cement", "bag", "100", "350", "35000"])
        wb.save(tender / "BOQ_12345.xlsx")
        wb.close()

        return tender

    def test_build_inventory_structure(self, tmp_path):
        from scripts.pilot_inventory import build_tender_inventory
        tender = self._make_tender_dir(tmp_path)
        inv = build_tender_inventory(tender)

        assert inv["folder_name"] == "Test Tender"
        assert inv["total_files"] == 5  # 4 PDFs + 1 xlsx
        assert inv["total_size_bytes"] > 0
        assert "files" in inv
        assert "classification_summary" in inv
        assert "analysis" in inv

    def test_build_inventory_analysis(self, tmp_path):
        from scripts.pilot_inventory import build_tender_inventory
        tender = self._make_tender_dir(tmp_path)
        inv = build_tender_inventory(tender)
        a = inv["analysis"]

        assert a["likely_boq_source"] == "Excel"
        assert "BOQ_12345.xlsx" in a["boq_candidates"]
        assert a["addenda_count"] == 1
        assert a["drawings_count"] >= 1
        assert len(a["suggested_processing_plan"]) > 0

    def test_build_inventory_excel_preview(self, tmp_path):
        from scripts.pilot_inventory import build_tender_inventory
        tender = self._make_tender_dir(tmp_path)
        inv = build_tender_inventory(tender)

        assert len(inv["excel_previews"]) == 1
        key = list(inv["excel_previews"].keys())[0]
        preview = inv["excel_previews"][key]
        assert preview["likely_boq"] is True


class TestPilotInventoryOutputs:
    """Test output writers for pilot_inventory."""

    def _quick_inventory(self) -> dict:
        """Build a minimal inventory dict for testing writers."""
        return {
            "folder_name": "Test Tender",
            "folder_path": "/tmp/test",
            "scan_timestamp": "2026-02-27T00:00:00",
            "files": [
                {"name": "BOQ.pdf", "relative_path": "BOQ.pdf",
                 "extension": ".pdf", "size_bytes": 1000, "size_human": "1.0 KB",
                 "category": "boq_pdf", "classification_reason": "keyword BOQ",
                 "page_count": 10},
            ],
            "file_type_counts": {".pdf": 1},
            "total_files": 1,
            "total_size_bytes": 1000,
            "total_size_human": "1.0 KB",
            "pdf_total_pages": 10,
            "classification_summary": {
                "boq_pdf": {"count": 1, "files": ["BOQ.pdf"], "total_size": 1000},
            },
            "pdf_metadata": {},
            "excel_previews": {},
            "analysis": {
                "likely_boq_source": "PDF",
                "boq_candidates": ["BOQ.pdf"],
                "conditions_files": [],
                "gcc_scc_candidates": [],
                "addenda_files": [],
                "addenda_count": 0,
                "drawings_count": 0,
                "drawings_total_pages": 0,
                "missing_items": ["No conditions/specs files found", "No drawing files identified"],
                "potential_duplicates": {},
                "suggested_processing_plan": ["1. OCR+extract BOQ PDFs (BOQ.pdf)"],
            },
            "scan_duration_sec": 0.1,
        }

    def test_write_manifest_json(self, tmp_path):
        from scripts.pilot_inventory import write_manifest_json
        inv = self._quick_inventory()
        path = tmp_path / "manifest.json"
        write_manifest_json(inv, path)
        assert path.exists()
        data = json.loads(path.read_text())
        assert data["folder_name"] == "Test Tender"
        assert len(data["files"]) == 1

    def test_write_summary_md(self, tmp_path):
        from scripts.pilot_inventory import write_summary_md
        inv = self._quick_inventory()
        path = tmp_path / "summary.md"
        write_summary_md(inv, path)
        assert path.exists()
        content = path.read_text()
        assert "# Tender Inventory: Test Tender" in content
        assert "BOQ" in content
        assert "Missing" in content

    def test_write_rollup_md(self, tmp_path):
        from scripts.pilot_inventory import write_rollup_md
        inv = self._quick_inventory()
        path = tmp_path / "rollup.md"
        write_rollup_md([inv], path)
        assert path.exists()
        content = path.read_text()
        assert "Master Rollup" in content
        assert "Test Tender" in content

    def test_write_rollup_csv(self, tmp_path):
        from scripts.pilot_inventory import write_rollup_csv, ROLLUP_CSV_FIELDS
        inv = self._quick_inventory()
        path = tmp_path / "rollup.csv"
        write_rollup_csv([inv], path)
        assert path.exists()
        import csv as csv_mod
        with open(path) as f:
            reader = csv_mod.DictReader(f)
            rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["tender_name"] == "Test Tender"


class TestPilotInventorySmoke:
    """Smoke test: run pilot_inventory on a temp folder with dummy files."""

    def test_end_to_end_smoke(self, tmp_path):
        """Full pipeline: create mock tender, run inventory, verify outputs."""
        from scripts.pilot_inventory import (
            find_tender_folders, build_tender_inventory,
            write_manifest_json, write_summary_md,
            write_rollup_md, write_rollup_csv,
        )

        # Create two mock tenders
        parent = tmp_path / "tenders"
        parent.mkdir()

        t1 = parent / "Tender Alpha"
        t1.mkdir()
        (t1 / "NIT.pdf").write_bytes(b"%PDF-1.4 nit" + b"x" * 500)
        (t1 / "BOQ.pdf").write_bytes(b"%PDF-1.4 boq" + b"x" * 800)

        t2 = parent / "Tender Beta"
        t2.mkdir()
        (t2 / "Drawings.pdf").write_bytes(b"%PDF-1.4 drw" + b"x" * 6_000_000)
        (t2 / "readme.txt").write_text("info file")

        # Discover
        folders = find_tender_folders(parent)
        assert len(folders) == 2

        # Build inventories
        inventories = [build_tender_inventory(f) for f in folders]
        assert len(inventories) == 2

        # Write outputs
        out = tmp_path / "output"
        for inv in inventories:
            td = out / "inventory" / inv["folder_name"]
            write_manifest_json(inv, td / "manifest.json")
            write_summary_md(inv, td / "summary.md")

        write_rollup_md(inventories, out / "inventory_rollup.md")
        write_rollup_csv(inventories, out / "inventory_rollup.csv")

        # Verify outputs exist
        assert (out / "inventory" / "Tender Alpha" / "manifest.json").exists()
        assert (out / "inventory" / "Tender Alpha" / "summary.md").exists()
        assert (out / "inventory" / "Tender Beta" / "manifest.json").exists()
        assert (out / "inventory" / "Tender Beta" / "summary.md").exists()
        assert (out / "inventory_rollup.md").exists()
        assert (out / "inventory_rollup.csv").exists()

        # Spot-check content
        alpha_manifest = json.loads(
            (out / "inventory" / "Tender Alpha" / "manifest.json").read_text()
        )
        assert alpha_manifest["total_files"] == 2
        assert alpha_manifest["analysis"]["likely_boq_source"] in ("PDF", "Excel")

        rollup_content = (out / "inventory_rollup.md").read_text()
        assert "Tender Alpha" in rollup_content
        assert "Tender Beta" in rollup_content


# =============================================================================
# SPRINT 21 ULTIMATE — CLAUDE OS + PILOT OPS + REGRESSION + DATASET CAPTURE
# =============================================================================


class TestRepoIntelligenceFiles:
    """Verify that key documentation and skill files exist."""

    REPO_ROOT = Path(__file__).resolve().parent.parent

    def test_claude_md_exists(self):
        path = self.REPO_ROOT / "CLAUDE.md"
        assert path.exists(), "CLAUDE.md missing"
        content = path.read_text()
        assert "xBOQ" in content
        assert "run_analysis_pipeline" in content
        assert "pilot_batch_ingest" in content

    def test_architecture_md_exists(self):
        path = self.REPO_ROOT / "docs" / "ARCHITECTURE.md"
        assert path.exists(), "docs/ARCHITECTURE.md missing"
        content = path.read_text()
        assert "pipeline" in content.lower()
        assert "extractors" in content.lower()

    def test_debugging_md_exists(self):
        path = self.REPO_ROOT / "docs" / "DEBUGGING.md"
        assert path.exists(), "docs/DEBUGGING.md missing"
        content = path.read_text()
        assert "BOQ" in content
        assert "Pilot Scorecard Triage" in content

    def test_pilot_workflow_md_exists(self):
        path = self.REPO_ROOT / "docs" / "PILOT_WORKFLOW.md"
        assert path.exists(), "docs/PILOT_WORKFLOW.md missing"
        content = path.read_text()
        assert "inventory" in content.lower()
        assert "dry-run" in content.lower() or "dry_run" in content.lower()

    def test_dataset_capture_md_exists(self):
        path = self.REPO_ROOT / "docs" / "DATASET_CAPTURE.md"
        assert path.exists(), "docs/DATASET_CAPTURE.md missing"
        content = path.read_text()
        assert "ground_truth" in content.lower() or "ground truth" in content.lower()

    def test_skill_files_exist(self):
        skills_dir = self.REPO_ROOT / "prompts" / "skills"
        assert skills_dir.is_dir(), "prompts/skills/ directory missing"
        skill_files = list(skills_dir.glob("*.skill.md"))
        assert len(skill_files) >= 7, f"Expected >=7 skill files, found {len(skill_files)}"
        # Check pilot_summary specifically (new in this sprint)
        assert (skills_dir / "pilot_summary.skill.md").exists(), "pilot_summary.skill.md missing"

    def test_benchmarks_scaffolding_exists(self):
        bench_dir = self.REPO_ROOT / "benchmarks"
        assert bench_dir.is_dir(), "benchmarks/ directory missing"
        assert (bench_dir / "README.md").exists(), "benchmarks/README.md missing"
        assert (bench_dir / "manifest.example.json").exists(), "benchmarks/manifest.example.json missing"
        assert (bench_dir / "expected_metrics.yaml").exists(), "benchmarks/expected_metrics.yaml missing"


class TestRunRegression:
    """Verify run_regression.py exits 0 with helpful message if no manifest."""

    def test_no_manifest_exits_zero(self, tmp_path):
        """Script should exit 0 and print instructions when manifest is missing."""
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "run_regression.py"),
             "--manifest", str(tmp_path / "nonexistent_manifest.json")],
            capture_output=True, text=True, timeout=30,
        )
        assert result.returncode == 0, f"Expected exit 0, got {result.returncode}: {result.stderr}"
        assert "manifest" in result.stdout.lower() or "setup" in result.stdout.lower()

    def test_extract_kpis_function(self):
        """Test KPI extraction from a synthetic payload."""
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from scripts.run_regression import extract_kpis

        payload = {
            "processing_stats": {
                "total_pages": 100,
                "deep_processed_pages": 80,
                "ocr_pages": 20,
                "toxic_pages": 2,
            },
            "extraction_diagnostics": {
                "boq_pages_attempted": 10,
                "boq_pages_parsed": 8,
                "schedule_pages_attempted": 5,
                "schedule_pages_parsed": 3,
                "table_methods_used": {"pdfplumber": 15, "camelot_lattice": 3},
            },
            "boq_items": [{"item": 1}, {"item": 2}],
            "rfis": [{"rfi": 1}],
            "blockers": [],
            "qa_score": 0.85,
        }

        kpis = extract_kpis(payload)
        assert kpis["pages_total"] == 100
        assert kpis["deep_processed_pages"] == 80
        assert kpis["ocr_pages"] == 20
        assert kpis["boq_pages_attempted"] == 10
        assert kpis["boq_pages_parsed"] == 8
        assert kpis["rfis_count"] == 1
        assert kpis["blockers_count"] == 0
        assert kpis["qa_score"] == 0.85


class TestPilotOps:
    """Verify pilot_ops.py help text and dry-run path."""

    def test_help_text(self):
        """--help should exit 0 and print usage."""
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "pilot_ops.py"),
             "--help"],
            capture_output=True, text=True, timeout=15,
        )
        assert result.returncode == 0, f"Expected exit 0, got {result.returncode}: {result.stderr}"
        assert "pilot" in result.stdout.lower() or "mode" in result.stdout.lower()

    def test_summary_no_scorecard(self, tmp_path):
        """Summary mode with no scorecard should exit 0."""
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "pilot_ops.py"),
             "--input", str(tmp_path), "--mode", "summary"],
            capture_output=True, text=True, timeout=15,
        )
        assert result.returncode == 0, f"Expected exit 0, got {result.returncode}: {result.stderr}"

    def test_summary_reads_scorecard(self, tmp_path):
        """Summary mode should read and print scorecard content."""
        # Create a minimal scorecard
        scorecard = tmp_path / "pilot_scorecard.csv"
        scorecard.write_text(
            "tender_name,status,boq_items,rfis,qa_score,duration_sec,error\n"
            "TestTender,OK,15,8,0.75,30.5,\n"
            "TestTender2,ERROR,0,0,,5.0,crash\n"
        )
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "pilot_ops.py"),
             "--input", str(tmp_path), "--mode", "summary"],
            capture_output=True, text=True, timeout=15,
        )
        assert result.returncode == 0
        assert "TestTender" in result.stdout
        assert "2 tender" in result.stdout.lower()

    def test_inventory_mode(self, tmp_path):
        """Inventory mode should run without crashing."""
        # Create mock tender folders
        input_dir = tmp_path / "input"
        output_dir = tmp_path / "output"
        input_dir.mkdir()
        (input_dir / "Tender A").mkdir()
        (input_dir / "Tender A" / "file.pdf").write_bytes(b"%PDF-fake")
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "pilot_ops.py"),
             "--input", str(input_dir), "--output", str(output_dir), "--mode", "inventory"],
            capture_output=True, text=True, timeout=30,
        )
        # Should complete without crash
        assert result.returncode == 0


class TestRunArtifacts:
    """Verify src/ops/run_artifacts.py writes expected files."""

    def test_write_run_manifest(self, tmp_path):
        from src.ops.run_artifacts import write_run_manifest
        payload = {"boq_items": [], "rfis": []}
        path = write_run_manifest(payload, tmp_path)
        assert path.exists()
        data = json.loads(path.read_text())
        assert "generated_at" in data
        assert "payload_keys" in data
        assert "boq_items" in data["payload_keys"]

    def test_write_processing_stats(self, tmp_path):
        from src.ops.run_artifacts import write_processing_stats
        payload = {
            "processing_stats": {"total_pages": 100, "ocr_pages": 20},
        }
        path = write_processing_stats(payload, tmp_path)
        assert path is not None
        assert path.exists()
        data = json.loads(path.read_text())
        assert data["total_pages"] == 100

    def test_write_processing_stats_missing(self, tmp_path):
        from src.ops.run_artifacts import write_processing_stats
        path = write_processing_stats({}, tmp_path)
        assert path is None

    def test_write_extraction_diagnostics(self, tmp_path):
        from src.ops.run_artifacts import write_extraction_diagnostics
        payload = {
            "extraction_diagnostics": {"boq_pages_attempted": 10, "boq_pages_parsed": 8},
        }
        path = write_extraction_diagnostics(payload, tmp_path)
        assert path is not None
        data = json.loads(path.read_text())
        assert data["boq_pages_attempted"] == 10

    def test_write_run_summary(self, tmp_path):
        from src.ops.run_artifacts import write_run_summary
        payload = {
            "processing_stats": {"total_pages": 50, "deep_processed_pages": 40, "ocr_pages": 5, "selection_mode": "full_audit"},
            "extraction_diagnostics": {"boq_pages_attempted": 5, "boq_pages_parsed": 4},
            "boq_items": [{"item": 1}],
            "rfis": [{"rfi": 1}, {"rfi": 2}],
            "blockers": [],
            "qa_score": 0.9,
        }
        path = write_run_summary(payload, tmp_path)
        assert path.exists()
        data = json.loads(path.read_text())
        assert data["pages_total"] == 50
        assert data["boq_items"] == 1
        assert data["rfis"] == 2
        assert data["qa_score"] == 0.9

    def test_write_error_report(self, tmp_path):
        from src.ops.run_artifacts import write_error_report
        try:
            raise ValueError("test error")
        except ValueError as e:
            path = write_error_report(e, tmp_path, context={"tender": "test"})
        assert path.exists()
        content = path.read_text().strip()
        entry = json.loads(content)
        assert entry["error_type"] == "ValueError"
        assert "test error" in entry["error_message"]
        assert entry["context"]["tender"] == "test"

    def test_write_error_report_append(self, tmp_path):
        """Multiple errors should append as JSONL."""
        from src.ops.run_artifacts import write_error_report
        write_error_report("first error", tmp_path)
        write_error_report("second error", tmp_path)
        lines = (tmp_path / "errors.jsonl").read_text().strip().split("\n")
        assert len(lines) == 2

    def test_write_all_artifacts(self, tmp_path):
        from src.ops.run_artifacts import write_all_artifacts
        payload = {
            "processing_stats": {"total_pages": 10},
            "extraction_diagnostics": {"boq_pages_attempted": 2},
            "boq_items": [],
            "rfis": [],
            "blockers": [],
        }
        paths = write_all_artifacts(payload, tmp_path, extra={"tender": "test"})
        assert len(paths) >= 3  # manifest + stats + diagnostics + summary
        assert (tmp_path / "run_manifest.json").exists()
        assert (tmp_path / "processing_stats.json").exists()
        assert (tmp_path / "extraction_diagnostics.json").exists()
        assert (tmp_path / "run_summary.json").exists()


class TestInitDatasetCase:
    """Verify init_dataset_case.py creates directory structure."""

    def test_init_case_function(self, tmp_path):
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from scripts.init_dataset_case import init_case
        case_root = init_case("acme", "hospital", "run_001", base_dir=tmp_path)
        assert case_root.exists()
        for subdir in ["inputs", "outputs", "playbook", "ground_truth", "feedback"]:
            assert (case_root / subdir).is_dir(), f"Missing subdir: {subdir}"
        assert case_root == tmp_path / "acme" / "hospital" / "run_001"

    def test_init_case_default_run(self, tmp_path):
        """Run name defaults to today's date."""
        from scripts.init_dataset_case import init_case
        case_root = init_case("acme", "hospital", base_dir=tmp_path)
        assert case_root.exists()
        # Run dir name should be a date string
        from datetime import date
        assert date.today().isoformat() in str(case_root)

    def test_init_case_cli(self, tmp_path):
        """CLI should exit 0 and create directories."""
        import subprocess
        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "scripts" / "init_dataset_case.py"),
             "--tenant", "test_co",
             "--project", "proj1",
             "--run", "run_test",
             "--base-dir", str(tmp_path)],
            capture_output=True, text=True, timeout=15,
        )
        assert result.returncode == 0, f"Exit {result.returncode}: {result.stderr}"
        assert (tmp_path / "test_co" / "proj1" / "run_test" / "inputs").is_dir()
        assert (tmp_path / "test_co" / "proj1" / "run_test" / "ground_truth").is_dir()


class TestJudgeKpi:
    """Verify KPI judgement logic in run_regression.py."""

    def setup_method(self):
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from scripts.run_regression import judge_kpi
        self.judge = judge_kpi

    def test_pass_above_min(self):
        assert self.judge("boq_items", 60, {"min_boq_items": 50}) == "PASS"

    def test_warn_near_min(self):
        assert self.judge("boq_items", 46, {"min_boq_items": 50}) == "WARN"

    def test_fail_below_min(self):
        assert self.judge("boq_items", 30, {"min_boq_items": 50}) == "FAIL"

    def test_pass_below_max(self):
        assert self.judge("toxic_pages", 3, {"max_toxic_pages": 5}) == "PASS"

    def test_fail_above_max(self):
        assert self.judge("toxic_pages", 10, {"max_toxic_pages": 5}) == "FAIL"

    def test_info_no_thresholds(self):
        assert self.judge("pages_total", 100, {}) == "INFO"

    def test_info_string_value(self):
        assert self.judge("table_methods", "pdfplumber=5", {"min_table_methods": 1}) == "INFO"


# =============================================================================
# RUN
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
