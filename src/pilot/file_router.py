"""
File Router — deterministic India tender heuristics for classifying
files inside a pilot delivery ZIP or folder.

Categories:
    drawings_pdf   — Architectural/structural/MEP drawing sets
    boq_pdf        — Bill of Quantities / Schedule of Quantities PDFs
    conditions_pdf — GCC, SCC, NIT, specifications, RFP, sections
    addenda_pdf    — Addenda, corrigenda, clarifications, amendments, pre-bid
    boq_xlsx       — XLSX with BOQ/price schedule sheets
    conditions_doc — DOC/DOCX conditions, letters, NIT documents
    unknown        — Anything else (logged for manual triage)

Heuristics are case-insensitive and use filename + parent directory +
(optionally) XLSX sheet names / column headers to classify.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ═════════════════════════════════════════════════════════════════════════
# KEYWORD BANKS
# ═════════════════════════════════════════════════════════════════════════

# BOQ / price schedule
BOQ_KEYWORDS = [
    "BOQ", "BILL OF QUANTITIES", "SCHEDULE OF QUANTITIES",
    "SOQ", "PRICE BID", "PRICE SCHEDULE", "RATE ANALYSIS",
    "BILL NO", "ABSTRACT OF COST", "ESTIMATE",
    "PBQ",  # Price BOQ (India)
]

# Conditions / specs / NIT
CONDITIONS_KEYWORDS = [
    "GCC", "SCC", "CONDITIONS", "NIT", "TENDER DOCUMENT",
    "TENDER NOTICE", "SPEC", "SPECIFICATION", "SCOPE OF WORK",
    "GENERAL CONDITIONS", "SPECIAL CONDITIONS", "CLAUSE",
    "INSTRUCTIONS TO BIDDERS", "ITB", "ELIGIBILITY",
    "INSTRUCTION", "NOTICE INVITING", "RFP",
    "DPR", "TENDER ACCEPTANCE", "SHE",
    "SECTION",  # Tender sections (Sec 1, Sec 2, etc.)
    "TENDERNOTICE", "TENDERDOCUMENT",  # Common India portal concatenations
]

# Addenda / corrigenda / clarifications
ADDENDA_KEYWORDS = [
    "ADDENDUM", "CORRIGENDUM", "CLARIFICATION",
    "AMENDMENT", "REVISED", "ERRATUM", "ERRATA",
    "EXTENSION OF DATE", "PRE-BID MINUTES",
    "AMDT", "AMMEND", "AMMENDENT",    # Common India misspellings
    "PREBID", "PRE BID", "PRE_BID",   # Pre-bid meeting minutes
    "MODIFIED",                         # Modified documents
]

# Drawing indicators (filenames + parent directories)
DRAWING_KEYWORDS = [
    "DRAWING", "DWG", "ARCH", "STR", "STRUCT",
    "ELECT", "ELEC", "MEP", "PLUMB", "HVAC", "FIRE",
    "CIVIL", "LANDSCAPE", "INTERIOR", "LAYOUT",
    "FLOOR PLAN", "ELEVATION", "SECTION", "DETAIL",
    "SITE PLAN", "ROOF PLAN", "FOUNDATION",
    "GAD", "SLD",  # General Arrangement Drawing, Single Line Diagram
]

# Sheet-like filenames (A-101, S-001, E-002 patterns)
_SHEET_RE = re.compile(
    r'(?:^|[_\-\s])([ASEMPCLFH])\s*[\-_]?\s*(\d{2,4})(?:[_\-\s.]|$)',
    re.IGNORECASE,
)

# XLSX BOQ sheet names
XLSX_BOQ_SHEET_KEYWORDS = ["BOQ", "SOQ", "PRICE", "BID", "RATE", "ESTIMATE", "BILL"]

# XLSX BOQ column headers (any of these suggests a pricing schedule)
XLSX_BOQ_COLUMN_KEYWORDS = ["QTY", "QUANTITY", "RATE", "AMOUNT", "UNIT", "TOTAL"]

# ═════════════════════════════════════════════════════════════════════════
# FILE CATEGORIES
# ═════════════════════════════════════════════════════════════════════════

CATEGORY_DRAWINGS_PDF = "drawings_pdf"
CATEGORY_BOQ_PDF = "boq_pdf"
CATEGORY_CONDITIONS_PDF = "conditions_pdf"
CATEGORY_ADDENDA_PDF = "addenda_pdf"
CATEGORY_BOQ_XLSX = "boq_xlsx"
CATEGORY_CONDITIONS_DOC = "conditions_doc"
CATEGORY_UNKNOWN = "unknown"

ALL_CATEGORIES = [
    CATEGORY_DRAWINGS_PDF,
    CATEGORY_BOQ_PDF,
    CATEGORY_CONDITIONS_PDF,
    CATEGORY_ADDENDA_PDF,
    CATEGORY_BOQ_XLSX,
    CATEGORY_CONDITIONS_DOC,
    CATEGORY_UNKNOWN,
]


# ═════════════════════════════════════════════════════════════════════════
# DATA MODEL
# ═════════════════════════════════════════════════════════════════════════

@dataclass
class ClassifiedFile:
    """A single file with its classification result."""
    path: Path
    category: str
    reason: str            # short explanation of why this category
    size_bytes: int = 0
    page_count: int = 0    # only for PDFs, 0 = unknown


@dataclass
class RoutingSummary:
    """Classification result for an entire tender delivery."""
    zip_name: str
    total_files: int = 0
    classified: List[ClassifiedFile] = field(default_factory=list)

    @property
    def by_category(self) -> Dict[str, List[ClassifiedFile]]:
        result: Dict[str, List[ClassifiedFile]] = {c: [] for c in ALL_CATEGORIES}
        for cf in self.classified:
            result.setdefault(cf.category, []).append(cf)
        return result

    @property
    def pdf_files(self) -> List[Path]:
        """PDF files with known categories (drawings + boq + conditions + addenda)."""
        cats = {CATEGORY_DRAWINGS_PDF, CATEGORY_BOQ_PDF,
                CATEGORY_CONDITIONS_PDF, CATEGORY_ADDENDA_PDF}
        return [cf.path for cf in self.classified if cf.category in cats]

    @property
    def all_pdf_files(self) -> List[Path]:
        """All PDF files regardless of classification (for full pipeline input)."""
        return [cf.path for cf in self.classified if cf.path.suffix.lower() == ".pdf"]

    def summary_table(self) -> str:
        """Human-readable routing summary."""
        lines = [f"{'Category':<20} {'Count':>5}  Files"]
        lines.append("-" * 60)
        by_cat = self.by_category
        for cat in ALL_CATEGORIES:
            files = by_cat.get(cat, [])
            if not files:
                continue
            names = ", ".join(f.path.name for f in files[:3])
            if len(files) > 3:
                names += f" +{len(files) - 3} more"
            lines.append(f"{cat:<20} {len(files):>5}  {names}")
        lines.append("-" * 60)
        lines.append(f"{'TOTAL':<20} {self.total_files:>5}")
        return "\n".join(lines)


# ═════════════════════════════════════════════════════════════════════════
# CLASSIFICATION ENGINE
# ═════════════════════════════════════════════════════════════════════════

def _normalize_name(name: str) -> str:
    """
    Normalize a filename/path component for keyword matching.

    Handles underscores, hyphens, and CamelCase:
      "NoticeInvitingTender" → "Notice Inviting Tender"
      "tech_prebid_67946"    → "tech prebid 67946"
      "Pre_Bid_Meeting"      → "Pre Bid Meeting"
    """
    # Replace underscores/hyphens with spaces
    s = name.replace("_", " ").replace("-", " ")
    # CamelCase split: insert space before uppercase preceded by lowercase
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
    return s


def _match_keywords(text: str, keywords: List[str]) -> Optional[str]:
    """Return the first matching keyword found in text (case-insensitive), or None."""
    text_upper = text.upper()
    for kw in keywords:
        if kw in text_upper:
            return kw
    return None


def _is_sheet_filename(name: str) -> bool:
    """Check if filename looks like a drawing sheet (A-101, S-002, etc.)."""
    return bool(_SHEET_RE.search(name))


def classify_file(path: Path, relative_root: Optional[Path] = None) -> ClassifiedFile:
    """
    Classify a single file based on name, extension, and parent directory.

    For XLSX files, also inspects sheet names if openpyxl is available.
    Names are normalized (underscores/CamelCase split) before keyword matching.
    Parent directory names (relative to root) are checked for drawing indicators.

    Args:
        path: File path to classify.
        relative_root: Root directory for computing relative parent paths.
            If provided, parent directory checks use paths relative to this root.
            If None, parent directory checks are skipped.
    """
    name = path.name
    stem = path.stem
    ext = path.suffix.lower()
    size = path.stat().st_size if path.exists() else 0

    # Normalized stem for better keyword matching
    norm_stem = _normalize_name(stem)

    # ── XLSX ─────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xls"):
        # Try to inspect sheet names
        xlsx_reason = _classify_xlsx(path)
        if xlsx_reason:
            return ClassifiedFile(path, CATEGORY_BOQ_XLSX, xlsx_reason, size)
        # Fallback: keyword match on filename
        kw = _match_keywords(norm_stem, BOQ_KEYWORDS + XLSX_BOQ_SHEET_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_BOQ_XLSX, f"filename keyword: {kw}", size)
        return ClassifiedFile(path, CATEGORY_UNKNOWN, "xlsx: no BOQ indicators", size)

    # ── DOC / DOCX ──────────────────────────────────────────────────
    if ext in (".doc", ".docx"):
        kw = _match_keywords(norm_stem, ADDENDA_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_CONDITIONS_DOC, f"addenda keyword: {kw}", size)
        kw = _match_keywords(norm_stem, CONDITIONS_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_CONDITIONS_DOC, f"conditions keyword: {kw}", size)
        kw = _match_keywords(norm_stem, BOQ_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_CONDITIONS_DOC, f"BOQ doc: {kw}", size)
        # Default DOC → conditions (most India tender DOCs are NIT/conditions)
        return ClassifiedFile(path, CATEGORY_CONDITIONS_DOC, "doc: default conditions", size)

    # ── PDF ──────────────────────────────────────────────────────────
    if ext == ".pdf":
        # Priority: addenda > boq > conditions > drawings > parent-dir drawings

        # Addenda
        kw = _match_keywords(norm_stem, ADDENDA_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_ADDENDA_PDF, f"addenda keyword: {kw}", size)

        # BOQ
        kw = _match_keywords(norm_stem, BOQ_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_BOQ_PDF, f"BOQ keyword: {kw}", size)

        # Conditions / spec
        kw = _match_keywords(norm_stem, CONDITIONS_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_CONDITIONS_PDF, f"conditions keyword: {kw}", size)

        # Drawings — keyword or sheet pattern in filename
        kw = _match_keywords(norm_stem, DRAWING_KEYWORDS)
        if kw:
            return ClassifiedFile(path, CATEGORY_DRAWINGS_PDF, f"drawing keyword: {kw}", size)
        if _is_sheet_filename(stem):
            return ClassifiedFile(path, CATEGORY_DRAWINGS_PDF, f"sheet pattern: {stem}", size)

        # Parent directory check — if parent dirs (relative to root) match drawing keywords
        if relative_root is not None:
            try:
                rel = path.relative_to(relative_root)
                parent_parts = rel.parts[:-1]  # directory components only
                if parent_parts:
                    parent_text = _normalize_name(" ".join(parent_parts))
                    kw = _match_keywords(parent_text, DRAWING_KEYWORDS)
                    if kw:
                        return ClassifiedFile(
                            path, CATEGORY_DRAWINGS_PDF,
                            f"parent directory keyword: {kw}", size,
                        )
            except ValueError:
                pass  # path is not relative to root

        # Large PDFs with no keyword match → likely drawings
        if size > 5_000_000:  # > 5 MB
            return ClassifiedFile(path, CATEGORY_DRAWINGS_PDF, "large PDF (>5MB): assumed drawings", size)

        # Default PDF → unknown
        return ClassifiedFile(path, CATEGORY_UNKNOWN, "pdf: no keyword match", size)

    # ── RAR — recognized but not processable ─────────────────────────
    if ext == ".rar":
        return ClassifiedFile(path, CATEGORY_UNKNOWN, "RAR archive: not extractable (use ZIP)", size)

    # ── Other extensions ─────────────────────────────────────────────
    return ClassifiedFile(path, CATEGORY_UNKNOWN, f"unrecognized extension: {ext}", size)


def _classify_xlsx(path: Path) -> Optional[str]:
    """Inspect XLSX sheet names and column headers. Returns reason or None."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            kw = _match_keywords(sheet_name, XLSX_BOQ_SHEET_KEYWORDS)
            if kw:
                wb.close()
                return f"sheet name '{sheet_name}' matches: {kw}"

        # Check first sheet's header row for BOQ-like columns
        ws = wb.active
        if ws:
            header_row = []
            for row in ws.iter_rows(min_row=1, max_row=3, max_col=20, values_only=True):
                header_row.extend(str(c or "").upper() for c in row)
            for kw in XLSX_BOQ_COLUMN_KEYWORDS:
                if any(kw in h for h in header_row):
                    wb.close()
                    return f"column header matches: {kw}"
        wb.close()
    except Exception:
        pass  # openpyxl not available or corrupted file
    return None


# ═════════════════════════════════════════════════════════════════════════
# BATCH CLASSIFICATION
# ═════════════════════════════════════════════════════════════════════════

# Skip hidden files and OS metadata
_SKIP_PATTERNS = {".DS_Store", "Thumbs.db", "__MACOSX", ".git"}
_SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".doc", ".docx", ".rar"}


def classify_directory(
    root: Path,
    zip_name: str = "",
) -> RoutingSummary:
    """
    Classify all supported files under a directory.

    Args:
        root: Directory containing extracted tender files.
        zip_name: Original ZIP filename (for audit trail).

    Returns:
        RoutingSummary with all classified files.
    """
    summary = RoutingSummary(zip_name=zip_name)

    if not root.exists() or not root.is_dir():
        return summary

    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        # Skip hidden / OS files
        if any(skip in path.parts for skip in _SKIP_PATTERNS):
            continue
        if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
            continue

        cf = classify_file(path, relative_root=root)
        summary.classified.append(cf)
        summary.total_files += 1

    return summary
