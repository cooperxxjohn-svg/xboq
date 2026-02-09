"""
Excel BOQ Generator - Generate priced BOQ in Excel format.
"""

from pathlib import Path
from typing import List, Dict


class ExcelBOQGenerator:
    """Generate priced BOQ in Excel format."""

    def generate(
        self,
        output_path: Path,
        priced_boq: List[Dict],
        prelims_items: List,
        project_info: Dict,
    ) -> Path:
        """Generate priced BOQ Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self._generate_csv_fallback(output_path, priced_boq, prelims_items, project_info)

        wb = openpyxl.Workbook()

        # Create main BOQ sheet
        ws_boq = wb.active
        ws_boq.title = "Priced BOQ"
        self._create_boq_sheet(ws_boq, priced_boq, project_info)

        # Create prelims sheet
        ws_prelims = wb.create_sheet("Preliminaries")
        self._create_prelims_sheet(ws_prelims, prelims_items)

        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, priced_boq, prelims_items, project_info)

        # Move summary to first position
        wb.move_sheet(ws_summary, offset=-2)

        # Save
        wb.save(output_path)
        return output_path

    def _create_boq_sheet(self, ws, priced_boq: List[Dict], project_info: Dict):
        """Create main BOQ sheet."""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            return

        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        currency_format = '₹#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = f"PRICED BILL OF QUANTITIES"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Project: {project_info.get('name', 'N/A')}"
        ws['A3'] = f"Location: {project_info.get('location', {}).get('city', 'N/A')}"

        # Headers
        headers = ["Sr. No.", "Item No.", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)", "Package"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Set column widths
        widths = [8, 12, 50, 8, 12, 15, 18, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        row = 6
        current_package = None
        package_start_row = row
        grand_total = 0

        # Group by package
        sorted_boq = sorted(priced_boq, key=lambda x: x.get("package", "zzz"))

        for i, item in enumerate(sorted_boq, 1):
            # Package header if changed
            pkg = item.get("package", "miscellaneous")
            if pkg != current_package:
                if current_package is not None:
                    # Insert subtotal for previous package
                    pass  # Could add subtotal rows here

                current_package = pkg
                # Package header row
                ws.cell(row=row, column=1, value="")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
                cell = ws.cell(row=row, column=2, value=pkg.replace("_", " ").upper())
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                row += 1

            # Item row
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=item.get("unified_item_no", item.get("item_no", ""))).border = thin_border
            ws.cell(row=row, column=3, value=item.get("description", "")[:100]).border = thin_border
            ws.cell(row=row, column=4, value=item.get("unit", "")).border = thin_border

            qty_cell = ws.cell(row=row, column=5, value=item.get("quantity", 0))
            qty_cell.border = thin_border
            qty_cell.number_format = '#,##0.00'

            rate_cell = ws.cell(row=row, column=6, value=item.get("rate", 0))
            rate_cell.border = thin_border
            rate_cell.number_format = currency_format

            amount = item.get("amount", 0)
            amount_cell = ws.cell(row=row, column=7, value=amount)
            amount_cell.border = thin_border
            amount_cell.number_format = currency_format
            grand_total += amount

            ws.cell(row=row, column=8, value=pkg).border = thin_border

            row += 1

        # Grand total row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        total_label = ws.cell(row=row, column=1, value="GRAND TOTAL (Works)")
        total_label.font = Font(bold=True, size=12)
        total_label.alignment = Alignment(horizontal='right')

        total_cell = ws.cell(row=row, column=7, value=grand_total)
        total_cell.font = Font(bold=True, size=12)
        total_cell.number_format = currency_format
        total_cell.border = thin_border

    def _create_prelims_sheet(self, ws, prelims_items: List):
        """Create preliminaries sheet."""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        currency_format = '₹#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = "PRELIMINARY COSTS"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        headers = ["Sr. No.", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)", "Category"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Set column widths
        widths = [8, 45, 10, 12, 15, 18, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data
        row = 4
        total = 0
        for i, item in enumerate(prelims_items, 1):
            # Handle both dataclass and dict
            if hasattr(item, 'description'):
                desc = item.description
                unit = item.unit
                qty = item.quantity
                rate = item.rate
                amount = item.amount
                category = item.category
            else:
                desc = item.get('description', '')
                unit = item.get('unit', '')
                qty = item.get('quantity', 0)
                rate = item.get('rate', 0)
                amount = item.get('amount', 0)
                category = item.get('category', '')

            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=desc).border = thin_border
            ws.cell(row=row, column=3, value=unit).border = thin_border
            ws.cell(row=row, column=4, value=qty).border = thin_border
            ws.cell(row=row, column=5, value=rate).border = thin_border
            ws.cell(row=row, column=5).number_format = currency_format
            ws.cell(row=row, column=6, value=amount).border = thin_border
            ws.cell(row=row, column=6).number_format = currency_format
            ws.cell(row=row, column=7, value=category).border = thin_border

            total += amount
            row += 1

        # Total
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="TOTAL PRELIMINARIES").font = Font(bold=True)
        ws.cell(row=row, column=6, value=total).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = currency_format

    def _create_summary_sheet(self, ws, priced_boq: List[Dict], prelims_items: List, project_info: Dict):
        """Create summary sheet."""
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            return

        bold_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14)
        currency_format = '₹#,##0.00'

        # Title
        ws['A1'] = "BID SUMMARY"
        ws['A1'].font = title_font

        ws['A3'] = "Project:"
        ws['B3'] = project_info.get('name', 'N/A')
        ws['A4'] = "Location:"
        ws['B4'] = f"{project_info.get('location', {}).get('city', '')}, {project_info.get('location', {}).get('state', '')}"

        # Calculate totals
        works_total = sum(item.get("amount", 0) for item in priced_boq)
        prelims_total = sum(item.amount if hasattr(item, 'amount') else item.get('amount', 0) for item in prelims_items)
        subtotal = works_total + prelims_total
        contingency = subtotal * 0.03
        margin = subtotal * 0.05
        total_before_gst = subtotal + contingency + margin
        gst = total_before_gst * 0.18
        grand_total = total_before_gst + gst

        # Summary table
        summary_data = [
            ("Works Total", works_total),
            ("Preliminaries", prelims_total),
            ("Sub-total", subtotal),
            ("Contingency (3%)", contingency),
            ("Margin (5%)", margin),
            ("Total before GST", total_before_gst),
            ("GST (18%)", gst),
            ("GRAND TOTAL", grand_total),
        ]

        row = 7
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = currency_format
            if label in ["Sub-total", "Total before GST", "GRAND TOTAL"]:
                ws.cell(row=row, column=1).font = bold_font
                cell.font = bold_font
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _generate_csv_fallback(
        self,
        output_path: Path,
        priced_boq: List[Dict],
        prelims_items: List,
        project_info: Dict,
    ) -> Path:
        """Generate CSV fallback if openpyxl not available."""
        import csv

        csv_path = output_path.with_suffix('.csv')

        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(["PRICED BILL OF QUANTITIES"])
            writer.writerow([f"Project: {project_info.get('name', 'N/A')}"])
            writer.writerow([])

            # BOQ headers
            writer.writerow(["Sr. No.", "Item No.", "Description", "Unit", "Quantity", "Rate", "Amount", "Package"])

            # BOQ items
            for i, item in enumerate(priced_boq, 1):
                writer.writerow([
                    i,
                    item.get("unified_item_no", ""),
                    item.get("description", ""),
                    item.get("unit", ""),
                    item.get("quantity", 0),
                    item.get("rate", 0),
                    item.get("amount", 0),
                    item.get("package", ""),
                ])

            # Totals
            writer.writerow([])
            writer.writerow(["", "", "", "", "", "Works Total:", sum(item.get("amount", 0) for item in priced_boq)])

        return csv_path
