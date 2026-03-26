"""Excel BOQ export — produces a 3-sheet workbook from a pipeline payload."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
FILL_HEADER = PatternFill("solid", fgColor="B8CCE4")
FILL_ROW_ALT = PatternFill("solid", fgColor="F2F2F2")
FILL_ROW_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_P1 = PatternFill("solid", fgColor="FCE4D6")
FILL_P2 = PatternFill("solid", fgColor="FFF2CC")
FILL_P3 = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER = Font(bold=True, name="Calibri", size=10)
FONT_BODY = Font(name="Calibri", size=10)
FONT_RED = Font(name="Calibri", size=10, bold=True, color="FF0000")
FONT_GREEN = Font(name="Calibri", size=10, bold=True, color="375623")
FONT_ORANGE = Font(name="Calibri", size=10, bold=True, color="C55A11")

THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

NUM_FMT = "#,##0.00"
MAX_COL_WIDTH = 60


def _auto_fit(ws) -> None:
    """Approximate column widths based on max content length (capped at MAX_COL_WIDTH)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, MAX_COL_WIDTH)


def _source_label(source: str) -> str:
    mapping = {"boq": "boq", "spec_item": "spec", "schedule_stub": "stub"}
    return mapping.get(source, source or "")


# ---------------------------------------------------------------------------
# Sheet 1 — Line Items
# ---------------------------------------------------------------------------
LINE_ITEMS_HEADERS = [
    "ID",
    "Item No",
    "Section",
    "Description",
    "Unit",
    "Qty",
    "Rate",
    "Amount",
    "Trade",
    "SOR Code",
    "Source",
    "Taxonomy Matched",
    "Qty Missing",
    "Rate Missing",
]

# Column indices (1-based) for special formatting
_COL_QTY = LINE_ITEMS_HEADERS.index("Qty") + 1       # 6
_COL_RATE = LINE_ITEMS_HEADERS.index("Rate") + 1      # 7
_COL_AMOUNT = LINE_ITEMS_HEADERS.index("Amount") + 1  # 8
_COL_TAX = LINE_ITEMS_HEADERS.index("Taxonomy Matched") + 1   # 12
_COL_QMI = LINE_ITEMS_HEADERS.index("Qty Missing") + 1        # 13
_COL_RMI = LINE_ITEMS_HEADERS.index("Rate Missing") + 1       # 14


def _build_line_items_sheet(ws, line_items: list[dict]) -> None:
    ws.title = "Line Items"

    # Header row
    for col_idx, header in enumerate(LINE_ITEMS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    for row_idx, item in enumerate(line_items, start=2):
        fill = FILL_ROW_ALT if row_idx % 2 == 0 else FILL_ROW_WHITE

        qty = item.get("qty")
        rate = item.get("rate")
        qty_missing = item.get("qty_missing", False)
        rate_missing = item.get("rate_missing", False)
        taxonomy_matched = item.get("taxonomy_matched", False)

        # Compute amount
        if qty is not None and rate is not None:
            amount = qty * rate
        else:
            amount = None

        row_data = [
            item.get("id", ""),
            item.get("item_no", ""),
            item.get("section", ""),
            item.get("description", ""),
            item.get("unit", ""),
            qty,
            rate,
            amount,
            item.get("trade", ""),
            item.get("sor_code", ""),
            _source_label(item.get("source", "")),
            None,  # Taxonomy Matched — handled below
            None,  # Qty Missing — handled below
            None,  # Rate Missing — handled below
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Number formats
            if col_idx in (_COL_QTY, _COL_RATE, _COL_AMOUNT):
                if value is not None:
                    cell.number_format = NUM_FMT

        # Taxonomy Matched
        tax_cell = ws.cell(row=row_idx, column=_COL_TAX)
        tax_cell.fill = fill
        tax_cell.border = THIN_BORDER
        tax_cell.alignment = Alignment(horizontal="center", vertical="center")
        if taxonomy_matched:
            tax_cell.value = "\u2713"
            tax_cell.font = FONT_GREEN
        else:
            tax_cell.value = "\u2717"
            tax_cell.font = FONT_ORANGE

        # Qty Missing
        qmi_cell = ws.cell(row=row_idx, column=_COL_QMI)
        qmi_cell.fill = fill
        qmi_cell.border = THIN_BORDER
        qmi_cell.alignment = Alignment(horizontal="center", vertical="center")
        if qty_missing:
            qmi_cell.value = "YES"
            qmi_cell.font = FONT_RED
        else:
            qmi_cell.value = ""
            qmi_cell.font = FONT_BODY

        # Rate Missing
        rmi_cell = ws.cell(row=row_idx, column=_COL_RMI)
        rmi_cell.fill = fill
        rmi_cell.border = THIN_BORDER
        rmi_cell.alignment = Alignment(horizontal="center", vertical="center")
        if rate_missing:
            rmi_cell.value = "YES"
            rmi_cell.font = FONT_RED
        else:
            rmi_cell.value = ""
            rmi_cell.font = FONT_BODY

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — BOQ Stats
# ---------------------------------------------------------------------------
def _build_boq_stats_sheet(ws, payload: dict) -> None:
    ws.title = "BOQ Stats"

    boq_stats = payload.get("boq_stats") or {}
    line_items_summary = payload.get("line_items_summary") or {}
    epc_mode = payload.get("epc_mode", False)
    epc_note = payload.get("epc_mode_note")
    by_trade = boq_stats.get("by_trade") or {}
    total_items = boq_stats.get("total_items") or 0

    row = 1

    # --- Trade breakdown ---
    ws.cell(row=row, column=1, value="Trade Breakdown").font = Font(bold=True, size=11)
    row += 1

    for col_idx, header in enumerate(["Trade", "Item Count", "% of Total"], start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row += 1

    for trade, count in sorted(by_trade.items()):
        pct = (count / total_items * 100) if total_items else 0.0
        for col_idx, value in enumerate([trade, count, f"{pct:.1f}%"], start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
        row += 1

    row += 1  # blank separator

    # --- Summary box ---
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=11)
    row += 1

    total_li = line_items_summary.get("total") or total_items or 0
    taxonomy_matched_count = line_items_summary.get("taxonomy_matched") or 0
    qty_missing_count = line_items_summary.get("qty_missing") or boq_stats.get("flagged_count") or 0
    rate_missing_count = line_items_summary.get("rate_missing") or 0
    tax_pct = (taxonomy_matched_count / total_li * 100) if total_li else 0.0

    summary_rows = [
        ("Total Items", total_li),
        ("Taxonomy Matched %", f"{tax_pct:.1f}%"),
        ("Qty Missing Count", qty_missing_count),
        ("Rate Missing Count", rate_missing_count),
    ]

    for label, value in summary_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True, name="Calibri", size=10)
        label_cell.border = THIN_BORDER
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = FONT_BODY
        value_cell.border = THIN_BORDER
        row += 1

    # --- EPC mode note ---
    if epc_mode:
        row += 1
        note_text = epc_note or "EPC mode active — contractor is responsible for full scope."
        note_cell = ws.cell(row=row, column=1, value=f"EPC Mode Note: {note_text}")
        note_cell.font = Font(bold=True, color="7030A0", name="Calibri", size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — RFIs
# ---------------------------------------------------------------------------
RFI_HEADERS = ["RFI ID", "Priority", "Trade", "Question", "Evidence"]


def _build_rfis_sheet(ws, rfis: list[dict]) -> None:
    ws.title = "RFIs"

    for col_idx, header in enumerate(RFI_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, rfi in enumerate(rfis, start=2):
        priority = (rfi.get("priority") or "P3").upper()
        if priority == "P1":
            fill = FILL_P1
        elif priority == "P2":
            fill = FILL_P2
        else:
            fill = FILL_P3

        row_data = [
            rfi.get("id", ""),
            priority,
            rfi.get("trade", ""),
            rfi.get("question", ""),
            rfi.get("evidence", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def export_boq_excel(payload: dict, output_path: Path) -> Path:
    """Export the pipeline payload to a 3-sheet Excel workbook.

    Args:
        payload: Pipeline output dict (see schema in module docstring).
        output_path: Destination file path (will be created/overwritten).

    Returns:
        output_path after writing.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet and add our three
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_li = wb.create_sheet("Line Items")
    ws_stats = wb.create_sheet("BOQ Stats")
    ws_rfis = wb.create_sheet("RFIs")

    line_items = payload.get("line_items") or []
    rfis = payload.get("rfis") or []

    _build_line_items_sheet(ws_li, line_items)
    _build_boq_stats_sheet(ws_stats, payload)
    _build_rfis_sheet(ws_rfis, rfis)

    wb.save(output_path)
    return output_path
