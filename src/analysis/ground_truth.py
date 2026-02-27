"""
Ground Truth Upload & Mapping — parse GT files and persist column mappings.

Storage:
    ~/.xboq/projects/<pid>/gt_mapping.json
    ~/.xboq/projects/<pid>/gt_data/<gt_type>.json

Pure module — no Streamlit dependency.
Sprint 20: Pilot Conversion + Paired Dataset Capture.
"""

import csv
import io
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# ── Canonical template schemas ──────────────────────────────────────────

GT_BOQ_COLUMNS = [
    "item_no", "description", "unit", "qty", "rate", "amount", "trade",
]

GT_SCHEDULES_DOORS_COLUMNS = [
    "mark", "size_w", "size_h", "type", "material", "qty", "remarks",
]

GT_QUANTITIES_COLUMNS = [
    "item", "unit", "qty", "source_type", "trade", "notes",
]


# ── Template generation ─────────────────────────────────────────────────

def generate_template_csv(template_name: str) -> str:
    """
    Generate a blank CSV template string with headers only.

    Args:
        template_name: One of "gt_boq", "gt_schedules_doors", "gt_quantities".

    Returns:
        CSV string with header row.
    """
    columns_map = {
        "gt_boq": GT_BOQ_COLUMNS,
        "gt_schedules_doors": GT_SCHEDULES_DOORS_COLUMNS,
        "gt_quantities": GT_QUANTITIES_COLUMNS,
    }
    cols = columns_map.get(template_name, [])
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(cols)
    return buf.getvalue()


# ── Excel parsing ────────────────────────────────────────────────────────

def parse_excel_sheets(file_bytes: bytes) -> List[str]:
    """
    Return list of sheet names from an xlsx file.

    Uses openpyxl (already a project dependency).
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_excel_sheet(
    file_bytes: bytes, sheet_name: str,
) -> Tuple[List[str], List[List[str]]]:
    """
    Read a specific sheet from xlsx. Returns (headers, rows).

    Args:
        file_bytes: Raw xlsx bytes.
        sheet_name: Name of the sheet to read.

    Returns:
        Tuple of (header_list, list_of_row_lists). All values are strings.
    """
    import pandas as pd
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, dtype=str)
    df = df.fillna("")
    headers = list(df.columns)
    rows = df.values.tolist()
    return headers, rows


# ── Column mapping ───────────────────────────────────────────────────────

def apply_column_mapping(
    headers: List[str],
    rows: List[List[str]],
    mapping: Dict[str, str],
) -> List[dict]:
    """
    Map raw Excel columns to canonical GT columns using a mapping dict.

    Args:
        headers: Raw column headers from the uploaded file.
        rows: Raw rows from the uploaded file.
        mapping: Dict of {canonical_col: raw_col_name}.

    Returns:
        List of dicts with canonical keys.
    """
    col_idx = {h: i for i, h in enumerate(headers)}
    result = []
    for row in rows:
        entry = {}
        for canonical, raw in mapping.items():
            idx = col_idx.get(raw)
            if idx is not None and idx < len(row):
                entry[canonical] = row[idx]
            else:
                entry[canonical] = ""
        result.append(entry)
    return result


# ── CSV parsing ──────────────────────────────────────────────────────────

def read_csv_file(file_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    """
    Read a CSV file. Returns (headers, rows).
    """
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_list = list(reader)
    if not rows_list:
        return [], []
    headers = rows_list[0]
    data_rows = rows_list[1:]
    return headers, data_rows


# ── Persistence ──────────────────────────────────────────────────────────

def save_gt_mapping(
    project_id: str,
    mapping: Dict[str, Any],
    projects_dir: Path,
) -> Path:
    """
    Save ground truth column mapping to project directory.

    Writes to: projects_dir/<pid>/gt_mapping.json

    The mapping dict should include:
        - "gt_type": "boq" | "schedules" | "quantities"
        - "source_file": original filename
        - "sheet_name": selected sheet name (or None for CSV)
        - "column_map": {canonical_col: raw_col_name}
    """
    proj_dir = Path(projects_dir) / project_id
    proj_dir.mkdir(parents=True, exist_ok=True)
    path = proj_dir / "gt_mapping.json"
    with open(path, "w") as f:
        json.dump(mapping, f, indent=2, default=str)
    return path


def load_gt_mapping(
    project_id: str,
    projects_dir: Path,
) -> Optional[Dict[str, Any]]:
    """Load GT mapping. Returns None if not found."""
    path = Path(projects_dir) / project_id / "gt_mapping.json"
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


def save_gt_data(
    project_id: str,
    gt_type: str,
    rows: List[dict],
    projects_dir: Path,
) -> Path:
    """
    Save parsed ground truth data as JSON.

    Writes to: projects_dir/<pid>/gt_data/<gt_type>.json
    """
    gt_dir = Path(projects_dir) / project_id / "gt_data"
    gt_dir.mkdir(parents=True, exist_ok=True)
    path = gt_dir / f"{gt_type}.json"
    with open(path, "w") as f:
        json.dump(rows, f, indent=2, default=str)
    return path


def load_gt_data(
    project_id: str,
    gt_type: str,
    projects_dir: Path,
) -> List[dict]:
    """Load parsed GT data. Returns [] if not found."""
    path = Path(projects_dir) / project_id / "gt_data" / f"{gt_type}.json"
    if not path.exists():
        return []
    with open(path) as f:
        return json.load(f)
