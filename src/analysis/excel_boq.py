"""
Excel BOQ Parser — Extract Bill of Quantities from .xlsx / .xls files.

Sprint 21C: India pilot tenders often include BOQ as Excel (Price Bid / SOQ).
This module detects BOQ sheets, maps columns, and returns structured boq_items
matching the same schema as PDF-based extract_boq.py output.

Sprint 25: Improved sheet detection for GeM/CPWD template format.
- Reads Schedule1/Schedule2/Schedule4 tabs (real items) not just BoQ1 (cover).
- max_header_row raised to 25 to skip GeM template preamble rows.
- Added DSR/SOR reference column and trade inference from schedule-letter prefix.
- Parent-item detection: letter + 3 digits with no qty = section header.

Pure module — no Streamlit dependency.
"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# =============================================================================
# COLUMN MAPPING — fuzzy header → canonical field
# =============================================================================

# Each canonical field has a list of header variants (case-insensitive match).
# Order matters: first match wins.
COLUMN_VARIANTS: Dict[str, List[str]] = {
    "item_no": [
        "item no", "item no.", "sl no", "sl no.", "sl.no", "sl.no.",
        "s.no", "s.no.", "s no", "sr no", "sr no.", "sr.no", "sr.no.",
        "serial no", "serial no.", "item", "sno", "sr", "sl",
    ],
    "description": [
        "description", "particulars", "name of item", "item description",
        "desc", "description of item", "description of work",
        "name of work", "specification", "details",
        "brief description", "item of work",
        # GeM / CPWD template variants
        "item particulars", "short description", "work description",
        "description of items", "work item",
    ],
    "unit": [
        "unit", "uom", "u/m", "unit of measurement", "units",
    ],
    "qty": [
        "qty", "quantity", "estimated qty", "est. qty", "est qty",
        "estimated quantity", "approx qty", "approx. qty",
        "tender qty", "bill qty",
        # GeM template variants
        "nos.", "nos", "quantity (nos)", "quantity in nos",
    ],
    "rate": [
        "rate", "unit rate", "rate (rs.)", "rate(rs)", "rate (rs)",
        "rate per unit", "quoted rate", "offered rate",
        "rate in rs", "rate in figures", "unit price",
        # GeM / CPWD variants — DSR/SOR rate column
        "basic unit rate", "basic rate", "dsr rate", "sor rate",
        "rate as per dsr", "rate as per sor", "basic unit rate as per dsr",
        "basic unit rate as per sor",
    ],
    "amount": [
        "amount", "total", "amount (rs.)", "amount(rs)", "amount (rs)",
        "total amount", "total cost", "total (rs.)", "value",
        # GeM template variants
        "total value", "total basic amount", "extended amount",
    ],
    # DSR / SOR reference column — common in CPWD / PWD BOQs
    "sor_code": [
        "dsr ref", "dsr reference", "sor ref", "sor reference",
        "reference of rate", "schedule of rate", "dsr item no",
        "sor item no", "dsr no", "sor no", "reference no",
        "rate reference", "dsr/sor", "dsr / sor",
        "specification reference", "spec ref",
    ],
}

# Sheet name keywords that indicate a BOQ sheet.
# "SCHEDULE" added for GeM/CPWD template tabs: Schedule1, Schedule2, Schedule4.
SHEET_NAME_KEYWORDS = [
    "BOQ", "SOQ", "PRICE", "BID", "RATE", "ESTIMATE", "BILL",
    "SCHEDULE OF QUANTITIES", "BILL OF QUANTITIES", "PRICE BID",
    "PRICE SCHEDULE", "ABSTRACT OF COST",
    # GeM portal template sheet names
    "SCHEDULE", "SCH", "ITEMS",
]

# Alpha parent-item pattern (letter + exactly 3 digits = section header in CPWD format)
_EXCEL_ALPHA_PARENT_RE = re.compile(r'^[A-Za-z]\d{3}$')

# Totals / summary row patterns (skip these rows)
_TOTALS_RE = re.compile(
    r'^\s*(?:sub[\-\s]?total|grand\s*total|'
    r'carried\s+over|brought\s+forward|c/?f|b/?f|'
    r'page\s+total|round\s*off|rounding|net\s+total)\s*$'
    r'|^\s*(?:total|abstract|summary|note)\s*$',
    re.IGNORECASE,
)

# Section / group header patterns (skip these rows).
# Three sub-patterns (any one sufficient):
#   1. Keyword prefix  – "Section A:", "Part 2 –", "Group C"
#   2. Numbered title  – "3.0 STRUCTURAL WORKS", "2) CIVIL WORKS"
#   3. ALL-CAPS title  – "FINISHING WORKS", "PLUMBING & SANITARY"
_SECTION_HEADER_RE = re.compile(
    r'^\s*(?:'
    r'(?:section|part|chapter|division|group|schedule)\s*[-:\s]'   # keyword prefix
    r'|(?:\d+[.)]\s+)[A-Z][A-Z\s/&,()–-]{6,}'                     # numbered heading
    r'|[A-Z][A-Z\s/&,()–-]{8,}$'                                   # ALL-CAPS line
    r')',
    re.IGNORECASE,
)


# =============================================================================
# HELPERS
# =============================================================================

def _cell_str(value: Any) -> str:
    """Convert a cell value to stripped string. None → ''."""
    if value is None:
        return ""
    return str(value).strip()


def _cell_float(value: Any) -> Optional[float]:
    """Try to parse a cell value as float, handling Indian number format."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (value != value):  # NaN check
            return None
        return float(value)
    s = _cell_str(value)
    if not s:
        return None
    # Strip currency symbols and whitespace
    s = re.sub(r'^[₹$Rs\.:\s]+', '', s).strip()
    if not s:
        return None
    # Use the existing Indian number parser logic: strip commas, parse
    cleaned = s.replace(",", "")
    # Remove trailing dash/hyphen (some BOQs use "-" for nil)
    if cleaned in ("-", "--", "nil", "NIL", "Nil", "N/A", "n/a", "NA", "na"):
        return None
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _match_column(header: str, canonical: str) -> bool:
    """Check if a header string matches any variant of a canonical field."""
    h = header.strip().lower()
    # Remove common suffixes/prefixes
    h = re.sub(r'[\(\)\[\]\{\}]', '', h).strip()
    h = re.sub(r'\s+', ' ', h)
    for variant in COLUMN_VARIANTS.get(canonical, []):
        if variant == h or variant in h:
            return True
    return False


def _is_totals_row(description: str) -> bool:
    """Check if a description looks like a totals/summary row."""
    return bool(_TOTALS_RE.match(description))


def _is_section_header(description: str) -> bool:
    """Check if a row is a section/group header."""
    return bool(_SECTION_HEADER_RE.match(description))


def _is_empty_data_row(row_data: Dict[str, Any]) -> bool:
    """Check if a data row has no meaningful content."""
    desc = _cell_str(row_data.get("description", ""))
    item_no = _cell_str(row_data.get("item_no", ""))
    qty = row_data.get("qty")
    rate = row_data.get("rate")
    amount = row_data.get("amount")
    # Row must have at least a description or item_no, plus at least one numeric
    has_text = bool(desc) or bool(item_no)
    has_number = qty is not None or rate is not None or amount is not None
    return not (has_text and has_number)


# =============================================================================
# XLSX READING (openpyxl)
# =============================================================================

def _read_xlsx_sheets(path: Path) -> Optional[List[Tuple[str, list]]]:
    """
    Read all sheets from an xlsx file using openpyxl.

    Returns list of (sheet_name, rows) where rows is list of list of cell values.
    Returns None if openpyxl is unavailable or file is unreadable.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available — cannot read .xlsx files")
        return None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"Cannot open {path.name}: {e}")
        return None

    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        try:
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
        except Exception as e:
            logger.warning(f"Error reading sheet '{sheet_name}' in {path.name}: {e}")
            continue
        result.append((sheet_name, rows))

    try:
        wb.close()
    except Exception as e:
        logger.warning(f"Failed to close workbook {path.name}: {e}")

    return result


def _read_xls_sheets(path: Path) -> Optional[List[Tuple[str, list]]]:
    """
    Read all sheets from an .xls file using pandas + xlrd fallback.

    Returns list of (sheet_name, rows) where rows is list of list of cell values.
    Returns None if pandas/xlrd are unavailable or file is unreadable.
    """
    try:
        import pandas as pd
    except ImportError:
        logger.warning(
            f"Cannot read .xls file {path.name}: pandas not available. "
            "Install pandas for .xls support."
        )
        return None

    try:
        xls = pd.ExcelFile(path, engine=None)  # let pandas choose engine
    except ImportError:
        logger.warning(
            f"Cannot read .xls file {path.name}: install xlrd for .xls support "
            "(pip install xlrd)."
        )
        return None
    except Exception as e:
        logger.warning(f"Cannot open .xls file {path.name}: {e}")
        return None

    result = []
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            rows = df.values.tolist()
            result.append((sheet_name, rows))
        except Exception as e:
            logger.warning(f"Error reading sheet '{sheet_name}' in {path.name}: {e}")
            continue

    try:
        xls.close()
    except Exception as e:
        logger.warning(f"Failed to close workbook {path.name}: {e}")

    return result


def _read_workbook(path: Path) -> Optional[List[Tuple[str, list]]]:
    """Read workbook sheets from xlsx or xls."""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx_sheets(path)
    elif ext == ".xls":
        # Try openpyxl first (won't work for .xls), then pandas fallback
        result = _read_xls_sheets(path)
        return result
    else:
        logger.warning(f"Unsupported Excel format: {ext}")
        return None


# =============================================================================
# SHEET + COLUMN DETECTION
# =============================================================================

def _detect_header_row(
    rows: list,
    max_header_row: int = 25,
) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Find the header row and build a column mapping.

    Scans rows 0..max_header_row for the row with the most canonical field matches.
    Requires at least 3 matches to qualify.

    Returns (header_row_index, {canonical_field: column_index}) or None.
    """
    best_row = None
    best_map: Dict[str, int] = {}
    best_score = 0

    for row_idx in range(min(max_header_row, len(rows))):
        row = rows[row_idx]
        col_map: Dict[str, int] = {}
        used_cols: set = set()

        for canonical in COLUMN_VARIANTS:
            for col_idx, cell_val in enumerate(row):
                if col_idx in used_cols:
                    continue
                cell_text = _cell_str(cell_val)
                if not cell_text:
                    continue
                if _match_column(cell_text, canonical):
                    col_map[canonical] = col_idx
                    used_cols.add(col_idx)
                    break

        score = len(col_map)
        if score > best_score:
            best_score = score
            best_map = col_map
            best_row = row_idx

    if best_score >= 3 and best_row is not None:
        return best_row, best_map
    return None


def detect_boq_sheets(path: Path) -> List[Dict[str, Any]]:
    """
    Detect BOQ sheets in an Excel workbook.

    Returns list of dicts:
        {sheet_name, header_row, column_map, score}
    where score indicates detection confidence.
    """
    sheets_data = _read_workbook(path)
    if not sheets_data:
        return []

    detected = []
    for sheet_name, rows in sheets_data:
        if not rows:
            continue

        # Score 1: sheet name keyword match
        name_score = 0
        sheet_upper = sheet_name.upper()
        for kw in SHEET_NAME_KEYWORDS:
            if kw in sheet_upper:
                name_score = 1
                break

        # Score 2: header row detection
        header_result = _detect_header_row(rows)
        if header_result is None:
            continue  # No usable header → skip this sheet

        header_row, column_map = header_result
        col_score = len(column_map)

        # Must have at least description + one numeric column
        has_desc = "description" in column_map
        has_numeric = any(f in column_map for f in ("qty", "rate", "amount"))
        if not (has_desc and has_numeric):
            continue

        total_score = name_score + col_score

        detected.append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "column_map": column_map,
            "score": total_score,
            "_rows": rows,  # carry rows for parsing (avoid re-reading file)
        })

    # Sort by score descending (best sheets first)
    detected.sort(key=lambda d: d["score"], reverse=True)
    return detected


# =============================================================================
# SHEET PARSING
# =============================================================================

def parse_boq_sheet(
    rows: list,
    header_row: int,
    column_map: Dict[str, int],
    source_file: str = "",
    source_sheet: str = "",
) -> Tuple[List[dict], int]:
    """
    Parse BOQ items from a single sheet's rows.

    Args:
        rows: All rows from the sheet (list of list of cell values).
        header_row: 0-indexed row where headers are.
        column_map: {canonical_field: column_index} mapping.
        source_file: Filename for traceability.
        source_sheet: Sheet name for traceability.

    Returns:
        (items, skipped_count)

    Sprint 25 additions:
    - Extracts sor_code from DSR/SOR reference column.
    - Infers trade from schedule-letter item_no prefix (A→civil, R→electrical, etc.).
    - Parent-item detection: alpha item_no (A101) with no qty = section header.
    """
    # Inline import to avoid circular dependency
    from .extractors.extract_boq import infer_boq_trade

    items = []
    skipped = 0
    data_start = header_row + 1
    current_section: str = ""   # tracks the most recent section header text

    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]

        # Extract raw values by column map
        raw: Dict[str, Any] = {}
        for field, col_idx in column_map.items():
            if col_idx < len(row):
                raw[field] = row[col_idx]
            else:
                raw[field] = None

        # Parse description
        description = _cell_str(raw.get("description", ""))

        # Skip empty rows
        if not description and not _cell_str(raw.get("item_no", "")):
            skipped += 1
            continue

        # Skip totals/summary rows
        if description and _is_totals_row(description):
            skipped += 1
            continue

        # Parse item number early (needed for parent-item detection)
        item_no = _cell_str(raw.get("item_no", ""))

        # Skip section headers (no numeric data); record text for child items.
        # Two cases:
        #   A. Description text looks like a section header (ALL-CAPS / keyword prefix)
        #   B. Alpha item_no like A101 (letter + exactly 3 digits) with no qty/rate
        qty_val = _cell_float(raw.get("qty"))
        rate_val = _cell_float(raw.get("rate"))
        amount_val = _cell_float(raw.get("amount"))

        is_parent_item = (
            item_no and _EXCEL_ALPHA_PARENT_RE.match(item_no)
            and qty_val is None and rate_val is None and amount_val is None
        )
        if is_parent_item:
            hdr = description or item_no
            current_section = f"{item_no} — {hdr[:70]}" if description else item_no
            skipped += 1
            continue

        if description and _is_section_header(description):
            if qty_val is None and rate_val is None and amount_val is None:
                current_section = description[:80]
                skipped += 1
                continue

        # Parse numeric fields
        qty = qty_val
        rate = rate_val
        amount = amount_val

        # Parse unit
        unit = _cell_str(raw.get("unit", "")) or None

        # Parse DSR/SOR reference (sor_code)
        raw_sor = _cell_str(raw.get("sor_code", ""))
        sor_code: Optional[str] = (
            raw_sor if raw_sor and raw_sor not in ("-", "–", "—", "N/A", "n/a") else None
        )

        # Skip rows with no text AND no numbers
        if not description and not item_no:
            skipped += 1
            continue
        if qty is None and rate is None and amount is None and not description:
            skipped += 1
            continue

        # Infer rate from amount/qty if missing
        if rate is None and qty and amount and qty > 0:
            rate = round(amount / qty, 2)

        # Trade inference — schedule prefix takes priority over keywords
        trade = infer_boq_trade(description, item_no)

        # Build boq_item matching the canonical schema
        item: Dict[str, Any] = {
            "item_no": item_no,
            "description": description,
            "unit": unit,
            "qty": qty,
            "rate": rate,
            "sor_code": sor_code,
            "trade": trade,
            "source_page": 0,          # Marker: not from a PDF page
            "confidence": 0.85,        # Structured Excel data = higher confidence
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_row": row_idx + 1,  # 1-indexed for human readability
            "section": current_section, # BOQ section header above this item (may be "")
        }

        items.append(item)

    return items, skipped


# =============================================================================
# TOP-LEVEL API
# =============================================================================

def parse_boq_excels(
    paths: List[Path],
) -> Tuple[List[dict], Dict[str, Any]]:
    """
    Parse BOQ items from one or more Excel files.

    Args:
        paths: List of Excel file paths (.xlsx or .xls).

    Returns:
        (combined_items, stats) where stats = {
            files_parsed, sheets_parsed, total_rows,
            skipped_rows, errors, files_skipped
        }
    """
    combined_items: List[dict] = []
    stats: Dict[str, Any] = {
        "files_parsed": 0,
        "sheets_parsed": 0,
        "total_rows": 0,
        "skipped_rows": 0,
        "errors": [],
        "files_skipped": 0,
    }

    for path in paths:
        path = Path(path)
        if not path.exists():
            stats["errors"].append(f"File not found: {path}")
            stats["files_skipped"] += 1
            continue

        try:
            detected_sheets = detect_boq_sheets(path)
        except Exception as e:
            err_msg = f"Error detecting sheets in {path.name}: {type(e).__name__}: {e}"
            logger.warning(err_msg)
            stats["errors"].append(err_msg)
            stats["files_skipped"] += 1
            continue

        if not detected_sheets:
            logger.info(f"No BOQ sheets detected in {path.name}")
            stats["files_skipped"] += 1
            continue

        stats["files_parsed"] += 1

        for sheet_info in detected_sheets:
            sheet_name = sheet_info["sheet_name"]
            header_row = sheet_info["header_row"]
            column_map = sheet_info["column_map"]
            sheet_rows = sheet_info.get("_rows", [])

            try:
                items, skipped = parse_boq_sheet(
                    rows=sheet_rows,
                    header_row=header_row,
                    column_map=column_map,
                    source_file=path.name,
                    source_sheet=sheet_name,
                )
                combined_items.extend(items)
                stats["sheets_parsed"] += 1
                stats["total_rows"] += len(items)
                stats["skipped_rows"] += skipped
                logger.info(
                    f"  {path.name} / {sheet_name}: "
                    f"{len(items)} items extracted, {skipped} rows skipped"
                )
            except Exception as e:
                err_msg = (
                    f"Error parsing sheet '{sheet_name}' in {path.name}: "
                    f"{type(e).__name__}: {e}"
                )
                logger.warning(err_msg)
                stats["errors"].append(err_msg)

    return combined_items, stats
