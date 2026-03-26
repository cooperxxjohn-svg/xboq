"""
Excel BOQ Exporter — xBOQ

Generates a professional, formatted Excel workbook from the full QTO payload.

Workbook structure:
  Sheet 0: Summary (project info + trade-wise cost summary)
  Sheet 1..N: One sheet per trade (trade items with rate + amount)

Each trade sheet columns:
  # | Description | Qty | Unit | Rate (INR) | Amount (INR) | Spec | Source

Requires: openpyxl >= 3.1
"""

from __future__ import annotations

import io
import logging
import math
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# =============================================================================
# CONSTANTS — colours / styles
# =============================================================================

_BRAND_BLUE   = "1B4F72"   # xBOQ header blue
_BRAND_LIGHT  = "D6EAF8"   # light blue fill for alternating rows
_HEADER_BG    = "1B4F72"
_HEADER_FG    = "FFFFFF"
_TRADE_BG     = "2E86C1"   # per-trade header row
_TRADE_FG     = "FFFFFF"
_SUBTOT_BG    = "D5F5E3"   # subtotal green
_GRANDTOT_BG  = "1B4F72"
_GRANDTOT_FG  = "FFFFFF"
_ALT_ROW      = "EBF5FB"   # alternate row fill

# Trade display order
_TRADE_ORDER = [
    "Civil / Site Work",
    "Structural",
    "Masonry",
    "Doors & Windows",
    "Finishes",
    "Waterproofing",
    "Painting",
    "Electrical",
    "Plumbing",
    "HVAC",
    "MEP",
    "External Works",
    "Miscellaneous",
]

_COLUMN_WIDTHS = [5, 55, 12, 10, 16, 18, 22, 14]
_COLUMN_HEADERS = ["#", "Description", "Qty", "Unit", "Rate (INR)", "Amount (INR)", "Spec / Ref", "Source"]


# =============================================================================
# HELPERS
# =============================================================================

def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_font(cell, bold=False, size=10, color="000000"):
    cell.font = Font(name="Calibri", bold=bold, size=size, color=color)


def _fill(cell, hex_color: str):
    cell.fill = PatternFill("solid", fgColor=hex_color)


def _fmt_inr(val: float) -> str:
    """Format as Indian numbering: 1,23,456.00"""
    if val == 0:
        return "—"
    s = f"{val:,.2f}"
    return s


def _normalise_trade(trade: str) -> str:
    """Map varied trade names to canonical display names."""
    t = trade.strip().title()
    mapping = {
        "Civil": "Civil / Site Work",
        "Civil / Site Work": "Civil / Site Work",
        "Site Work": "Civil / Site Work",
        "Sitework": "Civil / Site Work",
        "Earthwork": "Civil / Site Work",
        "Structural": "Structural",
        "Masonry": "Masonry",
        "Brickwork": "Masonry",
        "Finishes": "Finishes",
        "Finish": "Finishes",
        "Painting": "Painting",
        "Paint": "Painting",
        "Waterproofing": "Waterproofing",
        "Doors & Windows": "Doors & Windows",
        "Doors And Windows": "Doors & Windows",
        "Mep": "MEP",
        "Electrical": "Electrical",
        "Plumbing": "Plumbing",
        "Hvac": "HVAC",
        "External Works": "External Works",
    }
    return mapping.get(t, t)


def _sort_trades(trades: List[str]) -> List[str]:
    ordered = [t for t in _TRADE_ORDER if t in trades]
    remaining = [t for t in trades if t not in ordered]
    return ordered + sorted(remaining)


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def _apply_header_style(ws, row_num: int, num_cols: int, bg: str, fg: str,
                        bold=True, size=10):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        _fill(cell, bg)
        _set_font(cell, bold=bold, size=size, color=fg)
        cell.border = _thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_trade_sheet(wb: Workbook, trade: str, items: List[dict],
                       project_name: str):
    """Write one sheet for a trade with all its items."""
    # Sanitise sheet name (Excel: max 31 chars, no special chars)
    sheet_name = trade[:31].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "")
    ws = wb.create_sheet(title=sheet_name)

    num_cols = len(_COLUMN_HEADERS)

    # Set column widths
    for i, w in enumerate(_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Trade title ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{trade.upper()}  —  BILL OF QUANTITIES")
    _fill(title_cell, _TRADE_BG)
    _set_font(title_cell, bold=True, size=13, color=_TRADE_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Project name ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    proj_cell = ws.cell(row=2, column=1, value=f"Project: {project_name}")
    _fill(proj_cell, _BRAND_LIGHT)
    _set_font(proj_cell, bold=False, size=10, color="1B4F72")
    proj_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ─────────────────────────────────────────────
    for c, h in enumerate(_COLUMN_HEADERS, 1):
        cell = ws.cell(row=3, column=c, value=h)
    _apply_header_style(ws, 3, num_cols, _HEADER_BG, _HEADER_FG, bold=True, size=10)
    ws.row_dimensions[3].height = 22

    # Freeze top rows
    ws.freeze_panes = "A4"

    # ── Data rows ─────────────────────────────────────────────────────────
    row_num = 4
    trade_total = 0.0

    for idx, item in enumerate(items, 1):
        desc   = item.get("description", "")
        qty    = item.get("qty", 0) or 0
        unit   = item.get("unit", "")
        rate   = item.get("rate_inr", 0) or 0
        amount = item.get("amount_inr", 0) or 0
        spec   = item.get("spec", "") or item.get("spec_ref", "") or ""
        source = item.get("source", "")

        if amount == 0 and rate > 0 and qty > 0:
            amount = rate * qty

        trade_total += amount

        values = [idx, desc, round(qty, 3), unit,
                  rate if rate else None,
                  amount if amount else None,
                  spec, source]

        alt = (idx % 2 == 0)
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            if alt:
                _fill(cell, _ALT_ROW)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            _set_font(cell, bold=False, size=9)

            # Right-align numeric columns
            if c in (3, 5, 6):
                cell.alignment = Alignment(horizontal="right", vertical="top")
                if isinstance(val, (int, float)) and val is not None:
                    cell.number_format = '#,##0.00'

        ws.row_dimensions[row_num].height = 15 if len(desc) < 80 else 28
        row_num += 1

    # ── Subtotal row ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    sub_cell = ws.cell(row=row_num, column=1, value=f"TOTAL — {trade.upper()}")
    _fill(sub_cell, _SUBTOT_BG)
    _set_font(sub_cell, bold=True, size=10)
    sub_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=row_num, column=6, value=trade_total if trade_total else None)
    _fill(total_cell, _SUBTOT_BG)
    _set_font(total_cell, bold=True, size=10)
    total_cell.alignment = Alignment(horizontal="right")
    if trade_total:
        total_cell.number_format = '#,##0.00'
    for c in range(1, num_cols + 1):
        ws.cell(row=row_num, column=c).border = _thin_border()
    ws.row_dimensions[row_num].height = 20

    return trade_total


def _write_summary_sheet(wb: Workbook, trade_totals: Dict[str, float],
                         project_name: str, project_meta: dict,
                         all_items: List[dict]):
    """Write the first Summary sheet."""
    ws = wb.active
    ws.title = "Summary"

    num_cols = 6
    for i, w in enumerate([5, 40, 15, 15, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    t = ws.cell(row=1, column=1, value="BILL OF QUANTITIES — PROJECT COST SUMMARY")
    _fill(t, _HEADER_BG)
    _set_font(t, bold=True, size=14, color=_HEADER_FG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Project info ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    p = ws.cell(row=2, column=1, value=f"Project: {project_name}")
    _fill(p, _BRAND_LIGHT)
    _set_font(p, bold=True, size=11, color="1B4F72")
    p.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # Project meta (area, floors, date)
    meta_row = 3
    meta_items = [
        ("Total Floor Area", f"{project_meta.get('total_area_sqm', 0):.0f} sqm"),
        ("Floors", str(project_meta.get('floors', '—'))),
        ("Generated", datetime.now().strftime("%d %b %Y")),
        ("Source", "xBOQ AI Takeoff"),
    ]
    ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=num_cols)
    meta_str = "   |   ".join(f"{k}: {v}" for k, v in meta_items if v)
    mc = ws.cell(row=meta_row, column=1, value=meta_str)
    _fill(mc, "F8F9FA")
    _set_font(mc, bold=False, size=9, color="555555")
    mc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[meta_row].height = 16

    # ── Column headers ─────────────────────────────────────────────────────
    hdr_row = 5
    for c, h in enumerate(["#", "Trade / Division", "Item Count", "% of Total", "Amount (INR)", "Remarks"], 1):
        cell = ws.cell(row=hdr_row, column=c, value=h)
    _apply_header_style(ws, hdr_row, num_cols, _TRADE_BG, _TRADE_FG, bold=True, size=10)
    ws.row_dimensions[hdr_row].height = 20
    ws.freeze_panes = "A6"

    grand_total = sum(trade_totals.values())

    # ── Trade rows ─────────────────────────────────────────────────────────
    row_num = hdr_row + 1
    ordered_trades = _sort_trades(list(trade_totals.keys()))

    # Count items per trade
    items_by_trade: Dict[str, int] = {}
    for item in all_items:
        tr = _normalise_trade(item.get("trade", "Miscellaneous"))
        items_by_trade[tr] = items_by_trade.get(tr, 0) + 1

    for idx, trade in enumerate(ordered_trades, 1):
        amount = trade_totals.get(trade, 0)
        pct = (amount / grand_total * 100) if grand_total > 0 else 0
        count = items_by_trade.get(trade, 0)

        alt = (idx % 2 == 0)
        row_vals = [idx, trade, count,
                    f"{pct:.1f}%",
                    amount if amount else None,
                    "See trade sheet"]

        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            if alt:
                _fill(cell, _ALT_ROW)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")
            _set_font(cell, bold=False, size=10)
            if c == 5 and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # ── Grand total ────────────────────────────────────────────────────────
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    gt_label = ws.cell(row=row_num, column=1, value="GRAND TOTAL (ESTIMATED)")
    _fill(gt_label, _GRANDTOT_BG)
    _set_font(gt_label, bold=True, size=11, color=_GRANDTOT_FG)
    gt_label.alignment = Alignment(horizontal="right", vertical="center")

    gt_val = ws.cell(row=row_num, column=5, value=grand_total if grand_total else None)
    _fill(gt_val, _GRANDTOT_BG)
    _set_font(gt_val, bold=True, size=11, color=_GRANDTOT_FG)
    gt_val.alignment = Alignment(horizontal="right", vertical="center")
    if grand_total:
        gt_val.number_format = '#,##0.00'

    for c in range(1, num_cols + 1):
        ws.cell(row=row_num, column=c).border = _thin_border()
    ws.row_dimensions[row_num].height = 25
    row_num += 1

    # Per sqm cost
    area = project_meta.get("total_area_sqm", 0)
    if area and grand_total:
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        psm_label = ws.cell(row=row_num, column=1, value="Cost per sqm (BUA)")
        _fill(psm_label, _BRAND_LIGHT)
        _set_font(psm_label, bold=True, size=10, color="1B4F72")
        psm_label.alignment = Alignment(horizontal="right", vertical="center")
        psm_val = ws.cell(row=row_num, column=5, value=round(grand_total / area, 2))
        _fill(psm_val, _BRAND_LIGHT)
        _set_font(psm_val, bold=True, size=10, color="1B4F72")
        psm_val.alignment = Alignment(horizontal="right")
        psm_val.number_format = '#,##0.00'
        for c in range(1, num_cols + 1):
            ws.cell(row=row_num, column=c).border = _thin_border()
        ws.row_dimensions[row_num].height = 20

    # ── Disclaimer ────────────────────────────────────────────────────────
    row_num += 2
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    disc = ws.cell(
        row=row_num, column=1,
        value=(
            "DISCLAIMER: This estimate is generated by xBOQ AI from submitted drawings and specifications. "
            "Rates are indicative market rates for Tier-1 Indian cities (Q1 2025). "
            "Actual costs may vary by ±20–30%. This estimate does not replace a detailed quantity survey."
        )
    )
    _set_font(disc, bold=False, size=8, color="888888")
    disc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row_num].height = 35


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def export_to_excel(
    all_items: List[dict],
    project_name: str = "Untitled Project",
    project_meta: Optional[dict] = None,
) -> Optional[bytes]:
    """
    Generate an Excel workbook from all BOQ items.

    Args:
        all_items:     Combined list of all BOQ items from all QTO modules.
                       Each item: {description, qty, unit, trade, rate_inr, amount_inr, spec, source}
        project_name:  Display name for the project.
        project_meta:  Dict with optional keys: total_area_sqm, floors, project_type, location.

    Returns:
        Excel file as bytes (io.BytesIO content), or None if openpyxl not installed.
    """
    if not _HAS_OPENPYXL:
        logger.warning("openpyxl not installed — Excel export unavailable")
        return None

    if not all_items:
        logger.warning("No items to export")
        return None

    meta = project_meta or {}

    # Group items by trade
    by_trade: Dict[str, List[dict]] = {}
    for item in all_items:
        trade_raw = item.get("trade", "") or "Miscellaneous"
        trade = _normalise_trade(trade_raw)
        by_trade.setdefault(trade, []).append(item)

    wb = Workbook()

    # Compute trade totals (for summary sheet)
    trade_totals: Dict[str, float] = {}
    ordered_trades = _sort_trades(list(by_trade.keys()))

    for trade in ordered_trades:
        items = by_trade[trade]
        total = _write_trade_sheet(wb, trade, items, project_name)
        trade_totals[trade] = total

    # Summary sheet (written last but positioned first)
    _write_summary_sheet(wb, trade_totals, project_name, meta, all_items)

    # Move summary to position 0
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# =============================================================================
# PIPELINE HELPER
# =============================================================================

def export_from_payload(payload: dict) -> Optional[bytes]:
    """
    Convenience wrapper: extract all BOQ items from the full pipeline payload
    and export to Excel.

    Collects items from:
      - extraction_result.boq_items (raw extracted BOQ)
      - qto_summary embedded items (structural, finish, MEP, painting, etc.)
      - spec_items (all generated QTO items)
    """
    project_name = (
        payload.get("project_name")
        or payload.get("filename")
        or "Untitled Project"
    )

    qto = payload.get("qto_summary", {})
    meta = {
        "total_area_sqm": qto.get("vmeas_area_sqm") or qto.get("visual_area_sqm") or 0,
        "floors": payload.get("floor_count", 0),
        "location": payload.get("location", ""),
    }

    # Collect all spec items
    all_items: List[dict] = []

    spec_items = payload.get("spec_items", [])
    if spec_items:
        all_items.extend(spec_items)

    # If no spec_items, try raw BOQ items
    if not all_items:
        raw_boq = payload.get("boq_items", [])
        for item in raw_boq:
            all_items.append({
                "description": item.get("description", ""),
                "qty": item.get("qty", 0) or item.get("quantity", 0),
                "unit": item.get("unit", ""),
                "trade": item.get("trade", item.get("section", "Miscellaneous")),
                "spec": item.get("spec", ""),
                "source": "boq_extract",
                "rate_inr": item.get("rate_inr", 0) or item.get("rate", 0),
                "amount_inr": item.get("amount_inr", 0) or item.get("amount", 0),
            })

    if not all_items:
        logger.warning("export_from_payload: no items found in payload")
        return None

    return export_to_excel(all_items, project_name=project_name, project_meta=meta)
