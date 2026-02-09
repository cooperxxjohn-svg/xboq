"""
MEP Takeoff Generator

Generates bid-ready MEP takeoff with:
- Device schedules by system/room
- Connectivity summaries
- RFI lists for unknown specs
- CSV/Excel exports
"""

import csv
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import logging

from .device_detector import DetectedDevice
from .connectivity import Connection
from .systems import MEPSystem

logger = logging.getLogger(__name__)


@dataclass
class TakeoffLine:
    """A single line item in the MEP takeoff."""
    line_no: int
    system: str
    subsystem: str
    device_type: str
    description: str
    tag: str
    location: str
    qty: int
    unit: str

    # Spec details
    spec: str
    spec_notes: str = ""

    # Detection info
    detection_method: str = "inferred"  # measured, schedule, inferred
    confidence: float = 0.0

    # Provenance
    is_measured: bool = False
    source: str = ""

    # RFI
    rfi_required: bool = False
    rfi_fields: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "line_no": self.line_no,
            "system": self.system,
            "subsystem": self.subsystem,
            "device_type": self.device_type,
            "description": self.description,
            "tag": self.tag,
            "location": self.location,
            "qty": self.qty,
            "unit": self.unit,
            "spec": self.spec,
            "spec_notes": self.spec_notes,
            "detection_method": self.detection_method,
            "confidence": self.confidence,
            "is_measured": self.is_measured,
            "source": self.source,
            "rfi_required": self.rfi_required,
            "rfi_fields": self.rfi_fields,
        }


class MEPTakeoff:
    """
    MEP Takeoff generator.
    """

    def __init__(
        self,
        devices: List[DetectedDevice],
        connections: List[Connection] = None,
        systems: Dict[str, MEPSystem] = None,
        project_id: str = "",
    ):
        self.devices = devices
        self.connections = connections or []
        self.systems = systems or {}
        self.project_id = project_id

        self.takeoff_lines: List[TakeoffLine] = []
        self.connectivity_summary: Dict[str, Any] = {}
        self.rfi_list: List[Dict[str, Any]] = []

    def generate(self) -> Dict[str, Any]:
        """
        Generate complete MEP takeoff.

        Returns:
            Takeoff result with lines, summary, and RFIs
        """
        # Generate takeoff lines from devices
        self._generate_takeoff_lines()

        # Generate connectivity summary
        self._generate_connectivity_summary()

        # Generate RFI list
        self._generate_rfi_list()

        return {
            "project_id": self.project_id,
            "generated_at": datetime.now().isoformat(),
            "total_lines": len(self.takeoff_lines),
            "total_devices": sum(line.qty for line in self.takeoff_lines),
            "measured_count": sum(1 for line in self.takeoff_lines if line.is_measured),
            "inferred_count": sum(1 for line in self.takeoff_lines if not line.is_measured),
            "rfi_count": len(self.rfi_list),
            "takeoff_lines": [line.to_dict() for line in self.takeoff_lines],
            "connectivity": self.connectivity_summary,
            "rfis": self.rfi_list,
        }

    def _generate_takeoff_lines(self):
        """Generate takeoff lines from devices."""
        # Group devices by type and location for aggregation
        groups: Dict[str, List[DetectedDevice]] = {}

        for device in self.devices:
            # Create group key
            key = f"{device.device_type}|{device.room_name or 'Unassigned'}|{self._spec_key(device.spec)}"

            if key not in groups:
                groups[key] = []
            groups[key].append(device)

        # Create takeoff lines from groups
        line_no = 1
        for key, devices in groups.items():
            device = devices[0]  # Use first device as template
            qty = len(devices)

            # Format spec string
            spec_str = self._format_spec(device.spec)

            # Check if measured
            is_measured = device.detection_method in ["text", "symbol", "schedule"]
            detection = device.detection_method
            if detection == "text" or detection == "symbol":
                detection = "plan"

            line = TakeoffLine(
                line_no=line_no,
                system=device.system,
                subsystem=device.subsystem,
                device_type=device.device_type,
                description=device.device_type.replace("_", " ").title(),
                tag=device.tag if qty == 1 else f"{device.tag[:2]}*" if device.tag else "",
                location=device.room_name or "Unassigned",
                qty=qty,
                unit="NO",
                spec=spec_str,
                spec_notes=self._get_spec_notes(device.spec),
                detection_method=detection,
                confidence=device.confidence,
                is_measured=is_measured,
                source="schedule" if device.detection_method == "schedule" else "plan",
                rfi_required=bool(device.rfi_needed),
                rfi_fields=list(device.rfi_needed),
            )

            self.takeoff_lines.append(line)
            line_no += 1

    def _spec_key(self, spec: Dict[str, Any]) -> str:
        """Create hashable key from spec for grouping."""
        # Include key spec fields that affect takeoff
        key_fields = ["type", "wattage", "size", "capacity", "amp"]
        parts = []
        for field in key_fields:
            if field in spec and spec[field] and spec[field] != "TBD":
                parts.append(f"{field}={spec[field]}")
        return "|".join(parts) if parts else "default"

    def _format_spec(self, spec: Dict[str, Any]) -> str:
        """Format spec dict as readable string."""
        if not spec:
            return "TBD"

        parts = []
        for key, value in spec.items():
            if value and value != "TBD":
                parts.append(f"{value}")

        return ", ".join(parts) if parts else "TBD"

    def _get_spec_notes(self, spec: Dict[str, Any]) -> str:
        """Get notes about TBD spec fields."""
        tbd_fields = [k for k, v in spec.items() if v == "TBD"]
        if tbd_fields:
            return f"TBD: {', '.join(tbd_fields)}"
        return ""

    def _generate_connectivity_summary(self):
        """Generate connectivity summary."""
        # Cable summary
        cables = {}
        for conn in self.connections:
            if conn.connection_type == "electrical":
                spec = conn.medium_spec
                if spec not in cables:
                    cables[spec] = {"qty_m": 0, "runs": 0}
                cables[spec]["qty_m"] += conn.length_m
                cables[spec]["runs"] += 1

        # Pipe summary
        pipes = {}
        for conn in self.connections:
            if conn.connection_type == "plumbing":
                spec = conn.medium_spec
                if spec not in pipes:
                    pipes[spec] = {"qty_m": 0, "runs": 0}
                pipes[spec]["qty_m"] += conn.length_m
                pipes[spec]["runs"] += 1

        # Round quantities
        for spec in cables:
            cables[spec]["qty_m"] = round(cables[spec]["qty_m"], 1)
        for spec in pipes:
            pipes[spec]["qty_m"] = round(pipes[spec]["qty_m"], 1)

        self.connectivity_summary = {
            "electrical": {
                "cables": cables,
                "total_cable_m": round(sum(c["qty_m"] for c in cables.values()), 1),
                "total_runs": sum(c["runs"] for c in cables.values()),
            },
            "plumbing": {
                "pipes": pipes,
                "total_pipe_m": round(sum(p["qty_m"] for p in pipes.values()), 1),
                "total_runs": sum(p["runs"] for p in pipes.values()),
            },
        }

    def _generate_rfi_list(self):
        """Generate RFI list for unknown specs."""
        rfi_id = 1

        for device in self.devices:
            if not device.rfi_needed:
                continue

            for field in device.rfi_needed:
                rfi = {
                    "rfi_id": f"MEP-RFI-{rfi_id:03d}",
                    "device_id": device.id,
                    "device_type": device.device_type,
                    "location": device.room_name or "Unassigned",
                    "tag": device.tag,
                    "missing_field": field,
                    "question": self._generate_rfi_question(device.device_type, field),
                    "priority": "medium",
                    "status": "open",
                }

                self.rfi_list.append(rfi)
                rfi_id += 1

    def _generate_rfi_question(self, device_type: str, field: str) -> str:
        """Generate RFI question text."""
        device_name = device_type.replace("_", " ").title()

        questions = {
            "wattage": f"Please confirm wattage for {device_name}",
            "size": f"Please provide size/dimensions for {device_name}",
            "capacity": f"Please confirm capacity for {device_name}",
            "type": f"Please specify type of {device_name} (e.g., make/model)",
            "brand": f"Please confirm brand/make for {device_name}",
            "mounting_height": f"Please confirm mounting height for {device_name}",
            "cfm": f"Please confirm CFM rating for {device_name}",
            "color_temp": f"Please confirm color temperature for {device_name}",
            "cutout_size": f"Please confirm cutout size for {device_name}",
            "wc_type": f"Please confirm WC type (EWC/IWC)",
            "basin_type": f"Please confirm basin type (counter/wall-hung/pedestal)",
            "mixer_type": f"Please confirm mixer type",
            "ac_capacity": f"Please confirm AC capacity (tons/BTU)",
            "mcb_rating": f"Please confirm MCB rating",
            "ways": f"Please confirm number of ways for distribution board",
            "k_factor": f"Please confirm K-factor for sprinkler",
        }

        return questions.get(field, f"Please confirm {field} for {device_name}")


def generate_mep_takeoff(
    devices: List[DetectedDevice],
    connections: List[Connection] = None,
    systems: Dict[str, MEPSystem] = None,
    project_id: str = "",
) -> Dict[str, Any]:
    """
    Generate MEP takeoff from devices.

    Returns:
        Takeoff result dict
    """
    takeoff = MEPTakeoff(devices, connections, systems, project_id)
    return takeoff.generate()


def export_mep_csv(
    takeoff_result: Dict[str, Any],
    output_path: Path,
) -> Path:
    """
    Export MEP takeoff to CSV.

    Args:
        takeoff_result: Result from generate_mep_takeoff
        output_path: Output file path

    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Main takeoff CSV
    takeoff_lines = takeoff_result.get("takeoff_lines", [])

    fieldnames = [
        "line_no", "system", "subsystem", "device_type", "description",
        "tag", "location", "qty", "unit", "spec", "spec_notes",
        "detection_method", "confidence", "is_measured", "rfi_required"
    ]

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(takeoff_lines)

    # RFI list CSV
    rfi_path = output_path.parent / f"{output_path.stem}_rfis.csv"
    rfis = takeoff_result.get("rfis", [])

    if rfis:
        rfi_fields = ["rfi_id", "device_type", "location", "tag", "missing_field", "question", "priority", "status"]
        with open(rfi_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rfi_fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rfis)

    # Connectivity summary CSV
    conn_path = output_path.parent / f"{output_path.stem}_connectivity.csv"
    conn = takeoff_result.get("connectivity", {})

    with open(conn_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Category", "Item", "Spec", "Quantity", "Unit", "Runs"])

        # Cables
        for spec, data in conn.get("electrical", {}).get("cables", {}).items():
            writer.writerow(["Electrical", "Cable", spec, data["qty_m"], "m", data["runs"]])

        # Pipes
        for spec, data in conn.get("plumbing", {}).get("pipes", {}).items():
            writer.writerow(["Plumbing", "Pipe", spec, data["qty_m"], "m", data["runs"]])

    return output_path


def export_mep_excel(
    takeoff_result: Dict[str, Any],
    output_path: Path,
) -> Path:
    """
    Export MEP takeoff to Excel (falls back to CSV if openpyxl not available).

    Args:
        takeoff_result: Result from generate_mep_takeoff
        output_path: Output file path

    Returns:
        Path to created file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl not available, falling back to CSV")
        csv_path = output_path.with_suffix(".csv")
        return export_mep_csv(takeoff_result, csv_path)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: Device Takeoff
    ws1 = wb.active
    ws1.title = "MEP Takeoff"

    headers = [
        "Line", "System", "Subsystem", "Type", "Description",
        "Tag", "Location", "Qty", "Unit", "Spec", "Notes",
        "Method", "Confidence", "Measured", "RFI Required"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, line in enumerate(takeoff_result.get("takeoff_lines", []), 2):
        ws1.cell(row=row_idx, column=1, value=line.get("line_no"))
        ws1.cell(row=row_idx, column=2, value=line.get("system"))
        ws1.cell(row=row_idx, column=3, value=line.get("subsystem"))
        ws1.cell(row=row_idx, column=4, value=line.get("device_type"))
        ws1.cell(row=row_idx, column=5, value=line.get("description"))
        ws1.cell(row=row_idx, column=6, value=line.get("tag"))
        ws1.cell(row=row_idx, column=7, value=line.get("location"))
        ws1.cell(row=row_idx, column=8, value=line.get("qty"))
        ws1.cell(row=row_idx, column=9, value=line.get("unit"))
        ws1.cell(row=row_idx, column=10, value=line.get("spec"))
        ws1.cell(row=row_idx, column=11, value=line.get("spec_notes"))
        ws1.cell(row=row_idx, column=12, value=line.get("detection_method"))
        ws1.cell(row=row_idx, column=13, value=f"{line.get('confidence', 0):.0%}")
        ws1.cell(row=row_idx, column=14, value="Yes" if line.get("is_measured") else "No")
        ws1.cell(row=row_idx, column=15, value="Yes" if line.get("rfi_required") else "No")

    # Adjust column widths
    for col in range(1, 16):
        ws1.column_dimensions[chr(64 + col)].width = 15

    # Sheet 2: Connectivity Summary
    ws2 = wb.create_sheet("Connectivity")

    ws2.cell(row=1, column=1, value="ELECTRICAL - CABLES").font = Font(bold=True)
    ws2.cell(row=2, column=1, value="Spec")
    ws2.cell(row=2, column=2, value="Length (m)")
    ws2.cell(row=2, column=3, value="Runs")

    row = 3
    cables = takeoff_result.get("connectivity", {}).get("electrical", {}).get("cables", {})
    for spec, data in cables.items():
        ws2.cell(row=row, column=1, value=spec)
        ws2.cell(row=row, column=2, value=data["qty_m"])
        ws2.cell(row=row, column=3, value=data["runs"])
        row += 1

    row += 2
    ws2.cell(row=row, column=1, value="PLUMBING - PIPES").font = Font(bold=True)
    row += 1
    ws2.cell(row=row, column=1, value="Spec")
    ws2.cell(row=row, column=2, value="Length (m)")
    ws2.cell(row=row, column=3, value="Runs")
    row += 1

    pipes = takeoff_result.get("connectivity", {}).get("plumbing", {}).get("pipes", {})
    for spec, data in pipes.items():
        ws2.cell(row=row, column=1, value=spec)
        ws2.cell(row=row, column=2, value=data["qty_m"])
        ws2.cell(row=row, column=3, value=data["runs"])
        row += 1

    # Sheet 3: RFIs
    ws3 = wb.create_sheet("RFIs")

    rfi_headers = ["RFI ID", "Device Type", "Location", "Tag", "Missing Field", "Question", "Priority", "Status"]
    for col, header in enumerate(rfi_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, rfi in enumerate(takeoff_result.get("rfis", []), 2):
        ws3.cell(row=row_idx, column=1, value=rfi.get("rfi_id"))
        ws3.cell(row=row_idx, column=2, value=rfi.get("device_type"))
        ws3.cell(row=row_idx, column=3, value=rfi.get("location"))
        ws3.cell(row=row_idx, column=4, value=rfi.get("tag"))
        ws3.cell(row=row_idx, column=5, value=rfi.get("missing_field"))
        ws3.cell(row=row_idx, column=6, value=rfi.get("question"))
        ws3.cell(row=row_idx, column=7, value=rfi.get("priority"))
        ws3.cell(row=row_idx, column=8, value=rfi.get("status"))

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")

    ws4.cell(row=1, column=1, value="MEP TAKEOFF SUMMARY").font = Font(bold=True, size=14)
    ws4.cell(row=3, column=1, value="Project ID:")
    ws4.cell(row=3, column=2, value=takeoff_result.get("project_id"))
    ws4.cell(row=4, column=1, value="Generated:")
    ws4.cell(row=4, column=2, value=takeoff_result.get("generated_at"))
    ws4.cell(row=6, column=1, value="Total Line Items:")
    ws4.cell(row=6, column=2, value=takeoff_result.get("total_lines"))
    ws4.cell(row=7, column=1, value="Total Devices:")
    ws4.cell(row=7, column=2, value=takeoff_result.get("total_devices"))
    ws4.cell(row=8, column=1, value="Measured:")
    ws4.cell(row=8, column=2, value=takeoff_result.get("measured_count"))
    ws4.cell(row=9, column=1, value="Inferred:")
    ws4.cell(row=9, column=2, value=takeoff_result.get("inferred_count"))
    ws4.cell(row=10, column=1, value="RFIs Required:")
    ws4.cell(row=10, column=2, value=takeoff_result.get("rfi_count"))

    # Save
    wb.save(output_path)

    return output_path
