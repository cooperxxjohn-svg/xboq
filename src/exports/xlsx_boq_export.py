"""
src/exports/xlsx_boq_export.py

Excel (.xlsx) BOQ export with formatted sheets:
  - Sheet 1: BOQ Summary (trade grouped, subtotals, grand total)
  - Sheet 2: All Line Items (flat, sortable)
  - Sheet 3: RFI List (if rfis in payload)
  - Sheet 4: Gap Analysis (if gaps in payload)

Usage:
    from src.exports.xlsx_boq_export import export_boq_xlsx
    path = export_boq_xlsx(payload, output_path)
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL   = "1F4E79"   # dark blue
_TRADE_FILL    = "BDD7EE"   # light blue
_SUBTOTAL_FILL = "D9E1F2"   # pale blue
_TOTAL_FILL    = "FFC000"   # amber
_ALT_ROW_FILL  = "F2F2F2"   # light grey for alternating rows
_WHITE         = "FFFFFF"


def _header_style(ws, cell, text: str, bold: bool = True):
    cell.value = text
    cell.font = Font(bold=bold, color=_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _trade_style(ws, cell, text: str):
    cell.value = text
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor=_TRADE_FILL)


def _money(val) -> float:
    try:
        return round(float(val or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _qty(val) -> float:
    try:
        return round(float(val or 0), 3)
    except (TypeError, ValueError):
        return 0.0


def _sheet_boq_summary(wb, boq_items: list, project_name: str = "") -> None:
    """Sheet 1: Trade-grouped BOQ with subtotals."""
    ws = wb.create_sheet("BOQ Summary")
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Bill of Quantities — {project_name}" if project_name else "Bill of Quantities"
    title_cell.font = Font(bold=True, size=14, color=_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Column headers (row 2)
    headers = ["#", "Trade", "Description", "Qty", "Unit", "Rate (₹)", "Amount (₹)", "Confidence"]
    col_widths = [5, 15, 50, 10, 8, 12, 14, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _header_style(ws, ws.cell(row=2, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Group items by trade
    from collections import defaultdict
    trade_groups: dict = defaultdict(list)
    for item in boq_items:
        trade = str(item.get("trade", item.get("category", "General"))).title()
        trade_groups[trade].append(item)

    row = 3
    grand_total = 0.0
    item_num = 1
    for trade, items in sorted(trade_groups.items()):
        # Trade header row
        ws.merge_cells(f"A{row}:H{row}")
        tc = ws.cell(row=row, column=1)
        _trade_style(ws, tc, f"  {trade}")
        ws.row_dimensions[row].height = 18
        row += 1

        trade_total = 0.0
        for i, item in enumerate(items):
            qty = _qty(item.get("qty") or item.get("quantity"))
            rate = _money(item.get("rate_inr") or item.get("rate") or item.get("unit_rate"))
            amount = _money(item.get("total_inr") or item.get("amount") or (qty * rate))
            trade_total += amount
            grand_total += amount

            fill_color = _WHITE if i % 2 == 0 else _ALT_ROW_FILL
            row_fill = PatternFill("solid", fgColor=fill_color)

            values = [
                item_num,
                str(item.get("trade", "")).title(),
                str(item.get("description", "")),
                qty if qty else "",
                str(item.get("unit", "")),
                rate if rate else "",
                amount if amount else "",
                f"{int(float(item.get('confidence', 0) or 0) * 100)}%" if item.get("confidence") else "",
            ]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = row_fill
                c.font = Font(size=9)
                c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
                if col in (4, 6, 7):
                    c.number_format = '#,##0.00'
            ws.row_dimensions[row].height = 15
            row += 1
            item_num += 1

        # Trade subtotal
        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=_SUBTOTAL_FILL)
            c.font = Font(bold=True, size=9)
        ws.cell(row=row, column=3, value=f"Subtotal — {trade}").font = Font(bold=True, size=9, italic=True)
        st = ws.cell(row=row, column=7, value=trade_total)
        st.number_format = '#,##0.00'
        st.font = Font(bold=True, size=9)
        row += 1

    # Grand total
    for col in range(1, 9):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=_TOTAL_FILL)
        c.font = Font(bold=True, size=11)
    ws.cell(row=row, column=3, value="GRAND TOTAL").font = Font(bold=True, size=11)
    gt = ws.cell(row=row, column=7, value=grand_total)
    gt.number_format = '#,##0.00'
    gt.font = Font(bold=True, size=11)
    ws.row_dimensions[row].height = 22

    # Freeze top 2 rows
    ws.freeze_panes = "A3"


def _sheet_line_items(wb, boq_items: list) -> None:
    """Sheet 2: Flat line items (all trades, no grouping)."""
    ws = wb.create_sheet("Line Items")
    ws.sheet_properties.tabColor = "2E75B6"

    headers = ["#", "Trade", "Description", "Qty", "Unit", "Rate (₹)", "Amount (₹)",
               "Confidence", "Source Page", "Notes"]
    col_widths = [5, 15, 55, 10, 8, 12, 14, 12, 10, 30]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _header_style(ws, ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for i, item in enumerate(boq_items, start=1):
        qty = _qty(item.get("qty") or item.get("quantity"))
        rate = _money(item.get("rate_inr") or item.get("rate"))
        amount = _money(item.get("total_inr") or item.get("amount") or (qty * rate))
        fill = PatternFill("solid", fgColor=(_WHITE if i % 2 == 0 else _ALT_ROW_FILL))
        vals = [
            i,
            str(item.get("trade", "")).title(),
            str(item.get("description", "")),
            qty or "",
            str(item.get("unit", "")),
            rate or "",
            amount or "",
            f"{int(float(item.get('confidence', 0) or 0) * 100)}%" if item.get("confidence") else "",
            str(item.get("source_page", "") or item.get("page", "")),
            str(item.get("notes", "") or item.get("spec_ref", ""))[:100],
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i + 1, column=col, value=val)
            c.fill = fill
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (3, 10)))
            if col in (4, 6, 7):
                c.number_format = '#,##0.00'
        ws.row_dimensions[i + 1].height = 14


def _sheet_rfis(wb, rfis: list) -> None:
    """Sheet 3: RFI list."""
    if not rfis:
        return
    ws = wb.create_sheet("RFIs")
    ws.sheet_properties.tabColor = "C00000"

    headers = ["RFI #", "Trade", "Severity", "Question", "Evidence", "Action Required"]
    col_widths = [8, 15, 10, 60, 25, 35]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _header_style(ws, ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    sev_colors = {"CRITICAL": "C00000", "HIGH": "FF0000", "MEDIUM": "FFC000", "LOW": "92D050"}
    for i, rfi in enumerate(rfis, start=1):
        sev = str(rfi.get("severity", "MEDIUM")).upper()
        vals = [
            rfi.get("id") or rfi.get("rfi_id") or f"RFI-{i:03d}",
            str(rfi.get("trade", "general")).title(),
            sev,
            str(rfi.get("question") or rfi.get("description", "")),
            str(rfi.get("evidence", rfi.get("source_page", ""))),
            str(rfi.get("action_required") or rfi.get("action", "")),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i + 1, column=col, value=val)
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (4, 5, 6)))
        # Colour severity cell
        sev_cell = ws.cell(row=i + 1, column=3)
        sev_cell.fill = PatternFill("solid", fgColor=sev_colors.get(sev, "AAAAAA"))
        sev_cell.font = Font(bold=True, size=9, color=_WHITE)
        ws.row_dimensions[i + 1].height = 20


def _sheet_gaps(wb, gaps: list) -> None:
    """Sheet 4: Gap analysis."""
    if not gaps:
        return
    ws = wb.create_sheet("Gap Analysis")
    ws.sheet_properties.tabColor = "FF6600"

    headers = ["Gap #", "Trade", "Severity", "Description", "Action Required", "Cost Impact"]
    col_widths = [8, 15, 10, 60, 35, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _header_style(ws, ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    sev_colors = {"CRITICAL": "C00000", "HIGH": "FF4500", "MEDIUM": "FFC000", "LOW": "92D050"}
    for i, gap in enumerate(gaps, start=1):
        sev = str(gap.get("severity", "MEDIUM")).upper()
        vals = [
            gap.get("id") or f"GAP-{i:03d}",
            str(gap.get("trade", "general")).title(),
            sev,
            str(gap.get("description", "")),
            str(gap.get("action_required") or gap.get("action", "")),
            str(gap.get("cost_impact") or ""),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i + 1, column=col, value=val)
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (4, 5)))
        sev_cell = ws.cell(row=i + 1, column=3)
        sev_cell.fill = PatternFill("solid", fgColor=sev_colors.get(sev, "AAAAAA"))
        sev_cell.font = Font(bold=True, size=9, color=_WHITE)
        ws.row_dimensions[i + 1].height = 18


def export_boq_xlsx(
    payload: dict,
    output_path: Union[str, Path],
) -> Optional[Path]:
    """
    Export the analysis payload to a formatted Excel workbook.

    Parameters
    ----------
    payload : dict
        Full analysis payload from pipeline.py.
    output_path : str | Path
        Destination .xlsx file path.

    Returns
    -------
    Path | None
        Path to written file, or None if openpyxl unavailable.
    """
    if not _OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed — Excel export unavailable. pip install openpyxl")
        return None

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    boq_items = payload.get("boq_items") or []
    rfis      = payload.get("rfis") or []
    gaps      = payload.get("gaps") or []
    project   = str(payload.get("project_name") or payload.get("project_id") or "")

    wb = openpyxl.Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    _sheet_boq_summary(wb, boq_items, project_name=project)
    _sheet_line_items(wb, boq_items)
    _sheet_rfis(wb, rfis)
    _sheet_gaps(wb, gaps)

    wb.save(str(output_path))
    logger.info("Excel export: %s (%d items, %d RFIs, %d gaps)",
                output_path, len(boq_items), len(rfis), len(gaps))
    return output_path
