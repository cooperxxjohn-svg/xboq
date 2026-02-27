"""
Pricing Readiness Sheet Export

Generates a trade-by-trade view of what can be priced today,
what's blocked, and what actions are needed.
"""

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

from .models import PricingReadinessRow, Trade, RiskLevel, ConfidenceBand

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Default trade coverage assumptions when we can't compute
DEFAULT_TRADE_COVERAGE = {
    "civil": {"base": 40, "items": 8},
    "structural": {"base": 60, "items": 12},
    "architectural": {"base": 50, "items": 15},
    "mep": {"base": 20, "items": 10},
    "finishes": {"base": 45, "items": 20},
    "general": {"base": 70, "items": 5},
}


def compute_pricing_readiness(
    rfis: List[Dict],
    trade_summary: Dict[str, Dict],
    bid_gate: Dict,
    scope_summary: Optional[Dict] = None,
) -> List[PricingReadinessRow]:
    """
    Compute pricing readiness for each trade.

    Args:
        rfis: List of RFI dicts from analysis
        trade_summary: Trade-wise gap summary from RFI adapter
        bid_gate: Bid gate result dict
        scope_summary: Optional scope summary for item counts

    Returns:
        List of PricingReadinessRow for each trade
    """
    rows = []

    # Get overall score for baseline
    overall_score = bid_gate.get("score", 50)
    blockers = bid_gate.get("blockers", [])

    for trade_name in ["civil", "structural", "architectural", "mep", "finishes", "general"]:
        trade_data = trade_summary.get(trade_name, {})
        defaults = DEFAULT_TRADE_COVERAGE.get(trade_name, {"base": 50, "items": 10})

        # Get RFI counts for this trade
        rfi_count = trade_data.get("rfi_count", 0)
        high_priority = trade_data.get("high_priority", 0)
        gaps = trade_data.get("gaps", [])

        # Compute coverage based on gaps and RFIs
        base_coverage = defaults["base"]
        coverage_penalty = min(rfi_count * 5, 40)  # Each RFI reduces coverage
        scope_coverage = max(0, base_coverage - coverage_penalty)

        # Estimate priceable vs blocked items
        total_items = defaults["items"]
        blocked_ratio = min(0.8, rfi_count * 0.1 + high_priority * 0.15)
        blocked_items = int(total_items * blocked_ratio)
        priceable_items = total_items - blocked_items

        # Determine missing dependencies from gaps
        missing_deps = []
        for gap in gaps[:3]:
            gap_lower = gap.lower()
            if "schedule" in gap_lower:
                missing_deps.append("Schedule")
            elif "spec" in gap_lower:
                missing_deps.append("Specification")
            elif "scale" in gap_lower:
                missing_deps.append("Scale")
            elif "mep" in gap_lower or "electrical" in gap_lower or "plumbing" in gap_lower:
                missing_deps.append("MEP Drawings")
            elif "section" in gap_lower or "elevation" in gap_lower:
                missing_deps.append("Sections/Elevations")
            else:
                missing_deps.append(gap[:30])

        if not missing_deps and rfi_count > 0:
            missing_deps = ["Review RFIs"]

        # Determine confidence band
        if scope_coverage >= 70 and high_priority == 0:
            confidence = "high"
        elif scope_coverage >= 40 or high_priority <= 1:
            confidence = "medium"
        else:
            confidence = "low"

        # Determine cost risk
        if high_priority >= 2 or blocked_items > total_items * 0.5:
            cost_risk = "high"
        elif rfi_count >= 2 or blocked_items > total_items * 0.25:
            cost_risk = "medium"
        else:
            cost_risk = "low"

        # Determine schedule risk
        if "Schedule" in missing_deps or "MEP Drawings" in missing_deps:
            schedule_risk = "high"
        elif rfi_count >= 3:
            schedule_risk = "medium"
        else:
            schedule_risk = "low"

        # Determine next action
        if high_priority > 0:
            next_action = f"Resolve {high_priority} critical RFI(s)"
        elif rfi_count > 0:
            next_action = f"Address {rfi_count} RFI(s)"
        elif missing_deps:
            next_action = f"Obtain {missing_deps[0]}"
        else:
            next_action = "Ready for pricing"

        rows.append(PricingReadinessRow(
            trade=trade_name,
            scope_coverage_pct=scope_coverage,
            priceable_items=priceable_items,
            blocked_items=blocked_items,
            top_missing=missing_deps,
            next_action=next_action,
            confidence=confidence,
            cost_risk=cost_risk,
            schedule_risk=schedule_risk,
        ))

    return rows


def build_pricing_readiness_sheet(
    project_id: str,
    rfis: List[Dict],
    trade_summary: Dict[str, Dict],
    bid_gate: Dict,
    scope_summary: Optional[Dict] = None,
    output_dir: Optional[Path] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """
    Build Pricing Readiness Sheet as Excel and CSV.

    Args:
        project_id: Project identifier
        rfis: List of RFI dicts
        trade_summary: Trade-wise gap summary
        bid_gate: Bid gate result
        scope_summary: Optional scope data
        output_dir: Where to save files (defaults to out/{project_id}/exports/)

    Returns:
        Tuple of (xlsx_path, csv_path) or (None, None) on error
    """
    # Compute readiness rows
    rows = compute_pricing_readiness(rfis, trade_summary, bid_gate, scope_summary)

    if not rows:
        return None, None

    # Setup output directory
    if output_dir is None:
        output_dir = Path(__file__).parent.parent.parent / "out" / project_id / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate timestamp for filenames
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Export to CSV
    csv_path = output_dir / f"pricing_readiness_{timestamp}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Trade", "Scope Coverage %", "Priceable Items", "Blocked Items",
            "Top Missing", "Next Action", "Confidence", "Cost Risk", "Schedule Risk"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_dict())

    # Export to Excel if available
    xlsx_path = None
    if HAS_OPENPYXL:
        xlsx_path = output_dir / f"pricing_readiness_{timestamp}.xlsx"
        _write_pricing_xlsx(rows, xlsx_path, project_id, bid_gate)

    return str(xlsx_path) if xlsx_path else None, str(csv_path)


def _write_pricing_xlsx(
    rows: List[PricingReadinessRow],
    output_path: Path,
    project_id: str,
    bid_gate: Dict,
):
    """Write pricing readiness to styled Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Readiness"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    risk_fills = {
        "high": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "low": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    }

    # Title row
    ws.merge_cells('A1:I1')
    ws['A1'] = f"XBOQ Pricing Readiness - {project_id}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Summary row
    score = bid_gate.get("score", 0)
    status = bid_gate.get("status", "NO-GO")
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Bid Status: {status} | Score: {score}/100 | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        "Trade", "Scope Coverage %", "Priceable Items", "Blocked Items",
        "Top Missing", "Next Action", "Confidence", "Cost Risk", "Schedule Risk"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(rows, 5):
        data = row.to_dict()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data[header])
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx != 5 else 'left', wrap_text=True)

            # Apply risk coloring
            if header in ["Cost Risk", "Schedule Risk", "Confidence"]:
                risk_key = data[header].lower()
                if risk_key in risk_fills:
                    cell.fill = risk_fills[risk_key]

    # Column widths
    widths = [15, 18, 15, 15, 35, 30, 12, 12, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A5'

    wb.save(output_path)


def get_pricing_readiness_buffer(
    rfis: List[Dict],
    trade_summary: Dict[str, Dict],
    bid_gate: Dict,
    format: str = "xlsx",
) -> BytesIO:
    """
    Get pricing readiness as in-memory buffer for Streamlit download.

    Args:
        rfis: List of RFI dicts
        trade_summary: Trade-wise gap summary
        bid_gate: Bid gate result
        format: "xlsx" or "csv"

    Returns:
        BytesIO buffer with file contents
    """
    rows = compute_pricing_readiness(rfis, trade_summary, bid_gate)

    if format == "csv":
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "Trade", "Scope Coverage %", "Priceable Items", "Blocked Items",
            "Top Missing", "Next Action", "Confidence", "Cost Risk", "Schedule Risk"
        ])
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_dict())
        buffer = BytesIO(output.getvalue().encode('utf-8'))
        return buffer

    elif format == "xlsx" and HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pricing Readiness"

        # Simple headers
        headers = [
            "Trade", "Scope Coverage %", "Priceable Items", "Blocked Items",
            "Top Missing", "Next Action", "Confidence", "Cost Risk", "Schedule Risk"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, row in enumerate(rows, 2):
            data = row.to_dict()
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=data[header])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    return BytesIO()
