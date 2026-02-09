"""
Estimator Reconciliation Layer

Transforms raw BOQ extraction into estimator-ready output:
1. Merges measured and inferred quantities
2. Detects missing scope items
3. Generates confidence metrics
4. Applies estimator assumptions
5. Exports bid-ready workbook

This is the bridge between drawing parsing and actual estimation.
"""

import csv
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from enum import Enum

logger = logging.getLogger(__name__)


class QuantitySource(Enum):
    """Source of quantity data."""
    MEASURED = "measured"
    INFERRED = "inferred"
    ASSUMED = "assumed"
    MISSING = "missing"


@dataclass
class EstimatorAssumptions:
    """Estimator override assumptions."""
    wall_height_m: float = 3.0
    door_height_m: float = 2.1
    window_height_m: float = 1.2
    plaster_both_sides: bool = True
    floor_finish_all_rooms: bool = True
    skirting_height_mm: int = 100
    waterproof_wet_areas: bool = True
    frame_all_openings: bool = True

    @classmethod
    def from_cli_args(cls, args) -> "EstimatorAssumptions":
        """Create from CLI arguments."""
        return cls(
            wall_height_m=getattr(args, 'assume_wall_height', 3.0) or 3.0,
            door_height_m=getattr(args, 'assume_door_height', 2.1) or 2.1,
            plaster_both_sides=getattr(args, 'assume_plaster_both_sides', True),
            floor_finish_all_rooms=getattr(args, 'assume_floor_finish_all_rooms', True),
        )


@dataclass
class ReconciliationResult:
    """Result of BOQ reconciliation."""
    # Merged items
    estimator_view: List[Dict[str, Any]] = field(default_factory=list)

    # Missing scope
    missing_scope: List[Dict[str, Any]] = field(default_factory=list)
    missing_scope_rfis: List[Dict[str, Any]] = field(default_factory=list)

    # Confidence metrics
    confidence_by_page: List[Dict[str, Any]] = field(default_factory=list)

    # Summary stats
    total_measured: int = 0
    total_inferred: int = 0
    total_assumed: int = 0
    total_missing: int = 0
    needs_review_count: int = 0
    coverage_percent: float = 0.0


# =============================================================================
# MISSING SCOPE RULES
# =============================================================================

# Rules: (trigger_type, trigger_condition, required_items)
MISSING_SCOPE_RULES = [
    # Every room needs flooring and skirting
    ("room", "any", [
        ("flooring", "Flooring", "sqm", "area_sqm"),
        ("skirting", "Skirting 100mm", "rmt", "perimeter_m"),
        ("ceiling", "Ceiling finish", "sqm", "area_sqm"),
    ]),

    # Every opening needs frame and shutter
    ("opening", "door", [
        ("door_frame", "Door frame", "no", 1),
        ("door_shutter", "Door shutter", "no", 1),
        ("hardware", "Door hardware set", "set", 1),
    ]),
    ("opening", "window", [
        ("window_frame", "Window frame", "no", 1),
        ("window_shutter", "Window shutter/glass", "no", 1),
        ("window_grill", "Window grill", "no", 1),
    ]),

    # Wet areas need waterproofing
    ("room", "wet_area", [
        ("waterproofing_floor", "Waterproofing to floor", "sqm", "area_sqm"),
        ("waterproofing_wall", "Waterproofing to walls (up to 1.8m)", "sqm", "wet_wall_area"),
        ("anti_skid_tile", "Anti-skid floor tile", "sqm", "area_sqm"),
        ("wall_tile", "Wall tile dado", "sqm", "dado_area"),
    ]),

    # Every wall needs plaster and paint
    ("wall", "any", [
        ("plaster_internal", "Cement plaster 12mm (internal)", "sqm", "wall_area"),
        ("plaster_external", "Cement plaster 20mm (external)", "sqm", "wall_area"),
        ("putty", "Wall putty 2 coats", "sqm", "wall_area"),
        ("primer", "Primer coat", "sqm", "wall_area"),
        ("paint", "Emulsion paint 2 coats", "sqm", "wall_area"),
    ]),
]

WET_AREA_TYPES = {"toilet", "bathroom", "kitchen", "utility", "washroom", "wc", "lavatory"}


class EstimatorReconciler:
    """
    Reconciles measured and inferred BOQ into estimator-ready output.

    This is the core of the estimator assistant - taking raw extraction
    results and making them usable for actual cost estimation.
    """

    def __init__(
        self,
        output_dir: Path,
        assumptions: Optional[EstimatorAssumptions] = None,
    ):
        self.output_dir = Path(output_dir)
        self.assumptions = assumptions or EstimatorAssumptions()

        # Data containers
        self.measured_items: List[Dict[str, Any]] = []
        self.inferred_items: List[Dict[str, Any]] = []
        self.rooms: List[Dict[str, Any]] = []
        self.openings: List[Dict[str, Any]] = []

        # Results
        self.result = ReconciliationResult()

    def run(self) -> ReconciliationResult:
        """
        Run full reconciliation pipeline.

        Returns:
            ReconciliationResult with all outputs
        """
        logger.info("Starting estimator reconciliation...")

        # Step 1: Load data
        self._load_data()

        # Step 2: Merge measured and inferred
        self._merge_quantities()

        # Step 3: Detect missing scope
        self._detect_missing_scope()

        # Step 4: Generate confidence metrics
        self._generate_confidence_metrics()

        # Step 5: Write outputs
        self._write_outputs()

        logger.info(f"Reconciliation complete: {self.result.total_measured} measured, "
                   f"{self.result.total_inferred} inferred, {self.result.total_missing} missing")

        return self.result

    def _load_data(self):
        """Load measured, inferred BOQ and room/opening data."""
        boq_dir = self.output_dir / "boq"
        combined_dir = self.output_dir / "combined"

        # Load measured items
        measured_path = boq_dir / "boq_measured.csv"
        if measured_path.exists():
            with open(measured_path, newline='') as f:
                reader = csv.DictReader(f)
                self.measured_items = list(reader)
            logger.info(f"Loaded {len(self.measured_items)} measured items")

        # Load inferred items
        inferred_path = boq_dir / "boq_inferred.csv"
        if inferred_path.exists():
            with open(inferred_path, newline='') as f:
                reader = csv.DictReader(f)
                self.inferred_items = list(reader)
            logger.info(f"Loaded {len(self.inferred_items)} inferred items")

        # Load rooms
        rooms_path = combined_dir / "all_rooms.json"
        if rooms_path.exists():
            with open(rooms_path) as f:
                data = json.load(f)
                self.rooms = data.get("rooms", [])
            logger.info(f"Loaded {len(self.rooms)} rooms")

        # Load openings
        openings_path = combined_dir / "all_openings.json"
        if openings_path.exists():
            with open(openings_path) as f:
                data = json.load(f)
                self.openings = data.get("openings", [])
            logger.info(f"Loaded {len(self.openings)} openings")

    def _merge_quantities(self):
        """Merge measured and inferred into estimator view."""
        # Index measured items by (package, description, room)
        measured_index: Dict[Tuple[str, str, str], Dict] = {}
        for item in self.measured_items:
            key = (
                item.get("package", ""),
                item.get("description", ""),
                item.get("room", ""),
            )
            measured_index[key] = item

        # Index inferred items
        inferred_index: Dict[Tuple[str, str, str], Dict] = {}
        for item in self.inferred_items:
            key = (
                item.get("package", ""),
                item.get("description", ""),
                item.get("room", ""),
            )
            inferred_index[key] = item

        # All unique keys
        all_keys = set(measured_index.keys()) | set(inferred_index.keys())

        for key in sorted(all_keys):
            measured = measured_index.get(key)
            inferred = inferred_index.get(key)

            # Determine source and quantities
            if measured and inferred:
                # Both exist - prefer measured
                source = QuantitySource.MEASURED
                measured_qty = self._parse_qty(measured.get("qty", ""))
                inferred_qty = self._parse_qty(inferred.get("qty", ""))
                final_qty = measured_qty if measured_qty else inferred_qty
                confidence = float(measured.get("confidence", 0.8))
                needs_review = False
                base_item = measured
            elif measured:
                # Only measured
                source = QuantitySource.MEASURED
                measured_qty = self._parse_qty(measured.get("qty", ""))
                inferred_qty = None
                final_qty = measured_qty
                confidence = float(measured.get("confidence", 0.8))
                needs_review = False
                base_item = measured
            elif inferred:
                # Only inferred
                source = QuantitySource.INFERRED
                measured_qty = None
                inferred_qty = self._parse_qty(inferred.get("qty", ""))
                final_qty = inferred_qty
                confidence = float(inferred.get("confidence", 0.5))
                # Low confidence inferred needs review
                needs_review = confidence < 0.6
                base_item = inferred
            else:
                continue

            # Create estimator view item
            est_item = {
                "item_id": base_item.get("item_id", ""),
                "package": base_item.get("package", ""),
                "description": base_item.get("description", ""),
                "room": base_item.get("room", ""),
                "unit": base_item.get("unit", ""),
                "measured_qty": measured_qty if measured_qty else "",
                "inferred_qty": inferred_qty if inferred_qty else "",
                "final_qty": final_qty if final_qty else "",
                "source": source.value,
                "confidence": round(confidence, 2),
                "needs_review": "YES" if needs_review else "NO",
                "source_pages": base_item.get("source_pages", ""),
                "method": base_item.get("method", ""),
                "scale_basis": base_item.get("scale_basis", ""),
                "geometry_refs": base_item.get("geometry_refs", ""),
            }

            self.result.estimator_view.append(est_item)

            # Update stats
            if source == QuantitySource.MEASURED:
                self.result.total_measured += 1
            elif source == QuantitySource.INFERRED:
                self.result.total_inferred += 1

            if needs_review:
                self.result.needs_review_count += 1

        # Calculate coverage
        total = self.result.total_measured + self.result.total_inferred
        if total > 0:
            self.result.coverage_percent = (self.result.total_measured / total) * 100

        logger.info(f"Merged {len(self.result.estimator_view)} items, "
                   f"{self.result.needs_review_count} need review")

    def _parse_qty(self, qty_str: str) -> Optional[float]:
        """Parse quantity string to float."""
        if not qty_str or qty_str in ("", "None", "null", "no"):
            return None
        try:
            return float(qty_str)
        except (ValueError, TypeError):
            return None

    def _detect_missing_scope(self):
        """Detect missing scope items based on rules."""
        # Get existing BOQ descriptions for checking
        existing_descriptions: Set[str] = set()
        for item in self.result.estimator_view:
            desc = item.get("description", "").lower()
            existing_descriptions.add(desc)

        rfi_id = 1

        # Check room-based rules
        for room in self.rooms:
            room_id = room.get("id", "")
            room_label = room.get("label", "Unknown")
            room_type = room.get("room_type", "").lower()
            area = room.get("area_sqm", 15.0)
            perimeter = room.get("perimeter_m", 16.0)

            # Check if wet area
            is_wet_area = room_type in WET_AREA_TYPES

            # Apply room rules
            for trigger_type, trigger_cond, required_items in MISSING_SCOPE_RULES:
                if trigger_type != "room":
                    continue

                # Check trigger condition
                if trigger_cond == "any":
                    applies = True
                elif trigger_cond == "wet_area":
                    applies = is_wet_area
                else:
                    applies = False

                if not applies:
                    continue

                # Check each required item
                for item_type, item_desc, unit, qty_source in required_items:
                    # Check if already exists
                    desc_lower = item_desc.lower()
                    if any(desc_lower in existing for existing in existing_descriptions):
                        continue

                    # Calculate quantity
                    if qty_source == "area_sqm":
                        qty = area
                    elif qty_source == "perimeter_m":
                        qty = perimeter
                    elif qty_source == "wet_wall_area":
                        # Wall area up to 1.8m for waterproofing
                        qty = perimeter * 1.8
                    elif qty_source == "dado_area":
                        # Dado typically 1.2m height
                        qty = perimeter * 1.2
                    elif isinstance(qty_source, (int, float)):
                        qty = qty_source
                    else:
                        qty = 1

                    # Add to missing scope
                    missing_item = {
                        "item_type": item_type,
                        "description": f"{item_desc} in {room_label}",
                        "unit": unit,
                        "estimated_qty": round(qty, 2),
                        "room_id": room_id,
                        "room_label": room_label,
                        "rule": f"room_{trigger_cond}",
                        "priority": "HIGH" if is_wet_area and "waterproof" in item_type else "MEDIUM",
                    }
                    self.result.missing_scope.append(missing_item)

                    # Add RFI for high priority items
                    if missing_item["priority"] == "HIGH":
                        self.result.missing_scope_rfis.append({
                            "rfi_id": f"RFI-MS-{rfi_id:03d}",
                            "category": "Missing Scope",
                            "description": f"Confirm {item_desc} requirement for {room_label}",
                            "room": room_label,
                            "estimated_qty": f"{qty:.2f} {unit}",
                            "impact": "Cost impact if not included",
                        })
                        rfi_id += 1

        # Check opening-based rules
        for opening in self.openings:
            opening_id = opening.get("id", "")
            opening_type = opening.get("type", "door").lower()
            opening_label = opening.get("label", "Opening")
            width = opening.get("width_m", 0.9)
            height = opening.get("height_m", self.assumptions.door_height_m)

            for trigger_type, trigger_cond, required_items in MISSING_SCOPE_RULES:
                if trigger_type != "opening":
                    continue

                if trigger_cond not in opening_type and trigger_cond != "any":
                    continue

                for item_type, item_desc, unit, qty_source in required_items:
                    desc_lower = item_desc.lower()
                    if any(desc_lower in existing for existing in existing_descriptions):
                        continue

                    qty = qty_source if isinstance(qty_source, (int, float)) else 1

                    missing_item = {
                        "item_type": item_type,
                        "description": f"{item_desc} for {opening_label} ({width}m x {height}m)",
                        "unit": unit,
                        "estimated_qty": qty,
                        "opening_id": opening_id,
                        "opening_type": opening_type,
                        "rule": f"opening_{trigger_cond}",
                        "priority": "MEDIUM",
                    }
                    self.result.missing_scope.append(missing_item)

        self.result.total_missing = len(self.result.missing_scope)
        logger.info(f"Detected {self.result.total_missing} missing scope items, "
                   f"{len(self.result.missing_scope_rfis)} RFIs generated")

    def _generate_confidence_metrics(self):
        """Generate confidence metrics by page."""
        # Group items by source page
        page_stats: Dict[int, Dict[str, int]] = {}

        for item in self.result.estimator_view:
            pages_str = item.get("source_pages", "")
            if not pages_str:
                continue

            # Parse page numbers (could be "1;2;3" format)
            try:
                pages = [int(p.strip()) for p in str(pages_str).split(";") if p.strip().isdigit()]
            except (ValueError, AttributeError):
                pages = []

            source = item.get("source", "")

            for page in pages:
                if page not in page_stats:
                    page_stats[page] = {"measured": 0, "inferred": 0, "total": 0}

                page_stats[page]["total"] += 1
                if source == "measured":
                    page_stats[page]["measured"] += 1
                elif source == "inferred":
                    page_stats[page]["inferred"] += 1

        # Calculate percentages
        for page, stats in sorted(page_stats.items()):
            total = stats["total"]
            if total == 0:
                continue

            self.result.confidence_by_page.append({
                "page": page,
                "total_items": total,
                "measured_items": stats["measured"],
                "inferred_items": stats["inferred"],
                "measured_pct": round((stats["measured"] / total) * 100, 1),
                "inferred_pct": round((stats["inferred"] / total) * 100, 1),
                "confidence_score": round((stats["measured"] / total) * 100, 1),
            })

        logger.info(f"Generated confidence metrics for {len(self.result.confidence_by_page)} pages")

    def _write_outputs(self):
        """Write all output files."""
        estimator_dir = self.output_dir / "estimator"
        estimator_dir.mkdir(parents=True, exist_ok=True)

        # 1. Write estimator view CSV
        self._write_csv(
            estimator_dir / "boq_estimator_view.csv",
            self.result.estimator_view,
            [
                "item_id", "package", "description", "room", "unit",
                "measured_qty", "inferred_qty", "final_qty",
                "source", "confidence", "needs_review",
                "source_pages", "method", "scale_basis", "geometry_refs"
            ]
        )

        # 2. Write missing scope CSV
        self._write_csv(
            estimator_dir / "missing_scope.csv",
            self.result.missing_scope,
            ["item_type", "description", "unit", "estimated_qty",
             "room_id", "room_label", "rule", "priority"]
        )

        # 3. Write missing scope RFIs
        self._write_rfi_markdown(estimator_dir / "rfi_missing_scope.md")

        # 4. Write confidence by page
        self._write_csv(
            estimator_dir / "confidence_by_page.csv",
            self.result.confidence_by_page,
            ["page", "total_items", "measured_items", "inferred_items",
             "measured_pct", "inferred_pct", "confidence_score"]
        )

        # 5. Write assumptions used
        self._write_assumptions(estimator_dir / "assumptions_used.json")

        # 6. Write reconciliation summary
        self._write_summary(estimator_dir / "reconciliation_summary.md")

        logger.info(f"Outputs written to {estimator_dir}")

    def _write_csv(self, path: Path, data: List[Dict], fieldnames: List[str]):
        """Write data to CSV file."""
        if not data:
            # Write empty CSV with headers
            with open(path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
            return

        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(data)

    def _write_rfi_markdown(self, path: Path):
        """Write missing scope RFIs as markdown."""
        with open(path, "w") as f:
            f.write("# Missing Scope RFIs\n\n")
            f.write(f"Generated: {self._now()}\n\n")
            f.write(f"**Total Missing Scope Items**: {self.result.total_missing}\n")
            f.write(f"**RFIs Generated**: {len(self.result.missing_scope_rfis)}\n\n")
            f.write("---\n\n")

            if not self.result.missing_scope_rfis:
                f.write("*No high-priority missing scope items detected.*\n")
                return

            for rfi in self.result.missing_scope_rfis:
                f.write(f"## {rfi['rfi_id']}: {rfi['description']}\n\n")
                f.write(f"- **Category**: {rfi['category']}\n")
                f.write(f"- **Room/Location**: {rfi['room']}\n")
                f.write(f"- **Estimated Quantity**: {rfi['estimated_qty']}\n")
                f.write(f"- **Impact**: {rfi['impact']}\n\n")
                f.write("---\n\n")

    def _write_assumptions(self, path: Path):
        """Write assumptions used to JSON."""
        assumptions_dict = {
            "wall_height_m": self.assumptions.wall_height_m,
            "door_height_m": self.assumptions.door_height_m,
            "window_height_m": self.assumptions.window_height_m,
            "plaster_both_sides": self.assumptions.plaster_both_sides,
            "floor_finish_all_rooms": self.assumptions.floor_finish_all_rooms,
            "skirting_height_mm": self.assumptions.skirting_height_mm,
            "waterproof_wet_areas": self.assumptions.waterproof_wet_areas,
            "frame_all_openings": self.assumptions.frame_all_openings,
        }

        with open(path, "w") as f:
            json.dump(assumptions_dict, f, indent=2)

    def _write_summary(self, path: Path):
        """Write reconciliation summary markdown."""
        with open(path, "w") as f:
            f.write("# Estimator Reconciliation Summary\n\n")
            f.write(f"Generated: {self._now()}\n\n")

            f.write("## Quantity Sources\n\n")
            f.write("| Source | Count | Percentage |\n")
            f.write("|--------|-------|------------|\n")
            total = self.result.total_measured + self.result.total_inferred
            if total > 0:
                f.write(f"| Measured | {self.result.total_measured} | "
                       f"{(self.result.total_measured/total)*100:.1f}% |\n")
                f.write(f"| Inferred | {self.result.total_inferred} | "
                       f"{(self.result.total_inferred/total)*100:.1f}% |\n")
            f.write(f"| **Total** | **{total}** | 100% |\n\n")

            f.write("## Review Required\n\n")
            f.write(f"- Items needing review: **{self.result.needs_review_count}**\n")
            f.write(f"- Coverage (measured %): **{self.result.coverage_percent:.1f}%**\n\n")

            f.write("## Missing Scope\n\n")
            f.write(f"- Missing scope items detected: **{self.result.total_missing}**\n")
            f.write(f"- RFIs generated: **{len(self.result.missing_scope_rfis)}**\n\n")

            f.write("## Assumptions Used\n\n")
            f.write(f"- Wall height: {self.assumptions.wall_height_m}m\n")
            f.write(f"- Door height: {self.assumptions.door_height_m}m\n")
            f.write(f"- Plaster both sides: {self.assumptions.plaster_both_sides}\n")
            f.write(f"- Floor finish all rooms: {self.assumptions.floor_finish_all_rooms}\n")

    def _now(self) -> str:
        """Get current timestamp."""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M")


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_bid_ready_excel(
    output_dir: Path,
    result: ReconciliationResult,
    assumptions: EstimatorAssumptions,
) -> Path:
    """
    Export bid-ready Excel workbook with all tabs.

    Args:
        output_dir: Output directory
        result: ReconciliationResult from reconciler
        assumptions: EstimatorAssumptions used

    Returns:
        Path to generated Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV-only export")
        return _export_fallback_csvs(output_dir, result, assumptions)

    estimator_dir = output_dir / "estimator"
    estimator_dir.mkdir(parents=True, exist_ok=True)
    excel_path = estimator_dir / "bid_ready_boq.xlsx"

    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    needs_review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    measured_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Tab 1: Final BOQ
    ws1 = wb.active
    ws1.title = "Final BOQ"

    headers1 = ["Item ID", "Package", "Description", "Room", "Unit",
                "Final Qty", "Source", "Confidence", "Needs Review"]
    _write_sheet_header(ws1, headers1, header_font, header_fill, thin_border)

    for i, item in enumerate(result.estimator_view, start=2):
        row = [
            item.get("item_id", ""),
            item.get("package", ""),
            item.get("description", ""),
            item.get("room", ""),
            item.get("unit", ""),
            item.get("final_qty", ""),
            item.get("source", ""),
            item.get("confidence", ""),
            item.get("needs_review", ""),
        ]
        for j, val in enumerate(row, start=1):
            cell = ws1.cell(row=i, column=j, value=val)
            cell.border = thin_border

            # Highlight based on needs_review
            if item.get("needs_review") == "YES":
                cell.fill = needs_review_fill
            elif item.get("source") == "measured":
                if j == 7:  # Source column
                    cell.fill = measured_fill

    _autofit_columns(ws1)

    # Tab 2: Measured vs Inferred
    ws2 = wb.create_sheet("Measured vs Inferred")

    headers2 = ["Item ID", "Package", "Description", "Room", "Unit",
                "Measured Qty", "Inferred Qty", "Final Qty", "Source", "Method"]
    _write_sheet_header(ws2, headers2, header_font, header_fill, thin_border)

    for i, item in enumerate(result.estimator_view, start=2):
        row = [
            item.get("item_id", ""),
            item.get("package", ""),
            item.get("description", ""),
            item.get("room", ""),
            item.get("unit", ""),
            item.get("measured_qty", ""),
            item.get("inferred_qty", ""),
            item.get("final_qty", ""),
            item.get("source", ""),
            item.get("method", ""),
        ]
        for j, val in enumerate(row, start=1):
            cell = ws2.cell(row=i, column=j, value=val)
            cell.border = thin_border

    _autofit_columns(ws2)

    # Tab 3: Missing Scope
    ws3 = wb.create_sheet("Missing Scope")

    headers3 = ["Item Type", "Description", "Unit", "Est. Qty",
                "Room ID", "Room Label", "Rule", "Priority"]
    _write_sheet_header(ws3, headers3, header_font, header_fill, thin_border)

    for i, item in enumerate(result.missing_scope, start=2):
        row = [
            item.get("item_type", ""),
            item.get("description", ""),
            item.get("unit", ""),
            item.get("estimated_qty", ""),
            item.get("room_id", ""),
            item.get("room_label", ""),
            item.get("rule", ""),
            item.get("priority", ""),
        ]
        for j, val in enumerate(row, start=1):
            cell = ws3.cell(row=i, column=j, value=val)
            cell.border = thin_border

            if item.get("priority") == "HIGH":
                cell.fill = missing_fill

    _autofit_columns(ws3)

    # Tab 4: Confidence by Page
    ws4 = wb.create_sheet("Confidence by Page")

    headers4 = ["Page", "Total Items", "Measured", "Inferred",
                "Measured %", "Inferred %", "Confidence Score"]
    _write_sheet_header(ws4, headers4, header_font, header_fill, thin_border)

    for i, item in enumerate(result.confidence_by_page, start=2):
        row = [
            item.get("page", ""),
            item.get("total_items", ""),
            item.get("measured_items", ""),
            item.get("inferred_items", ""),
            item.get("measured_pct", ""),
            item.get("inferred_pct", ""),
            item.get("confidence_score", ""),
        ]
        for j, val in enumerate(row, start=1):
            cell = ws4.cell(row=i, column=j, value=val)
            cell.border = thin_border

    _autofit_columns(ws4)

    # Tab 5: Assumptions Used
    ws5 = wb.create_sheet("Assumptions")

    ws5.cell(row=1, column=1, value="Assumption").font = header_font
    ws5.cell(row=1, column=1).fill = header_fill
    ws5.cell(row=1, column=2, value="Value").font = header_font
    ws5.cell(row=1, column=2).fill = header_fill

    assumptions_list = [
        ("Wall Height", f"{assumptions.wall_height_m} m"),
        ("Door Height", f"{assumptions.door_height_m} m"),
        ("Window Height", f"{assumptions.window_height_m} m"),
        ("Plaster Both Sides", "Yes" if assumptions.plaster_both_sides else "No"),
        ("Floor Finish All Rooms", "Yes" if assumptions.floor_finish_all_rooms else "No"),
        ("Skirting Height", f"{assumptions.skirting_height_mm} mm"),
        ("Waterproof Wet Areas", "Yes" if assumptions.waterproof_wet_areas else "No"),
        ("Frame All Openings", "Yes" if assumptions.frame_all_openings else "No"),
    ]

    for i, (name, val) in enumerate(assumptions_list, start=2):
        ws5.cell(row=i, column=1, value=name).border = thin_border
        ws5.cell(row=i, column=2, value=val).border = thin_border

    _autofit_columns(ws5)

    # Save workbook
    wb.save(excel_path)
    logger.info(f"Exported bid-ready Excel to {excel_path}")

    return excel_path


def _write_sheet_header(ws, headers, font, fill, border):
    """Write styled header row."""
    for j, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=header)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')


def _autofit_columns(ws):
    """Auto-fit column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _export_fallback_csvs(
    output_dir: Path,
    result: ReconciliationResult,
    assumptions: EstimatorAssumptions,
) -> Path:
    """Fallback to CSV export if openpyxl not available."""
    estimator_dir = output_dir / "estimator"
    estimator_dir.mkdir(parents=True, exist_ok=True)

    # Already written by reconciler, just return the dir
    return estimator_dir


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def run_estimator_reconciliation(
    output_dir: Path,
    assumptions: Optional[EstimatorAssumptions] = None,
    export_excel: bool = True,
    project_id: str = "",
    apply_overrides: bool = False,
) -> ReconciliationResult:
    """
    Run full estimator reconciliation pipeline.

    Args:
        output_dir: Project output directory
        assumptions: Optional estimator assumptions
        export_excel: Whether to export Excel workbook
        project_id: Project identifier
        apply_overrides: Whether to apply overrides from YAML/Excel

    Returns:
        ReconciliationResult with all data
    """
    # Load estimator inputs from YAML if available
    inputs = None
    try:
        from src.estimator.inputs import load_estimator_inputs
        inputs = load_estimator_inputs(output_dir, project_id)

        # Override assumptions from YAML
        if inputs.loaded_from_file and inputs.assumptions:
            assumptions = EstimatorAssumptions(
                wall_height_m=inputs.assumptions.wall_height_m,
                door_height_m=inputs.assumptions.door_height_m,
                window_height_m=inputs.assumptions.window_height_m,
                plaster_both_sides=inputs.assumptions.plaster_both_sides,
                floor_finish_all_rooms=inputs.assumptions.floor_finish_all_rooms,
                skirting_height_mm=inputs.assumptions.skirting_height_mm,
                waterproof_wet_areas=inputs.assumptions.waterproof_wet_areas,
                frame_all_openings=inputs.assumptions.frame_all_openings,
            )
    except ImportError:
        pass

    # Load Excel overrides if applying
    excel_overrides = []
    if apply_overrides:
        try:
            from src.estimator.excel_interface import read_overrides_from_excel
            excel_path = output_dir / "estimator" / "bid_ready_boq.xlsx"
            excel_overrides = read_overrides_from_excel(excel_path)
            logger.info(f"Loaded {len(excel_overrides)} overrides from Excel")
        except ImportError:
            pass

    # Run reconciliation
    reconciler = EstimatorReconciler(output_dir, assumptions)
    result = reconciler.run()

    # Apply Excel overrides to result
    if excel_overrides:
        result = _apply_excel_overrides(result, excel_overrides)

    if export_excel:
        try:
            # Use new enhanced Excel interface
            from src.estimator.excel_interface import create_estimator_workbook

            assumptions_dict = {
                "wall_height_m": assumptions.wall_height_m if assumptions else 3.0,
                "door_height_m": assumptions.door_height_m if assumptions else 2.1,
                "plaster_both_sides": assumptions.plaster_both_sides if assumptions else True,
                "floor_finish_all_rooms": assumptions.floor_finish_all_rooms if assumptions else True,
            }

            create_estimator_workbook(
                output_dir=output_dir,
                estimator_view=result.estimator_view,
                missing_scope=result.missing_scope,
                confidence_by_page=result.confidence_by_page,
                assumptions=assumptions_dict,
                inputs=inputs,
            )
        except Exception as e:
            logger.warning(f"Excel export failed: {e}, falling back to legacy export")
            try:
                export_bid_ready_excel(output_dir, result, assumptions or EstimatorAssumptions())
            except Exception as e2:
                logger.warning(f"Legacy Excel export also failed: {e2}")

    # Generate bid gate report
    try:
        from src.estimator.bid_gate import assess_bid_gate, generate_bid_gate_report

        bid_result = assess_bid_gate(
            result.estimator_view,
            result.missing_scope,
            result.missing_scope_rfis,
        )
        generate_bid_gate_report(output_dir, bid_result, project_id)
        logger.info(f"Bid gate recommendation: {bid_result.recommendation.value} (score: {bid_result.score})")
    except Exception as e:
        logger.warning(f"Bid gate report generation failed: {e}")

    return result


def _apply_excel_overrides(
    result: ReconciliationResult,
    overrides: List,
) -> ReconciliationResult:
    """Apply overrides from Excel to reconciliation result."""
    # Create lookup by item_id
    override_lookup = {o.item_id: o for o in overrides}

    # Update estimator view
    for item in result.estimator_view:
        item_id = item.get("item_id", "")
        if item_id in override_lookup:
            override = override_lookup[item_id]
            if override.override_qty is not None:
                item["override_qty"] = override.override_qty
                item["final_qty"] = override.final_qty
                item["source"] = "override"
            if override.estimator_notes:
                item["estimator_notes"] = override.estimator_notes
            if override.approved:
                item["approved"] = "YES"

    return result
