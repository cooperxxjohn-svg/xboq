"""
Estimator Excel Interface - Editable Workbook

Creates an interactive Excel workbook that:
1. Color codes measured/inferred/missing items
2. Provides override columns for estimator input
3. Can be re-read to apply overrides on re-run
4. Tracks changes and approval status

This is the estimator's primary interface to XBOQ data.
"""

import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)


# =============================================================================
# COLOR SCHEME
# =============================================================================

# Color palette (RGB hex without #)
COLORS = {
    # Source colors
    "measured": "C6EFCE",       # Light green - reliable
    "inferred": "FFEB9C",       # Light yellow - needs review
    "missing": "FFC7CE",        # Light red - missing scope
    "override": "B4C7E7",       # Light blue - estimator override
    "excluded": "D9D9D9",       # Light gray - excluded

    # Header colors
    "header_main": "4472C4",    # Blue
    "header_input": "70AD47",   # Green - editable columns
    "header_calc": "ED7D31",    # Orange - calculated

    # Status colors
    "approved": "92D050",       # Bright green
    "pending": "FFC000",        # Yellow
    "rejected": "FF0000",       # Red
}


@dataclass
class ExcelOverride:
    """Override loaded from edited Excel."""
    item_id: str
    original_qty: Optional[float]
    override_qty: Optional[float]
    final_qty: float
    estimator_notes: str
    approved: bool


def create_estimator_workbook(
    output_dir: Path,
    estimator_view: List[Dict],
    missing_scope: List[Dict],
    confidence_by_page: List[Dict],
    assumptions: Dict,
    inputs: Optional[Any] = None,  # EstimatorInputs
) -> Path:
    """
    Create interactive estimator workbook.

    Args:
        output_dir: Output directory
        estimator_view: Merged BOQ data
        missing_scope: Missing scope items
        confidence_by_page: Page confidence metrics
        assumptions: Assumptions used
        inputs: Optional EstimatorInputs for overrides

    Returns:
        Path to created Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side,
            Protection, NamedStyle
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.comments import Comment
    except ImportError:
        logger.warning("openpyxl not installed, cannot create Excel workbook")
        return _create_fallback_csv(output_dir, estimator_view, missing_scope)

    estimator_dir = output_dir / "estimator"
    estimator_dir.mkdir(parents=True, exist_ok=True)
    excel_path = estimator_dir / "bid_ready_boq.xlsx"

    wb = openpyxl.Workbook()

    # Define styles
    styles = _create_styles()

    # =========================================================================
    # TAB 1: ESTIMATOR BOQ (Main editable interface)
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Estimator BOQ"

    # Headers with input columns
    headers = [
        ("A", "Item ID", "header_main", 12),
        ("B", "Package", "header_main", 15),
        ("C", "Description", "header_main", 45),
        ("D", "Room/Location", "header_main", 25),
        ("E", "Unit", "header_main", 8),
        ("F", "Measured Qty", "header_main", 12),
        ("G", "Inferred Qty", "header_main", 12),
        ("H", "System Qty", "header_calc", 12),
        ("I", "Override Qty", "header_input", 12),  # EDITABLE
        ("J", "Final Qty", "header_calc", 12),
        ("K", "Rate (₹)", "header_input", 12),  # EDITABLE
        ("L", "Amount (₹)", "header_calc", 14),
        ("M", "Source", "header_main", 10),
        ("N", "Confidence", "header_main", 10),
        ("O", "Needs Review", "header_main", 12),
        ("P", "Estimator Notes", "header_input", 30),  # EDITABLE
        ("Q", "Approved", "header_input", 10),  # EDITABLE
    ]

    # Write headers
    for col, (col_letter, header, style, width) in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color=COLORS[style],
            end_color=COLORS[style],
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws1.column_dimensions[col_letter].width = width

    # Add header row comment explaining editable columns
    ws1["I1"].comment = Comment(
        "EDITABLE: Enter override quantity here.\n"
        "Leave blank to use system quantity.",
        "XBOQ Estimator"
    )
    ws1["K1"].comment = Comment(
        "EDITABLE: Enter rate per unit.\n"
        "Amount will be calculated automatically.",
        "XBOQ Estimator"
    )
    ws1["P1"].comment = Comment(
        "EDITABLE: Add your notes here.",
        "XBOQ Estimator"
    )

    # Data validation for Approved column
    approval_dv = DataValidation(
        type="list",
        formula1='"YES,NO,PENDING"',
        allow_blank=True
    )
    approval_dv.error = "Please select YES, NO, or PENDING"
    approval_dv.errorTitle = "Invalid Selection"

    # Write data rows
    for i, item in enumerate(estimator_view, start=2):
        source = item.get("source", "inferred")
        needs_review = item.get("needs_review", "NO") == "YES"

        # Parse quantities
        measured = _parse_float(item.get("measured_qty"))
        inferred = _parse_float(item.get("inferred_qty"))
        system_qty = measured if measured else inferred

        # Apply override from inputs if available
        override_qty = None
        if inputs and hasattr(inputs, 'apply_override'):
            item_id = item.get("item_id", "")
            desc = item.get("description", "")
            if system_qty:
                overridden = inputs.apply_override(item_id, desc, system_qty)
                if overridden != system_qty:
                    override_qty = overridden

        # Calculate final qty
        final_qty = override_qty if override_qty else system_qty

        # Get row fill color based on source
        if source == "measured":
            row_fill = PatternFill(start_color=COLORS["measured"], end_color=COLORS["measured"], fill_type="solid")
        elif needs_review:
            row_fill = PatternFill(start_color=COLORS["inferred"], end_color=COLORS["inferred"], fill_type="solid")
        else:
            row_fill = None

        # Write cells
        row_data = [
            item.get("item_id", ""),
            item.get("package", ""),
            item.get("description", ""),
            item.get("room", ""),
            item.get("unit", ""),
            measured,
            inferred,
            system_qty,
            override_qty,  # Editable - starts blank or with override
            final_qty,
            None,  # Rate - editable
            None,  # Amount - formula
            source,
            item.get("confidence", ""),
            item.get("needs_review", "NO"),
            "",  # Notes - editable
            "PENDING" if needs_review else "",  # Approved
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=col, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Add formula for Final Qty: =IF(I{row}<>"", I{row}, H{row})
        ws1.cell(row=i, column=10).value = f'=IF(I{i}<>"",I{i},H{i})'

        # Add formula for Amount: =J{row}*K{row}
        ws1.cell(row=i, column=12).value = f'=J{i}*K{i}'

        # Mark editable cells with light blue background
        for col in [9, 11, 16, 17]:  # Override Qty, Rate, Notes, Approved
            cell = ws1.cell(row=i, column=col)
            if not cell.fill or cell.fill.start_color.rgb == "00000000":
                cell.fill = PatternFill(
                    start_color=COLORS["override"],
                    end_color=COLORS["override"],
                    fill_type="solid"
                )
            cell.protection = Protection(locked=False)

    # Apply data validation for Approved column
    if len(estimator_view) > 0:
        approval_dv.add(f"Q2:Q{len(estimator_view) + 1}")
        ws1.add_data_validation(approval_dv)

    # Freeze header row
    ws1.freeze_panes = "A2"

    # =========================================================================
    # TAB 2: MISSING SCOPE (Editable for inclusion)
    # =========================================================================
    ws2 = wb.create_sheet("Missing Scope")

    missing_headers = [
        ("A", "Item Type", 15),
        ("B", "Description", 45),
        ("C", "Unit", 8),
        ("D", "Estimated Qty", 12),
        ("E", "Room/Location", 25),
        ("F", "Rule", 15),
        ("G", "Priority", 10),
        ("H", "Include?", 10),  # EDITABLE
        ("I", "Override Qty", 12),  # EDITABLE
        ("J", "Rate (₹)", 12),  # EDITABLE
        ("K", "Amount (₹)", 14),
        ("L", "Notes", 30),  # EDITABLE
    ]

    for col, (col_letter, header, width) in enumerate(missing_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        is_editable = col in [8, 9, 10, 12]
        cell.fill = PatternFill(
            start_color=COLORS["header_input" if is_editable else "header_main"],
            end_color=COLORS["header_input" if is_editable else "header_main"],
            fill_type="solid"
        )
        ws2.column_dimensions[col_letter].width = width

    # Include validation
    include_dv = DataValidation(
        type="list",
        formula1='"YES,NO"',
        allow_blank=True
    )

    for i, item in enumerate(missing_scope, start=2):
        priority = item.get("priority", "MEDIUM")

        # Set fill based on priority
        if priority == "HIGH":
            row_fill = PatternFill(start_color=COLORS["missing"], end_color=COLORS["missing"], fill_type="solid")
        else:
            row_fill = PatternFill(start_color=COLORS["inferred"], end_color=COLORS["inferred"], fill_type="solid")

        row_data = [
            item.get("item_type", ""),
            item.get("description", ""),
            item.get("unit", ""),
            item.get("estimated_qty", ""),
            item.get("room_label", ""),
            item.get("rule", ""),
            priority,
            "YES" if priority == "HIGH" else "",  # Default include for high priority
            None,  # Override qty
            None,  # Rate
            None,  # Amount (formula)
            "",  # Notes
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=value)
            cell.fill = row_fill
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Amount formula: =IF(H{row}="YES", IF(I{row}<>"",I{row},D{row})*J{row}, 0)
        ws2.cell(row=i, column=11).value = f'=IF(H{i}="YES",IF(I{i}<>"",I{i},D{i})*J{i},0)'

    if len(missing_scope) > 0:
        include_dv.add(f"H2:H{len(missing_scope) + 1}")
        ws2.add_data_validation(include_dv)

    ws2.freeze_panes = "A2"

    # =========================================================================
    # TAB 3: SUMMARY DASHBOARD
    # =========================================================================
    ws3 = wb.create_sheet("Summary")

    # Calculate summary stats
    total_items = len(estimator_view)
    measured_items = sum(1 for x in estimator_view if x.get("source") == "measured")
    inferred_items = sum(1 for x in estimator_view if x.get("source") == "inferred")
    needs_review_items = sum(1 for x in estimator_view if x.get("needs_review") == "YES")
    missing_count = len(missing_scope)
    high_priority_missing = sum(1 for x in missing_scope if x.get("priority") == "HIGH")

    measured_pct = (measured_items / total_items * 100) if total_items > 0 else 0
    inferred_pct = (inferred_items / total_items * 100) if total_items > 0 else 0

    # Determine recommendation
    if measured_pct >= 70 and high_priority_missing == 0:
        recommendation = "GO"
        rec_color = COLORS["approved"]
    elif measured_pct >= 40 or high_priority_missing <= 5:
        recommendation = "REVIEW"
        rec_color = COLORS["pending"]
    else:
        recommendation = "NO-GO"
        rec_color = COLORS["rejected"]

    # Write summary
    summary_data = [
        ("ESTIMATOR BOQ SUMMARY", "", ""),
        ("", "", ""),
        ("QUANTITY EXTRACTION", "", ""),
        ("Total Items", total_items, ""),
        ("Measured (from geometry)", measured_items, f"{measured_pct:.1f}%"),
        ("Inferred (assumptions)", inferred_items, f"{inferred_pct:.1f}%"),
        ("Needs Review", needs_review_items, ""),
        ("", "", ""),
        ("MISSING SCOPE", "", ""),
        ("Missing Scope Items", missing_count, ""),
        ("High Priority Missing", high_priority_missing, ""),
        ("", "", ""),
        ("BID RECOMMENDATION", recommendation, ""),
        ("", "", ""),
        ("ASSUMPTIONS USED", "", ""),
        ("Wall Height", f"{assumptions.get('wall_height_m', 3.0)} m", ""),
        ("Door Height", f"{assumptions.get('door_height_m', 2.1)} m", ""),
        ("Plaster Both Sides", "Yes" if assumptions.get('plaster_both_sides', True) else "No", ""),
        ("Floor Finish All Rooms", "Yes" if assumptions.get('floor_finish_all_rooms', True) else "No", ""),
    ]

    for i, (label, value, extra) in enumerate(summary_data, 1):
        ws3.cell(row=i, column=1, value=label)
        ws3.cell(row=i, column=2, value=value)
        ws3.cell(row=i, column=3, value=extra)

        if label in ["ESTIMATOR BOQ SUMMARY", "QUANTITY EXTRACTION", "MISSING SCOPE", "BID RECOMMENDATION", "ASSUMPTIONS USED"]:
            ws3.cell(row=i, column=1).font = Font(bold=True, size=12)

        if label == "BID RECOMMENDATION":
            ws3.cell(row=i, column=2).fill = PatternFill(
                start_color=rec_color,
                end_color=rec_color,
                fill_type="solid"
            )
            ws3.cell(row=i, column=2).font = Font(bold=True, size=14)

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 15

    # =========================================================================
    # TAB 4: CONFIDENCE BY PAGE
    # =========================================================================
    ws4 = wb.create_sheet("Page Confidence")

    page_headers = ["Page", "Total Items", "Measured", "Inferred", "Measured %", "Inferred %", "Score"]
    for col, header in enumerate(page_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS["header_main"], end_color=COLORS["header_main"], fill_type="solid")

    for i, item in enumerate(confidence_by_page, start=2):
        # Handle string values from CSV
        measured_pct = _parse_float(item.get('measured_pct')) or 0
        inferred_pct = _parse_float(item.get('inferred_pct')) or 0
        confidence = _parse_float(item.get('confidence_score')) or 0

        row_data = [
            item.get("page"),
            item.get("total_items"),
            item.get("measured_items"),
            item.get("inferred_items"),
            f"{measured_pct:.1f}%",
            f"{inferred_pct:.1f}%",
            f"{confidence:.1f}",
        ]
        for col, value in enumerate(row_data, 1):
            ws4.cell(row=i, column=col, value=value)

    # =========================================================================
    # TAB 5: INSTRUCTIONS
    # =========================================================================
    ws5 = wb.create_sheet("Instructions")

    instructions = """
XBOQ ESTIMATOR WORKBOOK - INSTRUCTIONS
======================================

This workbook is your interface to the XBOQ extraction results.
You can review, override, and approve quantities here.

COLOR CODING:
- Green cells: Measured from geometry (high confidence)
- Yellow cells: Inferred from assumptions (needs review)
- Red cells: Missing scope items (high priority)
- Blue cells: EDITABLE - Enter your inputs here

EDITABLE COLUMNS:
1. Override Qty - Enter to override system quantity
2. Rate (₹) - Enter unit rate for pricing
3. Estimator Notes - Add your notes
4. Approved - Mark as YES/NO/PENDING
5. Include? (Missing Scope) - Include in estimate

WORKFLOW:
1. Review the "Estimator BOQ" tab
2. Check items marked "Needs Review"
3. Enter overrides where needed
4. Add rates for pricing
5. Mark items as Approved
6. Check "Missing Scope" tab for gaps
7. Include/exclude missing items as needed
8. Review "Summary" for bid recommendation

RE-RUN WITH OVERRIDES:
Save this file and run:
  python run_full_project.py --project_id <id> --apply_overrides

This will read your overrides and re-run reconciliation.

SUPPORT:
Report issues or questions to your XBOQ administrator.
""".strip()

    for i, line in enumerate(instructions.split("\n"), 1):
        ws5.cell(row=i, column=1, value=line)
        if line.endswith(":") or line.startswith("="):
            ws5.cell(row=i, column=1).font = Font(bold=True)

    ws5.column_dimensions["A"].width = 80

    # =========================================================================
    # SAVE WORKBOOK
    # =========================================================================

    # NOTE: Sheet protection disabled - allowing full editing
    # Uncomment to add advisory protection (not security)
    # ws1.protection.sheet = True
    # ws1.protection.enable()

    wb.save(excel_path)
    logger.info(f"Created estimator workbook at {excel_path}")

    return excel_path


def read_overrides_from_excel(excel_path: Path) -> List[ExcelOverride]:
    """
    Read overrides from edited Excel workbook.

    Args:
        excel_path: Path to edited workbook

    Returns:
        List of overrides to apply
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed")
        return []

    if not excel_path.exists():
        logger.info(f"No Excel file at {excel_path}")
        return []

    overrides = []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Read from Estimator BOQ tab
        if "Estimator BOQ" in wb.sheetnames:
            ws = wb["Estimator BOQ"]

            for row in range(2, ws.max_row + 1):
                item_id = ws.cell(row=row, column=1).value
                if not item_id:
                    continue

                original_qty = _parse_float(ws.cell(row=row, column=8).value)  # System Qty
                override_qty = _parse_float(ws.cell(row=row, column=9).value)  # Override Qty
                final_qty = _parse_float(ws.cell(row=row, column=10).value)  # Final Qty
                notes = ws.cell(row=row, column=16).value or ""
                approved = ws.cell(row=row, column=17).value

                # Only include if there's an override or approval
                if override_qty is not None or approved == "YES":
                    overrides.append(ExcelOverride(
                        item_id=str(item_id),
                        original_qty=original_qty,
                        override_qty=override_qty,
                        final_qty=final_qty or override_qty or original_qty or 0,
                        estimator_notes=str(notes),
                        approved=approved == "YES",
                    ))

        wb.close()
        logger.info(f"Read {len(overrides)} overrides from Excel")

    except Exception as e:
        logger.error(f"Error reading Excel overrides: {e}")

    return overrides


def _parse_float(value: Any) -> Optional[float]:
    """Parse value to float."""
    if value is None or value == "" or value == "None":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _create_styles() -> Dict:
    """Create reusable styles."""
    return {}


def _create_fallback_csv(
    output_dir: Path,
    estimator_view: List[Dict],
    missing_scope: List[Dict],
) -> Path:
    """Create fallback CSV if Excel not available."""
    estimator_dir = output_dir / "estimator"
    csv_path = estimator_dir / "estimator_boq_editable.csv"

    fieldnames = [
        "item_id", "package", "description", "room", "unit",
        "measured_qty", "inferred_qty", "system_qty",
        "override_qty", "final_qty", "rate", "amount",
        "source", "confidence", "needs_review",
        "estimator_notes", "approved"
    ]

    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for item in estimator_view:
            measured = _parse_float(item.get("measured_qty"))
            inferred = _parse_float(item.get("inferred_qty"))
            system_qty = measured if measured else inferred

            writer.writerow({
                "item_id": item.get("item_id", ""),
                "package": item.get("package", ""),
                "description": item.get("description", ""),
                "room": item.get("room", ""),
                "unit": item.get("unit", ""),
                "measured_qty": measured or "",
                "inferred_qty": inferred or "",
                "system_qty": system_qty or "",
                "override_qty": "",
                "final_qty": system_qty or "",
                "rate": "",
                "amount": "",
                "source": item.get("source", ""),
                "confidence": item.get("confidence", ""),
                "needs_review": item.get("needs_review", ""),
                "estimator_notes": "",
                "approved": "",
            })

    return csv_path
